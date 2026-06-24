import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import gspread
import re
import math
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from oauth2client.service_account import ServiceAccountCredentials

# =================================================================
# ⚡ MOTORES DE CACHÉ Y VELOCIDAD DE DATOS (Aislamiento en RAM)
# =================================================================

def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except:
        return None

def obtener_cliente_gspread_viejo():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_credentials" in st.secrets:
            creds_dict = dict(st.secrets["gcp_credentials"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except:
        return None

@st.cache_data(show_spinner=False)
def cargar_fuentes_maestras_bi(_descargar_matriz_rapida=None):
    gc_nuevo = obtener_cliente_gspread_unificado()
    
    try:
        boveda_act = gc_nuevo.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        datos_brutos_act = boveda_act.worksheet("TABLA 1").get_all_values()
    except:
        datos_brutos_act = []
    
    if len(datos_brutos_act) > 5:
        columnas_t1 = ["OS", "BLOQUE", "FINCA", "SECTOR", "AREA_BRUTA", "AREA_FUMIG", "COCTEL", "FECHA", "DIA", "SEMANA", "H_TOTAL", "GLN_HA", "VOL_TOTAL", "REND_HR", "REND_MIN", "PILOTO", "HK", "MODELO", "COSTO_AVION", "COSTO_HA", "DOMINICAL_HA", "COSTO_FINCA", "VALOR_FACTURAR", "PISTA", "INC_2026", "LIMITE", "ALERTA", "VAR_PCT", "COSTO_TOTAL", "PAGO_AVION"]
        filas_limpias = [r + [""]*(len(columnas_t1) - len(r)) for r in datos_brutos_act[5:]]
        df_vivos = pd.DataFrame([r[:len(columnas_t1)] for r in filas_limpias], columns=columnas_t1)
        
        df_vivos.rename(columns={'AREA_FUMIG': 'AREA_MAESTRA', 'COSTO_HA': 'AVION_MAESTRO', 'DOMINICAL_HA': 'DOMINIC_MAESTRO', 'FINCA': 'FINCA_MAESTRA', 'FECHA': 'FECHA_MAESTRA', 'OS': 'OS_MAESTRA', 'COCTEL': 'COCTEL_MAESTRO'}, inplace=True)
        df_vivos['ORIGEN_BI'] = 'ACTUAL'
    else:
        df_vivos = pd.DataFrame()

    datos_brutos_hist = []
    try:
        boveda_hist = gc_nuevo.open_by_url("https://docs.google.com/spreadsheets/d/16OZdiWwW7nLHyZBEnhiKlDTDttR7Tjhn37O9zm6wJOk/edit")
        datos_brutos_hist = boveda_hist.worksheet("Datos").get_all_values()
    except Exception as error_nuevo:
        try:
            gc_viejo = obtener_cliente_gspread_viejo()
            boveda_hist = gc_viejo.open_by_url("https://docs.google.com/spreadsheets/d/16OZdiWwW7nLHyZBEnhiKlDTDttR7Tjhn37O9zm6wJOk/edit")
            datos_brutos_hist = boveda_hist.worksheet("Datos").get_all_values()
        except Exception as error_viejo:
            st.error(f"🚨 **RADAR DE SEGURIDAD BI:** El archivo histórico rechazó ambas llaves.")
    
    if len(datos_brutos_hist) > 0:
        df_historico = pd.DataFrame(datos_brutos_hist[1:], columns=datos_brutos_hist[0])
        df_historico = estandarizar_base(limpiar_encabezados(df_historico))
        df_historico['ORIGEN_BI'] = 'HISTORICO'
    else:
        df_historico = pd.DataFrame()

    return df_vivos, df_historico

@st.cache_data(show_spinner=False)
def cargar_boveda_recetas_y_precios():
    try:
        gc = obtener_cliente_gspread_unificado()
        if not gc: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        boveda_recetas = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        
        # 💥 CORRECCIÓN TÁCTICA: Ahora busca "FINCA" o "PRODUCTO" en lugar de "COCTEL"
        data_mez = boveda_recetas.worksheet("DD_Mesclas").get_all_values()
        idx_mez = 0
        for i in range(min(5, len(data_mez))):
            fila_str = [str(c).upper() for c in data_mez[i]]
            if any('FINCA' in c for c in fila_str) or any('PRODUCTO' in c for c in fila_str):
                idx_mez = i; break
                
        df_mezclas = pd.DataFrame(data_mez[idx_mez+1:], columns=data_mez[idx_mez]) if len(data_mez) > idx_mez else pd.DataFrame()
        if not df_mezclas.empty:
            df_mezclas['COCTEL_CLEAN'] = df_mezclas.iloc[:,0].astype(str).str.upper().str.replace(" ", "")
            
        df_conf = pd.DataFrame(boveda_recetas.worksheet("Configuración").get_all_values()[1:], columns=boveda_recetas.worksheet("Configuración").get_all_values()[0])
        df_dicc = pd.DataFrame(boveda_recetas.worksheet("DICCIONARIO_SIGLAS").get_all_values()[1:], columns=boveda_recetas.worksheet("DICCIONARIO_SIGLAS").get_all_values()[0])
        df_t2 = pd.DataFrame(boveda_recetas.worksheet("TABLA 2").get_all_values()[1:], columns=boveda_recetas.worksheet("TABLA 2").get_all_values()[0])

        url_precios = "https://docs.google.com/spreadsheets/d/1qZ4av-DH2oCJdgllBX27gdA2jEhT9bt2yv_sboORfSg/edit"
        sh_precios = gc.open_by_url(url_precios)
        
        precios_consolidados = []
        for ws in sh_precios.worksheets():
            datos_hoja = ws.get_all_values()
            if not datos_hoja: continue
            idx_header, col_anio, col_prod = -1, -1, -1
            for i in range(min(10, len(datos_hoja))):
                fila_upper = [str(x).upper().strip() for x in datos_hoja[i]]
                if 'AÑO' in fila_upper and 'PRODUCTO' in fila_upper:
                    idx_header = i; col_anio = fila_upper.index('AÑO'); col_prod = fila_upper.index('PRODUCTO')
                    break
            if idx_header != -1:
                for row in datos_hoja[idx_header+1:]:
                    if len(row) > max(col_anio, col_prod):
                        anio_str, str_prod = str(row[col_anio]).strip().upper(), str(row[col_prod]).strip().upper()
                        if anio_str and str_prod:
                            col_inicio = max(col_anio, col_prod) + 1
                            vals = [float(str(v).strip().replace(',', '.')) for v in row[col_inicio:] if str(v).strip().replace(',', '.').replace('-','').replace('.','').isdigit()]
                            prom = sum(vals)/len(vals) if vals else 0.0
                            precios_consolidados.append({'AÑO': anio_str, 'PRODUCTO': str_prod, 'PRECIO_PROM': prom})

        df_precios = pd.DataFrame(precios_consolidados)
        return df_mezclas, df_conf, df_dicc, df_precios, df_t2
    except:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
# --- 🧪 APARTADO DE BARREDORAS Y AUXILIARES GLOBALES ---
def limpiar_encabezados(df):
    df.columns = [str(col).upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').strip() for col in df.columns]
    df = df.loc[:, ~df.columns.duplicated(keep='first')]
    if "" in df.columns: df = df.drop(columns=[""])
    return df
    
def estandarizar_base(df):
    renombres = {}
    for col in df.columns:
        col_u = str(col).upper().replace('\n', ' ').strip()
        if 'FINCA' in col_u and 'COSTO' in col_u: continue
            
        if 'FACTURAR' in col_u and 'PRODUCTOR' in col_u: renombres[col] = 'COSTO_MAESTRO'
        elif 'FUMIG' in col_u and 'AREA' in col_u: renombres[col] = 'AREA_MAESTRA'
        elif 'AVION' in col_u and '/HA' in col_u: renombres[col] = 'AVION_MAESTRO'
        elif 'DOMINIC' in col_u and '/HA' in col_u: renombres[col] = 'DOMINIC_MAESTRO'
        elif not ('FINCA_MAESTRA' in renombres.values()) and (col_u == 'FINCA' or col_u == 'PROPIEDAD'): renombres[col] = 'FINCA_MAESTRA'
        elif not ('FECHA_MAESTRA' in renombres.values()) and col_u == 'FECHA': renombres[col] = 'FECHA_MAESTRA'
        elif not ('OS_MAESTRA' in renombres.values()) and ("Nº ORDEN" in col_u or "ORDEN DE" in col_u or "OS" == col_u): renombres[col] = 'OS_MAESTRA'
        elif not ('COCTEL_MAESTRO' in renombres.values()) and col_u in ['COCTEL', 'CÓCTEL']: renombres[col] = 'COCTEL_MAESTRO'
            
    df.rename(columns=renombres, inplace=True)
    return df

def limpiar_area(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        if not v: return 0.0
        v = v.replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except: return 0.0

def limpiar_dinero(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        if not v: return 0.0
        v = v.replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        num = float(v) if v else 0.0
        if 5 < num < 2000: num = num * 1000
        return num
    except: return 0.0

# =================================================================
# 🔑 CEREBRO QUÍMICO HÍBRIDO (Lectura de Excel + Reglas Dinámicas)
# =================================================================
def extraer_receta_de_sigla_bi(coctel_sel, finca_sel, df_mezclas, df_dicc, df_t2):
    coctel_u = str(coctel_sel).upper().strip()
    coctel_norm = coctel_u.replace("+", " ").replace("-", " ")
    partes_coctel = coctel_norm.split()
    
    base_coctel = partes_coctel[0] if len(partes_coctel) > 0 else ""
    aditivos = partes_coctel[1:] if len(partes_coctel) > 1 else []
    
    dict_prods = {}
    es_organico = False
    
    try:
        if not df_t2.empty:
            match_f = df_t2[df_t2.iloc[:, 0].astype(str).str.upper().str.strip() == finca_sel.upper().strip()]
            if not match_f.empty and "ORGANIC" in str(match_f.iloc[0, 5]).upper(): 
                es_organico = True
    except: pass

    if es_organico and not base_coctel.endswith('O'):
        base_buscar = f"{base_coctel}O"
    else:
        base_buscar = base_coctel

    # 1. LECTURA CIEGA DEL ESQUELETO (DD_Mesclas)
    if not df_mezclas.empty:
        rb = df_mezclas[df_mezclas.iloc[:, 0].astype(str).str.upper().str.strip() == base_buscar]
        if rb.empty and es_organico: 
            rb = df_mezclas[df_mezclas.iloc[:, 0].astype(str).str.upper().str.strip() == base_coctel]
            
        for _, r in rb.iterrows():
            p = str(r.iloc[1]).strip().upper()
            d = limpiar_area(r.iloc[2])
            if d > 0 and p not in ['NAN', 'NONE', '']: 
                dict_prods[p] = d

    # 2. LECTURA CIEGA DE LOS FERTILIZANTES / ADITIVOS (DICCIONARIO_SIGLAS)
    if not df_dicc.empty and aditivos:
        for ad in aditivos:
            m_s = df_dicc[df_dicc['SIGLA'].astype(str).str.upper().str.strip() == ad]
            if not m_s.empty:
                p_ad = str(m_s.iloc[0]['PRODUCTO']).strip().upper()
                d_ad = limpiar_area(m_s.iloc[0]['DOSIS'])
                if d_ad > 0 and p_ad not in ['NAN', 'NONE', '']:
                    dict_prods[p_ad] = dict_prods.get(p_ad, 0.0) + d_ad

    # 3. REGLAS DE ORO DINÁMICAS
    for p in list(dict_prods.keys()):
        if "ACONDICIONADOR" in p:
            # EL NM NO ALTERA EL ACONDICIONADOR. Solo ZN, BT, ZT, ZITRON.
            if any(x in coctel_u for x in ["ZN", "BT", "ZT", "ZITRON"]):
                dict_prods[p] = 0.06
                
        elif "IMBIOSIL" in p.replace(" ", ""):
            if base_coctel.startswith("IN") or "IMBIOSIL" in base_coctel:
                dict_prods[p] = 1.5
                
        if es_organico and "ADHERENTE" in p:
            del dict_prods[p]

    if es_organico and not any("SPRAYFIX" in k for k in dict_prods.keys()):
        dict_prods["SPRAYFIX"] = 0.2

    return dict_prods

def calcular_frecuencia_por_finca(df_area, finca_seleccionada):
    if df_area.empty or 'FECHA_DT' not in df_area.columns: return 0, 0.0
        
    if finca_seleccionada != "TODAS":
        fechas = sorted(df_area['FECHA_DT'].dt.date.unique())
        if not fechas: return 0, 0.0
        ciclos = 1
        inicios_ciclo = [fechas[0]]
        for i in range(1, len(fechas)):
            if (fechas[i] - fechas[i-1]).days > 5:
                ciclos += 1
                inicios_ciclo.append(fechas[i])
        avg_int = sum([(inicios_ciclo[j] - inicios_ciclo[j-1]).days for j in range(1, ciclos)]) / (ciclos - 1) if ciclos > 1 else 0.0
        return ciclos, avg_int

    fincas_presentes = df_area['FINCA_MAESTRA'].unique()
    total_ciclos_todas = 0
    total_suma_dias = 0
    total_intervalos_contados = 0
    fincas_validas = 0
    
    for f in fincas_presentes:
        df_sub = df_area[df_area['FINCA_MAESTRA'] == f]
        fechas_f = sorted(df_sub['FECHA_DT'].dt.date.unique())
        if not fechas_f: continue
        
        c_f = 1
        inicios_c_f = [fechas_f[0]]
        for i in range(1, len(fechas_f)):
            if (fechas_f[i] - fechas_f[i-1]).days > 5:
                c_f += 1
                inicios_c_f.append(fechas_f[i])
                
        total_ciclos_todas += c_f
        fincas_validas += 1
        
        if c_f > 1:
            suma_dias_finca = sum([(inicios_c_f[j] - inicios_c_f[j-1]).days for j in range(1, c_f)])
            total_suma_dias += suma_dias_finca
            total_intervalos_contados += (c_f - 1)
            
    promedio_ciclos = int(round(total_ciclos_todas / fincas_validas)) if fincas_validas > 0 else 0
    promedio_intervalo = total_suma_dias / total_intervalos_contados if total_intervalos_contados > 0 else 0.0
            
    return promedio_ciclos, promedio_intervalo

# --- 📡 NÚCLEO OPERATIVO DEL DASHBOARD ESTRATÉGICO ---
def ejecutar(descargar_matriz_rapida, procesar_fecha_pesada, extraer_numero):
    st.header("", anchor="inicio_modulo")

    st.markdown("""
    <style>
    .titulo-principal { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; }
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] { border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; }
    
    .hud-bi {
        background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%);
        border-left: 5px solid #d4af37; padding: 15px; border-radius: 8px; color: white;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.15); margin-bottom: 25px;
    }
    .hud-bi-title { font-size: 11px; font-weight: bold; color: #d4af37; text-transform: uppercase; margin:0; letter-spacing: 1px; }
    .hud-bi-value { font-size: 22px; font-family: 'Arial Black'; margin: 5px 0 0 0; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-principal'>📊 Centro de Inteligencia Estratégica BI</h1>", unsafe_allow_html=True)

    try:
        df_vivos, df_historico = cargar_fuentes_maestras_bi(descargar_matriz_rapida)

        if df_vivos.empty and df_historico.empty:
            st.warning("⚠️ Los sistemas de almacenamiento están vacíos. (Ambas conexiones fallaron).")
            return

        super_base_bi = pd.concat([df_historico, df_vivos], ignore_index=True)

        if 'FINCA_MAESTRA' not in super_base_bi.columns or 'FECHA_MAESTRA' not in super_base_bi.columns:
            st.error("🚨 Columnas críticas estructurales ausentes en la Bóveda.")
            return

        for col_req in ['COSTO_MAESTRO', 'AVION_MAESTRO', 'DOMINIC_MAESTRO', 'AREA_MAESTRA', 'OS_MAESTRA']:
            if col_req not in super_base_bi.columns: super_base_bi[col_req] = 0.0

        super_base_bi['FINCA_MAESTRA'] = super_base_bi['FINCA_MAESTRA'].astype(str).str.strip().str.upper()
        super_base_bi['FECHA_DT'] = super_base_bi['FECHA_MAESTRA'].apply(procesar_fecha_pesada)
        super_base_bi = super_base_bi.dropna(subset=['FECHA_DT'])
        
        super_base_bi['AÑO'] = super_base_bi['FECHA_DT'].dt.year.astype(int)
        super_base_bi['MES'] = super_base_bi['FECHA_DT'].dt.month.astype(int)
        super_base_bi['TRIMESTRE'] = super_base_bi['FECHA_DT'].dt.quarter.astype(int)
        
        def calcular_costo_real(row):
            if row.get('ORIGEN_BI') == 'ACTUAL':
                return limpiar_dinero(row.get('VALOR_FACTURAR', 0))
            else:
                return limpiar_dinero(row.get('COSTO_MAESTRO', 0))

        super_base_bi['COSTO_NUM'] = super_base_bi.apply(calcular_costo_real, axis=1)
        super_base_bi['AREA_NUM'] = super_base_bi['AREA_MAESTRA'].apply(limpiar_area)
        super_base_bi['AVION_NUM'] = super_base_bi['AVION_MAESTRO'].apply(limpiar_dinero) + super_base_bi['DOMINIC_MAESTRO'].apply(limpiar_dinero)

        def purgar_avion(row):
            avion = row['AVION_NUM']
            area = row['AREA_NUM']
            if avion > 90000 and area > 0:
                avion = avion / area
                if avion > 90000: avion = 55000 
            return avion
            
        super_base_bi['AVION_NUM'] = super_base_bi.apply(purgar_avion, axis=1)

        total_ha_historicas = super_base_bi['AREA_NUM'].sum()
        costo_medio_historico = super_base_bi[super_base_bi['COSTO_NUM'] > 0]['COSTO_NUM'].mean()
        total_ordenes_auditadas = super_base_bi['OS_MAESTRA'].nunique()

        hb1, hb2, hb3 = st.columns(3)
        with hb1: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>Área Histórica Cubierta</p><p class='hud-bi-value'>🚜 {total_ha_historicas:,.1f} Ha</p></div>", unsafe_allow_html=True)
        with hb2: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>Costo Medio Consolidado</p><p class='hud-bi-value'>💰 $ {costo_medio_historico:,.0f}</p></div>", unsafe_allow_html=True)
        with hb3: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>Órdenes de Servicio Auditadas</p><p class='hud-bi-value'>🛰️ {total_ordenes_auditadas:,} OS</p></div>", unsafe_allow_html=True)

        fincas_disp = ["TODAS"] + sorted(super_base_bi['FINCA_MAESTRA'].dropna().unique().tolist())
        años_disp = sorted(super_base_bi['AÑO'].unique().tolist(), reverse=True)
        
        col_modelo = 'MODELO' if 'MODELO' in super_base_bi.columns else None
        modelos_disp = ["TODOS"] + sorted(super_base_bi[col_modelo].unique().tolist()) if col_modelo else ["TODOS"]
        
        f1, f2 = st.columns(2)
        finca_sel = f1.selectbox("📍 Objetivo Geográfico (Finca)", fincas_disp)
        modelo_sel = f2.selectbox("🚁 Escuadrón (Modelo/Tipo)", modelos_disp)
        
        # =====================================================================
        # 🎛️ PANEL TÁCTICO DE TIEMPO Y ESPACIO (Con Rango Personalizado)
        # =====================================================================
        t1, t2, t3, t4 = st.columns(4)
        año_base = t1.selectbox("📅 Año Base (Referencia)", años_disp, index=(1 if len(años_disp) > 1 else 0))
        año_comp = t2.selectbox("📆 Año Actual (Evaluar)", años_disp, index=0)
        tipo_periodo = t3.selectbox("⏱️ Lupa Temporal", ["AÑO COMPLETO", "POR TRIMESTRE", "POR MES", "RANGO PERSONALIZADO"])
        meses_dict = {1:'Ene', 2:'Feb', 3:'Mar', 4:'Abr', 5:'May', 6:'Jun', 7:'Jul', 8:'Ago', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dic'}
        
        if tipo_periodo == "POR TRIMESTRE":
            periodo_sel = t4.selectbox("📊 Seleccione Trimestre", [1, 2, 3, 4], format_func=lambda x: f"Q{x}")
            etiq_periodo = f"Q{periodo_sel}"
        elif tipo_periodo == "POR MES":
            periodo_sel = t4.selectbox("📅 Seleccione Mes", list(meses_dict.keys()), format_func=lambda x: meses_dict[x])
            etiq_periodo = meses_dict[periodo_sel]
        elif tipo_periodo == "RANGO PERSONALIZADO":
            c_fecha1, c_fecha2 = t4.columns(2)
            fecha_inicial_libre = c_fecha1.date_input("📅 Desde:", value=datetime.now().date(), key="bi_f_ini_libre")
            fecha_final_libre = c_fecha2.date_input("📅 Hasta:", value=datetime.now().date(), key="bi_f_fin_libre")
            etiq_periodo = f"{fecha_inicial_libre.strftime('%d/%m')} al {fecha_final_libre.strftime('%d/%m')}"
        else:
            t4.markdown("<br><span style='color:gray;'>Visión Anual Activada</span>", unsafe_allow_html=True)
            periodo_sel, etiq_periodo = "TODOS", "Total"

        df_finca = super_base_bi.copy()
        if finca_sel != "TODAS": df_finca = df_finca[df_finca['FINCA_MAESTRA'] == finca_sel]
        if col_modelo and modelo_sel != "TODOS": df_finca = df_finca[df_finca[col_modelo] == modelo_sel].copy()

        # =====================================================================
        # 🚀 MOTOR DE EXTRACCIÓN ESPEJO (PROYECCIÓN YoY DINÁMICA)
        # =====================================================================
        if tipo_periodo == "RANGO PERSONALIZADO":
            mes_inicio, dia_inicio = fecha_inicial_libre.month, fecha_inicial_libre.day
            mes_fin, dia_fin = fecha_final_libre.month, fecha_final_libre.day
            
            f_ini_b = datetime(int(año_comp), mes_inicio, dia_inicio)
            f_fin_b = datetime(int(año_comp), mes_fin, dia_fin)
            df_periodo_b = df_finca[(df_finca['FECHA_DT'] >= f_ini_b) & (df_finca['FECHA_DT'] <= f_fin_b)].copy()
            
            try:
                f_ini_a = datetime(int(año_base), mes_inicio, dia_inicio)
                f_fin_a = datetime(int(año_base), mes_fin, dia_fin)
            except ValueError:
                f_ini_a = datetime(int(año_base), mes_inicio, 28)
                f_fin_a = datetime(int(año_base), mes_fin, 28)
                
            df_periodo_a = df_finca[(df_finca['FECHA_DT'] >= f_ini_a) & (df_finca['FECHA_DT'] <= f_fin_a)].copy()
            
        else:
            df_periodo_a = df_finca[df_finca['AÑO'] == año_base].copy()
            df_periodo_b = df_finca[df_finca['AÑO'] == año_comp].copy()
            
            if tipo_periodo == "POR TRIMESTRE":
                df_periodo_a = df_periodo_a[df_periodo_a['TRIMESTRE'] == periodo_sel]
                df_periodo_b = df_periodo_b[df_periodo_b['TRIMESTRE'] == periodo_sel]
            elif tipo_periodo == "POR MES":
                df_periodo_a = df_periodo_a[df_periodo_a['MES'] == periodo_sel]
                df_periodo_b = df_periodo_b[df_periodo_b['MES'] == periodo_sel]

        df_area_a = df_periodo_a.copy()
        df_area_b = df_periodo_b.copy()

        area_a = df_area_a['AREA_NUM'].sum() if not df_area_a.empty else 0.0
        area_b = df_area_b['AREA_NUM'].sum() if not df_area_b.empty else 0.0
        costo_a = df_periodo_a['COSTO_NUM'].mean() if not df_periodo_a.empty else 0
        costo_b = df_periodo_b['COSTO_NUM'].mean() if not df_periodo_b.empty else 0
        delta_pct = ((costo_b - costo_a) / costo_a * 100) if costo_a > 0 else 0
        
        st.markdown(f"### 📌 Impacto General para {finca_sel} ({etiq_periodo})")
        k1, k2, k3 = st.columns(3)
        k1.metric(label=f"Costo Promedio Ha ({año_base})", value=f"$ {costo_a:,.0f}")
        k2.metric(label=f"Costo Promedio Ha ({año_comp})", value=f"$ {costo_b:,.0f}")
        k3.metric(label="Variación Total (%)", value=f"{delta_pct:+.2f} %", delta=f"{delta_pct:+.2f}%", delta_color="inverse")
        
        st.markdown("#### 🚜 Volumen Operativo (Hectáreas Aplicadas)")
        var_area = ((area_b - area_a) / area_a * 100) if area_a > 0 else 0

        h1, h2, h3 = st.columns(3)
        h1.metric(f"Total Hectáreas ({año_base})", f"{area_a:,.1f} Ha")
        h2.metric(f"Total Hectáreas ({año_comp})", f"{area_b:,.1f} Ha")
        h3.metric("Variación de Área", f"{var_area:+.1f} %" if area_a > 0 else "N/A", delta=f"{var_area:+.1f}%" if area_a > 0 else None)
        
        st.markdown("<br>", unsafe_allow_html=True)
        if delta_pct > 10: st.error(f"⚠️ **ALERTA DE DESVIACIÓN:** El costo operativo presenta un incremento del **{delta_pct:.1f}%**.")
        elif delta_pct < 0: st.success(f"✅ **EFICIENCIA:** Reducción detectada en el costo promedio del periodo.")
        else: st.info(f"⚖️ **ESTABILIDAD:** Los márgenes se mantienen balanceados.")
            
        st.markdown("#### ⏱️ Análisis de Frecuencia: Ciclos Reales e Intervalo Promedio")
        ciclos_a, int_a = calcular_frecuencia_por_finca(df_area_a, finca_sel)
        ciclos_b, int_b = calcular_frecuencia_por_finca(df_area_b, finca_sel)

        c1, c2, c3, c4 = st.columns(4)
        label_ciclo = "Ciclos Prom. / Finca" if finca_sel == "TODAS" else "Ciclos Totales"
        label_int = "Intervalo Prom. Zona" if finca_sel == "TODAS" else "Intervalo Promedio"

        c1.metric(f"{label_ciclo} ({año_base})", f"{ciclos_a} ciclos")
        c2.metric(f"{label_ciclo} ({año_comp})", f"{ciclos_b} ciclos", delta=f"{ciclos_b - ciclos_a} ciclos", delta_color="inverse")
        c3.metric(f"{label_int} ({año_base})", f"{int_a:.1f} días" if int_a > 0 else "N/A")
        c4.metric(f"{label_int} ({año_comp})", f"{int_b:.1f} días" if int_b > 0 else "N/A", delta=f"{int_b - int_a:+.1f} días" if (int_a > 0 and int_b > 0) else None)
        
        st.markdown("---")
        st.markdown("### 🧬 Análisis de Causa Raíz: Atribución de Variaciones")
        
        df_tendencia = pd.concat([df_periodo_a, df_periodo_b])
        if not df_tendencia.empty:
            if tipo_periodo in ["AÑO COMPLETO", "POR TRIMESTRE"]:
                tendencia_agrupa = df_tendencia.groupby(['AÑO', 'MES'])['COSTO_NUM'].mean().reset_index()
                tendencia_agrupa['EJE_X'] = tendencia_agrupa['MES'].map(meses_dict)
                tendencia_agrupa = tendencia_agrupa.sort_values('MES')
                titulo_x = "Meses Operativos"
            else:
                df_tendencia['DIA'] = df_tendencia['FECHA_DT'].dt.day
                tendencia_agrupa = df_tendencia.groupby(['AÑO', 'DIA'])['COSTO_NUM'].mean().reset_index()
                tendencia_agrupa['EJE_X'] = "Día " + tendencia_agrupa['DIA'].astype(str)
                tendencia_agrupa = tendencia_agrupa.sort_values('DIA')
                titulo_x = f"Días Operativos ({etiq_periodo})"
                
            tendencia_agrupa['AÑO'] = tendencia_agrupa['AÑO'].astype(str)
            fig_tendencia = px.line(tendencia_agrupa, x='EJE_X', y='COSTO_NUM', color='AÑO', markers=True, color_discrete_sequence=['#2F75B5', '#27AE60'])
            fig_tendencia.update_layout(yaxis_title="Costo Promedio ($ COP / Ha)", xaxis_title=titulo_x, plot_bgcolor='rgba(0,0,0,0)', hovermode="x unified")
            max_y = tendencia_agrupa['COSTO_NUM'].max() * 1.2
            if not pd.isna(max_y): fig_tendencia.update_yaxes(range=[0, max_y])
            fig_tendencia.update_traces(line=dict(width=3), marker=dict(size=8), texttemplate="$ %{y:,.0f}", textposition="top center", hovertemplate="<b>%{x}</b><br>Costo: $ %{y:,.0f} COP/Ha<extra></extra>")
            st.plotly_chart(fig_tendencia, use_container_width=True)
        else:
            st.warning("⚠️ No hay suficientes operaciones registradas para trazar la curva.")
            
        st.markdown("<hr>", unsafe_allow_html=True)
        
        vuelo_tot_a = (df_area_a['AVION_NUM'] * df_area_a['AREA_NUM']).sum()
        vuelo_tot_b = (df_area_b['AVION_NUM'] * df_area_b['AREA_NUM']).sum()
        costo_tot_a = (df_area_a['COSTO_NUM'] * df_area_a['AREA_NUM']).sum()
        costo_tot_b = (df_area_b['COSTO_NUM'] * df_area_b['AREA_NUM']).sum()
        
        vuelo_a = vuelo_tot_a / area_a if area_a > 0 else 0
        vuelo_b = vuelo_tot_b / area_b if area_b > 0 else 0
        costo_real_a = costo_tot_a / area_a if area_a > 0 else 0
        costo_real_b = costo_tot_b / area_b if area_b > 0 else 0
        
        insumos_a = max(0, costo_real_a - vuelo_a)
        insumos_b = max(0, costo_real_b - vuelo_b)
        
        categorias = [f'Análisis {año_base}', f'Análisis {año_comp}']
        
        st.markdown("#### 🛩️ vs 🧪 Distribución del Encarecimiento")
        tab_unit, tab_glob = st.tabs(["🎯 Impacto Unitario (Promedio / Ha)", "💰 Impacto Global (Presupuesto Total)"])
        
        with tab_unit:
            fig_unit = go.Figure(data=[
                go.Bar(name='Costo Avión / Ha', x=categorias, y=[vuelo_a, vuelo_b], marker_color='#2F75B5', text=[f"$ {vuelo_a:,.0f}", f"$ {vuelo_b:,.0f}"], textposition='auto'),
                go.Bar(name='Costo Insumos y Otros / Ha', x=categorias, y=[insumos_a, insumos_b], marker_color='#27AE60', text=[f"$ {insumos_a:,.0f}", f"$ {insumos_b:,.0f}"], textposition='auto')
            ])
            fig_unit.update_layout(barmode='stack', plot_bgcolor='rgba(0,0,0,0)', yaxis_title="Valor COP / Ha", margin=dict(t=20, b=20))
            fig_unit.update_xaxes(type='category')
            st.plotly_chart(fig_unit, use_container_width=True)
            
        with tab_glob:
            fig_glob = go.Figure(data=[
                go.Bar(name='Total Facturación Avión', x=categorias, y=[vuelo_tot_a, vuelo_tot_b], marker_color='#2F75B5', text=[f"$ {vuelo_tot_a:,.0f}", f"$ {vuelo_tot_b:,.0f}"], textposition='auto'),
                go.Bar(name='Total Consumo Insumos y Otros', x=categorias, y=[(costo_tot_a - vuelo_tot_a), (costo_tot_b - vuelo_tot_b)], marker_color='#27AE60', text=[f"$ {(costo_tot_a - vuelo_tot_a):,.0f}", f"$ {(costo_tot_b - vuelo_tot_b):,.0f}"], textposition='auto')
            ])
            fig_glob.update_layout(barmode='stack', plot_bgcolor='rgba(0,0,0,0)', yaxis_title="Valor Total COP", margin=dict(t=20, b=20))
            fig_glob.update_xaxes(type='category')
            st.plotly_chart(fig_glob, use_container_width=True)
        
        col_coctel = 'COCTEL' if 'COCTEL' in df_finca.columns else ('COCTEL_MAESTRO' if 'COCTEL_MAESTRO' in df_finca.columns else None)
        col_gln = 'GLN_HA' if 'GLN_HA' in df_finca.columns else None
        
        if col_coctel:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("#### 📋 Desglose Operativo: Cócteles y Variación")
            
            with st.expander("🛠️ RADAR DE DEPURACIÓN (LA TRAMPA) - Clic para inspeccionar datos", expanded=False):
                st.write(f"**Operaciones encontradas en {año_base}:**", len(df_periodo_a))
                st.write(f"**Operaciones encontradas en {año_comp}:**", len(df_periodo_b))
                if df_periodo_b.empty:
                    st.warning(f"¡ALERTA TÁCTICA! El sistema no encontró ningún vuelo para el año {año_comp} con los filtros actuales.")
            
            df_periodo_a.loc[:, col_coctel] = df_periodo_a[col_coctel].astype(str).str.strip().str.upper()
            df_periodo_b.loc[:, col_coctel] = df_periodo_b[col_coctel].astype(str).str.strip().str.upper()
            
            df_periodo_a['COSTO_NUM'] = pd.to_numeric(df_periodo_a['COSTO_NUM'], errors='coerce').fillna(0)
            df_periodo_b['COSTO_NUM'] = pd.to_numeric(df_periodo_b['COSTO_NUM'], errors='coerce').fillna(0)
            if col_gln:
                df_periodo_a[col_gln] = pd.to_numeric(df_periodo_a[col_gln], errors='coerce').fillna(0)
                df_periodo_b[col_gln] = pd.to_numeric(df_periodo_b[col_gln], errors='coerce').fillna(0)
            
            agg_dict = {'COSTO_NUM': 'mean'}
            if col_gln: agg_dict[col_gln] = 'mean'
            
            g_a = df_periodo_a.groupby(col_coctel).agg(agg_dict).reset_index()
            g_b = df_periodo_b.groupby(col_coctel).agg(agg_dict).reset_index()
            
            tabla_autopsia = pd.merge(g_a, g_b, on=col_coctel, how='outer', suffixes=('_BASE', '_ACTUAL')).fillna(0)
            
            tabla_autopsia['Variación ($)'] = tabla_autopsia['COSTO_NUM_ACTUAL'] - tabla_autopsia['COSTO_NUM_BASE']
            nombre_base = f'Costo/Ha ({año_base})'
            nombre_comp = f'Costo/Ha ({año_comp})' if año_base != año_comp else f'Costo/Ha ({año_comp} - Act)'
            
            dicc_renombres = {col_coctel: 'CÓCTEL APLICADO', 'COSTO_NUM_BASE': nombre_base, 'COSTO_NUM_ACTUAL': nombre_comp}
            if col_gln: 
                dicc_renombres[f'{col_gln}_BASE'] = f'Gln/Ha ({año_base})'
                dicc_renombres[f'{col_gln}_ACTUAL'] = f'Gln/Ha ({año_comp})' if año_base != año_comp else f'Gln/Ha ({año_comp} - Act)'
                
            tabla_autopsia.rename(columns=dicc_renombres, inplace=True)
                
            if not tabla_autopsia.empty and (tabla_autopsia[nombre_base].sum() > 0 or tabla_autopsia[nombre_comp].sum() > 0):
                st.markdown("##### 📊 Comparativo Histórico de Inversión por Cóctel")
                df_graf_coctel = pd.melt(tabla_autopsia, id_vars=['CÓCTEL APLICADO'], 
                                         value_vars=[nombre_base, nombre_comp],
                                         var_name='Periodo', value_name='Costo Promedio')
                
                fig_coctel = px.bar(df_graf_coctel, x='CÓCTEL APLICADO', y='Costo Promedio', color='Periodo', 
                                    barmode='group', color_discrete_sequence=['#2F75B5', '#27AE60'], text='Costo Promedio')
                
                fig_coctel.update_traces(texttemplate='$ %{text:,.0f}', textposition='outside', textangle=-90, textfont_size=11)
                
                max_val = df_graf_coctel['Costo Promedio'].max()
                if pd.notna(max_val) and max_val > 0:
                    fig_coctel.update_yaxes(range=[0, max_val * 1.35])
                    
                fig_coctel.update_layout(yaxis_title="Costo Operativo ($ COP / Ha)", xaxis_title="Estructura de la Receta", 
                                         plot_bgcolor='rgba(0,0,0,0)', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                                         margin=dict(t=50))
                st.plotly_chart(fig_coctel, use_container_width=True)

            df_vista = tabla_autopsia.copy()
            df_vista[nombre_base] = df_vista[nombre_base].map("$ {:,.0f}".format)
            df_vista[nombre_comp] = df_vista[nombre_comp].map("$ {:,.0f}".format)
            df_vista['Variación ($)'] = df_vista['Variación ($)'].map("$ {:,.0f}".format)
            st.dataframe(df_vista, use_container_width=True, hide_index=True)
            
            # =====================================================================
            # 🔬 NIVEL 2: CACHÉ ABSOLUTA DE BÓVEDA DE RECETAS
            # =====================================================================
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("### 🔬 Nivel 2: Composición del Cóctel y Variación Real de Insumos")
            
            cocteles_disponibles = sorted(list(set(df_periodo_a[col_coctel].dropna().unique()) | set(df_periodo_b[col_coctel].dropna().unique())))
            coctel_sel = st.selectbox("🎯 Seleccione un Cóctel para auditar su receta año vs año:", ["SELECCIONE UN CÓCTEL..."] + cocteles_disponibles)

            if coctel_sel != "SELECCIONE UN CÓCTEL...":
                with st.spinner("Extrayendo matrices químicas desde la caché en RAM..."):
                    df_mezclas, df_conf, df_dicc, df_precios, df_t2_cache = cargar_boveda_recetas_y_precios()
                    
                    if df_mezclas.empty or df_precios.empty:
                        st.error("🚨 Enlace roto con la bóveda de ingredientes históricos.")
                        st.stop()

                dict_prods_unicos = extraer_receta_de_sigla_bi(coctel_sel, finca_sel, df_mezclas, df_dicc, df_t2_cache)
                prods_receta = [{"PRODUCTO": k, "DOSIS": v} for k, v in dict_prods_unicos.items() if v > 0]
                
                if prods_receta:
                    matriz_mol = []
                    def obtener_precio_promedio(producto, anio_obj):
                        if not df_precios.empty:
                            match_df = df_precios[(df_precios['AÑO'] == str(anio_obj)) & (df_precios['PRODUCTO'] == producto)]
                            if match_df.empty: match_df = df_precios[(df_precios['AÑO'] == str(anio_obj)) & (df_precios['PRODUCTO'].str.contains(producto))]
                            if not match_df.empty and match_df['PRECIO_PROM'].mean() > 0: return match_df['PRECIO_PROM'].mean()
                        if (str(anio_obj) == str(año_comp) or str(anio_obj) == str(datetime.now().year)) and not df_conf.empty:
                            match_conf = df_conf[df_conf.iloc[:, 8].astype(str).str.upper().str.strip() == producto]
                            if not match_conf.empty: return limpiar_dinero(match_conf.iloc[0, 9])
                        return 0.0

                    costo_total_a, costo_total_b = 0.0, 0.0
                    for item in prods_receta:
                        prod, dosis = item["PRODUCTO"], item["DOSIS"]
                        precio_a, precio_b = obtener_precio_promedio(prod, año_base), obtener_precio_promedio(prod, año_comp)
                        costo_ha_a, costo_ha_b = dosis * precio_a, dosis * precio_b
                        costo_total_a += costo_ha_a
                        costo_total_b += costo_ha_b
                        matriz_mol.append({"INSUMO QUÍMICO": prod, "DOSIS/HA": f"{dosis:.3f}", f"P. Prom. ({año_base})": f"$ {precio_a:,.0f}", f"P. Prom. ({año_comp})": f"$ {precio_b:,.0f}", f"Costo/Ha ({año_base})": costo_ha_a, f"Costo/Ha ({año_comp})": costo_ha_b, "Variación ($)": costo_ha_b - costo_ha_a})

                    df_vista_mol = pd.DataFrame(matriz_mol).sort_values('Variación ($)', ascending=False)
                    df_vista_mol_print = df_vista_mol.copy()
                    df_vista_mol_print[f"Costo/Ha ({año_base})"] = df_vista_mol_print[f"Costo/Ha ({año_base})"].map("$ {:,.0f}".format)
                    df_vista_mol_print[f"Costo/Ha ({año_comp})"] = df_vista_mol_print[f"Costo/Ha ({año_comp})"].map("$ {:,.0f}".format)
                    df_vista_mol_print["Variación ($)"] = df_vista_mol_print["Variación ($)"].map("$ {:,.0f}".format)
                    st.dataframe(df_vista_mol_print, use_container_width=True, hide_index=True)
                    
                    c1, c2, c3 = st.columns(3)
                    c1.metric(f"Total Teórico ({año_base})", f"$ {costo_total_a:,.0f}")
                    c2.metric(f"Total Teórico ({año_comp})", f"$ {costo_total_b:,.0f}")
                    c3.metric("Variación Cóctel", f"$ {costo_total_b - costo_total_a:,.0f}", delta=f"$ {costo_total_b - costo_total_a:,.0f}", delta_color="inverse")
                    
                    if 'AVION_MAESTRO' in df_periodo_b.columns:
                        df_coctel_b = df_area_b[df_area_b[col_coctel] == coctel_sel]
                        costo_total_facturado_b = df_coctel_b['COSTO_NUM'].mean() if not df_coctel_b.empty else 0
                        vuelo_facturado_b = df_coctel_b['AVION_NUM'].mean() if not df_coctel_b.empty else 0
                        insumos_facturados_b = max(0, costo_total_facturado_b - vuelo_facturado_b)
                        
                        if costo_total_b > 0 and insumos_facturados_b > 0:
                            diff_b = insumos_facturados_b - costo_total_b
                            st.markdown("---")
                            st.markdown("### 🤖 Deliberador IA: Auditoría de Facturación SAP vs Receta Teórica")
                            if abs(diff_b) <= 2000: st.success(f"✅ **AUDITORÍA PERFECTA:** El costo de químicos facturados en SAP ($ {insumos_facturados_b:,.0f}) coincide con la receta ($ {costo_total_b:,.0f}).")
                            else:
                                st.warning(f"⚠️ **DISCREPANCIA DETECTADA:** Los insumos facturados ($ {insumos_facturados_b:,.0f}) no cuadran con el teórico ($ {costo_total_b:,.0f}). Diferencia: **$ {diff_b:,.0f} / Ha**")
                                st.markdown("#### 🔍 Conclusiones del Deliberador:")
                                if diff_b > 0: st.write("- 📈 **Sobrecosto:** Se cobró más de lo que indica la sigla. Es probable que se haya aplicado **SPRAYFIX**, **ADHERENTE** extra o mayor dosis de **ACEITE**.")
                                else: st.write("- 📉 **Ahorro/Faltante:** Se cobró menos. Si la finca es orgánica, se facturó correctamente (sin adherente), o hubo un error a favor en SAP.")
                                
                                candidatos_encontrados = False
                                for idx, p_row in df_precios[df_precios['AÑO'] == str(año_comp)].iterrows():
                                    precio_p = p_row['PRECIO_PROM']
                                    for d in [0.02, 0.06, 0.13, 0.2, 0.5, 1.0, 2.0]:
                                        costo_teorico = precio_p * d
                                        if costo_teorico > 0 and abs(costo_teorico - abs(diff_b)) <= (abs(diff_b) * 0.15 + 500):
                                            st.info(f"💡 ¿Se aplicó/omitió **{p_row['PRODUCTO']}** a dosis de **{d} L/Ha**? (Costo aprox: $ {costo_teorico:,.0f})")
                                            candidatos_encontrados = True; break
                                    if candidatos_encontrados: break
                else: st.info("No se encontraron ingredientes válidos para esta receta.")
                
        # =====================================================================
        # 📦 NIVEL 3: INTELIGENCIA LOGÍSTICA (CON CEBO DE TELEMETRÍA Y DIAGNÓSTICO)
        # =====================================================================
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown(f"### 📦 Nivel 3: Consumo Volumétrico de Insumos ({año_comp})")
        st.info(f"Cálculo volumétrico avanzado basado en el comportamiento individual de las siglas para **{area_b:,.1f} Ha**.")

        c_coctel = next((c for c in df_periodo_b.columns if 'COCTEL' in str(c).upper()), None)

        if not df_periodo_b.empty and c_coctel:
            with st.spinner("Lanzando cebo de telemetría en RAM..."):
                try:
                    df_m, df_c, df_d, df_p, df_t2_b = cargar_boveda_recetas_y_precios()
                    
                    # 💥 EL CEBO: PINTAR LA CAJA NEGRA DE DIAGNÓSTICO EN PANTALLA
                    with st.expander("🔎 RADAR DE TELEMETRÍA (EL CEBO) - Clic para ver interferencias", expanded=True):
                        st.write("📊 **Estado de la Bóveda DD_Mesclas:**")
                        st.write(f"- ¿Está vacía la tabla de recetas?: {'SÍ 🚨' if df_m.empty else 'NO ✅'}")
                        st.write(f"- Total de filas cargadas en memoria: {len(df_m)}")
                        if not df_m.empty:
                            st.write("- Columnas reales leídas del Excel:", list(df_m.columns))
                            st.write("- Primeros 10 cócteles en Columna A del Excel:", df_m.iloc[:, 0].dropna().astype(str).str.upper().str.strip().unique()[:10].tolist())
                        
                        st.write("---")
                        st.write("🚜 **Estado de las Operaciones (Tabla 1):**")
                        resumen_ha = df_periodo_b.groupby(c_coctel)['AREA_NUM'].sum().reset_index()
                        st.write("- Cócteles reales encontrados en este periodo:", resumen_ha[c_coctel].astype(str).tolist())
                        
                        # Rastreo de emparejamiento paso a paso
                        st.write("---")
                        st.write("🔬 **Simulación de Emparejamiento Quirúrgico:**")
                        trace_list = []
                        for _, fila in resumen_ha.iterrows():
                            c_original = str(fila[c_coctel]).upper().strip()
                            c_norm = c_original.replace("+", " ").replace("-", " ")
                            base_c = c_norm.split()[0] if c_norm.split() else ""
                            
                            # Simular búsqueda pura en columna A
                            if not df_m.empty:
                                coincidencia_pura = df_m[df_m.iloc[:, 0].astype(str).str.upper().str.strip() == base_c]
                                match_status = f"SÍ (Filas: {len(coincidencia_pura)})" if not coincidencia_pura.empty else "NO 🚨"
                            else:
                                match_status = "Bóveda vacía"
                                
                            trace_list.append({
                                "Cóctel en Tabla 1": c_original,
                                "Base Buscada": base_c,
                                "¿Hizo Match en Excel?": match_status
                            })
                        st.dataframe(pd.DataFrame(trace_list), use_container_width=True)

                    # Continuar con el proceso normal
                    consumo_log = {}
                    for _, fila in resumen_ha.iterrows():
                        nombre_coctel = str(fila[c_coctel]).upper().strip()
                        ha_aplicadas = fila['AREA_NUM']
                        if ha_aplicadas <= 0 or nombre_coctel in ["NAN", ""]: continue

                        dict_temp = extraer_receta_de_sigla_bi(nombre_coctel, finca_sel, df_m, df_d, df_t2_b)
                        for p, d in dict_temp.items():
                            consumo_log[p] = consumo_log.get(p, 0) + (d * ha_aplicadas)

                    if consumo_log:
                        st.markdown("#### 🔎 Auditoría de Consumo por Insumo")
                        lista_insumos = ["📦 VER TODOS LOS INSUMOS (RESUMEN GLOBAL)"] + sorted(list(consumo_log.keys()))
                        insumo_filtrado = st.selectbox("Seleccione el producto a auditar en el rango de fechas:", lista_insumos)

                        df_log = pd.DataFrame(list(consumo_log.items()), columns=["🧪 PRODUCTO", "📦 VOLUMEN ESTIMADO (L/Kg)"]).sort_values(by="📦 VOLUMEN ESTIMADO (L/Kg)", ascending=False)
                        
                        if insumo_filtrado == "📦 VER TODOS LOS INSUMOS (RESUMEN GLOBAL)":
                            df_vista = df_log.copy()
                            df_vista["📦 VOLUMEN ESTIMADO (L/Kg)"] = df_vista["📦 VOLUMEN ESTIMADO (L/Kg)"].map("{:,.1f}".format)
                            
                            c1, c2 = st.columns([1, 1.2])
                            with c1: st.dataframe(df_vista, use_container_width=True, hide_index=True)
                            with c2:
                                fig = px.bar(df_log.head(15), y="🧪 PRODUCTO", x="📦 VOLUMEN ESTIMADO (L/Kg)", orientation='h', color="📦 VOLUMEN ESTIMADO (L/Kg)", color_continuous_scale="GnBu", text_auto='.1f', title="Top 15 Insumos con Mayor Demanda")
                                fig.update_layout(yaxis={'categoryorder':'total ascending'}, plot_bgcolor='rgba(0,0,0,0)')
                                st.plotly_chart(fig, use_container_width=True)
                        else:
                            vol_especifico = consumo_log[insumo_filtrado]
                            st.markdown(f"""
                            <div style='background-color:#0d1b2a; padding:25px; border-radius:10px; border-left:8px solid #27AE60; text-align:center; box-shadow: 0 4px 8px rgba(0,0,0,0.2);'>
                                <h4 style='color:#27AE60; margin:0; text-transform: uppercase; letter-spacing: 2px;'>CONSUMO TOTAL EN EL PERIODO</h4>
                                <h1 style='color:white; margin:10px 0; font-size: 45px;'>{vol_especifico:,.1f}</h1>
                                <p style='color:#d4af37; margin:0; font-size: 18px; font-weight: bold;'>Litros o Kilos teóricos de {insumo_filtrado}</p>
                                </div>
                            """, unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"🚨 Error en el radar de inteligencia logística: {e}")
        else:
            st.info("Esperando datos de operaciones en el periodo evaluado...")
        # =====================================================================
        # --- 🤝 SIMULADOR DE NEGOCIACIÓN Y AUDITORÍA DE TARIFAS ---
        # =====================================================================
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("### 🤝 Simulador de Negociación (Tarifas de Aerofumigación)")
        st.info("💡 RADAR BLINDADO: Extracción estricta de Tarifas Unitarias (Columnas T y U).")

        col_pista_sim = next((c for c in super_base_bi.columns if any(k in str(c).upper() for k in ["PISTA", "ALMACEN", "CENTRO"])), None)
        pistas_sim_disp = ["TODAS"] + sorted(super_base_bi[col_pista_sim].dropna().astype(str).str.upper().unique().tolist()) if col_pista_sim else ["TODAS"]

        c_sim1, c_sim2, c_sim3 = st.columns(3)
        sim_fecha_inicio = c_sim1.date_input("📅 Fecha Inicial:", value=datetime.now().date(), key="sim_f_ini_f")
        sim_fecha_fin = c_sim2.date_input("📅 Fecha Final:", value=datetime.now().date(), key="sim_f_fin_f")
        sim_pista = c_sim3.selectbox("📍 Base / Pista:", pistas_sim_disp, key="sim_pista_f")

        st.markdown("<br>", unsafe_allow_html=True)
        c_sim_m1, c_sim_m2, c_sim_m3 = st.columns(3)
        margen_actual = c_sim_m1.number_input("📉 Margen Actual en Factura (%)", value=8.0, step=0.5, key="marg_act_f")
        margen_nuevo = c_sim_m2.number_input("📈 Nuevo Margen a Simular (%)", value=11.0, step=0.5, key="marg_nue_f")
        
        with c_sim_m3:
            st.markdown("<br>", unsafe_allow_html=True)
            btn_simular = st.button("🚀 EJECUTAR SIMULACIÓN", type="primary", use_container_width=True, key="btn_simular_f")

        if btn_simular:
            with st.spinner("Procesando auditoría de simulación..."):
                df_sim = super_base_bi[(super_base_bi['FECHA_DT'].dt.date >= sim_fecha_inicio) & (super_base_bi['FECHA_DT'].dt.date <= sim_fecha_fin)].copy()
                if col_pista_sim and sim_pista != "TODAS": df_sim = df_sim[df_sim[col_pista_sim].astype(str).str.upper() == sim_pista]
                df_sim = df_sim[df_sim['AREA_NUM'] > 0]

                if df_sim.empty:
                    st.warning("⚠️ No se encontraron Órdenes de Servicio para los parámetros elegidos.")
                else:
                    def red_excel(num): return math.floor(num + 0.5) if num >= 0 else math.ceil(num - 0.5)
                    col_finca, col_os = 'FINCA_MAESTRA', 'OS_MAESTRA'
                    
                    df_sim_unicos = df_sim.drop_duplicates(subset=['FECHA_DT', col_finca, col_os, 'COCTEL_MAESTRO', 'AREA_NUM'])

                    matriz_simulacion = []
                    for _, row in df_sim_unicos.iterrows():
                        os_val = str(row[col_os]).strip()
                        if os_val in ["", "nan"]: continue

                        finca_val, ha_val = str(row[col_finca]).upper().strip(), float(row['AREA_NUM'])
                        pista_val = str(row[col_pista_sim]).upper().strip() if col_pista_sim else "N/A"
                        fecha_val = row['FECHA_DT'].strftime('%d/%m/%Y')
                        semana_val = (row['FECHA_DT'] + pd.Timedelta(days=2)).isocalendar()[1]

                        tar_avion_raw = limpiar_dinero(row['AVION_MAESTRO'])
                        tar_dom_raw = limpiar_dinero(row['DOMINIC_MAESTRO'])
                        tarifa_unitaria_actual = tar_avion_raw + tar_dom_raw

                        if tarifa_unitaria_actual > 0 and ha_val > 0:
                            t_act_red = red_excel(tarifa_unitaria_actual)
                            base_neta_ha = tarifa_unitaria_actual / (1 + (margen_actual / 100))
                            t_nue_red = red_excel(base_neta_ha * (1 + (margen_nuevo / 100)))
                            resta_tarifas = t_nue_red - t_act_red
                            
                            matriz_simulacion.append({
                                "Nº OS": os_val, "FECHA": fecha_val, "SEMANA": int(semana_val) if str(semana_val).isdigit() else semana_val,
                                "FINCA": finca_val, "PISTA": pista_val, "HECTÁREAS": ha_val,
                                f"TARIFA ACTUAL / Ha ({margen_actual}%)": t_act_red, f"NUEVA TARIFA / Ha ({margen_nuevo}%)": t_nue_red,
                                "TOTAL ACTUAL ($)": red_excel(t_act_red * ha_val), "NUEVO TOTAL ($)": red_excel(t_nue_red * ha_val), "DIFERENCIA ($)": red_excel(resta_tarifas * ha_val)
                            })

                    if not matriz_simulacion:
                        st.warning("⚠️ Tarifas unitarias inválidas en la extracción.")
                    else:
                        df_resultados = pd.DataFrame(matriz_simulacion)
                        df_semanal = df_resultados.groupby("SEMANA").agg({"HECTÁREAS": "sum", "TOTAL ACTUAL ($)": "sum", "NUEVO TOTAL ($)": "sum", "DIFERENCIA ($)": "sum"}).reset_index().sort_values(by="SEMANA").reset_index(drop=True)

                        total_actual_global = df_resultados["TOTAL ACTUAL ($)"].sum()
                        total_simulado_global = df_resultados["NUEVO TOTAL ($)"].sum()
                        total_diferencia_global = df_resultados["DIFERENCIA ($)"].sum()

                        st.markdown("### 🎯 Impacto Financiero Real de la Simulación")
                        k1, k2, k3 = st.columns(3)
                        k1.metric(f"💰 Total Actual ({margen_actual}%)", f"$ {total_actual_global:,.0f}")
                        k2.metric(f"📈 Proyección ({margen_nuevo}%)", f"$ {total_simulado_global:,.0f}")
                        k3.metric("⚖️ Dinero Real en Juego", f"$ {abs(total_diferencia_global):,.0f}", delta=f"$ {total_diferencia_global:,.0f}", delta_color="normal" if total_diferencia_global > 0 else "inverse")

                        st.markdown("<br>", unsafe_allow_html=True)
                        tab_resumen, tab_detalle = st.tabs(["📊 1. Resumen Macroeconómico", "📋 2. Desglose Quirúrgico"])
                        
                        with tab_resumen:
                            df_sem_vista = df_semanal.copy()
                            df_sem_vista["HECTÁREAS"] = df_sem_vista["HECTÁREAS"].map("{:,.2f}".format)
                            for col in ["TOTAL ACTUAL ($)", "NUEVO TOTAL ($)", "DIFERENCIA ($)"]: df_sem_vista[col] = df_sem_vista[col].map("$ {:,.0f}".format)
                            st.dataframe(df_sem_vista, use_container_width=True, hide_index=True)
                        
                        with tab_detalle:
                            df_vista = df_resultados.copy()
                            df_vista["HECTÁREAS"] = df_vista["HECTÁREAS"].map("{:,.2f}".format)
                            for col in [f"TARIFA ACTUAL / Ha ({margen_actual}%)", f"NUEVA TARIFA / Ha ({margen_nuevo}%)", "TOTAL ACTUAL ($)", "NUEVO TOTAL ($)", "DIFERENCIA ($)"]: df_vista[col] = df_vista[col].map("$ {:,.0f}".format)
                            
                            def col_dif(val):
                                if isinstance(val, str) and "-" in val: return 'color: #721c24; background-color: #f8d7da; font-weight: bold; text-align: center;'
                                if isinstance(val, str) and "$" in val: return 'color: #155724; background-color: #d4edda; font-weight: bold; text-align: center;'
                                return ''
                            st.dataframe(df_vista.style.map(col_dif, subset=["DIFERENCIA ($)"]), use_container_width=True, hide_index=True)

                        # --- CONSTRUCCIÓN DE EXCEL EMBEBIDO ULTRA VELOZ ---
                        buffer_neg = io.BytesIO()
                        with pd.ExcelWriter(buffer_neg, engine='openpyxl') as writer:
                            df_semanal.to_excel(writer, sheet_name='Resumen_Semanal', index=False)
                            df_resultados.to_excel(writer, sheet_name='Detalle_OS', index=False)
                            
                            borde = Border(left=Side(style='thin', color='D1D1D1'), right=Side(style='thin', color='D1D1D1'), top=Side(style='thin', color='D1D1D1'), bottom=Side(style='thin', color='D1D1D1'))
                            fondo, blanca = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid"), Font(color="FFFFFF", bold=True)
                            
                            for name in ['Resumen_Semanal', 'Detalle_OS']:
                                ws = writer.sheets[name]
                                ws.sheet_view.showGridLines = True
                                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                    for cell in row:
                                        cell.border = borde
                                        if cell.row == 1: 
                                            cell.fill = fondo; cell.font = blanca; cell.alignment = Alignment(horizontal='center', vertical='center')
                                        else:
                                            if "HECTÁREAS" in str(ws.cell(1, cell.column).value): cell.number_format = '#,##0.00'
                                            elif "($" in str(ws.cell(1, cell.column).value) or "%" in str(ws.cell(1, cell.column).value): cell.number_format = '"$" #,##0'
                                for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col) + 4, 32)
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.download_button(label="📥 DESCARGAR INFORME DUAL (EXCEL OFICIAL)", data=buffer_neg.getvalue(), file_name=f"Auditoria_Tarifas_{sim_pista}_{sim_fecha_inicio}_a_{sim_fecha_fin}.xlsx", type="primary", use_container_width=True)

    except Exception as e:
        st.error(f"🚨 Falla crítica en los motores del Centro BI: {e}")

        st.markdown("""
            <a href="#inicio_modulo" target="_self" style="
                display: block; width: 100%; text-align: center; 
                background-color: #0d1b2a; color: #d4af37; border: 1px solid #d4af37; 
                padding: 12px; border-radius: 8px; text-decoration: none; font-weight: bold;
                box-shadow: 0px 4px 6px rgba(0,0,0,0.3); margin-bottom: 20px; font-size: 16px;
            ">
                ⬆️ VOLVER AL INICIO DEL MÓDULO ⬆️
            </a>
        """, unsafe_allow_html=True)

if __name__ == "__main__":
    pass
