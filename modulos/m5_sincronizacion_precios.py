import streamlit as st
import pandas as pd
import numpy as np
import gspread
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from oauth2client.service_account import ServiceAccountCredentials

# =================================================================
# ⚡ MOTORES DE CONEXIÓN Y ACCESO SATELITAL (ALTA VELOCIDAD)
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    """ Centraliza la autenticación unificada con Google Cloud una sola vez en RAM """
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

def purificar_y_convertir_precio(valor_crudo):
    if not valor_crudo:
        return 0.0
    val_str = str(valor_crudo).replace("$", "").replace("COP", "").replace(" ", "").strip()
    if "," in val_str and "." in val_str:
        if val_str.find(".") < val_str.find(","):
            val_str = val_str.replace(".", "").replace(",", ".")
        else:
            val_str = val_str.replace(",", "")
    elif "," in val_str:
        if len(val_str.split(",")[-1]) == 2:
            val_str = val_str.replace(",", ".")
        else:
            val_str = val_str.replace(",", "")
    try:
        return float(val_str)
    except ValueError:
        return 0.0

# =================================================================
# 📊 EXPORTADOR EXCEL PREMIUM: FORMATO VERTICAL GERENCIAL + DOSIS
# =================================================================
def generar_excel_gerencial(df_comp, dosis_dict):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        start_row = 3
        df_comp.to_excel(writer, sheet_name='Rentabilidad_Gerencial', index=False, startrow=start_row)
        ws = writer.sheets['Rentabilidad_Gerencial']

        # --- DIBUJAR TÍTULO Y FECHA DENTRO DEL EXCEL ---
        ws.cell(row=1, column=1, value="REPORTE GERENCIAL: MATRIZ DE RENTABILIDAD Y MÁRGENES")
        cell_titulo = ws.cell(row=1, column=1)
        cell_titulo.font = Font(name="Arial", size=14, bold=True, color="0D1B2A")

        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws.cell(row=2, column=1, value=f"Génesis Omega Pro | Auditoría Financiera Generada el: {fecha_actual}")
        cell_sub = ws.cell(row=2, column=1)
        cell_sub.font = Font(name="Arial", size=10, italic=True, color="555555")

        # --- ESTILOS DE TABLA ---
        header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        borde_fino = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        
        header_row = start_row + 1
        data_start_row = header_row + 1

        for col_num, col_name in enumerate(df_comp.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 24
            
            # --- 🎯 AGREGAR DOSIS EN LA FILA 3 (JUSTO ARRIBA DEL ENCABEZADO) ---
            for prod, dose in dosis_dict.items():
                if f"🧪 {prod}" == col_name:
                    cell_dose = ws.cell(row=start_row, column=col_num)
                    cell_dose.value = f"⚖️ Dosis: {dose:.2f} L/Kg/Ha"
                    cell_dose.font = Font(name="Arial", size=10, bold=True, color="0D1B2A")
                    cell_dose.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                    cell_dose.alignment = Alignment(horizontal='center', vertical='center')
                    cell_dose.border = borde_fino
            
            # Formato Encabezados (TODOS CON FONDO OSCURO Y LETRA DORADA EN EXCEL)
            cell_header = ws.cell(row=header_row, column=col_num)
            cell_header.fill = header_fill
            if ("DIF" in str(col_name) and "%" in str(col_name)) or "MARGEN REQ" in str(col_name):
                cell_header.font = Font(color="D4AF37", bold=True, size=11)
            else:
                cell_header.font = Font(color="FFFFFF", bold=True, size=11)
                
            cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_header.border = borde_fino
            
            # Formato de Datos
            for row_num in range(data_start_row, len(df_comp) + data_start_row):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = borde_fino
                cell.alignment = Alignment(vertical='center', horizontal='left')
                
                # Resaltar la fila de "Ganancia"
                concepto = str(ws.cell(row=row_num, column=2).value)
                is_ganancia = (concepto == 'Ganancia')
                
                val = cell.value
                if col_num in [1, 2]:
                    cell.font = Font(name="Arial", size=11, bold=True) if is_ganancia else Font(name="Arial", size=11)
                elif isinstance(val, (int, float)):
                    if ("DIF" in str(col_name) and "%" in str(col_name)) or "MARGEN REQ" in str(col_name):
                        cell.number_format = '0.00 "%";[Red]-0.00 "%";"0.00 %"'
                        if val < 0: cell.font = Font(name="Arial", size=11, bold=True, color="E74C3C")
                        elif "MARGEN REQ" in str(col_name): cell.font = Font(name="Arial", size=11, bold=True, color="8E44AD") # Púrpura estratégico
                        else: cell.font = Font(name="Arial", size=11, bold=is_ganancia)
                        cell.alignment = Alignment(vertical='center', horizontal='center')
                    else:
                        cell.number_format = '"$" #,##0;[Red]("-" "$" #,##0);"-"'
                        if "DIF" in str(col_name) and val < 0: 
                            cell.font = Font(name="Arial", size=11, bold=True, color="E74C3C")
                        else: 
                            cell.font = Font(name="Arial", size=11, bold=is_ganancia)
                        
                        if "DIF" in str(col_name): cell.alignment = Alignment(vertical='center', horizontal='center')
                else:
                    cell.font = Font(name="Arial", size=11, bold=is_ganancia)

    return buffer.getvalue()

# =================================================================
# 🔌 MEMORIA DINÁMICA DE MÁRGENES Y TARIFAS
# =================================================================
@st.cache_data(show_spinner=False, ttl=1800)
def obtener_tarifario_maestro_cached(_supabase_client):
    df = pd.DataFrame()
    if _supabase_client:
        try:
            respuesta = _supabase_client.table("PRECIOS_INSUMOS").select("*").execute()
            if respuesta.data: df = pd.DataFrame(respuesta.data)
        except Exception: df = pd.DataFrame()

    margenes = {"TERCERO": 1.451, "AFILIADO": 1.164, "COOPERATIVA / SOCIO": 1.112, "ORGANICO": 1.011}
    
    gc = obtener_cliente_gspread_unificado()
    if gc:
        try:
            sh = gc.open_by_url(URL_BOVEDA_MAESTRA)
            ws = sh.worksheet("Configuración")
            datos = ws.get_all_values()
            
            if len(datos) > 1:
                df_raw = pd.DataFrame(datos[1:], columns=datos[0])
                def parse_mult(v_str):
                    try:
                        v = str(v_str).strip().replace(",", ".")
                        if "%" in v: return 1 + (float(v.replace("%", "")) / 100.0)
                        vf = float(v)
                        if 1.0 <= vf <= 2.0: return vf
                        if 1000 <= vf <= 2000: return vf / 1000.0 
                    except: pass
                    return 0.0

                for idx, row in df_raw.iterrows():
                    grupo = str(row.iloc[0]).strip().upper()
                    if grupo in ["TERCERO", "AFILIADO", "SOCIO", "COOPERATIVA", "ORGANICO"]:
                        factor = 0.0
                        for c in range(1, 4):
                            f_val = parse_mult(row.iloc[c])
                            if f_val > 1.0:
                                factor = f_val
                                break
                        if factor > 0:
                            if grupo in ["SOCIO", "COOPERATIVA"]: margenes["COOPERATIVA / SOCIO"] = factor
                            elif grupo == "TERCERO": margenes["TERCERO"] = factor
                            elif grupo == "AFILIADO": margenes["AFILIADO"] = factor
                            elif grupo == "ORGANICO": margenes["ORGANICO"] = factor
                
                if df.empty and len(df_raw.columns) > 10:
                    df = df_raw.iloc[:, [8, 10]].copy()
                    df.columns = ['PRODUCTO', 'COSTO']
        except Exception: pass

    if df.empty or 'PRODUCTO' not in df.columns or 'COSTO' not in df.columns:
        return pd.DataFrame(), [], {}
        
    df['PRODUCTO'] = df['PRODUCTO'].astype(str).str.strip().str.upper()
    mask_validos = (df['PRODUCTO'].notna() & (df['PRODUCTO'] != "") & (df['PRODUCTO'] != "PRODUCTO") & (~df['PRODUCTO'].str.contains("INVENTARIO", na=False)))
    df = df[mask_validos].copy()
    
    df['COSTO BASE'] = df['COSTO'].apply(purificar_y_convertir_precio)
    df = df[df['COSTO BASE'] > 0].copy()
    
    # 💥 CURA TÁCTICA: PURIFICACIÓN DE DUPLICADOS (COMO EN MÓDULO 1)
    # Aplastamos las filas repetidas y conservamos solo la última actualización ('last')
    df = df.drop_duplicates(subset=['PRODUCTO'], keep='last').copy()
    
    if df.empty: return pd.DataFrame(), [], {}
    
    def fmt_pct(factor): return round((factor - 1.0) * 100, 2)
    
    col_ter = f"TERCERO (+{fmt_pct(margenes['TERCERO'])}%)"
    col_afi = f"AFILIADO (+{fmt_pct(margenes['AFILIADO'])}%)"
    col_soc = f"COOP/SOCIO (+{fmt_pct(margenes['COOPERATIVA / SOCIO'])}%)"
    col_org = f"ORGÁNICO (+{fmt_pct(margenes['ORGANICO'])}%)"

    df[col_ter] = (df['COSTO BASE'] * margenes['TERCERO']).round(0)
    df[col_afi] = (df['COSTO BASE'] * margenes['AFILIADO']).round(0)
    df[col_soc] = (df['COSTO BASE'] * margenes['COOPERATIVA / SOCIO']).round(0)
    df[col_org] = (df['COSTO BASE'] * margenes['ORGANICO']).round(0)
    
    cols = ["PRODUCTO", "COSTO BASE", col_ter, col_afi, col_soc, col_org]
    df_tarifario = df[cols].sort_values(by="PRODUCTO").reset_index(drop=True)
    
    return df_tarifario, cols[1:], margenes

# =================================================================
# 👑 PROCESAMIENTO PRINCIPAL DE TARIFAS
# =================================================================

def ejecutar(supabase_client, extraer_numero, fmt_sap, limpiar_texto_vba, val_seguro):
    st.markdown("""
    <style>
    .titulo-principal { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black', sans-serif; }
    div[data-testid="stDataFrame"] { border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; }
    .hud-tarifas { background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); border-left: 5px solid #d4af37; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.15); margin-bottom: 25px; display: flex; justify-content: space-between; align-items: center; }
    .hud-tarifas-item { text-align: center; flex: 1; }
    .hud-tarifas-title { font-size: 11px; font-weight: bold; color: #d4af37; text-transform: uppercase; margin:0; letter-spacing: 1px; }
    .hud-tarifas-value { font-size: 22px; font-family: 'Arial Black'; margin: 5px 0 0 0; }
    
    div[data-testid="stTextInput"] input, 
    div[data-testid="stNumberInput"] input, 
    div[data-testid="stSelectbox"] > div { 
        border: 2px solid #0d1b2a !important; 
        border-radius: 6px !important; 
        background-color: #ffffff !important;
        color: #0d1b2a !important;
        font-weight: 800 !important; 
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-principal'>Sincronización de Precios y Tarifas</h1>", unsafe_allow_html=True)
    gc = obtener_cliente_gspread_unificado()

    with st.container(border=True):
        st.markdown("### 🧮 Tarifario Maestro Dinámico (Visor y Cómputo de Perfiles)")
        
        col_t1, col_t2 = st.columns([3, 1])
        with col_t2:
            if st.button("🔄 RECARGAR TARIFARIO", use_container_width=True, type="secondary"):
                st.cache_data.clear()
                st.session_state.pop('df_tarifario', None)
                st.session_state.pop('opciones_cols_m5', None)
                st.session_state.pop('dict_margenes_m5', None)
                st.rerun()

        if 'df_tarifario' not in st.session_state or st.session_state['df_tarifario'].empty:
            df_tarifario_cached, opciones_cols, dict_m = obtener_tarifario_maestro_cached(supabase_client)
            if not df_tarifario_cached.empty:
                st.session_state['df_tarifario'] = df_tarifario_cached
                st.session_state['opciones_cols_m5'] = opciones_cols
                st.session_state['dict_margenes_m5'] = dict_m

        if 'df_tarifario' in st.session_state and not st.session_state['df_tarifario'].empty:
            df_t = st.session_state['df_tarifario']
            cols_dinamicas = st.session_state.get('opciones_cols_m5', [])
            dict_m = st.session_state.get('dict_margenes_m5', {})
            
            if not cols_dinamicas or len(cols_dinamicas) < 2:
                df_t, cols_dinamicas, dict_m = obtener_tarifario_maestro_cached(supabase_client)
                st.session_state['df_tarifario'] = df_t
                st.session_state['opciones_cols_m5'] = cols_dinamicas
                st.session_state['dict_margenes_m5'] = dict_m
            
            total_quimicos_tarifados = len(df_t)
            costo_maximo_comercial = df_t[cols_dinamicas[1]].max() if len(cols_dinamicas) > 1 else 0
            costo_medio_base = df_t['COSTO BASE'].mean() if 'COSTO BASE' in df_t.columns else 0
            
            st.markdown(f"""
            <div class="hud-tarifas">
                <div class="hud-tarifas-item">
                    <p class="hud-tarifas-title">Insumos en Línea</p>
                    <p class="hud-tarifas-value">🧪 {total_quimicos_tarifados} ítems</p>
                </div>
                <div class="hud-tarifas-item">
                    <p class="hud-tarifas-title">Costo Promedio</p>
                    <p class="hud-tarifas-value">💵 $ {costo_medio_base:,.0f}</p>
                </div>
                <div class="hud-tarifas-item">
                    <p class="hud-tarifas-title">Tope Máximo</p>
                    <p class="hud-tarifas-value">📈 $ {costo_maximo_comercial:,.0f}</p>
                </div>
            </div>
            """.replace(",", "."), unsafe_allow_html=True)
            
            t1, t2, t3, t4 = st.tabs([
                "💰 Visor General del Arsenal", 
                "📋 Copia Masiva (Por Margen)", 
                "🎯 Comparativo Gerencial (Utilidad/Ha)",
                "🔍 Consulta Multi-Producto"
            ])
            
            with t1:
                st.markdown("#### Matriz de Costos y Márgenes Oficiales")
                df_visual = df_t.copy()
                for col in df_visual.columns:
                    if col != "PRODUCTO":
                        df_visual[col] = df_visual[col].map("$ {:,.0f}".format).str.replace(",", ".")
                st.dataframe(df_visual, use_container_width=True, hide_index=True)
                
            with t2:
                st.markdown("#### 📟 Consola de Copiado Masivo para SAP")
                c_cop1, c_cop2 = st.columns([2, 1])
                with c_cop1:
                    lista_opciones = list(reversed(cols_dinamicas)) if cols_dinamicas else ["COSTO BASE"]
                    col_margen = st.selectbox("Seleccione el Perfil de Productor:", lista_opciones, key="sb_perfil_copia")
                with c_cop2:
                    st.write("") 
                    st.write("")
                    incluir_nombres = st.toggle("🏷️ Incluir Nombres de Productos", value=False, key="toggle_inc_nombres")

                if incluir_nombres:
                    max_len = max([len(str(p)) for p in df_t["PRODUCTO"]] + [35]) + 4
                    lista_textos = [f"{str(p).ljust(max_len)}│  {fmt_sap(v)}" for p, v in zip(df_t["PRODUCTO"], df_t[col_margen])]
                else:
                    lista_textos = [fmt_sap(x) for x in df_t[col_margen]]
                
                st.info(f"📊 **Listos para Transferencia:** {len(lista_textos)} registros procesados en columna alineada. Haga clic en 📋 para copiar.")
                st.code("\n".join(lista_textos), language="text")
                    
            with t3:
                st.markdown("#### 🎯 Análisis Comparativo Gerencial Tabular (Unpivoted)")
                st.info("💡 **MODELO VERTICAL EN CASCADA:** Visualización separada por Costo, Venta y Ganancia. Ideal para exportar a Excel, filtrar y analizar deltas financieros al detalle.")
                
                opciones_productos = df_t["PRODUCTO"].tolist()
                prods_sel = st.multiselect(
                    "🔍 Seleccione los Productos a Comparar (El primero es la Base):", 
                    options=opciones_productos,
                    default=[opciones_productos[0]] if opciones_productos else []
                )
                
                if prods_sel:
                    st.markdown("##### ⚖️ 1. Definir Dosis (L/Kg por Hectárea)")
                    df_dosis_base = pd.DataFrame({"PRODUCTO": prods_sel, "DOSIS (L/Kg/Ha)": [1.0] * len(prods_sel)})
                    
                    df_dosis = st.data_editor(
                        df_dosis_base, 
                        hide_index=True, 
                        use_container_width=True,
                        column_config={
                            "PRODUCTO": st.column_config.TextColumn("🧪 Producto", disabled=True),
                            "DOSIS (L/Kg/Ha)": st.column_config.NumberColumn("⚖️ Dosis/Ha (Doble clic para editar)", min_value=0.001, format="%.2f", step=0.1)
                        }
                    )
                    
                    dosis_dict = dict(zip(df_dosis["PRODUCTO"], df_dosis["DOSIS (L/Kg/Ha)"]))
                    
                    mapa_perfiles = [
                        ("TERCERO", cols_dinamicas[1]),
                        ("AFILIADO", cols_dinamicas[2]),
                        ("COOP/SOCIO", cols_dinamicas[3]),
                        ("ORGÁNICO", cols_dinamicas[4])
                    ]
                    
                    prod_base = prods_sel[0]
                    filas_excel = []
                    
                    for perfil_nombre, col_margen in mapa_perfiles:
                        datos_p1 = df_t[df_t["PRODUCTO"] == prod_base].iloc[0]
                        dosis_p1 = dosis_dict.get(prod_base, 1.0)
                        c_p1 = datos_p1["COSTO BASE"] * dosis_p1
                        v_p1 = datos_p1[col_margen] * dosis_p1
                        g_p1 = v_p1 - c_p1
                        
                        metricas = [("Costo", c_p1), ("Venta", v_p1), ("Ganancia", g_p1)]
                        
                        for idx_metrica, (nombre_metrica, val_p1) in enumerate(metricas):
                            perfil_display = perfil_nombre if idx_metrica == 0 else ""
                            
                            fila_ex = {
                                "🤝 PERFIL": perfil_display,
                                "📊 CONCEPTO": nombre_metrica,
                                f"🧪 {prod_base}": val_p1
                            }
                            
                            for prod_comp in prods_sel[1:]:
                                datos_p2 = df_t[df_t["PRODUCTO"] == prod_comp].iloc[0]
                                dosis_p2 = dosis_dict.get(prod_comp, 1.0)
                                c_p2 = datos_p2["COSTO BASE"] * dosis_p2
                                v_p2 = datos_p2[col_margen] * dosis_p2
                                g_p2 = v_p2 - c_p2
                                
                                if nombre_metrica == "Costo": val_p2 = c_p2
                                elif nombre_metrica == "Venta": val_p2 = v_p2
                                else: val_p2 = g_p2
                                
                                # Regla: Menor - Mayor
                                dif_pesos = min(val_p1, val_p2) - max(val_p1, val_p2)
                                dif_pct = (dif_pesos / max(val_p1, val_p2) * 100) if max(val_p1, val_p2) > 0 else 0.0
                                
                                # 💥 CURA TÁCTICA: CÁLCULO DE MARGEN REQUERIDO
                                # Margen necesario sobre COSTO Manzate para igualar VENTA Mancol
                                margen_req = ((v_p1 / c_p2) - 1.0) * 100.0 if c_p2 > 0 else 0.0
                                
                                fila_ex[f"🧪 {prod_comp}"] = val_p2
                                fila_ex[f"⚖️ DIFERENCIA ($) [{prod_comp} vs {prod_base}]"] = dif_pesos
                                fila_ex[f"📉 DIFERENCIA (%) [{prod_comp} vs {prod_base}]"] = dif_pct
                                fila_ex[f"🎯 MARGEN REQ. (%) [{prod_comp} -> Venta {prod_base}]"] = margen_req

                            filas_excel.append(fila_ex)

                    df_excel_export = pd.DataFrame(filas_excel)

                    st.markdown("##### 📊 2. Cuadro Comparativo (Vista Estructurada)")
                    
                    # Configuración de columnas para Streamlit
                    col_config = {
                        "🤝 PERFIL": st.column_config.TextColumn("🤝 PERFIL", width="small"),
                        "📊 CONCEPTO": st.column_config.TextColumn("📊 CONCEPTO", width="small")
                    }
                    for p in prods_sel:
                        col_config[f"🧪 {p}"] = st.column_config.NumberColumn(f"🧪 {p}", format="$ %d")
                    for p_comp in prods_sel[1:]:
                        col_config[f"⚖️ DIFERENCIA ($) [{p_comp} vs {prod_base}]"] = st.column_config.NumberColumn("⚖️ DIF. GANANCIA", format="$ %d")
                        col_config[f"📉 DIFERENCIA (%) [{p_comp} vs {prod_base}]"] = st.column_config.NumberColumn("📉 DIF. GANANCIA (%)", format="%.2f %%")
                        col_config[f"🎯 MARGEN REQ. (%) [{p_comp} -> Venta {prod_base}]"] = st.column_config.NumberColumn("🎯 MARGEN REQ. (%)", format="%.2f %%")

                    def aplicar_estilos_gerenciales(row):
                        estilos = [''] * len(row)
                        es_ganancia = (row['📊 CONCEPTO'] == 'Ganancia')
                        
                        base_style = 'background-color: #f8f9fa; font-weight: bold; color: #0d1b2a;' if es_ganancia else 'color: #333333;'
                        
                        for i, col in enumerate(row.index):
                            cell_style = base_style
                            val = row[col]
                            if "DIFERENCIA" in col and isinstance(val, (int, float)):
                                if val < 0: cell_style += ' color: #E74C3C; font-weight: bold;'
                                elif val > 0: cell_style += ' color: #27AE60; font-weight: bold;'
                                else: cell_style += ' color: #0d1b2a; font-weight: bold;'
                            elif "MARGEN REQ" in col and isinstance(val, (int, float)):
                                cell_style += ' color: #8E44AD; font-weight: bold;' # Púrpura estratégico
                            estilos[i] = cell_style
                        return estilos
                    st.dataframe(
                        df_excel_export.style.apply(aplicar_estilos_gerenciales, axis=1), 
                        use_container_width=True, 
                        hide_index=True, 
                        column_config=col_config
                    )
                    
                    # Botón de Descarga Excel
                    excel_data = generar_excel_gerencial(df_excel_export, dosis_dict)
                    st.download_button(
                        label="📥 DESCARGAR MODELO GERENCIAL (EXCEL VIP)", 
                        data=excel_data, 
                        file_name=f"Comparativo_Vertical_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                        use_container_width=True,
                        type="primary"
                    )

            with t4:
                st.markdown("#### Búsqueda Rápida de Costos y Márgenes")
                opciones_productos_t4 = df_t["PRODUCTO"].tolist()
                prods_sel_t4 = st.multiselect(
                    "🔍 Seleccione uno o varios Productos para consultar rápidamente:", 
                    options=opciones_productos_t4,
                    default=[opciones_productos_t4[0]] if opciones_productos_t4 else [],
                    key="multiselect_tab4_rapida"
                )
                
                for prod_sel in prods_sel_t4:
                    datos_prod = df_t[df_t["PRODUCTO"] == prod_sel].iloc[0]
                    
                    st.markdown(f"#### 🧪 Arsenal: `{prod_sel}`")
                    c1, c2, c3, c4, c5 = st.columns(5)
                    caja_titulo = "height: 45px; display: flex; align-items: flex-end; margin-bottom: 5px;"
                    estilo_etiqueta = "font-size: 11px; font-weight: 900; color: #0d1b2a; margin: 0; line-height: 1.2;"
                    
                    with c1: 
                        st.markdown(f"<div style='{caja_titulo}'><p style='{estilo_etiqueta}'>🏷️ COSTO BASE</p></div>", unsafe_allow_html=True)
                        st.code(fmt_sap(datos_prod[cols_dinamicas[0]]))
                    with c2: 
                        st.markdown(f"<div style='{caja_titulo}'><p style='{estilo_etiqueta}'>🌱 ORGÁNICO<br></p></div>", unsafe_allow_html=True)
                        st.code(fmt_sap(datos_prod[cols_dinamicas[4]]))
                    with c3: 
                        st.markdown(f"<div style='{caja_titulo}'><p style='{estilo_etiqueta}'>🤝 SOCIO/COOP<br></p></div>", unsafe_allow_html=True)
                        st.code(fmt_sap(datos_prod[cols_dinamicas[3]]))
                    with c4: 
                        st.markdown(f"<div style='{caja_titulo}'><p style='{estilo_etiqueta}'>🏢 AFILIADO<br></p></div>", unsafe_allow_html=True)
                        st.code(fmt_sap(datos_prod[cols_dinamicas[2]]))
                    with c5: 
                        st.markdown(f"<div style='{caja_titulo}'><p style='{estilo_etiqueta}'>👤 TERCERO<br></p></div>", unsafe_allow_html=True)
                        st.code(fmt_sap(datos_prod[cols_dinamicas[1]]))
                    st.markdown("<hr style='border:1px dashed #d4af37; margin-top:5px; margin-bottom:20px;'/>", unsafe_allow_html=True)

        else:
            st.warning("⚠️ No se detectaron datos en el tarifario. Haga clic en 'Recargar Tarifario' para sincronizar con la nube.")

    # --- 🚀 SECCIÓN INFERIOR: OMEGA V12 ---
    st.markdown("---")
    st.markdown("### 🚀 Sincronización Automática a la Macro (Omega V12)")
    
    with st.container(border=True):
        c_url1, c_url2 = st.columns(2)
        with c_url1:
            st.text_input("🔗 1. Base de Origen Activa:", value="DATABASE: Supabase Cloud / Drive Fallback [PRECIOS_INSUMOS]", disabled=True)
        with c_url2:
            url_dest = st.text_input("🎯 2. URL de Sábana Destino (Google Sheets):", placeholder="Pegue el enlace completo aquí...")
        
        semana_target = st.number_input("🔢 Digite la Semana a actualizar (1 a 53):", min_value=1, max_value=53, value=24, step=1)
        
        c_btn1, c_btn2 = st.columns(2)
        
        dict_precios = {}
        if 'df_tarifario' in st.session_state and not st.session_state['df_tarifario'].empty:
            df_m = st.session_state['df_tarifario']
            dict_precios = dict(zip(df_m['PRODUCTO'], df_m['COSTO BASE']))

        with c_btn1:
            if st.button("📊 PREVISUALIZAR COMPORTAMIENTO DE PRECIOS POR DOSIS", use_container_width=True, type="secondary"):
                if gc is None:
                    st.error("🚨 Enlace satelital roto con Google Cloud.")
                elif not url_dest or "http" not in url_dest:
                    st.error("❌ Ingrese la URL de destino para previsualizar.")
                elif not dict_precios:
                    st.error("⚠️ El tarifario no está cargado. Cargue el tarifario antes de previsualizar.")
                else:
                    try:
                        with st.spinner("🕵️‍♂️ Calculando Comportamiento Operativo desde la RAM..."):
                            sh_dest = gc.open_by_url(url_dest)
                            ws_datos = sh_dest.worksheet("DATOS")
                            datos_dest = ws_datos.get_all_values(value_render_option='UNFORMATTED_VALUE')
                            
                            idx_fila_semanas = 6
                            for idx, r in enumerate(datos_dest[:12]):
                                r_str = [str(cell).strip().split('.')[0] for cell in r]
                                if any(w in r_str for w in ["11", "12", "13", "18"]):
                                    idx_fila_semanas = idx
                                    break
                            
                            filas_comp = []
                            for r_idx, row in enumerate(datos_dest):
                                n_fila = r_idx + 1
                                if n_fila < (idx_fila_semanas + 2): continue
                                
                                row_padded = row + [""] * (15 - len(row)) if len(row) < 15 else row
                                tipo_tabla = limpiar_texto_vba(row_padded[1]).upper().strip() 
                                producto_dest = limpiar_texto_vba(row_padded[3]).upper().strip()
                                
                                if producto_dest in dict_precios:
                                    precio_pleno = dict_precios[producto_dest]
                                    dosis_valor = extraer_numero(row_padded[0])
                                    
                                    if "DOSIS-HA" in tipo_tabla.replace(" ", ""):
                                        valor_dosis = precio_pleno * dosis_valor if dosis_valor > 0 else 0
                                        formula = f"{dosis_valor} Dosis × $ {precio_pleno:,.0f}"
                                    else:
                                        valor_dosis = precio_pleno
                                        formula = "Precio Unitario Directo"
                                        
                                    filas_comp.append({
                                        "Fila SAP": n_fila,
                                        "Tipo de Registro": tipo_tabla,
                                        "Insumo / Producto": producto_dest,
                                        "Precio Final Calculado": valor_dosis,
                                        "Lógica de Impacto": formula
                                    })
                            
                            if filas_comp:
                                df_vis = pd.DataFrame(filas_comp).copy()
                                df_vis["Precio Final Calculado"] = df_vis["Precio Final Calculado"].map("$ {:,.0f}".format).str.replace(",", ".")
                                st.dataframe(df_vis, use_container_width=True, hide_index=True)
                            else:
                                st.warning("⚠️ No se encontraron coincidencias de insumos en la hoja DATOS.")
                    except Exception as e:
                        st.error(f"🚨 Falla en análisis: {e}")

        with c_btn2:
            if st.button("🚀 EJECUTAR SINCRONIZACIÓN OMEGA V12", use_container_width=True, type="primary"):
                if gc is None:
                    st.error("🚨 Enlace satelital roto con Google Cloud.")
                    return
                elif not url_dest or "http" not in url_dest:
                    st.error("❌ Digite una URL válida de destino.")
                    return
                elif not dict_precios:
                    st.error("⚠️ Tarifario no disponible en memoria.")
                    return
                try:
                    with st.status("🕵️‍♂️ CONECTANDO CON CÉLULA DE PRECIOS Y DESTINO...", expanded=True) as status:
                        sh_dest = gc.open_by_url(url_dest)
                        ws_datos = sh_dest.worksheet("DATOS")
                        datos_dest = ws_datos.get_all_values(value_render_option='UNFORMATTED_VALUE')
                        
                        idx_fila_semanas = 6
                        for idx, r in enumerate(datos_dest[:12]):
                            r_str = [str(cell).strip().split('.')[0] for cell in r]
                            if any(w in r_str for w in ["11", "12", "13", "18"]):
                                idx_fila_semanas = idx
                                break
                        
                        col_semana = -1
                        for i, v in enumerate(datos_dest[idx_fila_semanas]):
                            if str(v).strip().split('.')[0] == str(semana_target):
                                col_semana = i + 1
                                break
                        
                        if col_semana == -1: col_semana = int(semana_target) + 5
                        
                        updates = [{'range': gspread.utils.rowcol_to_a1(idx_fila_semanas + 1, col_semana), 'values': [[int(semana_target)]]}]
                        
                        for r_idx, row in enumerate(datos_dest):
                            n_fila = r_idx + 1
                            if n_fila < (idx_fila_semanas + 2): continue
                            
                            row_padded = row + [""] * (max(col_semana + 2, 15) - len(row)) if len(row) < max(col_semana + 2, 15) else row
                            tipo_tabla = limpiar_texto_vba(row_padded[1]).upper().strip() 
                            producto_dest = limpiar_texto_vba(row_padded[3]).upper().strip()
                            
                            if producto_dest in dict_precios:
                                precio_unitario = dict_precios[producto_dest]
                                if "DOSIS-HA" in tipo_tabla.replace(" ", ""):
                                    dosis_valor = extraer_numero(row_padded[0])
                                    valor_final = precio_unitario * dosis_valor if dosis_valor > 0 else 0
                                else:
                                    valor_final = precio_unitario
                                    
                                updates.append({'range': gspread.utils.rowcol_to_a1(n_fila, col_semana), 'values': [[valor_final]]})

                        if len(updates) > 1:
                            ws_datos.batch_update(updates, value_input_option='USER_ENTERED')
                            status.update(label="🎯 ¡MÓDULO DE DOSIS AJUSTADO!", state="complete")
                            st.success(f"🎉 Precios impactados en la columna {col_semana} de la macro para la semana {semana_target}.")
                            st.balloons()
                        else:
                            st.warning("⚠️ No se generaron actualizaciones de precios.")
                except Exception as e:
                    st.error(f"🚨 FALLA EN LA INYECCIÓN: {e}")

if __name__ == "__main__":
    pass
