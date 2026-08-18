import streamlit as st
import pandas as pd
import gspread
import re
import math
import io
import openpyxl
from datetime import datetime, timedelta
import streamlit.components.v1 as components
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =================================================================
# 🔌 CONEXIÓN GOOGLE CLOUD Y MOTORES DE EXTRACCIÓN
# =================================================================

@st.cache_resource(show_spinner=False)
def inicializar_cliente_gspread():
    try:
        if "gcp_service_account" in st.secrets:
            return gspread.service_account_from_dict(dict(st.secrets["gcp_service_account"]))
        return gspread.service_account(filename='credenciales.json')
    except Exception: return None

# 💥 URL MAESTRA DEL COMANDANTE
URL_BASE_AFOROS = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

@st.cache_data(show_spinner=False, ttl=300)
def extraer_tablas_aforo():
    gc = inicializar_cliente_gspread()
    if not gc: return pd.DataFrame(), "Sin conexión a Google Cloud"
    try:
        sh = gc.open_by_url(URL_BASE_AFOROS)
        ws = sh.worksheet("TABLAS_AFORO")
        datos = ws.get_all_records()
        df_aforo = pd.DataFrame(datos)
        
        df_aforo.columns = [str(c).strip().upper() for c in df_aforo.columns]
        
        if 'PISTA' in df_aforo.columns and 'TANQUE' in df_aforo.columns:
            df_aforo['PISTA'] = df_aforo['PISTA'].astype(str).str.strip().str.upper()
            df_aforo['TANQUE'] = df_aforo['TANQUE'].astype(str).str.strip().str.upper()
            df_aforo = df_aforo[(df_aforo['PISTA'] != '') & (df_aforo['TANQUE'] != '')]
        
        def purificador_numeros(x):
            if pd.isna(x) or x is None: return 0.0
            if isinstance(x, (int, float)): return float(x)
            s = str(x).strip().replace(' ', '').replace("'", "")
            if not s or s.lower() in ['nan', 'none', '']: return 0.0
            if '.' in s and ',' in s:
                if s.rfind(',') > s.rfind('.'): s = s.replace('.', '').replace(',', '.')
                else: s = s.replace(',', '')
            elif ',' in s: s = s.replace(',', '.')
            try: return float(s)
            except: return 0.0

        if 'VOLUMEN_GAL' in df_aforo.columns: df_aforo['VOLUMEN_GAL'] = df_aforo['VOLUMEN_GAL'].apply(purificador_numeros)
        if 'INCREMENTO_MM' in df_aforo.columns:
            df_aforo['INCREMENTO_MM'] = df_aforo['INCREMENTO_MM'].apply(purificador_numeros)
            df_aforo['INCREMENTO_MM'] = df_aforo['INCREMENTO_MM'].apply(lambda x: x / 1000 if x > 10 else x)
        if 'CM' in df_aforo.columns: df_aforo['CM'] = pd.to_numeric(df_aforo['CM'], errors='coerce').fillna(0).astype(int)
            
        return df_aforo, None
    except Exception as e:
        return pd.DataFrame(), f"Error al extraer aforos: {str(e)}"

# 💥 NUEVO MOTOR: EXTRACCIÓN DE CONSUMOS TEÓRICOS (TABLA 1)
@st.cache_data(show_spinner=False, ttl=300)
def extraer_consumos_teoricos():
    gc = inicializar_cliente_gspread()
    if not gc: return pd.DataFrame(), pd.DataFrame(), "Sin conexión a Google Cloud"
    try:
        sh = gc.open_by_url(URL_BASE_AFOROS)
        hojas = [w.title for w in sh.worksheets()]
        nombre_hoja = next((h for h in hojas if "TABLA 1" in h.upper()), None)
        if not nombre_hoja: return pd.DataFrame(), pd.DataFrame(), "No se encontró la pestaña TABLA 1 en el Excel Maestro."
        
        ws = sh.worksheet(nombre_hoja)
        datos = ws.get_all_values()
        
        idx_head = next((i for i, row in enumerate(datos[:15]) if any("COCTEL" in str(x).upper() for x in row)), 4)
        
        # 💥 LIMPIEZA DE ESPACIOS PARA ENCONTRAR COLUMNAS EXACTAS
        encabezados = [re.sub(r'\s+', ' ', str(x).strip().upper()) for x in datos[idx_head]]
        df_vuelos = pd.DataFrame(datos[idx_head+1:], columns=encabezados)
        
        c_finca = next((c for c in df_vuelos.columns if "FINCA" in c), None)
        c_area = next((c for c in df_vuelos.columns if "FUMIG" in c), None)
        c_coctel = next((c for c in df_vuelos.columns if "COCTEL" in c), None)
        c_sem = next((c for c in df_vuelos.columns if c in ["SEM", "SEMANA"]), None)
        c_pista_raw = next((c for c in df_vuelos.columns if c == "PISTA"), None)
        
        if not all([c_finca, c_area, c_coctel, c_sem, c_pista_raw]):
            faltantes = [n for n, v in zip(["FINCA", "ÁREA FUMIG", "COCTEL", "SEM", "PISTA"], [c_finca, c_area, c_coctel, c_sem, c_pista_raw]) if v is None]
            return pd.DataFrame(), pd.DataFrame(), f"Faltan estas columnas exactas en la fila de encabezados de TABLA 1: {faltantes}"

        # 🎯 REGLA TÁCTICA 1: Extracción del PRIMER DÍGITO absoluto (Galones)
        def extraer_dosis(coctel):
            coctel_str = str(coctel).strip()
            match = re.search(r'\d', coctel_str)
            if match: return float(match.group())
            return 0.0
        
        # 🎯 REGLA TÁCTICA 2: Mapeo de Pistas y Excepción Lucila Marina
        def mapear_pista_oficial(row):
            p = str(row[c_pista_raw]).strip().upper()
            f = str(row[c_finca]).strip().upper()
            if 'GENESYS' in p:
                if 'LUCILA MARINA' in f: return 'LUCI'
                return 'PLUC' 
            if 'FUMIGARAY' in p: return 'PLUC'
            if 'AEROPENORT' in p: return 'PORI'
            if 'ASA' in p: return 'PDIV'
            return p

        df_vuelos['DOSIS_GAL'] = df_vuelos[c_coctel].apply(extraer_dosis)
        df_vuelos['PISTA_OFICIAL'] = df_vuelos.apply(mapear_pista_oficial, axis=1)
        
        def purificar_area(x):
            s = str(x).strip().replace(' ', '').replace(',', '.')
            try: return float(s)
            except: return 0.0
            
        df_vuelos['AREA_LIMPIA'] = df_vuelos[c_area].apply(purificar_area)
        
        # 💥 CALCULAR EL CONSUMO TEÓRICO (DE GALONES A LITROS)
        df_vuelos['CONSUMO_TEORICO_GAL'] = df_vuelos['AREA_LIMPIA'] * df_vuelos['DOSIS_GAL']
        df_vuelos['CONSUMO_TEORICO_L'] = df_vuelos['CONSUMO_TEORICO_GAL'] * 3.78541
        
        # 💥 BLINDAJE DE LA SEMANA
        df_vuelos['SEMANA_LIMPIA'] = df_vuelos[c_sem].astype(str).str.replace('.0', '', regex=False).str.strip()
        
        df_teorico = df_vuelos.groupby(['SEMANA_LIMPIA', 'PISTA_OFICIAL'], as_index=False)['CONSUMO_TEORICO_L'].sum()
        df_teorico.columns = ['SEMANA', 'PISTA', 'LITROS_TEORICOS']
        
        # 💥 GENERAR TABLA DE AUDITORÍA DETALLADA
        df_auditoria = df_vuelos[[c_finca, c_coctel, 'DOSIS_GAL', 'AREA_LIMPIA', 'CONSUMO_TEORICO_GAL', 'CONSUMO_TEORICO_L', 'SEMANA_LIMPIA', 'PISTA_OFICIAL']].copy()
        df_auditoria.columns = ['FINCA', 'COCTEL', 'DOSIS (Gal)', 'ÁREA (Ha)', 'TOTAL GALONES', 'TOTAL LITROS', 'SEMANA', 'PISTA']
        
        return df_teorico, df_auditoria, None
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), f"Error al extraer vuelos de TABLA 1: {str(e)}"

# =================================================================
# ⚡ COMPILADORES DE ESTRUCTURA Y ESTILOS 
# =================================================================
def compilar_excel_maestro(cruce_final, semana):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        f_df = cruce_final.drop(columns=['LOTE_KEY'], errors='ignore')
        f_df[f_df['ESTADO'] == "❌ DISCREPANCIA"].to_excel(writer, index=False, sheet_name='Diferencias')
        f_df.to_excel(writer, index=False, sheet_name='Total')
        borde_fino = Border(left=Side(style='thin', color='D1D1D1'), right=Side(style='thin', color='D1D1D1'), top=Side(style='thin', color='D1D1D1'), bottom=Side(style='thin', color='D1D1D1'))
        fondo_navy = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        texto_blanco = Font(color="FFFFFF", bold=True, name="Arial", size=11)
        for sheetname in writer.sheets:
            worksheet = writer.sheets[sheetname]
            worksheet.auto_filter.ref = worksheet.dimensions 
            for r_idx in range(2, worksheet.max_row + 1):
                worksheet.cell(row=r_idx, column=7).value = f"=F{r_idx}-E{r_idx}"
                worksheet.cell(row=r_idx, column=8).value = f'=IF(ABS(G{r_idx})<=0.05, "✅ OK", "❌ DISCREPANCIA")'
                worksheet.cell(row=r_idx, column=5).number_format = '0.000'
                worksheet.cell(row=r_idx, column=6).number_format = '0.000'
                worksheet.cell(row=r_idx, column=7).number_format = '0.000'
            for row_cells in worksheet.iter_rows():
                for cell in row_cells:
                    cell.border = borde_fino
                    if cell.row == 1:
                        cell.fill = fondo_navy; cell.font = texto_blanco; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif cell.column in [5, 6, 7]: cell.alignment = Alignment(horizontal='right')
                    elif cell.column == 8: cell.alignment = Alignment(horizontal='center')
            for col in worksheet.columns:
                max_len = max(len(str(c.value or '')) for c in col)
                worksheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 4, 12), 42)
    return buffer.getvalue()

def compilar_html_pdf(cruce_final, semana, css_vip):
    pistas = sorted(cruce_final['PISTA'].unique())
    html_out = f"""<html><head><script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script><script>function imprimir() {{ window.print(); }} function descargarPDF() {{ var element = document.getElementById('contenido-reporte'); var opt = {{ margin: [10, 10, 10, 10], filename: 'Reporte_Arqueo_Semana_{semana}.pdf', image: {{ type: 'jpeg', quality: 0.98 }}, html2canvas: {{ scale: 2, useCORS: true }}, jsPDF: {{ unit: 'mm', format: 'a4', orientation: 'landscape' }}, pagebreak: {{ mode: ['css', 'legacy'] }} }}; html2pdf().set(opt).from(element).save(); }}</script>{css_vip}</head><body><div class="no-print" style="position: sticky; top: 0; background: white; padding: 10px; z-index: 100; border-bottom: 2px solid #0d1b2a; text-align: right;"><button onclick="descargarPDF()" style="background:#0d1b2a; color:#d4af37; border:2px solid #d4af37; padding:10px 20px; cursor:pointer; border-radius:6px; font-weight:bold; font-family:'Arial Black'; margin-right: 10px;">📥 DESCARGAR PDF DIRECTO</button><button onclick="imprimir()" style="background:#4a5568; color:white; border:2px solid #4a5568; padding:10px 20px; cursor:pointer; border-radius:6px; font-weight:bold; font-family:'Arial Black';">🖨️ PANEL DE IMPRESIÓN</button></div><div id="contenido-reporte">"""
    for i, pista in enumerate(pistas):
        df_pista = cruce_final[cruce_final['PISTA'] == pista]
        salto = "salto-pagina" if i < len(pistas) - 1 else ""
        html_out += f"""<div class="b-print {salto}"><p class="title">REPORTE OFICIAL DE ARQUEO DE INVENTARIOS</p><p class="subtitle">BASE OPERATIVA: {pista} | SEMANA AUDITADA: {semana}</p><table><tr><th style="width:10%;">CÓDIGO</th><th style="width:30%;">PRODUCTO</th><th style="width:15%;">LOTE</th><th style="width:10%;">S. SAP</th><th style="width:10%;">S. FÍSICO</th><th style="width:10%;">DIF.</th><th style="width:15%;">ESTADO</th></tr>"""
        for _, row in df_pista.iterrows():
            st_color = "#155724" if "OK" in str(row['ESTADO']) else "#721c24"
            bg_color = "#d4edda" if "OK" in str(row['ESTADO']) else "#f8d7da"
            val_dif = f"+{row['DIFERENCIA']:.3f}" if row['DIFERENCIA'] > 0 else f"{row['DIFERENCIA']:.3f}"
            html_out += f"<tr><td>{row['ITEM']}</td><td class='td-left'>{row['PRODUCTO']}</td><td>{row['LOTE']}</td><td>{row['SALDO_SAP']:.3f}</td><td>{row['SALDO_FISICO']:.3f}</td><td style='color:{st_color};'><b>{val_dif}</b></td><td style='color:{st_color}; background-color:{bg_color}; font-weight:bold;'>{row['ESTADO']}</td></tr>"
        html_out += """</table><div class='firmas-container'><div class='firma-box'>FIRMA SUPERVISOR DE PISTA</div><div class='firma-box'>FIRMA AUDITOR DE INVENTARIOS</div></div></div>"""
    html_out += "</div></body></html>"
    return html_out

def obtener_hora_colombia():
    return datetime.utcnow() + timedelta(hours=-5)

# =================================================================
# 🛢️ RENDERIZADOR DEL RADAR DE PLOMADAS Y CONCILIACIÓN
# =================================================================
def renderizar_radar_plomadas():
    
    with st.spinner("Sincronizando Bóveda de Aforos y Tabla de Vuelos (Teórico)..."):
        df_aforo, error_aforo = extraer_tablas_aforo()
        df_teorico, df_auditoria, error_teorico = extraer_consumos_teoricos()
        
    if error_aforo:
        st.error(f"🚨 No se pudo conectar a la Bóveda de Aforos. Detalle: {error_aforo}")
        return
        
    if error_teorico:
        st.error(f"🚨 Error en la lectura de TABLA 1: {error_teorico}")
        
    col_radar, col_cruce = st.columns([1, 1], gap="large")
    
    with col_radar:
        st.markdown("<h3 style='color: #0d1b2a; border-bottom: 2px solid #d4af37;'>📍 1. Medición de Plomada (Físico)</h3>", unsafe_allow_html=True)
        
        pistas_disponibles = sorted(df_aforo['PISTA'].dropna().unique().tolist())
        
        c_f, c_p = st.columns(2)
        fecha_plomada = c_f.date_input("🗓️ Fecha de Medición", value=obtener_hora_colombia().date())
        semana_calculada = str(fecha_plomada.isocalendar()[1]).strip()
        
        pista_sel = c_p.selectbox("📍 Pista", pistas_disponibles)
        
        tanques_disponibles = sorted(df_aforo[df_aforo['PISTA'] == pista_sel]['TANQUE'].dropna().unique().tolist())
        tanque_sel = st.selectbox("🛢️ Tanque a Medir", tanques_disponibles)
        
        st.markdown("---")
        df_tanque_especifico = df_aforo[(df_aforo['PISTA'] == pista_sel) & (df_aforo['TANQUE'] == str(tanque_sel))]
        max_cm = int(df_tanque_especifico['CM'].max()) if not df_tanque_especifico.empty else 200
        
        c_cm, c_mm = st.columns(2)
        cm_input = c_cm.number_input("📏 CM de la cinta:", min_value=0, max_value=max_cm, step=1, value=0)
        mm_input = c_mm.number_input("🤏 MM extra:", min_value=0, max_value=9, step=1, value=0)
        
        litros_totales_actuales = 0.0
        
        if not df_tanque_especifico.empty:
            fila_medida = df_tanque_especifico[df_tanque_especifico['CM'] == cm_input]
            if not fila_medida.empty:
                vol_gal_base = float(fila_medida['VOLUMEN_GAL'].values[0])
                inc_mm = float(fila_medida['INCREMENTO_MM'].values[0])
                galones_totales = vol_gal_base + (mm_input * inc_mm)
                litros_totales_actuales = galones_totales * 3.78541
                
                k1, k2 = st.columns(2)
                k1.markdown(f"<div class='hud-arqueo' style='padding: 10px;'><div class='hud-arqueo-item'><p class='hud-arqueo-title'>GALONES</p><p class='hud-arqueo-value'>{galones_totales:,.2f}</p></div></div>", unsafe_allow_html=True)
                k2.markdown(f"<div class='hud-arqueo' style='padding: 10px;'><div class='hud-arqueo-item'><p class='hud-arqueo-title'>LITROS</p><p class='hud-arqueo-value hud-arqueo-ok'>{litros_totales_actuales:,.2f}</p></div></div>", unsafe_allow_html=True)
                
                btn_registrar_plomada = st.button("💾 GUARDAR LECTURA FÍSICA EN BÓVEDA", type="secondary", use_container_width=True)
                if btn_registrar_plomada:
                    with st.spinner("Guardando en REGISTRO_PLOMADAS..."):
                        try:
                            gc = inicializar_cliente_gspread()
                            sh = gc.open_by_url(URL_BASE_AFOROS)
                            try: ws_reg = sh.worksheet("REGISTRO_PLOMADAS")
                            except:
                                ws_reg = sh.add_worksheet(title="REGISTRO_PLOMADAS", rows="100", cols="9")
                                ws_reg.append_row(["FECHA", "SEMANA", "PISTA", "TANQUE", "CM", "MM", "GALONES", "LITROS", "USUARIO"])
                            
                            usuario_actual = st.session_state.get('usuario_nombre', 'Comandante')
                            ws_reg.append_row([fecha_plomada.strftime("%d/%m/%Y"), semana_calculada, pista_sel, tanque_sel, int(cm_input), int(mm_input), f"{galones_totales:.3f}".replace('.', ','), f"{litros_totales_actuales:.3f}".replace('.', ','), usuario_actual])
                            st.success(f"✅ ¡Guardado!")
                        except Exception as e: st.error(f"🚨 Error: {e}")
            else: st.error(f"🚨 El tanque no tiene un registro para {cm_input} cm.")

    # 💥 PANEL DE CONCILIACIÓN
    with col_cruce:
        st.markdown("<h3 style='color: #d4af37; border-bottom: 2px solid #0d1b2a;'>⚖️ 2. Cruce de Consumo Semanal</h3>", unsafe_allow_html=True)
        
        if df_teorico.empty:
            st.warning("⚠️ No hay datos teóricos procesados. Revisa la TABLA 1.")
        else:
            df_teorico_filtrado = df_teorico[(df_teorico['SEMANA'] == semana_calculada) & (df_teorico['PISTA'] == pista_sel)]
            
            litros_teoricos_sap = 0.0
            if not df_teorico_filtrado.empty:
                litros_teoricos_sap = df_teorico_filtrado['LITROS_TEORICOS'].sum()
                
            st.info(f"**Semana Auditada:** {semana_calculada} | **Base:** {pista_sel}")
            st.markdown(f"#### ✈️ Consumo Teórico (TABLA 1): `{litros_teoricos_sap:,.2f} L`")
            st.caption("Fórmula: (Área x 1er Dígito Cóctel) = Galones ➔ x 3.78541 = Litros Reales.")
            
            if litros_teoricos_sap == 0:
                semanas_bd = df_teorico['SEMANA'].unique().tolist()
                st.warning("⚠️ Diagnóstico del Escáner Teórico:")
                if semana_calculada in semanas_bd:
                    pistas_sem = df_teorico[df_teorico['SEMANA'] == semana_calculada]['PISTA'].unique().tolist()
                    st.caption(f"- ✅ La semana **{semana_calculada}** sí existe en la TABLA 1.")
                    st.caption(f"- 🚨 Pero en esa semana solo se registraron vuelos en estas pistas: **{pistas_sem}**.")
                else:
                    st.caption(f"- 🚨 La semana **{semana_calculada}** NO tiene ningún vuelo registrado en toda la TABLA 1.")
            else:
                # 💥 LA TABLA DE LA VERDAD (Desglosa las filas y verifica la multiplicación por 3.78541)
                with st.expander("🔬 Ver vuelos sumados por el sistema (Auditoría)"):
                    vuelos_auditoria = df_auditoria[(df_auditoria['SEMANA'] == semana_calculada) & (df_auditoria['PISTA'] == pista_sel)]
                    vuelos_auditoria = vuelos_auditoria[vuelos_auditoria['TOTAL LITROS'] > 0]
                    st.caption(f"El sistema sumó estos **{len(vuelos_auditoria)} vuelos** (incluso los ocultos por filtros en Excel):")
                    st.dataframe(vuelos_auditoria.drop(columns=['SEMANA', 'PISTA']), use_container_width=True, hide_index=True)
            
            st.markdown("---")
            st.markdown("**🔍 Datos para el Consumo Físico:**")
            col_in1, col_in2 = st.columns(2)
            litros_iniciales = col_in1.number_input("📦 Litros Iniciales (Apertura):", min_value=0.0, value=0.0, step=10.0)
            litros_ingresos = col_in2.number_input("📥 Litros Ingresados (Nuevos):", min_value=0.0, value=0.0, step=10.0)
            
            litros_finales = st.number_input("🏁 Litros Finales (Plomada Actual):", min_value=0.0, value=float(litros_totales_actuales), step=10.0)
            
            consumo_fisico = (litros_iniciales + litros_ingresos) - litros_finales
            merma_litros = consumo_fisico - litros_teoricos_sap
            
            st.markdown("---")
            st.markdown(f"#### 📏 Consumo Físico Real: `{consumo_fisico:,.2f} L`")
            
            color_merma = "#00ff66" if abs(merma_litros) <= 5.0 else "#ff3333"
            icono_merma = "✅ OK" if abs(merma_litros) <= 5.0 else "🚨 DESCUADRE CRÍTICO"
            
            st.markdown(f"""
            <div style="background-color: #0d1b2a; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid {color_merma};">
                <p style="color: white; font-weight: bold; margin: 0;">DESCUADRE (Físico vs SAP):</p>
                <h2 style="color: {color_merma}; margin: 5px 0;">{merma_litros:,.2f} L</h2>
                <p style="color: {color_merma}; font-weight: bold; margin: 0;">{icono_merma}</p>
            </div>
            """, unsafe_allow_html=True)

# =================================================================
# 👑 INTERFAZ PRINCIPAL 
# =================================================================

def ejecutar(quitar_tildes, purificar_lote):
    st.markdown("""
    <style>
    .titulo-principal { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black', sans-serif; }
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] { border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; }
    div[data-testid="stTextInput"] input, div[data-testid="stDateInput"] input, div[data-testid="stSelectbox"] > div, div[data-testid="stSelectbox"] div[data-baseweb="select"] { border: 2px solid #0d1b2a !important; border-radius: 6px !important; background-color: #ffffff !important; color: #0d1b2a !important; font-weight: 800 !important; font-size: 15px !important; }
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div { background-color: transparent !important; border: none !important; }
    div[data-testid="stSelectbox"] * { color: #000000 !important; font-weight: bold !important; }
    div[data-testid="stFileUploader"] section { background-color: #ffffff !important; border: 2px dashed #0d1b2a !important; border-radius: 8px !important; padding: 10px !important; }
    .hud-arqueo { background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); border-left: 5px solid #d4af37; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.15); margin-bottom: 25px; display: flex; justify-content: space-between; align-items: center; }
    .hud-arqueo-item { text-align: center; flex: 1; }
    .hud-arqueo-title { font-size: 11px; font-weight: bold; color: #d4af37; text-transform: uppercase; margin:0; letter-spacing: 1px; }
    .hud-arqueo-value { font-size: 22px; font-family: 'Arial Black'; margin: 5px 0 0 0; }
    .hud-arqueo-ok { color: #00ff66; font-family: 'Arial Black'; }
    .hud-arqueo-fail { color: #ff3333; font-family: 'Arial Black'; }
    div[data-testid="stTabs"] button[role="tab"] { font-family: 'Arial Black', sans-serif; font-size: 14px; text-transform: uppercase; color: #0d1b2a; }
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { border-bottom-color: #d4af37; background-color: rgba(212, 175, 55, 0.1); }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-principal'>⚖️ 7. Arqueo de Inventarios y Plomadas</h1>", unsafe_allow_html=True)
    
    tab_arqueo, tab_plomada = st.tabs(["📊 1. CRUCE SAP VS FÍSICO (Auditoría)", "🛢️ 2. RADAR DE PLOMADAS (Aforo)"])
    
    with tab_plomada:
        renderizar_radar_plomadas()

    with tab_arqueo:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("### 📁 1. Sábana SAP")
            archivo_sap = st.file_uploader("1️⃣ Sábana de SAP", type=['xlsx', 'csv'])
        with c2:
            st.markdown("### 📋 2. Reportes Físicos")
            archivos_sup = st.file_uploader("2️⃣ Reportes Supervisores (.xlsx)", type=['xlsx'], accept_multiple_files=True)
        with c3:
            st.markdown("### 🎯 3. Objetivo")
            semana_obj = st.text_input("Semana a Auditar (Ej: 29):", placeholder="Escriba la semana aquí...")

        if "arqueo_procesado" not in st.session_state: st.session_state.arqueo_procesado = False
        if "observaciones_memoria" not in st.session_state: st.session_state.observaciones_memoria = {}
        if "historial_fusiones" not in st.session_state: st.session_state.historial_fusiones = []
        if "centro_pdf_activo" not in st.session_state: st.session_state.centro_pdf_activo = False

        def limpiar_numeros_generico(x):
            if pd.isna(x) or x is None: return 0.0
            if isinstance(x, (int, float)): return float(x)
            s = str(x).strip().replace(' ', '')
            if not s or s.lower() in ['nan', 'none', '']: return 0.0
            if '.' in s and ',' in s:
                if s.rfind(',') > s.rfind('.'): s = s.replace('.', '').replace(',', '.')
                else: s = s.replace(',', '')
            elif ',' in s: s = s.replace(',', '.')
            try: return float(s)
            except: return 0.0

        def generar_cruce():
            cruce = pd.merge(st.session_state.df_sap_grouped, st.session_state.df_sup_grouped, on=['PISTA', 'LOTE_KEY'], how='outer')
            cruce['ITEM'] = cruce['ITEM'].fillna("---")
            cruce['PRODUCTO'] = cruce['PRODUCTO'].fillna(cruce['PRODUCTO_SUP']).fillna("N/A")
            cruce['LOTE'] = cruce['LOTE'].fillna(cruce['LOTE_SUP'])
            cruce['SALDO_SAP'] = cruce['SALDO_SAP'].fillna(0).round(3)
            cruce['SALDO_FISICO'] = cruce['SALDO_FISICO'].fillna(0).round(3)
            cruce = cruce[~((cruce['SALDO_SAP'] == 0) & (cruce['SALDO_FISICO'] == 0))]
            cruce['DIFERENCIA'] = (cruce['SALDO_FISICO'] - cruce['SALDO_SAP']).round(3)
            cruce['ESTADO'] = cruce['DIFERENCIA'].apply(lambda x: "✅ OK" if abs(x) <= 0.05 else "❌ DISCREPANCIA")
            cruce['OBSERVACIONES'] = ""
            
            comentarios_cloud = {}
            if 'supabase' in st.session_state and semana_obj:
                try:
                    resp_obs = st.session_state['supabase'].table("arqueos_observaciones").select("lote_pista_key, observacion").eq("semana", str(semana_obj).strip()).execute()
                    if resp_obs.data: comentarios_cloud = {r["lote_pista_key"]: r["observacion"] for r in resp_obs.data}
                except Exception: pass

            for idx, row in cruce.iterrows():
                key = f"{row['PISTA']}_{row['LOTE_KEY']}"
                if key in st.session_state.observaciones_memoria: cruce.at[idx, 'OBSERVACIONES'] = st.session_state.observaciones_memoria[key]
                elif key in comentarios_cloud:
                    cruce.at[idx, 'OBSERVACIONES'] = comentarios_cloud[key]
                    st.session_state.observaciones_memoria[key] = comentarios_cloud[key]
                elif row['SALDO_SAP'] > 0 and row['SALDO_FISICO'] == 0: cruce.at[idx, 'OBSERVACIONES'] = "SUGERIDO: Entrega / Traslado / Pendiente por Facturar"
                    
            st.session_state.cruce_final = cruce[['PISTA', 'ITEM', 'PRODUCTO', 'LOTE_KEY', 'LOTE', 'SALDO_SAP', 'SALDO_FISICO', 'DIFERENCIA', 'ESTADO', 'OBSERVACIONES']].sort_values(by=['PISTA', 'PRODUCTO']).reset_index(drop=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col_act1, col_act2 = st.columns([2, 1])
        
        with col_act2:
            if st.button("🔄 RESETEAR Y PURGAR MEMORIA", use_container_width=True):
                for k in ["arqueo_procesado", "df_sap_grouped", "df_sup_grouped", "df_sup_grouped_virgen", "cruce_final", "observaciones_memoria", "df_sap_raw", "historial_fusiones"]:
                    if k in st.session_state: del st.session_state[k]
                st.toast("✅ Memoria purgada.", icon="🔄")
                st.rerun()

        with col_act1:
            btn_iniciar = st.button("🚀 INICIAR ARQUEO ESTRATÉGICO", type="primary", use_container_width=True)

        if btn_iniciar:
            if not archivo_sap or not archivos_sup or not semana_obj: st.error("❌ Suministros insuficientes para el cruce maestro.")
            else:
                try:
                    with st.spinner("Desplegando analista algorítmico y escáner anti-filas ocultas..."):
                        st.session_state.observaciones_memoria = {}
                        st.session_state.historial_fusiones = []
                        
                        sap_file = archivo_sap[0] if isinstance(archivo_sap, list) else archivo_sap
                        sap_file.seek(0)
                        if sap_file.name.lower().endswith('.xlsx') or sap_file.name.lower().endswith('.xls'): df_sap = pd.read_excel(sap_file)
                        else:
                            try: df_sap = pd.read_csv(sap_file, sep=None, engine='python', encoding='utf-8')
                            except UnicodeDecodeError: 
                                sap_file.seek(0)
                                df_sap = pd.read_csv(sap_file, sep=None, engine='python', encoding='latin1')

                        columnas_originales = list(df_sap.columns)
                        headers = [quitar_tildes(str(c)).strip().lower() for c in columnas_originales]
                        
                        idx_item, idx_desc, idx_pista, idx_lote, idx_saldo = -1, -1, -1, -1, -1
                        pistas_conocidas = ["TEHO", "PLUC", "PORI", "LUCI", "PDIV", "Z-1", "Z-2"]
                        for i in range(len(df_sap.columns)):
                            vals = df_sap.iloc[:, i].dropna().astype(str).str.upper().head(50).tolist()
                            for v in vals:
                                if v.strip() in pistas_conocidas:
                                    idx_pista = i
                                    break
                            if idx_pista != -1: break

                        for i, h in enumerate(headers):
                            if idx_item == -1 and any(k in h for k in ['material', 'codigo', 'item']): idx_item = i
                            if idx_desc == -1 and any(k in h for k in ['descripcion', 'texto', 'producto', 'nombre']): idx_desc = i
                            if idx_pista == -1 and any(k in h for k in ['almacen', 'alm', 'pista', 'ubicacion']) and 'centro' not in h: idx_pista = i
                            if idx_lote == -1 and 'lote' in h: idx_lote = i
                            if idx_saldo == -1 and any(k in h for k in ['libre', 'saldo', 'cantidad', 'stock']): idx_saldo = i
                            
                        if idx_item == -1: idx_item = 0
                        if idx_desc == -1: idx_desc = 1
                        if idx_pista == -1: idx_pista = 2
                        if idx_lote == -1: idx_lote = 3
                        if idx_saldo == -1: idx_saldo = 4

                        df_sap_clean = df_sap[[columnas_originales[idx_item], columnas_originales[idx_desc], columnas_originales[idx_pista], columnas_originales[idx_lote], columnas_originales[idx_saldo]]].copy()
                        df_sap_clean.columns = ['ITEM', 'PRODUCTO', 'PISTA', 'LOTE', 'SALDO_SAP']
                        df_sap_clean['LOTE_KEY'] = df_sap_clean['LOTE'].apply(purificar_lote)
                        df_sap_clean['PISTA'] = df_sap_clean['PISTA'].astype(str).str.strip().str.upper()
                        df_sap_clean['SALDO_SAP'] = df_sap_clean['SALDO_SAP'].apply(limpiar_numeros_generico)
                        
                        st.session_state.df_sap_raw = df_sap_clean 
                        st.session_state.df_sap_grouped = df_sap_clean.groupby(['PISTA', 'LOTE_KEY'], as_index=False).agg({'ITEM': 'first', 'PRODUCTO': 'first', 'LOTE': 'first', 'SALDO_SAP': 'sum'})

                        lista_sup = []
                        sem_num = str(semana_obj).strip()
                        
                        for file in archivos_sup:
                            file.seek(0)
                            dict_dfs = pd.read_excel(file, sheet_name=None, header=None, dtype=str)
                            file.seek(0)
                            try: wb = openpyxl.load_workbook(file, data_only=True)
                            except: wb = None
                                
                            target_sheet = None
                            for sheet_name in dict_dfs.keys():
                                s_clean = quitar_tildes(str(sheet_name)).upper().strip()
                                if re.search(r'\b' + re.escape(sem_num) + r'\b', s_clean) or s_clean == sem_num:
                                    target_sheet = sheet_name
                                    break 

                            if target_sheet:
                                df_raw = dict_dfs[target_sheet]
                                if wb and target_sheet in wb.sheetnames:
                                    ws = wb[target_sheet]
                                    filas_ocultas = [r - 1 for r, dim in ws.row_dimensions.items() if dim.hidden]
                                    if filas_ocultas: df_raw = df_raw.drop(index=filas_ocultas, errors='ignore').reset_index(drop=True)

                                h_idx = -1
                                for i in range(min(30, len(df_raw))):
                                    row_v = [quitar_tildes(str(x)).upper() for x in df_raw.iloc[i].values if pd.notna(x)]
                                    if any("LOTE" in val for val in row_v) and any("SALDO" in val for val in row_v):
                                        h_idx = i
                                        break
                                        
                                if h_idx != -1:
                                    df_s = df_raw.iloc[h_idx + 1:].copy()
                                    raw_headers = [str(x) for x in df_raw.iloc[h_idx]]
                                    df_s.columns = [f"{quitar_tildes(x)}_{idx}" for idx, x in enumerate(raw_headers)]
                                    
                                    c_p = next((c for c in df_s.columns if any(k in c.upper() for k in ["PRODUC", "DESCRI", "TEXTO", "ARTICULO"])), None)
                                    c_a = next((c for c in df_s.columns if any(k in c.upper() for k in ["ALMAC", "PISTA", "CENTRO", "UBICAC"])), None)
                                    c_l = next((c for c in df_s.columns if "LOTE" in c.upper() and "SALDO" not in c.upper()), None)
                                    
                                    cand_saldos = [c for c in df_s.columns if "SALDO" in c.upper() and not any(ex in c.upper() for ex in ["INIC", "INICIAL", "INGRES", "ENTRA", "SALID"])]
                                    c_v = cand_saldos[-1] if cand_saldos else next((c for c in df_s.columns if "SALDO" in c.upper() and "INIC" not in c.upper()), None)
                                    
                                    if all([c_p, c_a, c_l, c_v]):
                                        df_s_c = df_s[[c_p, c_a, c_l, c_v]].copy()
                                        df_s_c.columns = ['PRODUCTO_SUP', 'PISTA', 'LOTE_SUP', 'SALDO_FISICO']
                                        df_s_c = df_s_c.dropna(subset=['PRODUCTO_SUP', 'LOTE_SUP'])
                                        df_s_c = df_s_c[df_s_c['PRODUCTO_SUP'].astype(str).str.strip() != '']
                                        df_s_c = df_s_c[~df_s_c['PRODUCTO_SUP'].astype(str).str.upper().str.contains("TOTAL|SUMA", regex=True)]
                                        df_s_c['PISTA'] = df_s_c['PISTA'].astype(str).str.strip().str.upper().replace('NAN', None).ffill().bfill()
                                        df_s_c['LOTE_KEY'] = df_s_c['LOTE_SUP'].apply(purificar_lote)
                                        df_s_c['SALDO_FISICO'] = df_s_c['SALDO_FISICO'].apply(limpiar_numeros_generico)
                                        lista_sup.append(df_s_c)

                        if lista_sup:
                            st.session_state.df_sup_grouped_virgen = pd.concat(lista_sup, ignore_index=True).groupby(['PISTA', 'LOTE_KEY'], as_index=False).agg({'PRODUCTO_SUP': 'first', 'LOTE_SUP': 'first', 'SALDO_FISICO': 'sum'})
                            st.session_state.df_sup_grouped = st.session_state.df_sup_grouped_virgen.copy()
                            st.session_state.semana_actual = semana_obj
                            generar_cruce()
                            st.session_state.arqueo_procesado = True
                            
                            if 'supabase' in st.session_state:
                                try:
                                    payload_cruce = [{"semana": str(semana_obj).strip(), "pista": str(r["PISTA"]), "item_codigo": str(r["ITEM"]), "producto": str(r["PRODUCTO"]), "lote": str(r["LOTE"]), "saldo_sap": float(r["SALDO_SAP"]), "saldo_fisico": float(r["SALDO_FISICO"]), "diferencia": float(r["DIFERENCIA"]), "estado": str(r["ESTADO"])} for _, r in st.session_state.cruce_final.iterrows()]
                                    if payload_cruce:
                                        st.session_state['supabase'].table("arqueos_inventario_maestro").delete().eq("semana", str(semana_obj).strip()).execute()
                                        st.session_state['supabase'].table("arqueos_inventario_maestro").insert(payload_cruce).execute()
                                except Exception: pass
                        else: st.error("❌ No se localizaron hojas válidas en los reportes físicos.")
                except Exception as e: st.error(f"🚨 Error estructural: {e}")
                    
        if st.session_state.arqueo_procesado:
            f_df_cruce = st.session_state.cruce_final
            total_sku_arqueados = len(f_df_cruce)
            coincidencias_ok = len(f_df_cruce[f_df_cruce['ESTADO'] == "✅ OK"])
            desfases_criticos = len(f_df_cruce[f_df_cruce['ESTADO'] == "❌ DISCREPANCIA"])
            volumen_desfase_neto = f_df_cruce['DIFERENCIA'].sum()
            fail_class = "hud-arqueo-fail" if desfases_criticos > 0 else "hud-arqueo-ok"
            fail_icon = "⚠️" if desfases_criticos > 0 else "✅"
            balance_color = "#ff3333" if volumen_desfase_neto < 0 else "#00ff66"
            balance_sign = "+" if volumen_desfase_neto > 0 else ""
            
            st.markdown(f"""
             <div class="hud-arqueo">
                 <div class="hud-arqueo-item"><p class="hud-arqueo-title">Lotes Arqueados</p><p class="hud-arqueo-value">⚖️ {total_sku_arqueados} Ítems</p></div>
                 <div class="hud-arqueo-item"><p class="hud-arqueo-title">Cuadrados con SAP</p><p class="hud-arqueo-value hud-arqueo-ok">🟢 {coincidencias_ok} OK</p></div>
                 <div class="hud-arqueo-item"><p class="hud-arqueo-title">Desfases Críticos</p><p class="hud-arqueo-value {fail_class}">{fail_icon} {desfases_criticos} Alarmas</p></div>
                 <div class="hud-arqueo-item"><p class="hud-arqueo-title">Balance Neto Físico</p><p class="hud-arqueo-value" style="color: {balance_color};">{balance_sign}{volumen_desfase_neto:,.2f} L/Kg</p></div>
             </div>
            """, unsafe_allow_html=True)
    
            tab1, tab2, tab3 = st.tabs(["⚠️ Discrepancias", "🛠️ Conciliador", "📋 Inventario Completo"])
            with tab1:
                df_err = st.session_state.cruce_final[st.session_state.cruce_final['ESTADO'] == "❌ DISCREPANCIA"].copy()
                if df_err.empty: st.success("✅ ¡Felicidades Comandante! Todo el arsenal químico se encuentra perfectamente cuadrado con SAP.")
                else:
                    edited_df = st.data_editor(df_err.drop(columns=['LOTE_KEY'], errors='ignore'), use_container_width=True, hide_index=True, disabled=["PISTA", "ITEM", "PRODUCTO", "LOTE", "SALDO_SAP", "SALDO_FISICO", "DIFERENCIA", "ESTADO"], column_config={"SALDO_SAP": st.column_config.NumberColumn("SALDO SAP", format="%.3f"), "SALDO_FISICO": st.column_config.NumberColumn("SALDO FÍSICO", format="%.3f"), "DIFERENCIA": st.column_config.NumberColumn("DIFERENCIA", format="%.3f"), "OBSERVACIONES": st.column_config.TextColumn("📝 OBSERVACIONES (Editable)", width="large")})
                    payload_obs_cloud = []
                    for _, row in edited_df.iterrows():
                        lote_purificado = purificar_lote(row['LOTE'])
                        key = f"{row['PISTA']}_{lote_purificado}"
                        st.session_state.observaciones_memoria[key] = row['OBSERVACIONES']
                        idx_m = st.session_state.cruce_final[(st.session_state.cruce_final['PISTA'] == row['PISTA']) & (st.session_state.cruce_final['LOTE_KEY'] == lote_purificado)].index
                        if not idx_m.empty: st.session_state.cruce_final.at[idx_m[0], 'OBSERVACIONES'] = row['OBSERVACIONES']
                        payload_obs_cloud.append({"semana": str(st.session_state.semana_actual).strip(), "lote_pista_key": str(key), "observacion": str(row['OBSERVACIONES']), "fecha_auditoria": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
                    if 'supabase' in st.session_state and payload_obs_cloud:
                        try: st.session_state['supabase'].table("arqueos_observaciones").upsert(payload_obs_cloud, on_conflict="semana,lote_pista_key").execute()
                        except Exception: pass
    
            with tab2:
                st.markdown("#### 🛠️ Conciliador e Historial de Correcciones")
                mask_fantasmas = ((st.session_state.cruce_final['ESTADO'] == "❌ DISCREPANCIA") & (st.session_state.cruce_final['SALDO_SAP'] == 0) & (st.session_state.cruce_final['SALDO_FISICO'] > 0) & (~st.session_state.cruce_final['OBSERVACIONES'].astype(str).str.contains("FÍSICO REAL|JUSTIFICADO", case=False, na=False)))
                err_fantasmas = st.session_state.cruce_final[mask_fantasmas]
                if not err_fantasmas.empty:
                    opciones = err_fantasmas.apply(lambda x: f"{x['PISTA']} | Prod: {x['PRODUCTO']} | Lote Físico: {x['LOTE']} ({x['SALDO_FISICO']} L/Kg)", axis=1).tolist()
                    sel = st.selectbox("1️⃣ Seleccione el error de digitación a corregir:", opciones)
                    if sel:
                        row_s = err_fantasmas.iloc[opciones.index(sel)]
                        df_sap_pista = st.session_state.df_sap_raw[st.session_state.df_sap_raw['PISTA'] == row_s['PISTA']].copy()
                        df_sap_pista['PROD_CLEAN'] = df_sap_pista['PRODUCTO'].apply(lambda x: re.sub(r'\s+', ' ', str(x)).strip().upper())
                        prod_fisico_clean = re.sub(r'\s+', ' ', str(row_s['PRODUCTO'])).strip().upper()
                        
                        df_exact = df_sap_pista[df_sap_pista['PROD_CLEAN'] == prod_fisico_clean]
                        if df_exact.empty: df_exact = df_sap_pista[df_sap_pista['PROD_CLEAN'].apply(lambda x: x in prod_fisico_clean or prod_fisico_clean in x)]
                        if df_exact.empty:
                            p_word = prod_fisico_clean.split()[0] if prod_fisico_clean else ""
                            if len(p_word) >= 3: df_exact = df_sap_pista[df_sap_pista['PROD_CLEAN'].str.contains(p_word, regex=False)]
                        if df_exact.empty:
                            import difflib
                            matches = difflib.get_close_matches(prod_fisico_clean, df_sap_pista['PROD_CLEAN'].unique().tolist(), n=3, cutoff=0.4)
                            if matches: df_exact = df_sap_pista[df_sap_pista['PROD_CLEAN'].isin(matches)]
                        
                        c_tog, _ = st.columns([1, 1])
                        mostrar_todos = c_tog.toggle("🔄 Ver todo el arsenal de la pista (Ignorar filtro inteligente)", value=False)
                        
                        opciones_dest = []
                        if not df_exact.empty and not mostrar_todos: 
                            st.success(f"🎯 Producto localizado automáticamente en SAP: **{df_exact.iloc[0]['PRODUCTO']}**")
                            opciones_dest = sorted(df_exact.apply(lambda x: f"{x['PRODUCTO']} | Lote: {x['LOTE']}", axis=1).unique().tolist())
                        else: 
                            st.warning("⚠️ Mostrando el inventario completo de la base para selección manual.")
                            opciones_dest = sorted(df_sap_pista.apply(lambda x: f"{x['PRODUCTO']} | Lote: {x['LOTE']}", axis=1).unique().tolist())
                        
                        opcion_na = "🚫 N/A - NO EXISTE EN SAP (MARCAR COMO FÍSICO REAL)"
                        opciones_dest.insert(0, opcion_na)
                        lote_ok_str = st.selectbox(f"2️⃣ Seleccione el Lote destino en SAP para unificarlos:", opciones_dest)
                        
                        if lote_ok_str == opcion_na:
                            if st.button("💾 JUSTIFICAR Y OCULTAR DEL CONCILIADOR", type="primary"):
                                txt_obs = "FÍSICO REAL - Pendiente de ingreso/traslado en SAP"
                                key_obs = f"{row_s['PISTA']}_{row_s['LOTE_KEY']}"
                                st.session_state.observaciones_memoria[key_obs] = txt_obs
                                idx_m = st.session_state.cruce_final[(st.session_state.cruce_final['PISTA'] == row_s['PISTA']) & (st.session_state.cruce_final['LOTE_KEY'] == row_s['LOTE_KEY'])].index
                                if not idx_m.empty: st.session_state.cruce_final.at[idx_m[0], 'OBSERVACIONES'] = txt_obs
                                if 'supabase' in st.session_state:
                                    try: st.session_state['supabase'].table("arqueos_observaciones").upsert([{"semana": str(st.session_state.semana_actual).strip(), "lote_pista_key": str(key_obs), "observacion": txt_obs, "fecha_auditoria": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}], on_conflict="semana,lote_pista_key").execute()
                                    except Exception: pass
                                st.toast("✅ Lote justificado exitosamente.", icon="✅")
                                generar_cruce()
                                st.rerun()
                        else:
                            if st.button("⚡ FUSIONAR Y CORREGIR LOTE", type="primary"):
                                prod_sap, lote_sap = lote_ok_str.split(" | Lote: ")[0].strip(), lote_ok_str.split(" | Lote: ")[1].strip()
                                mask = (st.session_state.df_sup_grouped['PISTA'] == row_s['PISTA']) & (st.session_state.df_sup_grouped['LOTE_KEY'] == row_s['LOTE_KEY'])
                                txt_obs = f"Corrección unificada con SAP ({prod_sap} - {lote_sap})"
                                key_obs = f"{row_s['PISTA']}_{purificar_lote(lote_sap)}"
                                st.session_state.observaciones_memoria[key_obs] = txt_obs
                                st.session_state.historial_fusiones.append({"pista": row_s['PISTA'], "lote_erroneo": row_s['LOTE'], "lote_key_erroneo": row_s['LOTE_KEY'], "lote_destino": lote_sap, "producto": prod_sap, "volumen": row_s['SALDO_FISICO']})
                                st.session_state.df_sup_grouped.loc[mask, 'LOTE_SUP'] = lote_sap
                                st.session_state.df_sup_grouped.loc[mask, 'LOTE_KEY'] = purificar_lote(lote_sap)
                                st.session_state.df_sup_grouped.loc[mask, 'PRODUCTO_SUP'] = prod_sap
                                st.session_state.df_sup_grouped = st.session_state.df_sup_grouped.groupby(['PISTA', 'LOTE_KEY'], as_index=False).agg({'PRODUCTO_SUP': 'first', 'LOTE_SUP': 'first', 'SALDO_FISICO': 'sum'})
                                if 'supabase' in st.session_state:
                                    try: st.session_state['supabase'].table("arqueos_log_fusiones").insert({"semana": str(st.session_state.semana_actual).strip(), "pista": str(row_s['PISTA']), "lote_erroneo": str(row_s['LOTE']), "lote_corregido_sap": str(lote_sap), "insumo": str(prod_sap), "fecha_correccion": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}).execute()
                                    except Exception: pass
                                generar_cruce()
                                st.rerun()
                else: st.success("✅ No se detectan lotes pendientes por fusionar.")

                if st.session_state.historial_fusiones:
                    st.markdown("---")
                    with st.expander(f"↩️ Ver / Gestionar Fusiones Realizadas ({len(st.session_state.historial_fusiones)})", expanded=False):
                        if st.button("🧹 LIMPIAR TODO EL HISTORIAL DE FUSIONES", type="secondary", use_container_width=True):
                            st.session_state.historial_fusiones = []
                            st.session_state.df_sup_grouped = st.session_state.df_sup_grouped_virgen.copy()
                            generar_cruce()
                            st.toast("✅ Historial limpiado.", icon="🧹")
                            st.rerun()
                        st.markdown("<br>", unsafe_allow_html=True)
                        for idx_f, f_item in enumerate(st.session_state.historial_fusiones):
                            c_f1, c_f2 = st.columns([3, 1])
                            c_f1.info(f"📍 **{f_item['pista']}** | Lote Creado: `{f_item['lote_erroneo']}` ➔ Sumado a Lote SAP: `{f_item['lote_destino']}` (+{f_item['volumen']} L/Kg)")
                            if c_f2.button(f"↩️ DESHACER", key=f"btn_undo_{idx_f}"):
                                st.session_state.historial_fusiones.pop(idx_f)
                                st.session_state.df_sup_grouped = st.session_state.df_sup_grouped_virgen.copy()
                                for f_rest in st.session_state.historial_fusiones:
                                    mask_r = (st.session_state.df_sup_grouped['PISTA'] == f_rest['pista']) & (st.session_state.df_sup_grouped['LOTE_KEY'] == f_rest['lote_key_erroneo'])
                                    st.session_state.df_sup_grouped.loc[mask_r, 'LOTE_SUP'] = f_rest['lote_destino']
                                    st.session_state.df_sup_grouped.loc[mask_r, 'LOTE_KEY'] = purificar_lote(f_rest['lote_destino'])
                                    st.session_state.df_sup_grouped.loc[mask_r, 'PRODUCTO_SUP'] = f_rest['producto']
                                st.session_state.df_sup_grouped = st.session_state.df_sup_grouped.groupby(['PISTA', 'LOTE_KEY'], as_index=False).agg({'PRODUCTO_SUP': 'first', 'LOTE_SUP': 'first', 'SALDO_FISICO': 'sum'})
                                generar_cruce()
                                st.toast("✅ Fusión deshecha con éxito.", icon="↩️")
                                st.rerun()
    
            with tab3:
                st.dataframe(st.session_state.cruce_final.drop(columns=['LOTE_KEY'], errors='ignore').style.map(lambda x: 'background-color: #d4edda; color: #155724' if x == "✅ OK" else '', subset=['ESTADO']), use_container_width=True, hide_index=True, column_config={"SALDO_SAP": st.column_config.NumberColumn("SALDO SAP", format="%.3f"), "SALDO_FISICO": st.column_config.NumberColumn("SALDO FÍSICO", format="%.3f"), "DIFERENCIA": st.column_config.NumberColumn("DIFERENCIA", format="%.3f")})
    
            st.markdown("---")
            col_dw1, col_dw2 = st.columns(2)
            with col_dw1:
                excel_binario = compilar_excel_maestro(st.session_state.cruce_final, st.session_state.semana_actual)
                st.download_button("📊 DESCARGAR EXCEL VIVO", excel_binario, f"Arqueo_Excel_Semana_{st.session_state.semana_actual}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with col_dw2:
                st.session_state.centro_pdf_activo = st.toggle("📄 ACTIVAR CENTRO DE EMISIÓN DE PDF", value=st.session_state.centro_pdf_activo)
                if st.session_state.centro_pdf_activo:
                    css_vip = """<style>body { font-family: Helvetica, sans-serif; background: white; color: black; font-size: 11px; } .b-print { padding: 20px; } table { width: 100%; border-collapse: collapse; margin-bottom: 20px; } th { background-color: #0d1b2a; color: #d4af37; border: 1px solid #000; padding: 6px; text-align: center; font-size: 12px; } td { border: 1px solid #000; padding: 4px; text-align: center; } .td-left { text-align: left; } .title { font-size: 20px; color: #0d1b2a; font-weight: bold; text-align: center; margin: 0; } .subtitle { font-size: 14px; color: #d4af37; text-align: center; margin: 0 0 20px 0; font-weight: bold; } .firmas-container { display: flex; justify-content: space-around; margin-top: 50px; page-break-inside: avoid; } .firma-box { text-align: center; width: 40%; border-top: 2px solid #0d1b2a; padding-top: 5px; font-weight: bold; color: #0d1b2a; } @media print { @page { size: A4 landscape; margin: 10mm; } body { background: white; -webkit-print-color-adjust: exact; print-color-adjust: exact; } .no-print { display: none !important; } .salto-pagina { page-break-after: always; } }</style>"""
                    html_reporte_masivo = compilar_html_pdf(st.session_state.cruce_final, st.session_state.semana_actual, css_vip)
                    st.info("💡 **Coordenada Activada:** Use el botón azul de adentro del visor inferior para descargar el PDF directo a su disco local.")
                    components.html(html_reporte_masivo, height=600, scrolling=True)

if __name__ == "__main__":
    pass
