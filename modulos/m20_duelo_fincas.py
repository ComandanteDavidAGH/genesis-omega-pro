import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import gspread
import requests
import io
import re
import math
import json
from datetime import datetime, timedelta, date

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# 🔌 CONEXIÓN Y FORMATO
# =================================================================

def obtener_hora_colombia():
    return datetime.utcnow() + timedelta(hours=-5)

def formato_latino(numero, decimales=0):
    if pd.isna(numero) or numero is None: return "0"
    try:
        num = float(numero)
        if num == 0: return "0"
        if decimales == 0: texto_us = f"{num:,.0f}"
        else: texto_us = f"{num:,.{decimales}f}"
        return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0"

@st.cache_resource(show_spinner=False)
def inicializar_cliente_gspread():
    try:
        if "gcp_service_account" in st.secrets:
            return gspread.service_account_from_dict(dict(st.secrets["gcp_service_account"]))
        elif "gcp_credentials" in st.secrets:
            return gspread.service_account_from_dict(dict(st.secrets["gcp_credentials"]))
        return gspread.service_account(filename='credenciales.json')
    except Exception: return None

def limpiar_numeros_universales(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        if not v or v in ['-', 'N/A', '']: return 0.0
        
        v = v.replace('$', '').replace('COP', '').replace(' ', '')
        has_dot = '.' in v
        has_comma = ',' in v
        
        if has_dot and has_comma:
            if v.rfind(',') > v.rfind('.'):
                v = v.replace('.', '').replace(',', '.')
            else:
                v = v.replace(',', '')
        elif has_comma:
            partes = v.split(',')
            if len(partes) == 2 and len(partes[1]) != 3:
                v = v.replace(',', '.') 
            else:
                v = v.replace(',', '') 
        elif has_dot:
            partes = v.split('.')
            if len(partes) == 2 and len(partes[1]) != 3:
                pass 
            else:
                v = v.replace('.', '') 
        
        v = re.sub(r'[^\d\.\-]', '', v)
        return float(v) if v else 0.0
    except:
        return 0.0

def limpiar_tiempo(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        if not v: return 0.0
        if ':' in v:
            partes = v.split(':')
            return float(partes[0]) + (float(partes[1]) / 60.0)
        v = v.replace(',', '.')
        v = re.sub(r'[^\d\.]', '', v)
        return float(v) if v else 0.0
    except: return 0.0

def procesar_fecha_estricta(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() in ["none", "nan", "nat", "<na>"]: return pd.NaT
    s = str(val).strip().lower()
    if s.replace('.', '', 1).isdigit(): return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    meses_es = {'enero':1, 'febrero':2, 'marzo':3, 'abril':4, 'mayo':5, 'junio':6, 'julio':7, 'agosto':8, 'septiembre':9, 'octubre':10, 'noviembre':11, 'diciembre':12}
    mes_encontrado = next((meses_es[m] for m in meses_es if m in s), None)
    if mes_encontrado:
        numeros = re.findall(r'\d+', s)
        if len(numeros) >= 2:
            n1, n2 = int(numeros[0]), int(numeros[1])
            anio = n1 if n1 > 1000 else (n2 if n2 > 1000 else (2000 + n2 if n2 < 100 else n2))
            dia = n2 if n1 > 1000 else (n1 if n2 > 1000 else n1)
            try: return pd.Timestamp(year=anio, month=mes_encontrado, day=dia)
            except: pass
    s = s.replace(',', '').replace(' de ', '/').replace('-', '/').strip()
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: 
        res = pd.to_datetime(s, dayfirst=True)
        return pd.NaT if pd.isna(res) else res
    except: return pd.NaT 

def extraer_diccionario_flota(gc):
    try:
        sh = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        for ws in sh.worksheets():
            try:
                data = ws.get_all_values()
                for row in data[:20]:
                    row_upper = [str(x).upper().strip() for x in row]
                    if 'MATRICULA' in row_upper and ('TIPO AVION' in row_upper or 'MODELO' in row_upper):
                        idx_hk = row_upper.index('MATRICULA')
                        idx_mod = row_upper.index('TIPO AVION') if 'TIPO AVION' in row_upper else row_upper.index('MODELO')
                        flota = {}
                        for r in data[data.index(row)+1:]:
                            if len(r) > max(idx_hk, idx_mod):
                                hk = str(r[idx_hk]).strip().upper()
                                mod = str(r[idx_mod]).strip().upper()
                                if hk: flota[hk] = mod
                        return flota
            except: continue
    except: pass
    return {}

@st.cache_data(show_spinner=False, ttl=600)
def cargar_fuentes_maestras_duelo_v4():
    gc = inicializar_cliente_gspread()
    if not gc: return pd.DataFrame()
    
    flota_dict = extraer_diccionario_flota(gc)

    try:
        boveda_act = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        datos_brutos_act = boveda_act.worksheet("TABLA 1").get_all_values()
    except:
        datos_brutos_act = []
    
    df_vivos = pd.DataFrame()
    if len(datos_brutos_act) > 5:
        columnas_t1 = ["OS", "BLOQUE", "FINCA", "SECTOR", "AREA_BRUTA", "AREA_FUMIG", "COCTEL", "FECHA", "DIA", "SEMANA", "H_TOTAL", "GLN_HA", "VOL_TOTAL", "REND_HR", "REND_MIN", "PILOTO", "HK", "MODELO", "COSTO_AVION", "COSTO_HA", "DOMINICAL_HA", "COSTO_FINCA", "VALOR_FACTURAR", "PISTA", "INC_2026", "LIMITE", "ALERTA", "VAR_PCT", "COSTO_TOTAL", "PAGO_AVION"]
        filas_limpias = [r + [""]*(len(columnas_t1) - len(r)) for r in datos_brutos_act[5:]]
        df_vivos = pd.DataFrame([r[:len(columnas_t1)] for r in filas_limpias], columns=columnas_t1)
        df_vivos.rename(columns={'AREA_FUMIG': 'AREA_MAESTRA', 'COSTO_HA': 'COSTO_HA_BASE', 'VALOR_FACTURAR': 'VALOR_FACTURAR', 'COSTO_TOTAL': 'COSTO_TOTAL', 'FINCA': 'FINCA_MAESTRA', 'FECHA': 'FECHA_MAESTRA', 'PISTA': 'PISTA_MAESTRA', 'OS': 'OS_MAESTRA'}, inplace=True)
        df_vivos['ORIGEN'] = 'ACTUAL'
        df_vivos = df_vivos.loc[:, ~df_vivos.columns.duplicated()]

    datos_brutos_hist = []
    try:
        boveda_hist = gc.open_by_url("https://docs.google.com/spreadsheets/d/16OZdiWwW7nLHyZBEnhiKlDTDttR7Tjhn37O9zm6wJOk/edit")
        datos_brutos_hist = boveda_hist.worksheet("Datos").get_all_values()
    except: pass
    
    df_historico = pd.DataFrame()
    if len(datos_brutos_hist) > 0:
        df_historico = pd.DataFrame(datos_brutos_hist[1:], columns=datos_brutos_hist[0])
        df_historico.columns = [str(c).upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').strip() for c in df_historico.columns]
        
        df_historico = df_historico.loc[:, ~df_historico.columns.duplicated()]
        df_historico = df_historico.loc[:, [c for c in df_historico.columns if c != ""]]

        renombres = {}
        for col in df_historico.columns:
            if 'FUMIG' in col and 'AREA' in col: renombres[col] = 'AREA_MAESTRA'
            elif 'FACTURAR' in col and 'VALOR' in col: renombres[col] = 'VALOR_FACTURAR'
            elif 'COSTO' in col and 'HA' in col: renombres[col] = 'COSTO_HA_BASE'
            elif 'COSTO' in col and 'TOTAL' in col: renombres[col] = 'COSTO_TOTAL'
            elif col in ['FINCA', 'PROPIEDAD']: renombres[col] = 'FINCA_MAESTRA'
            elif col == 'FECHA': renombres[col] = 'FECHA_MAESTRA'
            elif col == 'PISTA': renombres[col] = 'PISTA_MAESTRA'
            elif "ORDEN" in col or "OS" == col: renombres[col] = 'OS_MAESTRA'
            elif "REND" in col and "HR" in col: renombres[col] = 'REND_HR'
            elif ("H" in col or "HORA" in col) and "TOTAL" in col: renombres[col] = 'H_TOTAL'
            elif col in ['AERONAVE', 'AVION', 'MATRICULA', 'HK']: renombres[col] = 'HK'
            elif col in ['MODELO', 'TIPO AVION', 'TIPO']: renombres[col] = 'MODELO'
        df_historico.rename(columns=renombres, inplace=True)
        df_historico['ORIGEN'] = 'HISTORICO'

    super_base = pd.concat([df_historico, df_vivos], ignore_index=True)
    
    if not super_base.empty and 'FINCA_MAESTRA' in super_base.columns:
        super_base['FINCA_MAESTRA'] = super_base['FINCA_MAESTRA'].astype(str).str.strip().str.upper()
        super_base['PISTA_MAESTRA'] = super_base['PISTA_MAESTRA'].astype(str).str.strip().str.upper()
        super_base['FECHA_DT'] = super_base['FECHA_MAESTRA'].apply(procesar_fecha_estricta)
        super_base = super_base.dropna(subset=['FECHA_DT'])
        
        super_base['AREA_NUM'] = super_base.get('AREA_MAESTRA', 0).apply(limpiar_numeros_universales)
        super_base['VALOR_FACTURAR_NUM'] = super_base.get('VALOR_FACTURAR', 0).apply(limpiar_numeros_universales) 
        super_base['COSTO_HA_NUM'] = super_base.get('COSTO_HA_BASE', 0).apply(limpiar_numeros_universales) 
        super_base['COSTO_TOTAL_NUM'] = super_base.get('COSTO_TOTAL', 0).apply(limpiar_numeros_universales) 
        
        if 'H_TOTAL' not in super_base.columns: super_base['H_TOTAL'] = 0
        super_base['H_TOTAL_NUM'] = super_base['H_TOTAL'].apply(limpiar_tiempo)
        
        def mapear_modelo(row):
            mod = str(row.get('MODELO', '')).strip().upper()
            hk = str(row.get('HK', '')).strip().upper()
            if mod and mod not in ["", "NAN", "NONE"]: return mod
            if hk in flota_dict: return flota_dict[hk]
            return hk if hk else "NO REGISTRADO"
            
        super_base['MODELO_FINAL'] = super_base.apply(mapear_modelo, axis=1)
        
        super_base = super_base[super_base['AREA_NUM'] > 0]
        super_base = super_base[super_base['FINCA_MAESTRA'] != ""]
        
        return super_base
    else:
        return pd.DataFrame()

# =================================================================
# 👑 INTERFAZ PRINCIPAL (EL COLISEO LOGÍSTICO)
# =================================================================

def ejecutar():
    st.markdown("""
    <style>
    .titulo-mod { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; text-transform: uppercase; }
    .kpi-vs { background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); color: white; padding: 20px; border-radius: 12px; border-left: 6px solid #d4af37; box-shadow: 0 8px 16px rgba(0,0,0,0.3); text-align: center; margin-bottom: 15px; }
    .kpi-vs-title { font-size: 13px; font-weight: bold; color: #d4af37; text-transform: uppercase; margin-bottom: 5px; letter-spacing: 1px; }
    .kpi-vs-value { font-size: 32px; font-weight: 900; margin: 0; color: #ffffff; font-family: 'Arial Black', sans-serif; }
    .victoria { border: 3px solid #28a745 !important; box-shadow: 0 0 15px rgba(40, 167, 69, 0.5) !important; }
    
    /* 🎯 ESTILO ESTABLE CORPORATIVO PARA INPUTS 🎯 */
    .stSelectbox > div > div, .stDateInput > div > div, .stMultiSelect > div > div {
        border: 2px solid #d4af37 !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1) !important;
    }
    
    .stSelectbox label, .stDateInput label, .stMultiSelect label {
        color: #0d1b2a !important;
        font-weight: bold !important;
    }

    .lista-hk { 
        text-align: left; background-color: #ffffff; padding: 15px; border-radius: 8px; 
        border: 2px solid #0d1b2a; border-left: 6px solid #d4af37; box-shadow: 0 4px 8px rgba(0,0,0,0.1); margin-bottom: 15px;
    }
    .lista-hk ul { list-style-type: none; padding-left: 0; margin: 0; }
    .lista-hk li { font-size: 14px; margin-bottom: 8px; color: #0d1b2a; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-mod'>⚔️ 20. Duelo Logístico (Pista vs Pista)</h1>", unsafe_allow_html=True)
    st.write("Analiza el rendimiento de una misma finca operando desde dos bases distintas. Compara el Costo Integral, la Tarifa de Avión y los Ciclos Reales agrupados por operación.")

    with st.spinner("Desplegando el radar sobre la Bóveda Maestra de Vuelos..."):
        df_base = cargar_fuentes_maestras_duelo_v4()

    if df_base.empty:
        st.error("🚨 No se encontró información en la base de datos o hubo un error de conexión.")
        return

    st.markdown("### 🎯 Configuración del Duelo")
    lista_fincas = sorted(df_base['FINCA_MAESTRA'].unique().tolist())
    
    # --- CONTENEDOR TÁCTICO AL 50% ---
    c_finca_mitad, _ = st.columns([1, 1])
    with c_finca_mitad:
        finca_sel = st.selectbox("🏡 SELECCIONE LA FINCA A ANALIZAR", lista_fincas)

    df_finca_global = df_base[df_base['FINCA_MAESTRA'] == finca_sel]
    lista_pistas = sorted(df_finca_global['PISTA_MAESTRA'].unique().tolist())

    if not lista_pistas:
        st.warning("⚠️ No hay pistas registradas para esta finca.")
        return

    col_filt1, col_filt2, col_filt3, col_filt4 = st.columns([1.2, 1.2, 1, 1])
    with col_filt1: pista_A = st.selectbox("🔴 PISTA RETADORA A", lista_pistas, index=0)
    with col_filt2: pista_B = st.selectbox("🔵 PISTA RETADORA B", lista_pistas, index=1 if len(lista_pistas) > 1 else 0)
    
    min_date_allowed = date(2020, 1, 1)
    max_date_allowed = date(2026, 12, 31)
    
    with col_filt3: start_date = st.date_input("📅 Fecha Inicial", value=date(2026, 1, 1), min_value=min_date_allowed, max_value=max_date_allowed)
    with col_filt4: end_date = st.date_input("📅 Fecha Final", value=date(2026, 12, 31), min_value=min_date_allowed, max_value=max_date_allowed)

    df_finca = df_finca_global.copy()
    if start_date and end_date:
        if start_date > end_date:
            st.error("🚨 La Fecha Inicial no puede ser mayor que la Fecha Final.")
            return
        df_finca = df_finca[(df_finca['FECHA_DT'].dt.date >= start_date) & (df_finca['FECHA_DT'].dt.date <= end_date)]

    df_A = df_finca[df_finca['PISTA_MAESTRA'] == pista_A]
    df_B = df_finca[df_finca['PISTA_MAESTRA'] == pista_B]

    st.markdown("<hr style='border: 1px solid #d4af37;'>", unsafe_allow_html=True)

    if df_A.empty and df_B.empty:
        st.warning("⚠️ La finca no operó desde ninguna de estas pistas en el rango de fechas seleccionado.")
        return

    def calcular_metricas(df_pista):
        if df_pista.empty: return 0, 0, 0, 0, "", ""
        total_ha = df_pista['AREA_NUM'].sum()
        total_facturado = df_pista['COSTO_TOTAL_NUM'].sum()
        costo_integral_ha = df_pista['VALOR_FACTURAR_NUM'].mean()
        costo_tarifa_avion = df_pista['COSTO_HA_NUM'].mean()
        
        html_mod = "<div class='lista-hk'><p style='font-weight:900; margin-bottom:5px; color:#0d1b2a;'>✈️ RENDIMIENTO DISCRIMINADO POR AERONAVE:</p><ul>"
        df_mod = df_pista.groupby('MODELO_FINAL').agg(REGISTROS=('OS_MAESTRA', 'count'), HORAS=('H_TOTAL_NUM', 'sum'), AREA=('AREA_NUM', 'sum')).reset_index()
        for _, row in df_mod.iterrows():
            mod_nombre = str(row['MODELO_FINAL'])
            registros, horas, area = row['REGISTROS'], row['HORAS'], row['AREA']
            rend_ha_hr = area / horas if horas > 0 else 0
            html_mod += f"<li><b>{mod_nombre}:</b> {formato_latino(rend_ha_hr, 1)} Ha / Hora <i>({registros} líneas voladas)</i></li>"
        html_mod += "</ul></div>"
        
        fechas_unicas = sorted(df_pista['FECHA_DT'].dropna().unique())
        total_ciclos_reales = 0
        if len(fechas_unicas) > 0:
            total_ciclos_reales = 1
            inicio_ciclo = fechas_unicas[0]
            for f in fechas_unicas[1:]:
                if (f - inicio_ciclo).days > 5:
                    total_ciclos_reales += 1
                    inicio_ciclo = f
        
        html_ciclos = f"<div class='kpi-vs' style='padding: 10px; margin-top: 5px; border: 2px dashed #d4af37;'><p class='kpi-vs-title'>TOTAL CICLOS REALES (Agrupados a 5 días)</p><p class='kpi-vs-value' style='font-size:26px;'>{total_ciclos_reales} Ciclos</p></div>"
        return total_ha, total_facturado, costo_integral_ha, costo_tarifa_avion, html_mod, html_ciclos

    ha_A, inv_A, costo_integral_A, costo_os_A, html_mod_A, html_ciclos_A = calcular_metricas(df_A)
    ha_B, inv_B, costo_integral_B, costo_os_B, html_mod_B, html_ciclos_B = calcular_metricas(df_B)

    tiene_datos_A, tiene_datos_B = costo_integral_A > 0, costo_integral_B > 0
    clase_win_A = "victoria" if tiene_datos_A and (not tiene_datos_B or costo_integral_A < costo_integral_B) else ""
    clase_win_B = "victoria" if tiene_datos_B and (not tiene_datos_A or costo_integral_B < costo_integral_A) else ""

    col_A, col_vs, col_B = st.columns([4, 1, 4])
    with col_A:
        st.markdown(f"<h3 style='text-align:center; color:#dc3545;'>🔴 SALIENDO DESDE: {pista_A}</h3>", unsafe_allow_html=True)
        st.markdown(f"<div class='kpi-vs {clase_win_A}'><p class='kpi-vs-title'>COSTO PROMEDIO X HECTÁREA</p><p class='kpi-vs-value'>$ {formato_latino(costo_integral_A, 0)}</p></div>", unsafe_allow_html=True)
        m_a1, m_a2 = st.columns(2)
        m_a1.markdown(f"<div class='kpi-vs' style='padding: 10px;'><p class='kpi-vs-title'>COSTO PROMEDIO X OS</p><p class='kpi-vs-value' style='font-size:20px;'>$ {formato_latino(costo_os_A, 0)}</p></div>", unsafe_allow_html=True)
        m_a2.markdown(f"<div class='kpi-vs' style='padding: 10px;'><p class='kpi-vs-title'>COSTO TOTAL FACTURADO</p><p class='kpi-vs-value' style='font-size:20px;'>$ {formato_latino(inv_A, 0)}</p></div>", unsafe_allow_html=True)
        if df_A.empty: st.info("Sin operaciones registradas.")
        else: st.markdown(html_mod_A, unsafe_allow_html=True)

    with col_vs:
        st.markdown("<br><br><h1 style='text-align:center; color:#d4af37; font-size: 50px; font-family:Arial Black;'>VS</h1>", unsafe_allow_html=True)

    with col_B:
        st.markdown(f"<h3 style='text-align:center; color:#2F75B5;'>🔵 SALIENDO DESDE: {pista_B}</h3>", unsafe_allow_html=True)
        st.markdown(f"<div class='kpi-vs {clase_win_B}'><p class='kpi-vs-title'>COSTO PROMEDIO X HECTÁREA</p><p class='kpi-vs-value'>$ {formato_latino(costo_integral_B, 0)}</p></div>", unsafe_allow_html=True)
        m_b1, m_b2 = st.columns(2)
        m_b1.markdown(f"<div class='kpi-vs' style='padding: 10px;'><p class='kpi-vs-title'>COSTO PROMEDIO X OS</p><p class='kpi-vs-value' style='font-size:20px;'>$ {formato_latino(costo_os_B, 0)}</p></div>", unsafe_allow_html=True)
        m_b2.markdown(f"<div class='kpi-vs' style='padding: 10px;'><p class='kpi-vs-title'>COSTO TOTAL FACTURADO</p><p class='kpi-vs-value' style='font-size:20px;'>$ {formato_latino(inv_B, 0)}</p></div>", unsafe_allow_html=True)
        if df_B.empty: st.info("Sin operaciones registradas.")
        else: st.markdown(html_mod_B, unsafe_allow_html=True)

    col_ciclos_A, col_ciclos_vs, col_ciclos_B = st.columns([4, 1, 4])
    with col_ciclos_A:
        if not df_A.empty: st.markdown(html_ciclos_A, unsafe_allow_html=True)
    with col_ciclos_B:
        if not df_B.empty: st.markdown(html_ciclos_B, unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("### 🛫 Desglose Logístico, Financiero y de Aeronaves")
    df_ambos = pd.concat([df_A, df_B])
    
    if not df_ambos.empty:
        df_pistas = df_ambos.groupby('PISTA_MAESTRA').agg(COSTO_PROMEDIO_HECTAREA=('VALOR_FACTURAR_NUM', 'mean'), COSTO_PROMEDIO_OS=('COSTO_HA_NUM', 'mean')).reset_index()
        c_graf1, c_graf2 = st.columns(2)
        with c_graf1:
            fig_ha = px.bar(df_pistas, x='PISTA_MAESTRA', y='COSTO_PROMEDIO_HECTAREA', color='PISTA_MAESTRA', text='COSTO_PROMEDIO_HECTAREA', color_discrete_map={pista_A: '#dc3545', pista_B: '#2F75B5'}, title="Costo Promedio X Hectárea (Integral)")
            fig_ha.update_traces(texttemplate='$ %{text:,.0f}', textposition='outside', textfont=dict(family="Arial Black"))
            fig_ha.update_layout(yaxis_title="$ / Ha", xaxis_title="Pista", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='#ffffff', showlegend=False)
            st.plotly_chart(fig_ha, use_container_width=True)
        with c_graf2:
            fig_os = px.bar(df_pistas, x='PISTA_MAESTRA', y='COSTO_PROMEDIO_OS', color='PISTA_MAESTRA', text='COSTO_PROMEDIO_OS', color_discrete_map={pista_A: '#dc3545', pista_B: '#2F75B5'}, title="Costo Promedio X OS (Tarifa Avión)")
            fig_os.update_traces(texttemplate='$ %{text:,.0f}', textposition='outside', textfont=dict(family="Arial Black"))
            fig_os.update_layout(yaxis_title="$ / OS", xaxis_title="Pista", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='#ffffff', showlegend=False)
            st.plotly_chart(fig_os, use_container_width=True)

        st.markdown("#### ✈️ Rendimiento Operativo (Ha/Hora) por Modelo de Aeronave")
        df_mod_graf = df_ambos.groupby(['PISTA_MAESTRA', 'MODELO_FINAL']).agg(HORAS=('H_TOTAL_NUM', 'sum'), AREA=('AREA_NUM', 'sum')).reset_index()
        df_mod_graf['RENDIMIENTO_HA_HR'] = np.where(df_mod_graf['HORAS'] > 0, df_mod_graf['AREA'] / df_mod_graf['HORAS'], 0)

        fig_rend = px.bar(df_mod_graf, x='MODELO_FINAL', y='RENDIMIENTO_HA_HR', color='PISTA_MAESTRA', barmode='group', text='RENDIMIENTO_HA_HR', color_discrete_map={pista_A: '#dc3545', pista_B: '#2F75B5'}, title="Rendimiento (Ha / Hora) por Aeronave")
        fig_rend.update_traces(texttemplate='%{text:,.1f} Ha/Hr', textposition='outside', textfont=dict(family="Arial Black"))
        fig_rend.update_layout(yaxis_title="Hectáreas por Hora", xaxis_title="Modelo de Aeronave", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='#ffffff', legend_title_text='Pista')
        st.plotly_chart(fig_rend, use_container_width=True)

    # 💥 CENTRO DE EXTRACCIÓN CON DISEÑO PROFESIONAL (ESTILOS EXCEL) 💥
    st.markdown("<hr style='border: 1px solid #d4af37;'>", unsafe_allow_html=True)
    st.markdown("### 📥 Centro de Extracción de Datos (Reporte Ejecutivo)")
    
    activar_descargas = st.toggle("🛸 HABILITAR PANEL DE EXPORTACIÓN", value=False)
    if activar_descargas:
        st.info("💡 Seleccione las fincas a exportar. El sistema generará el comparativo Gerencial y la Auditoría de vuelos con formato corporativo.")
        fincas_a_descargar = st.multiselect("🚜 Seleccionar Fincas a Exportar:", lista_fincas, default=[finca_sel])
        
        if fincas_a_descargar:
            df_filtrado_fecha = df_base[(df_base['FECHA_DT'].dt.date >= start_date) & (df_base['FECHA_DT'].dt.date <= end_date)]
            df_export = df_filtrado_fecha[df_filtrado_fecha['FINCA_MAESTRA'].isin(fincas_a_descargar)].copy()
            
            resumen_data = []
            rendimiento_data = []
            
            for finca in fincas_a_descargar:
                df_f = df_export[df_export['FINCA_MAESTRA'] == finca]
                pistas = df_f['PISTA_MAESTRA'].unique()
                for pista in pistas:
                    df_p = df_f[df_f['PISTA_MAESTRA'] == pista]
                    if df_p.empty: continue
                    
                    fechas_unicas = sorted(df_p['FECHA_DT'].dropna().unique())
                    ciclos = 0
                    if len(fechas_unicas) > 0:
                        ciclos = 1
                        inicio = fechas_unicas[0]
                        for f in fechas_unicas[1:]:
                            if (f - inicio).days > 5:
                                ciclos += 1
                                inicio = f
                    
                    total_ha = df_p['AREA_NUM'].sum()
                    avg_ha = df_p['VALOR_FACTURAR_NUM'].mean()
                    avg_os = df_p['COSTO_HA_NUM'].mean()
                    tot_fac = df_p['COSTO_TOTAL_NUM'].sum()
                    
                    resumen_data.append({
                        "FINCA": finca,
                        "PISTA BASE": pista,
                        "TOTAL HECTÁREAS": float(total_ha),
                        "CICLOS REALES (5 Días)": int(ciclos),
                        "COSTO PROMEDIO X HA (Integral)": float(avg_ha),
                        "COSTO PROMEDIO X OS (Avión)": float(avg_os),
                        "TOTAL FACTURADO": float(tot_fac)
                    })
                    
                    df_mod = df_p.groupby('MODELO_FINAL').agg(VUELOS=('OS_MAESTRA', 'count'), AREA=('AREA_NUM', 'sum'), HORAS=('H_TOTAL_NUM', 'sum')).reset_index()
                    for _, r_mod in df_mod.iterrows():
                        rend = r_mod['AREA'] / r_mod['HORAS'] if r_mod['HORAS'] > 0 else 0
                        rendimiento_data.append({
                            "FINCA": finca,
                            "PISTA": pista,
                            "MODELO AERONAVE": r_mod['MODELO_FINAL'],
                            "VUELOS REALIZADOS": int(r_mod['VUELOS']),
                            "HECTÁREAS APLICADAS": float(r_mod['AREA']),
                            "RENDIMIENTO (Ha / Hora)": float(rend)
                        })

            df_resumen = pd.DataFrame(resumen_data)
            df_rendimiento = pd.DataFrame(rendimiento_data)
            
            cols_sistema = [c for c in df_export.columns if c.endswith('_NUM') or c.endswith('_DT') or c == 'MODELO_FINAL']
            cols_inutiles = ['BLOQUE', 'SECTOR', 'AREA_BRUTA', 'LIMITE', 'ALERTA', 'VAR_PCT', 'INC_2026', 'PAGO_AVION']
            cols_a_matar = cols_sistema + [c for c in cols_inutiles if c in df_export.columns]
            
            df_auditoria = df_export.drop(columns=cols_a_matar).copy()
            df_auditoria.columns = [str(c).replace('\n', ' ').replace('\r', '').strip() for c in df_auditoria.columns]
            
            df_auditoria.replace(r'^\s*$', np.nan, regex=True, inplace=True)
            df_auditoria.dropna(axis=1, how='all', inplace=True)
            df_auditoria = df_auditoria.fillna("")
            
            num_cols_auditoria = ['AREA_MAESTRA', 'VALOR_FACTURAR', 'COSTO_HA_BASE', 'COSTO_TOTAL', 'H_TOTAL', 'GLN_HA', 'VOL_TOTAL', 'REND_HR', 'REND_MIN', 'COSTO_AVION', 'DOMINICAL_HA', 'COSTO_FINCA']
            for c in num_cols_auditoria:
                if c in df_auditoria.columns:
                    df_auditoria[c] = df_auditoria[c].apply(limpiar_numeros_universales)

            c_preview1, c_preview2 = st.columns(2)
            with c_preview1:
                st.markdown("**📊 Vista Previa: Resumen Gerencial**")
                st.dataframe(df_resumen, use_container_width=True)
            with c_preview2:
                st.markdown("**✈️ Vista Previa: Rendimiento Máquinas**")
                st.dataframe(df_rendimiento, use_container_width=True)
                
            try:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    if not df_resumen.empty:
                        df_resumen.to_excel(writer, sheet_name='Resumen Gerencial', index=False, startrow=2)
                    if not df_rendimiento.empty:
                        df_rendimiento.to_excel(writer, sheet_name='Rendimiento Aeronaves', index=False, startrow=2)
                    if not df_auditoria.empty:
                        df_auditoria.to_excel(writer, sheet_name='Auditoría Vuelos', index=False, startrow=2)
                    
                    titulos_hojas = {
                        'Resumen Gerencial': 'REPORTE GERENCIAL - COMPARATIVO DE FINCAS',
                        'Rendimiento Aeronaves': 'REPORTE GERENCIAL - RENDIMIENTO DE AERONAVES',
                        'Auditoría Vuelos': 'AUDITORÍA OPERATIVA - DATOS DE VUELO'
                    }
                    
                    str_fecha = f"Período Analizado: {start_date.strftime('%d de %B %Y')} - {end_date.strftime('%d de %B %Y')} | Fincas Evaluadas: {', '.join(fincas_a_descargar)}"
                    
                    fill_titulo = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                    font_titulo = Font(color="FFFFFF", bold=True, size=14)
                    
                    fill_sub = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    font_sub = Font(color="000000", italic=True, size=11)
                    
                    fill_header = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                    font_header = Font(color="000000", bold=True, size=11)
                    
                    fill_zebra1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    fill_zebra2 = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    
                    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
                    
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")
                    align_right = Alignment(horizontal="right", vertical="center")
                    
                    for sheet_name in writer.sheets:
                        ws = writer.sheets[sheet_name]
                        max_col = ws.max_column
                        max_row = ws.max_row
                        
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                        c_tit = ws.cell(row=1, column=1, value=titulos_hojas.get(sheet_name, 'REPORTE'))
                        c_tit.fill = fill_titulo
                        c_tit.font = font_titulo
                        c_tit.alignment = align_center
                        
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
                        c_sub = ws.cell(row=2, column=1, value=str_fecha)
                        c_sub.fill = fill_sub
                        c_sub.font = font_sub
                        c_sub.alignment = align_left
                        
                        for col_idx in range(1, max_col + 1):
                            h_cell = ws.cell(row=3, column=col_idx)
                            h_cell.fill = fill_header
                            h_cell.font = font_header
                            h_cell.alignment = align_center
                            h_cell.border = thin_border
                        
                        if max_row >= 3:
                            ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{max_row}"
                        
                        for r_idx in range(4, max_row + 1):
                            fill_c = fill_zebra1 if r_idx % 2 == 0 else fill_zebra2
                            for c_idx in range(1, max_col + 1):
                                cell = ws.cell(row=r_idx, column=c_idx)
                                cell.fill = fill_c
                                cell.border = thin_border
                                
                                if isinstance(cell.value, str):
                                    cell.alignment = align_left
                                elif isinstance(cell.value, (int, float)):
                                    header_val = str(ws.cell(row=3, column=c_idx).value).upper()
                                    if any(p in header_val for p in ['COSTO', 'VALOR', 'FACTURADO', 'PAGO', 'LIMITE', '$']):
                                        cell.number_format = '"$"#,##0'
                                    elif any(p in header_val for p in ['HECTÁREAS', 'RENDIMIENTO', 'AREA']):
                                        cell.number_format = '#,##0.00'
                                    cell.alignment = align_right
                                    
                        for c_idx in range(1, max_col + 1):
                            max_len = 0
                            for r_idx in range(3, max_row + 1):
                                val_str = str(ws.cell(row=r_idx, column=c_idx).value or "")
                                if len(val_str) > max_len: max_len = len(val_str)
                            ws.column_dimensions[get_column_letter(c_idx)].width = min(max((max_len * 1.2) + 2, 12), 45)
                
                st.download_button(
                    label="💾 DESCARGAR REPORTE EJECUTIVO (.xlsx)",
                    data=buffer.getvalue(),
                    file_name=f"Reporte_Gerencial_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            except Exception as e:
                st.warning(f"⚠️ Servidor sin openpyxl nativo. Generando CSV: {e}")
                csv = df_resumen.to_csv(index=False, sep=';', decimal=',').encode('utf-8-sig')
                st.download_button(
                    label="💾 DESCARGAR REPORTE EN CSV",
                    data=csv,
                    file_name=f"Auditoria_Resumen_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    type="primary"
                )

if __name__ == "__main__":
    pass
