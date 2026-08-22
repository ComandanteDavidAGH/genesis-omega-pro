import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
from datetime import datetime
import re
import io
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =================================================================
# ⚙️ MOTOR DE CONEXIÓN UNIFICADO (V42 VIP)
# =================================================================
@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

# =================================================================
# 🛡️ UTILIDADES DE PURIFICACIÓN Y FORMATO
# =================================================================
def a_numero_limpio(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip().replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except: return 0.0

def parsear_precio(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        v = re.sub(r'[^\d\.,\-]', '', v)
        if not v: return 0.0
        if ',' in v and '.' in v:
            if v.rfind(',') > v.rfind('.'): v = v.replace('.', '').replace(',', '.')
            else: v = v.replace(',', '')
        elif ',' in v:
            if v.count(',') > 1: v = v.replace(',', '')
            else:
                if len(v.split(',')[1]) == 3: v = v.replace(',', '') 
                else: v = v.replace(',', '.')
        elif '.' in v:
            if v.count('.') > 1: v = v.replace('.', '')
            else:
                if len(v.split('.')[1]) == 3: v = v.replace('.', '') 
                return float(v)
        return float(v)
    except: return 0.0

def procesar_fecha_pesada(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    s = str(val).strip()
    if s.replace('.', '', 1).isdigit(): 
        return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: return pd.to_datetime(s, errors='coerce')
    except: return pd.NaT

def fmt_latino(val, decimales=1):
    try: return f"{float(val):,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except: return str(val)

def extraer_receta_rapida(coctel_sel, dict_bases, dict_aditivos_dosis, dict_fertilizantes_dinamico):
    coctel_u = str(coctel_sel).upper().strip().replace("+", " ").replace("-", " ")
    partes = coctel_u.split()
    base_coctel = partes[0] if len(partes) > 0 else ""
    aditivos = partes[1:] if len(partes) > 1 else []
    
    dict_prods = dict_bases.get(base_coctel, {}).copy()

    for aditivo in aditivos:
        nombre_fert = dict_fertilizantes_dinamico.get(aditivo)
        if nombre_fert:
            dosis_fert = dict_aditivos_dosis.get(nombre_fert, 0.5)
            dict_prods[nombre_fert] = dict_prods.get(nombre_fert, 0.0) + dosis_fert
        else:
            if "NM" in aditivo: dict_prods["NATURAMIN WSP"] = dict_prods.get("NATURAMIN WSP", 0.0) + 0.2
            elif "ZN" in aditivo: dict_prods["ZINTRAC X LITRO SV"] = dict_prods.get("ZINTRAC X LITRO SV", 0.0) + 0.5
            elif "BT" in aditivo: dict_prods["BANATREL SC"] = dict_prods.get("BANATREL SC", 0.0) + 0.5
    
    if not any("ADHERENTE" in k for k in dict_prods.keys()): dict_prods["ADHERENTE SV"] = 0.13
    if not any("ACONDICIONADOR" in k for k in dict_prods.keys()): 
        dict_prods["ACONDICIONADOR SV"] = 0.06 if any(x in coctel_u for x in ["ZN", "BT", "ZT", "ZITRON"]) else 0.02
    if base_coctel.startswith("IN") or "IMBIOSIL" in base_coctel: 
        dict_prods["IMBIOSIL O"] = 1.5

    return dict_prods

# =================================================================
# 💾 EXTRACCIÓN Y PREPROCESAMIENTO
# =================================================================
@st.cache_data(show_spinner=False, ttl=7200) 
def descargar_y_masticar_bases():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame(), {}, {}, {}, pd.DataFrame(), pd.DataFrame()
    
    try:
        boveda = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        t1_vals = boveda.worksheet("TABLA 1").get_all_values()
        mz_vals = boveda.worksheet("DD_Mesclas").get_all_values()
        cfg_vals = boveda.worksheet("Configuración").get_all_values()
        
        df_t1 = pd.DataFrame(t1_vals[5:], columns=[str(c).upper().strip() for c in t1_vals[4]])
        df_mezclas = pd.DataFrame(mz_vals[1:], columns=[str(c).upper().strip() for c in mz_vals[0]])
        df_cfg = pd.DataFrame(cfg_vals[1:], columns=[str(c).upper().strip() for c in cfg_vals[0]])
        
        col_fecha = next((c for c in df_t1.columns if 'FECHA' in c), 'FECHA')
        col_ha = next((c for c in df_t1.columns if 'NETA' in c or 'FUMIG' in c or 'HECT' in c), None)
        col_coctel = next((c for c in df_t1.columns if 'COCTEL' in c or 'CÓCTEL' in c or 'MEZCLA' in c), None)
        col_pista_t1 = next((c for c in df_t1.columns if 'PISTA' in c or 'BASE' in c), None)

        if col_fecha and col_ha and col_pista_t1 and col_coctel:
            df_t1['FECHA_DT'] = df_t1[col_fecha].apply(procesar_fecha_pesada)
            df_t1 = df_t1.dropna(subset=['FECHA_DT'])
            df_t1['MES'] = df_t1['FECHA_DT'].dt.month
            df_t1['AÑO'] = df_t1['FECHA_DT'].dt.year
            df_t1['HA_CALCULO'] = df_t1[col_ha].apply(a_numero_limpio)
            df_t1['PISTA_OPERATIVA'] = df_t1[col_pista_t1].astype(str).str.upper().str.strip()
            df_t1['COCTEL_NOM'] = df_t1[col_coctel].astype(str).str.upper().str.strip()
        
        dict_bases = {}
        dict_aditivos_dosis = {}
        dict_fert = {}

        if not df_mezclas.empty:
            col_0_limpia = df_mezclas.iloc[:, 0].astype(str).str.upper().str.strip()
            for base_name in col_0_limpia.unique():
                if base_name in ["NAN", "", "NONE"]: continue
                rb = df_mezclas[col_0_limpia == base_name]
                prods = {}
                for _, r in rb.iterrows():
                    p = str(r.iloc[1]).strip().upper()
                    d = a_numero_limpio(r.iloc[2])
                    if d > 0 and p not in ['NAN', 'NONE', '']: prods[p] = d
                dict_bases[base_name] = prods
            
            for col_idx in range(len(df_mezclas.columns) - 1):
                for row_idx in range(len(df_mezclas)):
                    val_name = str(df_mezclas.iloc[row_idx, col_idx]).strip().upper()
                    if val_name and val_name not in ['NAN', 'NONE', '']:
                        val_dosis = a_numero_limpio(df_mezclas.iloc[row_idx, col_idx+1])
                        if val_dosis > 0: dict_aditivos_dosis[val_name] = val_dosis

        if len(df_mezclas.columns) > 13:
            for _, row in df_mezclas.iterrows():
                f_n = str(row.iloc[12]).strip().upper() 
                f_s = str(row.iloc[13]).strip().upper() 
                if f_s and f_n not in ["", "NAN", "NONE", "FERTILIZANTES", "SIGLAS"]:
                    dict_fert[f_s] = f_n

        df_precios_master = pd.DataFrame()
        try:
            sh_precios = gc.open_by_url("https://docs.google.com/spreadsheets/d/1qZ4av-DH2oCJdgllBX27gdA2jEhT9bt2yv_sboORfSg/edit")
            precios_consolidados = []
            for ws in sh_precios.worksheets():
                datos_hoja = ws.get_all_values()
                if not datos_hoja: continue
                
                idx_header, col_anio, col_prod, col_precio_tipo = -1, -1, -1, -1
                for i in range(min(10, len(datos_hoja))):
                    fila_upper = [str(x).upper().strip() for x in datos_hoja[i]]
                    if 'AÑO' in fila_upper and 'PRODUCTO' in fila_upper:
                        idx_header = i
                        col_anio = fila_upper.index('AÑO')
                        col_prod = fila_upper.index('PRODUCTO')
                        col_precio_tipo = next((idx for idx, val in enumerate(fila_upper) if 'PRECIO' in val), -1)
                        break
                
                if idx_header != -1:
                    for row in datos_hoja[idx_header+1:]:
                        if col_precio_tipo != -1 and len(row) > col_precio_tipo:
                            if "DOSIS" in str(row[col_precio_tipo]).upper(): continue
                                
                        if len(row) > max(col_anio, col_prod):
                            anio_str = str(row[col_anio]).strip()
                            str_prod = str(row[col_prod]).strip().upper()
                            if anio_str.isdigit() and str_prod:
                                col_inicio = max(col_anio, col_prod) + 1
                                vals = [parsear_precio(v) for v in row[col_inicio:] if str(v).strip() != ""]
                                vals = [v for v in vals if v > 0]
                                prom = sum(vals)/len(vals) if vals else 0.0
                                if prom > 0:
                                    precios_consolidados.append({
                                        'AÑO': int(anio_str), 
                                        'PRODUCTO': str_prod, 
                                        'PROD_CLEAN': re.sub(r'[^\w]', '', str_prod), 
                                        'PRECIO': prom
                                    })
            df_precios_master = pd.DataFrame(precios_consolidados)
        except Exception: pass
            
        return df_t1, dict_bases, dict_aditivos_dosis, dict_fert, df_cfg, df_precios_master
    except Exception as e:
        st.error(f"🚨 Error de Extracción: {e}")
        return pd.DataFrame(), {}, {}, {}, pd.DataFrame(), pd.DataFrame()

def extraer_precios_maestros(df_cfg):
    precios = {}
    if df_cfg.empty: return precios
    c_p_i, c_c_i = 8, 9
    for i in range(min(5, len(df_cfg))):
        r_c = [str(x).upper().strip() for x in df_cfg.iloc[i].tolist()]
        if 'PRODUCTO' in r_c and 'COSTO' in r_c:
            c_p_i, c_c_i = r_c.index('PRODUCTO'), r_c.index('COSTO')
            break
    for r in range(len(df_cfg)):
        p = str(df_cfg.iloc[r, c_p_i]).upper().strip()
        c = parsear_precio(df_cfg.iloc[r, c_c_i])
        if p and p not in ["NAN", "NONE", ""]: precios[p] = c
    return precios

# =================================================================
# 🚀 EJECUCIÓN PRINCIPAL DEL SIMULADOR
# =================================================================
def ejecutar(purificar_lote, extraer_numero):
    VERDE_INTENSO = '#143521'
    DORADO = '#d4af37'
    
    st.markdown(f"""
    <style>
    .titulo-presupuesto {{ color: #0d1b2a; border-bottom: 3px solid {DORADO}; padding-bottom: 5px; font-family: 'Arial Black'; text-transform: uppercase; }}
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {{ border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.08) !important; }}
    
    .kpi-presupuesto {{ background-color: #0d1b2a; color: white; padding: 20px; border-radius: 10px; border-left: 6px solid #d4af37; box-shadow: 0 4px 6px rgba(0,0,0,0.2); margin-bottom: 15px; transition: transform 0.3s ease, box-shadow 0.3s ease;}}
    .kpi-presupuesto:hover {{ transform: translateY(-5px) scale(1.02); box-shadow: 0 10px 20px rgba(212, 175, 55, 0.3); border: 1px solid #d4af37;}}
    .kpi-titulo {{ color: #d4af37; font-weight: bold; font-size: 14px; margin-bottom: 5px; text-transform: uppercase; }}
    .kpi-valor {{ font-size: 28px; font-weight: 900; margin: 0; }}
    
    [data-testid="stPlotlyChart"] {{ transition: transform 0.3s ease, box-shadow 0.3s ease !important; border-radius: 8px; }}
    [data-testid="stPlotlyChart"]:hover {{ transform: translateY(-4px) scale(1.015) !important; box-shadow: 0 12px 25px rgba(212, 175, 55, 0.25) !important; z-index: 10; }}

    [data-testid="column"] {{
        display: flex !important;
        flex-direction: column !important;
        justify-content: flex-start !important;
        align-items: stretch !important;
    }}

    div[data-testid="stSelectbox"] > div, div[data-testid="stSelectbox"] div[data-baseweb="select"], div[data-testid="stNumberInput"] > div, div[data-testid="stTextInput"] > div {{
        background-color: #ffffff !important; border: 2px solid {VERDE_INTENSO} !important; border-radius: 8px !important; box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important;
    }}
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div, div[data-testid="stNumberInput"] div[data-baseweb="input"], div[data-testid="stTextInput"] div[data-baseweb="input"] {{ 
        background-color: transparent !important; border: none !important; 
    }}
    div[data-testid="stSelectbox"] *, div[data-testid="stNumberInput"] input, div[data-testid="stTextInput"] input {{
        color: #0d1b2a !important; font-weight: 900 !important; 
    }}
    div[data-testid="stMainBlockContainer"] label p {{ color: #0d1b2a !important; font-weight: 800 !important; text-transform: uppercase !important; }}
    </style>
    """, unsafe_allow_html=True)

    c_tit, c_btn = st.columns([3, 1])
    c_tit.markdown(f"<h1 class='titulo-presupuesto'>💰 Simulador Estratégico <span style='color:{DORADO}; font-size:16px;'>[V.GERENCIAL 42.1]</span></h1>", unsafe_allow_html=True)
    
    # 🎯 INICIALIZACIÓN DE LA MEMORIA DE EDICIÓN PERMANENTE
    if 'ediciones_usuario' not in st.session_state:
        st.session_state['ediciones_usuario'] = {}

    if c_btn.button("🧹 REINICIAR MEMORIA", type="primary", use_container_width=True):
        st.cache_data.clear()
        for key in list(st.session_state.keys()):
            if key in ['lab_raw_data', 'ediciones_usuario', 'laboratorio_v49', 'total_inercial_base', 'anio_presupuesto_guardado']:
                del st.session_state[key]
        st.toast("✅ Memoria de simulación purgada desde la raíz.", icon="🧹")
        st.rerun()

    st.write("Laboratorio interactivo para proyectar flujos de efectivo. Mueve la Inflación o Frecuencia y observa el impacto **en tiempo real**.")

    # ==========================================
    # FASE 1: PARÁMETROS BASE Y EXTRACCIÓN
    # ==========================================
    with st.expander("⚙️ 1. Definir Escenario Base (Cruce Histórico)", expanded=not bool('lab_raw_data' in st.session_state)):
        f1, f2, f3 = st.columns(3)
        meses_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
        opciones_mes = ["📊 AÑO COMPLETO (TODOS)"] + list(meses_dict.values())
        
        mes_sel = f1.selectbox("📅 Mes a Proyectar:", opciones_mes)
        pista_sel = f2.selectbox("📍 Base Operativa:", ["TODAS", "PLUC", "PORI", "PDIV", "TEHO", "LUCI"])
        anio_presupuesto = f3.selectbox("🎯 Año Objetivo:", [2026, 2027, 2028, 2029, 2030], index=1)
        
        f4, f5 = st.columns([1, 2])
        profundidad_sel = f4.selectbox("🔍 Base Histórica:", ["Último Año", "Últimos 2 Años", "Últimos 3 Años", "Histórico Completo"])
        
        with f5:
            st.markdown("<br>", unsafe_allow_html=True)
            btn_generar = st.button("🧬 1. EXTRAER BASE HISTÓRICA", type="primary", use_container_width=True)

    if btn_generar:
        with st.spinner("Compilando historia y cruzando precios en la Bóveda..."):
            df_t1_base, dict_bases, dict_aditivos_dosis, dict_fert, df_cfg, df_precios_master = descargar_y_masticar_bases()
            
            if df_t1_base.empty:
                st.error("🚨 No se pudo establecer conexión con las Bóvedas de Datos.")
                st.stop()

            # Limpiar ediciones anteriores al generar nueva base
            st.session_state['ediciones_usuario'] = {}

            df_t1 = df_t1_base.copy()
            dict_precios_backup = extraer_precios_maestros(df_cfg)
            anio_actual = datetime.now().year
            año_base = 2026 

            if profundidad_sel == "Último Año": df_t1 = df_t1[df_t1['AÑO'] >= (anio_actual - 1)]
            elif profundidad_sel == "Últimos 2 Años": df_t1 = df_t1[df_t1['AÑO'] >= (anio_actual - 2)]
            elif profundidad_sel == "Últimos 3 Años": df_t1 = df_t1[df_t1['AÑO'] >= (anio_actual - 3)]

            if mes_sel != "📊 AÑO COMPLETO (TODOS)":
                mes_num = next(k for k, v in meses_dict.items() if v == mes_sel)
                df_t1 = df_t1[df_t1['MES'] == mes_num]
            
            total_anios_boveda = df_t1['AÑO'].nunique()
            if total_anios_boveda == 0: total_anios_boveda = 1

            traductor_pistas = {"PLUC": "FUMIGARAY", "PORI": "AEROPENOR", "LUCI": "GENESYS", "TEHO": "AVIL", "PDIV": "ASA"}
            if pista_sel != "TODAS":
                pista_t1_esperada = traductor_pistas.get(pista_sel, pista_sel)
                df_t1 = df_t1[df_t1['PISTA_OPERATIVA'].str.contains(pista_t1_esperada, na=False)]

            consumo_esperado = {} 
            if not df_t1.empty:
                ha_total_por_coctel = df_t1.groupby(['PISTA_OPERATIVA', 'COCTEL_NOM'])['HA_CALCULO'].sum().reset_index()
                ha_total_por_coctel['HA_PROYECTADA'] = ha_total_por_coctel['HA_CALCULO'] / total_anios_boveda

                for _, row_c in ha_total_por_coctel.iterrows():
                    coctel_completo = str(row_c['COCTEL_NOM'])
                    ha_proyectada = row_c['HA_PROYECTADA']
                    receta_dict = extraer_receta_rapida(coctel_completo, dict_bases, dict_aditivos_dosis, dict_fert)
                    for prod_quimico, dosis in receta_dict.items():
                        consumo_esperado[prod_quimico] = consumo_esperado.get(prod_quimico, 0) + (dosis * ha_proyectada)

            resultados_raw = []
            precios_records = df_precios_master.to_dict('records') if not df_precios_master.empty else []

            for producto, volumen in consumo_esperado.items():
                if volumen > 0:
                    precio_hist_base = 0.0
                    anio_hist_match = año_base
                    p_clean = re.sub(r'[^\w]', '', producto.upper().strip())
                    
                    for r_db in precios_records:
                        if r_db['AÑO'] == año_base:
                            if p_clean in r_db['PROD_CLEAN'] or r_db['PROD_CLEAN'] in p_clean:
                                precio_hist_base = float(r_db['PRECIO'])
                                anio_hist_match = año_base
                                break
                    
                    if precio_hist_base == 0.0 and precios_records:
                        matches_hist = [r for r in precios_records if r['AÑO'] < año_base and (p_clean in r['PROD_CLEAN'] or r['PROD_CLEAN'] in p_clean)]
                        if matches_hist:
                            best_match = max(matches_hist, key=lambda x: x['AÑO'])
                            precio_hist_base = float(best_match['PRECIO'])
                            anio_hist_match = int(best_match['AÑO'])

                    if precio_hist_base == 0.0:
                        precio_bk = float(dict_precios_backup.get(producto, 0.0))
                        if precio_bk < 1000:
                            for p_bk, val_bk in dict_precios_backup.items():
                                bk_clean = re.sub(r'[^\w]', '', p_bk.upper().strip())
                                if p_clean in bk_clean or bk_clean in p_clean:
                                    if val_bk >= 1000: 
                                        precio_bk = float(val_bk)
                                        break
                        if precio_bk >= 1000:
                            precio_hist_base = precio_bk
                            anio_hist_match = anio_actual

                    resultados_raw.append({
                        "🧪 Insumo Químico": producto,
                        "vol_base_num": float(volumen),
                        "precio_base_num": float(precio_hist_base),
                        "anio_hist_match": anio_hist_match
                    })
            
            st.session_state['lab_raw_data'] = sorted(resultados_raw, key=lambda x: x["🧪 Insumo Químico"])
            st.session_state['anio_presupuesto_guardado'] = anio_presupuesto
            if "laboratorio_v49" in st.session_state:
                del st.session_state["laboratorio_v49"]

    # ==========================================
    # FASE 2: MOTOR DINÁMICO EN TIEMPO REAL
    # ==========================================
    if 'lab_raw_data' in st.session_state and st.session_state['lab_raw_data']:
        st.markdown("### 🧪 2. Variables Macroeconómicas y Laboratorio Interactivo")
        
        c_mac1, c_mac2 = st.columns(2)
        frecuencia_vuelos = c_mac1.number_input("✈️ Ajuste de Frecuencia/Ciclos (%)", min_value=-80, max_value=300, value=0, step=5, help="Impacta el volumen de todos los químicos.")
        inflacion_sel = c_mac2.number_input("💸 Inflación Anual Estimada (%)", min_value=0.0, max_value=50.0, value=8.0, step=1.0, help="Proyecta los precios históricos al año objetivo.")

        llave_editor = "laboratorio_v49"

        # 1. Construir la base matemática
        df_build = pd.DataFrame(st.session_state['lab_raw_data'])
        anio_presupuesto_actual = st.session_state.get('anio_presupuesto_guardado', datetime.now().year + 1)
        
        # 2. Aplicar Macros Dinámicos (Frecuencia e Inflación)
        factor_frec = 1 + (frecuencia_vuelos / 100.0)
        df_build['vol_sist_num'] = df_build['vol_base_num'] * factor_frec
        
        def calc_precio_inflado(row):
            anios = max(0, anio_presupuesto_actual - row['anio_hist_match'])
            return row['precio_base_num'] * ((1 + inflacion_sel / 100.0) ** anios)
            
        df_build['precio_sist_num'] = df_build.apply(calc_precio_inflado, axis=1)

        # 3. Preparar Columnas UI (Valores por Defecto)
        df_build['✅ Activo'] = True
        df_build['📦 Vol. Sist. (Base)'] = df_build['vol_sist_num']
        df_build['🎯 Ajuste Vol. (%)'] = 0.0
        df_build['💵 Precio Base (Histórico)'] = df_build['precio_base_num']
        df_build['📈 Precio Sist. (+Inflación)'] = df_build['precio_sist_num']
        df_build['🎯 Precio Irreal (Modificable)'] = df_build['precio_sist_num'].round(0)

        # 🎯 4. CAPTURAR Y APLICAR EDICIONES PERMANENTES DEL USUARIO
        if llave_editor in st.session_state:
            nuevas_ediciones = st.session_state[llave_editor].get("edited_rows", {})
            for r_idx_str, changes in nuevas_ediciones.items():
                r_idx = int(r_idx_str)
                if r_idx not in st.session_state['ediciones_usuario']:
                    st.session_state['ediciones_usuario'][r_idx] = {}
                st.session_state['ediciones_usuario'][r_idx].update(changes)

        # Aplicamos la memoria guardada sobre el DataFrame antes de calcular los totales
        for r_idx, changes in st.session_state['ediciones_usuario'].items():
            for col, val in changes.items():
                if col in df_build.columns:
                    df_build.loc[r_idx, col] = val

        # 5. Cálculos Finales con la data ya editada
        df_build['vol_final_num'] = df_build['📦 Vol. Sist. (Base)'] * (1 + pd.to_numeric(df_build['🎯 Ajuste Vol. (%)'], errors='coerce').fillna(0) / 100.0)
        df_build['vol_final_num'] = df_build.apply(lambda r: float(r['vol_final_num']) if r['✅ Activo'] else 0.0, axis=1)
        df_build['subtotal_num'] = df_build['vol_final_num'] * pd.to_numeric(df_build['🎯 Precio Irreal (Modificable)'], errors='coerce').fillna(0)

        df_build['📊 Vol. Final (Calc)'] = df_build['vol_final_num']
        df_build['💰 Subtotal (Calc)'] = df_build['subtotal_num']

        cols_ordenadas = [
            "✅ Activo", "🧪 Insumo Químico", 
            "📦 Vol. Sist. (Base)", "🎯 Ajuste Vol. (%)", "📊 Vol. Final (Calc)", 
            "💵 Precio Base (Histórico)", "📈 Precio Sist. (+Inflación)", "🎯 Precio Irreal (Modificable)", 
            "💰 Subtotal (Calc)", 
            "vol_sist_num", "precio_sist_num", "precio_base_num", "vol_final_num", "subtotal_num"
        ]
        
        st.caption("💡 Apaga el interruptor '✅ Activo' para eliminar una molécula del presupuesto. Modifica el **Ajuste Vol.** o el **Precio Modificable** para simular negociaciones.")

        df_editado = st.data_editor(
            df_build[cols_ordenadas],
            column_config={
                "vol_sist_num": None,
                "precio_sist_num": None,
                "precio_base_num": None,
                "vol_final_num": None,
                "subtotal_num": None,
                "✅ Activo": st.column_config.CheckboxColumn("Activo"),
                "🧪 Insumo Químico": st.column_config.TextColumn("Molécula / Insumo"),
                "📦 Vol. Sist. (Base)": st.column_config.NumberColumn("📦 Vol. Sist. (Base)", format="%,.1f"),
                "🎯 Ajuste Vol. (%)": st.column_config.NumberColumn("🎯 Ajuste Vol. (%)", format="%d %%", step=1.0),
                "📊 Vol. Final (Calc)": st.column_config.NumberColumn("📊 Vol. Proyectado", format="%,.1f"),
                "💵 Precio Base (Histórico)": st.column_config.NumberColumn("💵 Precio Histórico", format="$ %,.0f"),
                "📈 Precio Sist. (+Inflación)": st.column_config.NumberColumn("📈 Precio (+Inflación)", format="$ %,.0f"),
                "🎯 Precio Irreal (Modificable)": st.column_config.NumberColumn("🎯 Precio Modificable", min_value=0.0, format="$ %,.0f"),
                "💰 Subtotal (Calc)": st.column_config.NumberColumn("💰 Subtotal", format="$ %,.0f")
            },
            disabled=["🧪 Insumo Químico", "📦 Vol. Sist. (Base)", "📊 Vol. Final (Calc)", "💵 Precio Base (Histórico)", "📈 Precio Sist. (+Inflación)", "💰 Subtotal (Calc)"],
            hide_index=True,
            use_container_width=True,
            key=llave_editor
        )

        with st.expander("➕ Inyectar Nueva Molécula al Escenario"):
            i1, i2, i3, i4 = st.columns([2, 1, 1, 1])
            nuevo_nombre = i1.text_input("Nombre del Insumo", placeholder="Ej: NUEVO FUNGICIDA X")
            nuevo_vol = i2.number_input("Volumen Proyectado", min_value=0.0, value=100.0, step=10.0)
            nuevo_precio = i3.number_input("Precio Estimado", min_value=0.0, value=50000.0, step=5000.0)
            
            with i4:
                st.markdown("<br>", unsafe_allow_html=True) 
                if st.button("Inyectar Molécula", type="secondary", use_container_width=True):
                    if nuevo_nombre:
                        vol_base_inyectado = float(nuevo_vol) / (1 + (frecuencia_vuelos / 100.0))
                        st.session_state['lab_raw_data'].append({
                            "🧪 Insumo Químico": nuevo_nombre.upper(),
                            "vol_base_num": vol_base_inyectado,
                            "precio_base_num": float(nuevo_precio),
                            "anio_hist_match": anio_presupuesto_actual 
                        })
                        st.rerun()

        # ==========================================
        # FASE 3: DASHBOARD DE CONTRASTE GERENCIAL
        # ==========================================
        st.markdown("---")
        st.markdown("### 📊 3. Panel de Contraste Gerencial")
        
        total_inercial = (df_build['vol_sist_num'] * df_build['precio_sist_num']).sum()
        
        df_estrategico = df_editado[df_editado["✅ Activo"] == True].copy()
        total_estrategico = df_estrategico['subtotal_num'].sum()
        
        diferencia = total_estrategico - total_inercial
        pct_dif = (diferencia / total_inercial) * 100 if total_inercial > 0 else 0.0

        col_k1, col_k2, col_k3 = st.columns(3)
        
        col_k1.markdown(f"""
        <div class='kpi-presupuesto' style='border-left-color: #6c757d;'>
            <div class='kpi-titulo' style='color: #a0aec0;'>Presupuesto Inercial (Tradicional)</div>
            <p class='kpi-valor'>$ {fmt_latino(total_inercial, 0)}</p>
        </div>
        """, unsafe_allow_html=True)

        col_k2.markdown(f"""
        <div class='kpi-presupuesto' style='border-left-color: #d4af37;'>
            <div class='kpi-titulo'>Presupuesto Estratégico (Modificado)</div>
            <p class='kpi-valor'>$ {fmt_latino(total_estrategico, 0)}</p>
        </div>
        """, unsafe_allow_html=True)

        color_dif = "#28a745" if diferencia <= 0 else "#dc3545"
        texto_dif = "AHORRO" if diferencia <= 0 else "SOBRECOSTO"
        signo_dif = "" if diferencia <= 0 else "+"
        
        col_k3.markdown(f"""
        <div class='kpi-presupuesto' style='border-left-color: {color_dif};'>
            <div class='kpi-titulo' style='color: {color_dif};'>Brecha ({texto_dif})</div>
            <p class='kpi-valor' style='color: {color_dif};'>$ {fmt_latino(abs(diferencia), 0)}</p>
            <p style='margin:0; font-size:14px; font-weight:bold; color: {color_dif};'>{signo_dif}{pct_dif:.1f}% vs Inercial</p>
        </div>
        """, unsafe_allow_html=True)

        g_col1, g_col2 = st.columns(2)
        
        df_comp = pd.DataFrame({
            "Escenario": ["Tradicional (Inercial)", "Estratégico (Modificado)"],
            "Presupuesto": [total_inercial, total_estrategico]
        })
        fig_bar = px.bar(df_comp, x="Escenario", y="Presupuesto", text=df_comp['Presupuesto'].apply(lambda x: f"${fmt_latino(x,0)}"),
                         color="Escenario", color_discrete_map={"Tradicional (Inercial)": "#6c757d", "Estratégico (Modificado)": "#d4af37"},
                         title="<b>Comparativo Global</b>")
        fig_bar.update_traces(textposition='auto', textfont_size=14, textfont_color="white")
        fig_bar.update_layout(showlegend=False, yaxis_title="COP", xaxis_title="", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        g_col1.plotly_chart(fig_bar, use_container_width=True)

        df_pie = df_estrategico.copy().sort_values('subtotal_num', ascending=False)
        if len(df_pie) > 7:
            top_6 = df_pie.head(6)
            otros_sub = df_pie.iloc[6:]['subtotal_num'].sum()
            fila_otros = pd.DataFrame([{"🧪 Insumo Químico": "OTROS INSUMOS", "subtotal_num": otros_sub}])
            df_pie = pd.concat([top_6, fila_otros], ignore_index=True)
            
        fig_pie = px.pie(df_pie, values='subtotal_num', names='🧪 Insumo Químico', hole=0.45, 
                         title="<b>Peso Financiero por Molécula (Estratégico)</b>", color_discrete_sequence=px.colors.qualitative.Prism)
        fig_pie.update_traces(textposition='inside', textinfo='percent+label')
        fig_pie.update_layout(showlegend=False, margin=dict(t=40, b=0, l=0, r=0), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
        g_col2.plotly_chart(fig_pie, use_container_width=True)

        st.markdown("---")
        
        # 💥 EXPORTACIÓN A EXCEL VIP
        df_export = df_estrategico.copy()
        df_export = df_export[[
            "🧪 Insumo Químico", 
            "vol_final_num", 
            "precio_base_num", 
            "🎯 Precio Irreal (Modificable)", 
            "subtotal_num"
        ]]
        
        df_export.columns = [
            "PRODUCTO", 
            "VOLUMEN PROYECTADO (L/Kg)", 
            "PRECIO BASE HISTÓRICO", 
            "PRECIO UNITARIO PROYECTADO", 
            "PRESUPUESTO TOTAL"
        ]
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Estrategia_Presupuesto', startrow=2, index=False)
            
            workbook = writer.book
            ws = workbook['Estrategia_Presupuesto']
            
            ws['A1'] = f"REPORTE ESTRATÉGICO DE PRESUPUESTO - {anio_presupuesto_actual}"
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="143521", end_color="143521", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A1:E2')
            
            header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
            header_font = Font(bold=True, color="D4AF37")
            borde_fino = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                                top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

            for col in range(1, 6):
                cell = ws.cell(row=3, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = borde_fino
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
            max_row = ws.max_row
            for row in range(4, max_row + 1):
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                for col in [3, 4, 5]:
                    ws.cell(row=row, column=col).number_format = '"$" #,##0'
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = borde_fino
            
            ws.cell(row=max_row + 1, column=1, value="TOTAL GENERAL").font = Font(bold=True, size=12)
            ws.cell(row=max_row + 1, column=1).alignment = Alignment(horizontal='right')
            ws.cell(row=max_row + 1, column=5, value=f"=SUM(E4:E{max_row})").font = Font(bold=True, size=12)
            ws.cell(row=max_row + 1, column=5).number_format = '"$" #,##0'
            ws.cell(row=max_row + 1, column=5).border = borde_fino
            
            for col_letter, width in zip(["A", "B", "C", "D", "E"], [35, 30, 25, 30, 25]):
                ws.column_dimensions[col_letter].width = width
                for cell in ws[col_letter]:
                    if cell.row > 3:
                        if col_letter == "A":
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.freeze_panes = 'A4'
        
        st.download_button(
            label="💾 DESCARGAR ESTRATEGIA EN EXCEL (PARA JUNTA DIRECTIVA)",
            data=buffer.getvalue(),
            file_name=f"Estrategia_Presupuesto_Gerencia_{anio_presupuesto_actual}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

if __name__ == "__main__":
    pass
