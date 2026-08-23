import streamlit as st
import pandas as pd
import gspread
from datetime import datetime, timedelta, date
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- 🔌 CONEXIÓN Y TIEMPO ---
def obtener_hora_colombia():
    return datetime.utcnow() + timedelta(hours=-5)

@st.cache_resource(show_spinner=False)
def inicializar_cliente_gspread():
    try:
        if "gcp_service_account" in st.secrets:
            return gspread.service_account_from_dict(dict(st.secrets["gcp_service_account"]))
        return gspread.service_account(filename='credenciales.json')
    except Exception: return None

@st.cache_data(show_spinner=False, ttl=300)
def obtener_datos_bovedas():
    gc = inicializar_cliente_gspread()
    if not gc: return None, None, None, None, "No hay conexión con Google Cloud"
    URL_ING = "https://docs.google.com/spreadsheets/d/1G_bt4nFudeqqTmRbK-pF52w_9-L_Jf5uNCFeQKIPuO0/edit"
    URL_TRA = "https://docs.google.com/spreadsheets/d/1JV-f8zzGuhGNlqvrSjeKYN4eBdshAN5EOkfDHMi1WIs/edit"
    try:
        sh_ing = gc.open_by_url(URL_ING)
        ws_ing = sh_ing.worksheets()[0]
        datos_ing = ws_ing.get_all_values()
        try: datos_dicc = sh_ing.worksheet("DICCIONARIO").get_all_values()
        except: datos_dicc = []
        sh_tras = gc.open_by_url(URL_TRA)
        ws_tras = sh_tras.worksheets()[0] 
        datos_tras = ws_tras.get_all_values()
        titulo_tras = ws_tras.title 
        return datos_ing, datos_dicc, datos_tras, titulo_tras, None
    except Exception as e: return None, None, None, None, str(e)

# --- 🔍 RASTREADOR DE MATERIALES 100% ESTRICTO ---
def extraer_mapeo_materiales():
    gc = inicializar_cliente_gspread()
    if not gc: return {"ERROR": "Sin conexión a los servidores de Google."}
    try:
        sh = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        ws = sh.worksheet("Plantilla")
        datos = ws.get_all_values()
        mapeo = {}
        for row in datos[2:]:
            if len(row) >= 11:
                mat = str(row[0]).strip()
                desc_j = str(row[9]).strip().upper()
                desc_k = str(row[10]).strip().upper()
                if mat and mat != "MATERIAL":
                    if desc_k: mapeo[re.sub(r'\s+', ' ', desc_k)] = mat
                    if desc_j: mapeo[re.sub(r'\s+', ' ', desc_j)] = mat
        return mapeo
    except Exception as e: 
        return {"ERROR": str(e)}

# 💥 BÚSQUEDA EXACTA: Cero aproximaciones permitidas
def buscar_codigo_material(producto_nombre, mapeo):
    if "ERROR" in mapeo: return "S/N"
    prod_clean = re.sub(r'\s+', ' ', str(producto_nombre).strip().upper())
    if not prod_clean or not mapeo: return "S/N"
    if prod_clean in mapeo: return mapeo[prod_clean]
    return "S/N"

# 💥 RADAR CRONOLÓGICO Y ANTI-FALLOS
def procesar_fecha_estricta(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() in ["none", "nan", "nat", "<na>"]: return pd.NaT
    s = str(val).strip().lower()
    if s.replace('.', '', 1).isdigit(): return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    meses_es = {'enero':1, 'febrero':2, 'marzo':3, 'abril':4, 'mayo':5, 'junio':6, 'julio':7, 'agosto':8, 'septiembre':9, 'octubre':10, 'noviembre':11, 'diciembre':12}
    mes_encontrado = next((meses_es[m] for m in meses_es if m in s), None)
    if mes_encontrado:
        numeros = re.findall(r'\d+', s)
        if len(numeros) >= 2:
            n1, n2 = int(numeros[0]), int(numeros[1])
            anio = n1 if n1 > 1000 else (n2 if n2 > 1000 else (2000 + n2 if n2 < 100 else n2))
            dia = n2 if n1 > 1000 else (n1 if n2 > 1000 else n1)
            try: return pd.Timestamp(year=anio, month=mes_encontrado, day=dia)
            except: pass
    for dia_sem in ['lunes','martes','miércoles','miercoles','jueves','viernes','sábado','sabado','domingo']: s = s.replace(dia_sem, '')
    s = s.replace(',', '').replace(' de ', '/').replace('-', '/').strip()
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: 
        res = pd.to_datetime(s, dayfirst=True)
        return pd.NaT if pd.isna(res) else res
    except: return pd.NaT 

def formatear_numero_sap(val):
    try:
        f_val = float(str(val).replace(",", ""))
        if f_val.is_integer(): return f"{int(f_val):,}".replace(",", ".")
        val_str = f"{f_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return val_str[:-3] if val_str.endswith(",00") else val_str
    except: return str(val)

def estandarizar_pista(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() in ["none", "nan", "nat", "<na>"]: return ""
    return str(val).strip().upper().replace("ORIHUECA", "PORI").replace("DIVAS", "PDIV").replace("TEHOBROMINA", "TEHO").replace("LUCHA", "PLUC")

@st.cache_data(show_spinner=False, ttl=3600)
def extraer_catalogo_oficial_sap():
    gc = inicializar_cliente_gspread()
    if not gc: return []
    try:
        sh_config = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        ws = sh_config.worksheet("Configuración") if "Configuración" in [w.title for w in sh_config.worksheets()] else sh_config.worksheet("DD_Mesclas")
        datos = ws.get_all_values()
        if not datos: return []
        productos_oficiales = set()
        idx_prod = -1
        for i in range(min(10, len(datos))):
            if 'PRODUCTO' in [str(x).upper().strip() for x in datos[i]]:
                idx_prod = [str(x).upper().strip() for x in datos[i]].index('PRODUCTO')
                break
        if idx_prod != -1:
            for row in datos[i+1:]:
                if len(row) > idx_prod:
                    p_nombre = re.sub(r'\s+', ' ', str(row[idx_prod]).strip().upper())
                    if p_nombre and len(p_nombre) > 2 and p_nombre != "0": productos_oficiales.add(p_nombre)
        return list(productos_oficiales)
    except: return []

DICT_BASE_PRODUCTOS = {"ACEITE DICAM": "ROYAL BIOCHEM", "ACONDICIONADOR SV": "SYS TECNOLOGIES", "ADHERENTE SV": "SYS TECNOLOGIES", "BANADAK": "PLANDAK", "BANANO Y PLATANO * LT": "INVESA S.A.S.", "BANATREL SC": "YARA S.A.S.", "BOSCALID 50 WG": "DVA COLOMBIA", "CERAQUINT SP": "CERADIS COLOMBIA", "CEROSTRESS SV * LT": "MICROFERTIZA", "COMPER SV": "ADAMA", "EPOXICONAZOLE DEL MONTE": "DEL MONTE SAS", "FENTRIUPH AGRO 88 OL": "DEL MONTE SAS", "FOSFOSTRESS SV": "MICROFERTIZA", "GLOBAFOL nf": "SYNGENTA", "IMBIOSIL O": "INBIOMA", "KURDO 250 EC": "INVESA S.A.S.", "KYVENTIQ": "CORTEVA", "LONSELOR 30 SC": "BASF QUÍMICA", "MANCOL 430 SC": "CASAGRO", "NATURAMIN WSP": "AGRIANDES DAINSA", "OPORTO": "ADAMA", "OPUS 12 EC": "BASF QUÍMICA", "POLYTHION SC": "UPL", "POWMYL SV": "SUMITOMO", "QUELAMIX": "INGEPLANT", "REFLECT": "SYNGENTA", "ROUTINE SC": "BAYER", "SEEKER": "SYNGENTA", "SICO": "SYNGENTA", "SIGANEX 60 SC": "BAYER", "SPRAYFIX": "AGRIANDES DAINSA", "THIOPRON 825 SC": "UPL", "TIMOREX PRO": "ADAMA", "XILOTROM": "AGRIFOL", "ZINTRAC x LITRO SV": "YARA S.A.S."}

# --- 🚀 EJECUCIÓN DEL MÓDULO ---
def ejecutar():
    hoy_colombia = obtener_hora_colombia().date()
    if 'form_key_m19' not in st.session_state: st.session_state['form_key_m19'] = 0
    if 'form_key_m19_traslados' not in st.session_state: st.session_state['form_key_m19_traslados'] = 0

    def limpiar_campos_operativos(): st.session_state['form_key_m19'] += 1
    def limpiar_campos_traslados(): st.session_state['form_key_m19_traslados'] += 1

    st.markdown("<div id='inicio-modulo-19'></div>", unsafe_allow_html=True)
    
    # 💥 REPARACIÓN CSS: SELECTORES CON BORDES FORZADOS EXACTOS
    st.markdown("""
    <style>
    .titulo-mod { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; text-transform: uppercase; }
    .kpi-card { background-color: #0d1b2a; color: white; padding: 20px; border-radius: 10px; border-left: 6px solid #d4af37; box-shadow: 0 4px 6px rgba(0,0,0,0.2); margin-bottom: 15px; }
    .kpi-rojo { border-left-color: #dc3545; } .kpi-amarillo { border-left-color: #ffc107; } .kpi-verde { border-left-color: #28a745; } .kpi-azul { border-left-color: #2F75B5; }
    .kpi-titulo { font-weight: bold; font-size: 14px; margin-bottom: 5px; text-transform: uppercase; color: #a0aec0; }
    .kpi-valor { font-size: 28px; font-weight: 900; margin: 0; color: white; }
    
    div[data-testid="stExpander"] { border: 2px solid #0d1b2a !important; border-radius: 8px !important; box-shadow: 0 4px 6px rgba(0,0,0,0.15) !important; background-color: #ffffff !important; margin-bottom: 20px !important; }
    div[data-testid="stExpander"] summary { background-color: #0d1b2a !important; border-radius: 6px 6px 0px 0px !important; padding: 10px 15px !important; }
    div[data-testid="stExpander"] summary p { color: #d4af37 !important; font-family: 'Arial Black', sans-serif !important; font-size: 15px !important; text-transform: uppercase !important; margin: 0 !important; }
    div[data-testid="stMainBlockContainer"] label p { color: #0d1b2a !important; font-weight: 900 !important; text-transform: uppercase !important; font-size: 13px !important; }
    
    /* 💥 CAJAS EXTERNAS: TEXTO, NÚMEROS Y FECHAS */
    div[data-testid="stTextInput"] > div, 
    div[data-testid="stNumberInput"] > div, 
    div[data-testid="stDateInput"] > div { 
        background-color: #ffffff !important; 
        border: 2px solid #0d1b2a !important; 
        border-radius: 6px !important; 
    }
    
    /* INTERIORES TRANSPARENTES PARA NO DUPLICAR BORDES */
    div[data-testid="stTextInput"] input, 
    div[data-testid="stNumberInput"] input, 
    div[data-testid="stDateInput"] input { 
        color: #0d1b2a !important; 
        font-weight: 900 !important; 
        border: none !important; 
        background-color: transparent !important; 
        box-shadow: none !important;
    }
    
    /* 💥 CAJAS EXTERNAS: SELECTORES DROPDOWN FORZADOS AL 100% */
    div[data-baseweb="select"] {
        border: 2px solid #0d1b2a !important; 
        border-radius: 6px !important; 
        background-color: #ffffff !important; 
    }
    
    div[data-baseweb="select"] > div {
        border: none !important; 
        background-color: transparent !important; 
    }
    
    div[data-baseweb="select"] * { 
        color: #0d1b2a !important; 
        font-weight: 900 !important; 
    }
    
    div[data-testid="stDataEditor"], div[data-testid="stDataFrame"] { border: 3px solid #0d1b2a !important; border-radius: 8px !important; box-shadow: 0px 6px 15px rgba(0,0,0,0.15) !important; background-color: #ffffff !important; }
    div[data-testid="stTabs"] button[role="tab"] { font-family: 'Arial Black', sans-serif; font-size: 14px; text-transform: uppercase; color: #0d1b2a; }
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { border-bottom-color: #d4af37; background-color: rgba(212, 175, 55, 0.1); }
    .btn-ascensor { display: block; width: 100%; text-align: center; background-color: #15283c; color: #d4af37 !important; padding: 12px; border-radius: 8px; text-decoration: none !important; font-weight: 900; border: 2px solid #d4af37; margin-bottom: 20px; box-shadow: 0px 4px 6px rgba(0,0,0,0.2); transition: all 0.3s ease; }
    .btn-ascensor:hover { background-color: #0d1b2a; box-shadow: 0px 0px 10px rgba(212, 175, 55, 0.8); }
    .banner-anulacion { background-color: #721c24; color: #ffffff; padding: 12px; border-radius: 8px; border-left: 6px solid #f5c6cb; font-weight: bold; margin-bottom: 15px; }
    </style>
    """, unsafe_allow_html=True)

    URL_SHEET_INGRESOS = "https://docs.google.com/spreadsheets/d/1G_bt4nFudeqqTmRbK-pF52w_9-L_Jf5uNCFeQKIPuO0/edit"
    URL_SHEET_TRASLADOS = "https://docs.google.com/spreadsheets/d/1JV-f8zzGuhGNlqvrSjeKYN4eBdshAN5EOkfDHMi1WIs/edit"

    c_tit, c_btn1, c_btn2 = st.columns([2, 1, 1])
    c_tit.markdown("<h1 class='titulo-mod'>📦 19. Centro Logístico Unificado</h1>", unsafe_allow_html=True)
    if c_btn1.button("🔄 REFRESCAR RADARES", type="primary", use_container_width=True):
        st.cache_data.clear(); st.rerun()
    if c_btn2.button("🧹 PURGAR DICCIONARIO", type="secondary", use_container_width=True):
        with st.spinner("Aniquilando diccionario oculto..."):
            gc = inicializar_cliente_gspread()
            if gc:
                try:
                    sh_local = gc.open_by_url(URL_SHEET_INGRESOS)
                    ws_dicc = sh_local.worksheet("DICCIONARIO")
                    sh_local.del_worksheet(ws_dicc)
                    st.cache_data.clear(); st.success("✅ ¡Fantasma destruido!"); st.rerun()
                except Exception: st.info("No había basura. La lista ya estaba limpia.")

    with st.spinner("📡 Sincronizando Bóvedas de Ingresos y Traslados (Caché Activo)..."):
        datos_ing_crudos, datos_dicc_crudos, datos_tras_crudos, titulo_ws_traslados, error_api = obtener_datos_bovedas()
        if error_api:
            st.error("🚨 **LÍMITE DE PROTECCIÓN DE GOOGLE (Error 429)**: Espera de 1 a 2 minutos y presiona 'Refrescar Radares'." if "429" in error_api else f"🚨 Error de acceso a las Bóvedas. Detalle: {error_api}")
            return

        mapeo_materiales = extraer_mapeo_materiales()
        
        dict_operativo = {k.upper().strip(): v.upper().strip() for k, v in DICT_BASE_PRODUCTOS.items()}
        for row in datos_dicc_crudos[1:]:
            if len(row) >= 2 and str(row[0]).strip():
                p_nube = re.sub(r'\s+', ' ', str(row[0]).strip().upper())
                p_prov = str(row[1]).strip().upper()
                if len(p_nube) > 3: dict_operativo[p_nube] = p_prov
                
        productos_precio_sap = extraer_catalogo_oficial_sap()
        for prod_sap in productos_precio_sap:
            if prod_sap.strip() not in dict_operativo: dict_operativo[prod_sap.strip()] = "" 
        lista_autorizada = set([re.sub(r'\s+', ' ', str(k).strip().upper()) for k in dict_operativo.keys()])
        for p in productos_precio_sap: lista_autorizada.add(re.sub(r'\s+', ' ', str(p).strip().upper()))
            
        cache_mapeo = {}
        
        def estandarizar_y_marcar_inteligente(prod):
            p_clean = re.sub(r'\s+', ' ', str(prod).strip().upper())
            if p_clean in ["", "NONE", "NAN", "NAT", "<NA>"]: return ""
            if p_clean in cache_mapeo: return cache_mapeo[p_clean]
            
            if p_clean in lista_autorizada: 
                resultado = p_clean
            else:
                resultado = f"{p_clean} 🛑 [NO RECONOCIDO]"
                
            cache_mapeo[p_clean] = resultado
            return resultado

        # ================= PROCESAMIENTO TRASLADOS =================
        df_traslados = pd.DataFrame()
        encabezados_limpios_tras = []
        if datos_tras_crudos:
            columnas_oficiales = ["CONSECUTIVO", "FECHA", "PRODUCTO", "CANTIDAD", "UNIDAD", "PISTA", "SEMANA", "OBSERVACION", "LOTE"]
            idx_head_t = next((i for i, row in enumerate(datos_tras_crudos[:15]) if "CONSECUTIVO" in [str(x).upper().strip() for x in row]), -1)
            if idx_head_t != -1:
                enc_brutos = [str(x).strip().upper() for x in datos_tras_crudos[idx_head_t]]
                for j, h in enumerate(enc_brutos):
                    if "OBSERVAC" in h or ("COLUMNA1" in h and "OBSERVACION" not in encabezados_limpios_tras): h = "OBSERVACION"
                    elif "LOTE" in h: h = "LOTE"
                    elif h == "": h = f"COL_VACIA_{j}"
                    encabezados_limpios_tras.append(h)
                
                if "OBSERVACION" not in encabezados_limpios_tras: encabezados_limpios_tras.append("OBSERVACION")
                if "LOTE" not in encabezados_limpios_tras: encabezados_limpios_tras.append("LOTE")
                
                max_cols_t = len(encabezados_limpios_tras)
                datos_recortados = []
                for row in datos_tras_crudos[idx_head_t+1:]:
                    pad_row = row + [""] * (max_cols_t - len(row))
                    pad_row = ["" if str(x).strip().lower() in ["none", "nan", "null", "nat"] else str(x).strip() for x in pad_row[:max_cols_t]]
                    datos_recortados.append(pad_row)
                    
                df_t = pd.DataFrame(datos_recortados, columns=encabezados_limpios_tras)
                df_t = df_t.loc[:, ~df_t.columns.str.startswith('COL_VACIA_')]
                df_t = df_t.loc[:, ~df_t.columns.duplicated()]
                df_t['FILA_EXCEL'] = range(idx_head_t + 2, len(df_t) + idx_head_t + 2)
                cols_presentes = [c for c in columnas_oficiales if c in df_t.columns] + ['FILA_EXCEL']
                df_traslados = df_t[cols_presentes]
                if "PRODUCTO" in df_traslados.columns: df_traslados["PRODUCTO"] = df_traslados["PRODUCTO"].apply(estandarizar_y_marcar_inteligente)
                if "PISTA" in df_traslados.columns: df_traslados["PISTA"] = df_traslados["PISTA"].apply(estandarizar_pista)

        # ================= PROCESAMIENTO INGRESOS =================
        df_ingresos = pd.DataFrame()
        encabezados_limpios_ing = []
        if datos_ing_crudos:
            idx_head_i = next((i for i, row in enumerate(datos_ing_crudos[:5]) if "ESTADO / OBSERVACIÓN" in [str(x).upper().strip() for x in row] or "PRODUCTO" in [str(x).upper().strip() for x in row]), 0)
            enc_brutos_i = [str(x).strip().upper() for x in datos_ing_crudos[idx_head_i]]
            for j, h in enumerate(enc_brutos_i):
                if h == "": h = f"COL_VACIA_{j}"
                encabezados_limpios_ing.append(h)
                
            max_cols_ing = len(encabezados_limpios_ing)
            datos_ing_recortados = []
            for row in datos_ing_crudos[idx_head_i+1:]:
                pad_row = row + [""] * (max_cols_ing - len(row))
                pad_row = ["" if str(x).strip().lower() in ["none", "nan", "null", "nat"] else str(x).strip() for x in pad_row[:max_cols_ing]]
                datos_ing_recortados.append(pad_row)
                
            df_i = pd.DataFrame(datos_ing_recortados, columns=encabezados_limpios_ing)
            df_i = df_i.loc[:, ~df_i.columns.str.startswith('COL_VACIA_')]
            df_i = df_i.loc[:, ~df_i.columns.duplicated()]
            df_i['FILA_EXCEL'] = range(idx_head_i + 2, len(df_i) + idx_head_i + 2)
            if "PRODUCTO" in df_i.columns: df_i["PRODUCTO"] = df_i["PRODUCTO"].apply(estandarizar_y_marcar_inteligente)
            if "PISTA" in df_i.columns: df_i["PISTA"] = df_i["PISTA"].apply(estandarizar_pista)
            df_ingresos = df_i

    tab_ingresos, tab_traslados = st.tabs(["📥 1. INGRESOS (COMPRAS/PROVEEDOR)", "🚚 2. MOVIMIENTOS INTERNOS (TRASLADOS)"])

    # ========================================================================
    # 📥 PESTAÑA 1: INGRESOS 
    # ========================================================================
    with tab_ingresos:
        st.markdown(f"<a href='{URL_SHEET_INGRESOS}' target='_blank' class='btn-ascensor' style='background-color:#1e4620; border-color:#2e7d32; color:#ffffff !important;'>👁️ VER BASE DE INGRESOS EN GOOGLE SHEETS</a>", unsafe_allow_html=True)
        st.write("Panel táctico de ingresos. Registra lotes de proveedores externos cruzando información oficial con la Base de Precios SAP.")

        if df_ingresos.empty:
            st.warning("La base de datos de ingresos parece estar vacía.")
        else:
            df = df_ingresos
            COL_ESTADO = "ESTADO / OBSERVACIÓN"
            col_producto = next((c for c in df.columns if "PRODUCTO" in c), None)
            col_pista = next((c for c in df.columns if "PISTA" in c), None)
            col_fecha_ingreso = next((c for c in df.columns if "FECHA DE INGRESO" in c), None)
            col_fv = next((c for c in df.columns if c in ["F/V", "FECHA VENCIMIENTO", "VENCIMIENTO"]), None)

            if COL_ESTADO not in df.columns: st.error(f"🚨 FALTA COLUMNA TÁCTICA: No se encontró la columna **{COL_ESTADO}**.")
            else:
                idx_col_estado = encabezados_limpios_ing.index(COL_ESTADO) + 1 
                st.markdown("### 📡 Radares de Vencimiento")
                limite_90_dias = pd.to_datetime(hoy_colombia) + pd.to_timedelta(90, unit='D')
                hoy_ts = pd.to_datetime(hoy_colombia)
                
                lotes_vencidos, lotes_riesgo = 0, 0
                if col_fv:
                    df['FECHA_VENC_DT'] = df[col_fv].apply(procesar_fecha_estricta)
                    df_activos = df[~df[COL_ESTADO].str.contains("ANULADO|ELIMINAR", na=False, case=False)]
                    lotes_vencidos = df_activos[df_activos['FECHA_VENC_DT'] < hoy_ts].shape[0]
                    lotes_riesgo = df_activos[(df_activos['FECHA_VENC_DT'] >= hoy_ts) & (df_activos['FECHA_VENC_DT'] <= limite_90_dias)].shape[0]

                k1, k2, k3 = st.columns(3)
                k1.markdown(f"<div class='kpi-card kpi-verde'><div class='kpi-titulo'>📦 Ingresos Registrados</div><p class='kpi-valor'>{len(df)}</p></div>", unsafe_allow_html=True)
                k2.markdown(f"<div class='kpi-card kpi-rojo'><div class='kpi-titulo'>🚨 Lotes Vencidos</div><p class='kpi-valor'>{lotes_vencidos}</p></div>", unsafe_allow_html=True)
                k3.markdown(f"<div class='kpi-card kpi-amarillo'><div class='kpi-titulo'>⚠️ Por Vencer (90 Días)</div><p class='kpi-valor'>{lotes_riesgo}</p></div>", unsafe_allow_html=True)

                st.markdown("---")
                st.markdown("""<a href="#seccion-auditoria" class="btn-ascensor">👇 SALTAR DIRECTO A LA MATRIZ DE AUDITORÍA 👇</a>""", unsafe_allow_html=True)
                st.markdown("### ➕ Inyector de Nuevos Ingresos & Anulaciones SAP")

                with st.expander("🧪 1. IDENTIFICACIÓN DEL QUÍMICO OFICIAL Y MODO DE OPERACIÓN", expanded=True):
                    c_tog1, c_tog2 = st.columns(2)
                    es_nuevo_producto = c_tog1.toggle("✨ Ingresar un Producto Totalmente NUEVO")
                    modificar_prov = False
                    
                    st.markdown("<hr style='margin: 5px 0px 15px 0px; border: 1px dashed #d4af37;'>", unsafe_allow_html=True)
                    es_anulacion_sap = st.toggle("🔴 REGISTRAR ANULACIÓN / DEVOLUCIÓN DE INGRESO (SAP)", help="Activa este interruptor para registrar el Consecutivo de Anulación oficial generado por SAP.")

                    if es_anulacion_sap:
                        st.markdown("<div class='banner-anulacion'>🚨 MODO ANULACIÓN SAP ACTIVO: Ingrese el Consecutivo de Anulación oficial de SAP y vincúlelo a los lotes anulados.</div>", unsafe_allow_html=True)

                    c_prod, c_mat, c_prov = st.columns([2, 1, 2])
                    
                    if es_nuevo_producto:
                        n_prod = c_prod.text_input("🧪 Nombre del Nuevo Producto")
                        mat_item_ing = c_mat.text_input("🔢 Cód. Material", placeholder="Ej: 4591234", key=f"mat_ing_nuevo_{st.session_state['form_key_m19']}")
                        if not mat_item_ing: mat_item_ing = "S/N"
                        n_prov = c_prov.text_input("🏭 Nombre del Proveedor")
                        with c_tog2:
                            st.markdown("<div style='margin-top: -5px;'></div>", unsafe_allow_html=True)
                            if st.button("💾 GUARDAR EN DICCIONARIO", type="secondary", use_container_width=True):
                                prod_limpio = str(n_prod).strip().upper(); prov_limpio = str(n_prov).strip().upper()
                                if not prod_limpio or not prov_limpio: st.warning("⚠️ Escribe nombre y proveedor.")
                                else:
                                    with st.spinner("Registrando nuevo químico..."):
                                        try:
                                            gc_temp = inicializar_cliente_gspread()
                                            sh_temp = gc_temp.open_by_url(URL_SHEET_INGRESOS)
                                            try: ws_dicc = sh_temp.worksheet("DICCIONARIO")
                                            except:
                                                ws_dicc = sh_temp.add_worksheet(title="DICCIONARIO", rows="100", cols="2")
                                                ws_dicc.append_row(["PRODUCTO", "PROVEEDOR"])
                                            datos_d = ws_dicc.get_all_values()
                                            fila_a_actualizar = next((i + 1 for i, r in enumerate(datos_d) if len(r)>0 and str(r[0]).strip().upper() == prod_limpio), -1)
                                            if fila_a_actualizar != -1: ws_dicc.update_cell(fila_a_actualizar, 2, prov_limpio)
                                            else: ws_dicc.append_row([prod_limpio, prov_limpio])
                                            st.success(f"✅ ¡Producto {prod_limpio} registrado!")
                                            st.cache_data.clear(); st.rerun()
                                        except Exception as e: st.error(f"🚨 Fallo al guardar: {e}")
                    else:
                        modificar_prov = c_tog2.toggle("✏️ Corregir / Modificar Proveedor")
                        lista_prods_limpia = set([p for p in lista_autorizada if len(p) > 3 and "🛑" not in p])
                        n_prod = c_prod.selectbox("🧪 Producto (Integrado SAP)", sorted(list(lista_prods_limpia)))
                        
                        mat_item_ing = buscar_codigo_material(n_prod, mapeo_materiales)
                        c_mat.text_input("🔢 Cód. Material", value=mat_item_ing, disabled=True, key=f"mat_ing_exist_{st.session_state['form_key_m19']}_{n_prod}")

                        proveedor_asignado = dict_operativo.get(n_prod, "")
                        debe_desbloquear = modificar_prov or not bool(proveedor_asignado.strip())
                        n_prov = c_prov.text_input("🏭 Proveedor", value=proveedor_asignado, disabled=not debe_desbloquear, placeholder="Digite proveedor")
                        
                        if debe_desbloquear:
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button("💾 GUARDAR PROVEEDOR PERMANENTEMENTE", type="secondary", use_container_width=True):
                                prod_limpio = str(n_prod).strip().upper(); prov_limpio = str(n_prov).strip().upper()
                                if not prov_limpio: st.warning("⚠️ Escribe un proveedor.")
                                elif prov_limpio == proveedor_asignado.upper(): st.info("ℹ️ Proveedor es igual.")
                                else:
                                    with st.spinner("Actualizando Diccionario..."):
                                        try:
                                            gc_temp = inicializar_cliente_gspread()
                                            sh_temp = gc_temp.open_by_url(URL_SHEET_INGRESOS)
                                            try: ws_dicc = sh_temp.worksheet("DICCIONARIO")
                                            except:
                                                ws_dicc = sh_temp.add_worksheet(title="DICCIONARIO", rows="100", cols="2")
                                                ws_dicc.append_row(["PRODUCTO", "PROVEEDOR"])
                                            datos_d = ws_dicc.get_all_values()
                                            fila_a_actualizar = next((i + 1 for i, r in enumerate(datos_d) if len(r)>0 and str(r[0]).strip().upper() == prod_limpio), -1)
                                            if fila_a_actualizar != -1: ws_dicc.update_cell(fila_a_actualizar, 2, prov_limpio)
                                            else: ws_dicc.append_row([prod_limpio, prov_limpio])
                                            st.success("✅ ¡Diccionario Actualizado!")
                                            st.cache_data.clear(); st.rerun()
                                        except Exception as e: st.error(f"🚨 Fallo: {e}")

                with st.expander("⚙️ 2. DATOS OPERATIVOS Y TRAZABILIDAD (ENTRADAS Y ANULACIONES)", expanded=True):
                    col_espacio, col_limpiar = st.columns([3, 1])
                    col_limpiar.button("🧹 VACIAR CASILLAS", on_click=limpiar_campos_operativos, use_container_width=True)
                    f1, f2, f3 = st.columns(3)
                    n_fecha_ing = f2.date_input("🗓️ Fecha de Contabilización SAP", value=hoy_colombia)
                    semana_calculada = n_fecha_ing.isocalendar()[1]
                    n_semana = f1.text_input("📅 Semana del Año (Auto)", value=str(semana_calculada), disabled=True)
                    n_pista = f3.selectbox("📍 Almacén SAP (Pista)", ["PORI", "LUCI", "PLUC", "PDIV", "TEHO"])
                    
                    fk = st.session_state['form_key_m19']
                    
                    if es_anulacion_sap:
                        f8, f9, f10, f11 = st.columns(4)
                        n_consecutivo_anulacion = f8.text_input("🔢 CONSECUTIVO ANULACIÓN SAP *", placeholder="Ej: 5000799557", key=f"in_anul_sap_{fk}")
                        n_consecutivo_orig = f9.text_input("📄 Doc. Ingreso Original (Opcional)", placeholder="Ej: 5000799100", key=f"in_orig_sap_{fk}")
                        n_factura = f10.text_input("🧾 Factura / Nota", key=f"in_factura_anul_{fk}")
                        n_pedido = f11.text_input("🛒 Pedido SAP", key=f"in_pedido_anul_{fk}")
                        
                        n_motivo_anulacion = st.selectbox("📝 Motivo de Anulación SAP", [
                            "❌ ANULADO: ERROR EN PRECIOS",
                            "❌ ANULADO: ERROR DE CANTIDAD",
                            "❌ ANULADO: DEVOLUCIÓN A PROVEEDOR",
                            "❌ ANULADO: ERROR EN LOTE/FECHAS",
                            "❌ ANULADO: OTRO MOTIVO"
                        ], key=f"motivo_anul_sel_{fk}")
                        
                        n_consecutivo = n_consecutivo_anulacion
                    else:
                        f8, f9, f10 = st.columns(3)
                        n_factura = f8.text_input("🧾 Factura", key=f"in_factura_{fk}")
                        n_pedido = f9.text_input("🛒 Pedido", key=f"in_pedido_{fk}")
                        n_consecutivo = f10.text_input("🔢 Consecutivo SAP", key=f"in_consecutivo_{fk}")
                        n_consecutivo_anulacion = ""
                        n_motivo_anulacion = "✅ VIGENTE"

                    st.markdown("<hr style='margin: 10px 0px; border: 1px solid #d4af37;'>", unsafe_allow_html=True)
                    st.markdown("#### 📦 Detalle de Lotes a Procesar")
                    st.caption("Añade tantas filas como líneas tenga el documento SAP de Ingreso o Anulación.")
                    df_lotes_base = pd.DataFrame([{"CANTIDAD": 0.0, "LOTE": "", "F_FABRICACION": hoy_colombia, "F_VENCIMIENTO": hoy_colombia}])
                    
                    config_lotes = {
                        "CANTIDAD": st.column_config.NumberColumn("⚖️ Cantidad", min_value=0.0, format="%.2f"),
                        "LOTE": st.column_config.TextColumn("📦 Lote"),
                        "F_FABRICACION": st.column_config.DateColumn("⚙️ F. Fabricación (F/F)", format="YYYY-MM-DD"),
                        "F_VENCIMIENTO": st.column_config.DateColumn("⏳ F. Vencimiento (F/V)", format="YYYY-MM-DD")
                    }
                    
                    lotes_editados = st.data_editor(df_lotes_base, num_rows="dynamic", column_config=config_lotes, hide_index=True, use_container_width=True, key=f"multi_lote_{fk}")
                    
                    lotes_validos = [row for _, row in lotes_editados.iterrows() if float(row["CANTIDAD"]) > 0 and str(row["LOTE"]).strip() != ""]

                    st.markdown("<hr style='margin: 15px 0px; border: 1px solid #d4af37;'>", unsafe_allow_html=True)
                    st.markdown("<p style='color: #0d1b2a; font-size: 14px; font-weight: 900; text-transform: uppercase;'>📋 Panel de Copiado Rápido (1-Clic para SAP)</p>", unsafe_allow_html=True)
                    
                    if not lotes_validos:
                        st.info("⚠️ Ingresa cantidades y lotes válidos en la tabla superior para generar el panel de copiado.")
                    else:
                        if es_anulacion_sap:
                            h1, h2, h3, h4, h5 = st.columns(5)
                            h1.caption("🔢 MATERIAL")
                            h2.caption("⚖️ CANT. ANULADA")
                            h3.caption("📦 LOTE")
                            h4.caption("🔴 DOC. ANULACIÓN")
                            h5.caption("📝 ESTADO / MOTIVO")
                            
                            for row_lote in lotes_validos:
                                c1, c2, c3, c4, c5 = st.columns(5)
                                c1.code(str(mat_item_ing), language="text")
                                with c2: st.code(formatear_numero_sap(row_lote['CANTIDAD']), language="text")
                                with c3: st.code(str(row_lote['LOTE']).strip(), language="text")
                                with c4: st.code(n_consecutivo_anulacion if n_consecutivo_anulacion else "...", language="text")
                                with c5: st.code(n_motivo_anulacion, language="text")
                        else:
                            h1, h2, h3, h4, h5 = st.columns(5)
                            h1.caption("🔢 MATERIAL")
                            h2.caption("⚖️ CANTIDAD")
                            h3.caption("📦 LOTE")
                            h4.caption("🧾 FACTURA")
                            h5.caption("🛒 PEDIDO")
                            
                            for row_lote in lotes_validos:
                                c1, c2, c3, c4, c5 = st.columns(5)
                                c1.code(str(mat_item_ing), language="text")
                                with c2: st.code(formatear_numero_sap(row_lote['CANTIDAD']), language="text")
                                with c3: st.code(str(row_lote['LOTE']).strip(), language="text")
                                with c4: st.code(n_factura if n_factura else "...", language="text")
                                with c5: st.code(n_pedido if n_pedido else "...", language="text")

                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    if es_anulacion_sap:
                        btn_guardar_nuevo = st.button("🔴 INYECTAR ANULACIÓN DE LOTE(S) A LA BÓVEDA", type="primary", use_container_width=True)
                    else:
                        btn_guardar_nuevo = st.button("🚀 INYECTAR LOTE(S) A LA BÓVEDA", type="primary", use_container_width=True)
                    
                    if btn_guardar_nuevo:
                        if not n_prod or str(n_prod).strip() == "": 
                            st.error("🚨 El nombre del producto no puede estar vacío.")
                        elif es_anulacion_sap and not n_consecutivo_anulacion.strip():
                            st.error("🚨 Debes ingresar el Consecutivo de Anulación SAP obligatoriamente.")
                        elif not lotes_validos:
                            st.error("🚨 Debes ingresar al menos una fila válida con Cantidad mayor a 0 y Lote.")
                        else:
                            prod_limpio = str(n_prod).strip().upper(); prov_limpio = str(n_prov).strip().upper()
                            if es_nuevo_producto or ((modificar_prov or not bool(proveedor_asignado.strip())) and prov_limpio and prov_limpio != proveedor_asignado.upper()):
                                try:
                                    gc_temp = inicializar_cliente_gspread()
                                    sh_temp = gc_temp.open_by_url(URL_SHEET_INGRESOS)
                                    try: ws_dicc = sh_temp.worksheet("DICCIONARIO")
                                    except:
                                        ws_dicc = sh_temp.add_worksheet(title="DICCIONARIO", rows="100", cols="2")
                                        ws_dicc.append_row(["PRODUCTO", "PROVEEDOR"])
                                    datos_d = ws_dicc.get_all_values()
                                    fila_a_actualizar = next((i + 1 for i, r in enumerate(datos_d) if len(r)>0 and str(r[0]).strip().upper() == prod_limpio), -1)
                                    if fila_a_actualizar != -1: ws_dicc.update_cell(fila_a_actualizar, 2, prov_limpio)
                                    else: ws_dicc.append_row([prod_limpio, prov_limpio])
                                except Exception as e: st.warning(f"Se guardó el diccionario: {e}")

                            nuevas_filas_bulk = []
                            for row_lote in lotes_validos:
                                lote_ing_inject = f"'{str(row_lote['LOTE']).strip()}"
                                cant = row_lote['CANTIDAD']
                                ff = pd.to_datetime(row_lote['F_FABRICACION']).strftime("%d/%m/%Y") if pd.notnull(row_lote['F_FABRICACION']) else hoy_colombia.strftime("%d/%m/%Y")
                                fv = pd.to_datetime(row_lote['F_VENCIMIENTO']).strftime("%d/%m/%Y") if pd.notnull(row_lote['F_VENCIMIENTO']) else hoy_colombia.strftime("%d/%m/%Y")
                                
                                estado_final_inyect = n_motivo_anulacion if es_anulacion_sap else "✅ VIGENTE"
                                consecutivo_final_inyect = f"{n_consecutivo_anulacion} (ANULACIÓN)" if es_anulacion_sap else str(n_consecutivo)
                                
                                nueva_fila_drive = []
                                for header in encabezados_limpios_ing:
                                    h = header.upper()
                                    if "SEMANA" in h: nueva_fila_drive.append(str(semana_calculada))
                                    elif "PROV" in h: nueva_fila_drive.append(prov_limpio)
                                    elif "INGRESO" in h: nueva_fila_drive.append(n_fecha_ing.strftime("%d/%m/%Y"))
                                    elif "PROD" in h: nueva_fila_drive.append(prod_limpio)
                                    elif "PISTA" in h: nueva_fila_drive.append(str(n_pista))
                                    elif "CANT" in h: nueva_fila_drive.append(str(cant))
                                    elif "LOTE" in h: nueva_fila_drive.append(lote_ing_inject)
                                    elif "F/F" in h: nueva_fila_drive.append(ff)
                                    elif "F/V" in h: nueva_fila_drive.append(fv)
                                    elif "FACT" in h: nueva_fila_drive.append(str(n_factura))
                                    elif "PEDIDO" in h: nueva_fila_drive.append(str(n_pedido))
                                    elif "CONSECUT" in h: nueva_fila_drive.append(consecutivo_final_inyect)
                                    elif "ESTADO" in h: nueva_fila_drive.append(estado_final_inyect)
                                    elif any(k in h for k in ["ITEM", "MATERIAL", "CÓDIGO", "COD"]): nueva_fila_drive.append(str(mat_item_ing))
                                    else: nueva_fila_drive.append("") 
                                nuevas_filas_bulk.append(nueva_fila_drive)
                            
                            try:
                                with st.spinner(f"Inyectando {len(nuevas_filas_bulk)} registros con láser matemático..."):
                                    gc_temp = inicializar_cliente_gspread()
                                    sh_temp = gc_temp.open_by_url(URL_SHEET_INGRESOS)
                                    ws_write_ing = sh_temp.worksheets()[0]
                                    try: idx_col_prod = encabezados_limpios_ing.index("PRODUCTO") + 1
                                    except: idx_col_prod = 4 
                                    col_prod_data = ws_write_ing.col_values(idx_col_prod)
                                    last_row_ing = len(col_prod_data)
                                    while last_row_ing > 0 and str(col_prod_data[last_row_ing-1]).strip() == "": last_row_ing -= 1
                                    
                                    fila_destino = last_row_ing + 1
                                    rango_inyeccion = f"A{fila_destino}:{get_column_letter(len(encabezados_limpios_ing))}{fila_destino + len(nuevas_filas_bulk) - 1}"
                                    
                                    try: ws_write_ing.update(range_name=rango_inyeccion, values=nuevas_filas_bulk, value_input_option='USER_ENTERED')
                                    except: ws_write_ing.update(rango_inyeccion, nuevas_filas_bulk, value_input_option='USER_ENTERED')
                                    
                                st.success(f"✅ ¡{len(nuevas_filas_bulk)} registro(s) de {prod_limpio} inyectados exitosamente a la Bóveda!")
                                st.session_state['form_key_m19'] += 1
                                st.cache_data.clear(); st.rerun()
                            except Exception as e: st.error(f"Error al inyectar datos: {e}")

                # --- GENERADOR DE REPORTE CORREO ---
                st.markdown("---")
                st.markdown("### 📧 Reporte Rápido para Correo (Copy & Paste)")
                st.info("💡 **Filtro Anti-Infiltración:** Los registros anulados se ocultan por defecto.")
                
                col_fecha_rep, col_vacia = st.columns([1, 3])
                fecha_reporte = col_fecha_rep.date_input("Fecha a reportar:", value=hoy_colombia)
                
                if col_fecha_ingreso:
                    df['FECHA_ING_TEMP'] = df[col_fecha_ingreso].apply(procesar_fecha_estricta)
                    mask = df['FECHA_ING_TEMP'].apply(lambda x: x.date() if pd.notna(x) else None) == fecha_reporte
                    df_correo = df[mask].copy()
                    if COL_ESTADO in df_correo.columns: df_correo = df_correo[~df_correo[COL_ESTADO].str.contains("ANULADO|ELIMINAR", na=False, case=False)]
                    if not df_correo.empty:
                        df_correo = df_correo.sort_values(by='FILA_EXCEL', ascending=False)
                        df_correo.insert(0, "✅ INCLUIR", True)
                        cols_ed = [c for c in df_correo.columns if c not in ["FILA_EXCEL", "FECHA_ING_TEMP", "FECHA_VENC_DT", COL_ESTADO]]
                        st.markdown("👇 **Paso 1: Desmarca los registros que NO quieres enviar en el correo:**")
                        df_editado_correo = st.data_editor(df_correo[cols_ed], column_config={"✅ INCLUIR": st.column_config.CheckboxColumn("✅ INCLUIR", default=True)}, disabled=[c for c in cols_ed if c != "✅ INCLUIR"], hide_index=True, use_container_width=True, key=f"editor_correo_{fecha_reporte}")
                        df_correo_final = df_editado_correo[df_editado_correo["✅ INCLUIR"] == True].copy()
                        if not df_correo_final.empty:
                            mapa_columnas = {}
                            for col_excel in df_correo_final.columns:
                                c_up = str(col_excel).upper()
                                if "SEMANA" in c_up: mapa_columnas[col_excel] = "SEMANA"
                                elif "PROV" in c_up: mapa_columnas[col_excel] = "PROVEEDOR"
                                elif "INGRESO" in c_up: mapa_columnas[col_excel] = "F. INGRESO"
                                elif "PROD" in c_up: mapa_columnas[col_excel] = "PRODUCTO"
                                elif "PISTA" in c_up: mapa_columnas[col_excel] = "PISTA"
                                elif "CANT" in c_up: mapa_columnas[col_excel] = "CANTIDAD"
                                elif "LOTE" in c_up: mapa_columnas[col_excel] = "LOTE"
                                elif "F/F" in c_up: mapa_columnas[col_excel] = "F/F"
                                elif "F/V" in c_up: mapa_columnas[col_excel] = "F/V"
                                elif "FACT" in c_up: mapa_columnas[col_excel] = "FACTURA"
                                elif "PEDIDO" in c_up: mapa_columnas[col_excel] = "PEDIDO"
                                elif "CONSECUT" in c_up: mapa_columnas[col_excel] = "CONSECUTIVO"
                            df_correo_limpio = df_correo_final[list(mapa_columnas.keys())].rename(columns=mapa_columnas)
                            orden_ideal = ["PROVEEDOR", "PRODUCTO", "PISTA", "CANTIDAD", "LOTE", "F/F", "F/V", "FACTURA", "PEDIDO", "CONSECUTIVO"]
                            df_correo_limpio = df_correo_limpio[[col for col in orden_ideal if col in df_correo_limpio.columns]]
                            html_manual = "<table style='border-collapse: collapse; width: 100%; font-family: Arial, Helvetica, sans-serif; font-size: 11px; border: 2px solid #0d1b2a; margin-top: 10px; background-color: #ffffff;'><thead><tr>"
                            for col_name in df_correo_limpio.columns: html_manual += f"<th style='background-color: #0d1b2a; color: #d4af37; padding: 8px 6px; border: 2px solid #0d1b2a; text-align: center; font-weight: 900; text-transform: uppercase; white-space: nowrap;'>{col_name}</th>"
                            html_manual += "</tr></thead><tbody>"
                            for _, row in df_correo_limpio.iterrows():
                                html_manual += "<tr>"
                                for col_name in df_correo_limpio.columns:
                                    val = row[col_name]
                                    val_str = "" if pd.isna(val) or str(val).strip() == "" else (formatear_numero_sap(str(val).strip()) if col_name == "CANTIDAD" else str(val).strip())
                                    if val_str.startswith("'"): val_str = val_str[1:]
                                    html_manual += f"<td style='padding: 8px 6px; border: 1px solid #0d1b2a; text-align: center; color: #000000; font-weight: bold; white-space: nowrap;'>{val_str}</td>"
                                html_manual += "</tr>"
                            html_manual += "</tbody></table>"
                            st.markdown("👇 **Paso 2: Copia la tabla a continuación y pégala en tu correo:**")
                            st.markdown(html_manual, unsafe_allow_html=True)
                        else: st.info("Has desmarcado todos los registros. La tabla final está vacía.")
                    else: st.warning(f"No se encontraron ingresos válidos en la bóveda con la fecha {fecha_reporte.strftime('%d/%m/%Y')}.")

                st.markdown("---")
                st.markdown("<div id='seccion-auditoria'></div>", unsafe_allow_html=True)
                st.markdown("### 🔍 Escáner de Auditoría (Filtros)")
                
                f_col1, f_col2 = st.columns(2)
                filtro_seleccionado = f_col1.selectbox("🛡️ Estado Operativo:", ["🌐 Mostrar Todos", "✅ Solo Vigentes", "🚨 Solo Vencidos", "⚠️ Por Vencer (90 Días)", "❌ Solo Anulados SAP"])
                
                lista_productos_tabla = ["TODOS"] + sorted(list(set([str(x).strip().upper() for x in df[col_producto].dropna() if str(x).strip() != ""]))) if col_producto else ["TODOS"]
                producto_filtro = f_col2.selectbox("🧪 Filtrar por Producto:", lista_productos_tabla)

                st.markdown("<br>", unsafe_allow_html=True)
                f_col3, f_col4, _ = st.columns([1, 1, 1.5])
                fecha_ini_filtro = f_col3.date_input("📅 Ingreso Desde:", value=date(2021, 1, 1))
                fecha_fin_filtro = f_col4.date_input("📅 Ingreso Hasta:", value=hoy_colombia)
                
                df[COL_ESTADO] = df[COL_ESTADO].replace(r'^\s*$', '✅ VIGENTE', regex=True).fillna('✅ VIGENTE')
                df_filtrado = df.copy()
                
                if col_fecha_ingreso:
                    fechas_dt = df_filtrado[col_fecha_ingreso].apply(procesar_fecha_estricta)
                    mask_fechas = fechas_dt.apply(lambda x: x.date() if pd.notna(x) else None)
                    df_filtrado = df_filtrado[(mask_fechas >= fecha_ini_filtro) & (mask_fechas <= fecha_fin_filtro)]

                if producto_filtro != "TODOS" and col_producto: df_filtrado = df_filtrado[df_filtrado[col_producto].str.upper() == producto_filtro]
                if filtro_seleccionado == "✅ Solo Vigentes": df_filtrado = df_filtrado[df_filtrado[COL_ESTADO].str.contains("VIGENTE", case=False, na=False)]
                elif filtro_seleccionado == "❌ Solo Anulados SAP": df_filtrado = df_filtrado[df_filtrado[COL_ESTADO].str.contains("ANULADO", case=False, na=False)]
                elif filtro_seleccionado == "🚨 Solo Vencidos" and col_fv: df_filtrado = df_filtrado[(~df_filtrado[COL_ESTADO].str.contains("ANULADO|ELIMINAR", na=False)) & (df_filtrado['FECHA_VENC_DT'] < hoy_ts)]
                elif filtro_seleccionado == "⚠️ Por Vencer (90 Días)" and col_fv: df_filtrado = df_filtrado[(~df_filtrado[COL_ESTADO].str.contains("ANULADO|ELIMINAR", na=False)) & (df_filtrado['FECHA_VENC_DT'] >= hoy_ts) & (df_filtrado['FECHA_VENC_DT'] <= limite_90_dias)]

                if col_fecha_ingreso:
                    df_filtrado['FECHA_SORT'] = df_filtrado[col_fecha_ingreso].apply(procesar_fecha_estricta)
                    df_filtrado = df_filtrado.sort_values(by=['FECHA_SORT', 'FILA_EXCEL'], ascending=[False, False])
                else: df_filtrado = df_filtrado.sort_values(by=['FILA_EXCEL'], ascending=[False])

                st.markdown("### 🛠️ Matriz de Anulaciones y Edición Masiva")
                st.caption("🔒 Haz doble clic en las columnas para **Copiar, Pegar o Editar**. Usa el Estado para anular o ELIMINAR el registro físicamente.")
                
                columnas_editables_nombres = ["CONSECUTIVO", "PEDIDO", "FACTURA", "LOTE", "CANTIDAD"]
                columnas_editables_reales = [COL_ESTADO]
                
                for col in df_filtrado.columns:
                    for nom_edit in columnas_editables_nombres:
                        if nom_edit in str(col).upper() and col not in columnas_editables_reales:
                            columnas_editables_reales.append(col)
                            
                cols_disabled = [col for col in df_filtrado.columns if col not in columnas_editables_reales and col not in ['FILA_EXCEL', 'FECHA_VENC_DT', 'FECHA_ING_TEMP', 'FECHA_SORT']]
                
                opciones_estado = ["✅ VIGENTE", "❌ ANULADO: ERROR EN PRECIOS", "❌ ANULADO: ERROR DE CANTIDAD", "❌ ANULADO: DEVOLUCIÓN A PROVEEDOR", "❌ ANULADO: ERROR EN LOTE/FECHAS", "❌ ANULADO: OTRO MOTIVO", "💥 ELIMINAR REGISTRO (BORRADO FÍSICO)"]

                columnas_vista = [c for c in df_filtrado.columns if c not in ['FILA_EXCEL', 'FECHA_VENC_DT', 'FECHA_ING_TEMP', 'FECHA_SORT']]
                df_vista = df_filtrado[columnas_vista].copy()
                
                if "LOTE" in df_vista.columns: df_vista["LOTE"] = df_vista["LOTE"].astype(str).str.lstrip("'")
                
                col_config = {COL_ESTADO: st.column_config.SelectboxColumn("🛡️ ESTADO / OBSERVACIÓN", help="Doble clic para anular o cambiar estado.", width="large", options=opciones_estado, required=True)}
                for c in df_vista.columns:
                    c_up = c.upper()
                    if "SEMANA" in c_up: col_config[c] = st.column_config.TextColumn("📅 SEMANA", width="small")
                    elif "PROV" in c_up: col_config[c] = st.column_config.TextColumn("🏭 PROVEEDOR", width="medium")
                    elif "INGRESO" in c_up: col_config[c] = st.column_config.TextColumn("🗓️ INGRESO SAP", width="medium")
                    elif "PROD" in c_up: col_config[c] = st.column_config.TextColumn("🧪 PRODUCTO", width="large")
                    elif "PISTA" in c_up: col_config[c] = st.column_config.TextColumn("📍 BASE", width="small")
                    elif "CANT" in c_up: col_config[c] = st.column_config.TextColumn("⚖️ CANTIDAD (Editable)", width="medium")
                    elif "LOTE" in c_up: col_config[c] = st.column_config.TextColumn("📦 LOTE (Editable)", width="medium")
                    elif "F/F" in c_up: col_config[c] = st.column_config.TextColumn("⚙️ F/F", width="small")
                    elif "F/V" in c_up: col_config[c] = st.column_config.TextColumn("⏳ F/V", width="small")
                    elif "FACT" in c_up: col_config[c] = st.column_config.TextColumn("🧾 FACTURA (Editable)", width="medium")
                    elif "PEDIDO" in c_up: col_config[c] = st.column_config.TextColumn("🛒 PEDIDO (Editable)", width="medium")
                    elif "CONSECUT" in c_up: col_config[c] = st.column_config.TextColumn("🔢 CONSECUTIVO (Editable)", width="medium")
                    
                df_editado = st.data_editor(df_vista, column_config=col_config, disabled=cols_disabled, hide_index=True, use_container_width=True, key="editor_ingresos")

                st.markdown("<br>", unsafe_allow_html=True)
                
                if st.button("💾 SINCRONIZAR CAMBIOS Y ELIMINACIONES EN DRIVE", type="primary"):
                    cambios_actualizacion = []
                    eliminaciones = []
                    
                    idx_cols = {}
                    for col in columnas_editables_reales:
                        for idx_h, h in enumerate(encabezados_limpios_ing):
                            if col.upper() == h.upper() or (col == COL_ESTADO and "ESTADO" in h.upper()):
                                idx_cols[col] = idx_h + 1
                                break

                    for i in range(len(df_filtrado)):
                        estado_nuevo = str(df_editado.iloc[i][COL_ESTADO]).strip()
                        fila_excel = int(df_filtrado.iloc[i]['FILA_EXCEL'])
                        
                        if "ELIMINAR REGISTRO" in estado_nuevo:
                            eliminaciones.append(fila_excel)
                        else:
                            for col in columnas_editables_reales:
                                if col in df_filtrado.columns and col in df_editado.columns:
                                    val_orig = str(df_filtrado.iloc[i][col]).strip()
                                    val_nuevo = str(df_editado.iloc[i][col]).strip()
                                    
                                    if val_orig != val_nuevo:
                                        val_inyectar = f"'{val_nuevo}" if "LOTE" in col.upper() else val_nuevo
                                        if col in idx_cols:
                                            cambios_actualizacion.append({'fila': fila_excel, 'col_idx': idx_cols[col], 'nuevo': val_inyectar})
                    
                    if cambios_actualizacion or eliminaciones:
                        gc_temp = inicializar_cliente_gspread()
                        sh_temp = gc_temp.open_by_url(URL_SHEET_INGRESOS)
                        ws_write_ing = sh_temp.worksheets()[0]
                        
                        with st.spinner(f"Sincronizando la nube con un solo paquete de datos..."):
                            try:
                                if cambios_actualizacion:
                                    celdas_a_enviar = [gspread.Cell(act['fila'], act['col_idx'], act['nuevo']) for act in cambios_actualizacion]
                                    ws_write_ing.update_cells(celdas_a_enviar, value_input_option='USER_ENTERED')
                                
                                if eliminaciones:
                                    eliminaciones = sorted(list(set(eliminaciones)), reverse=True)
                                    peticiones_borrado = []
                                    for eli in eliminaciones:
                                        peticiones_borrado.append({
                                            "deleteDimension": {
                                                "range": {
                                                    "sheetId": ws_write_ing.id,
                                                    "dimension": "ROWS",
                                                    "startIndex": eli - 1, 
                                                    "endIndex": eli
                                                }
                                            }
                                        })
                                    sh_temp.batch_update({"requests": peticiones_borrado})
                                    
                                st.success("✅ ¡Misión Cumplida! Base de datos sincronizada y purgada exitosamente en una sola operación maestra.")
                                st.cache_data.clear(); st.rerun()
                            except Exception as e:
                                st.error(f"🚨 Error crítico en la sincronización masiva: {e}")
                    else: st.info("No se detectaron cambios ni órdenes de eliminación.")

                st.markdown("---")
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_editado.to_excel(writer, sheet_name='Auditoria_Ingresos', index=False)
                    ws_excel = writer.sheets['Auditoria_Ingresos']
                    header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                    for cell in ws_excel[1]: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', vertical='center')
                    for col in ws_excel.columns:
                        max_length = 0
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(cell.value)
                            except: pass
                        ws_excel.column_dimensions[col[0].column_letter].width = (max_length + 2)

                st.download_button("💾 DESCARGAR REPORTE DE AUDITORÍA (EXCEL)", data=buffer.getvalue(), file_name=f"Reporte_Auditoria_Ingresos_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # ========================================================================
    # 🚚 PESTAÑA 2: MOVIMIENTOS INTERNOS (TRASLADOS)
    # ========================================================================
    with tab_traslados:
        st.markdown(f"<a href='{URL_SHEET_TRASLADOS}' target='_blank' class='btn-ascensor' style='background-color:#2F75B5; border-color:#1d4e7a; color:#ffffff !important;'>👁️ VER BASE DE TRASLADOS EN GOOGLE SHEETS</a>", unsafe_allow_html=True)
        st.write("Panel táctico de logística interna. Registra los movimientos de inventario entre las diferentes bases operativas.")

        sabana_guardada = st.session_state.get('mem_sabana')
        if sabana_guardada is None or not isinstance(sabana_guardada, pd.DataFrame) or sabana_guardada.empty:
            st.info("💡 **Conexión Directa:** No necesitas ir al Módulo 2. Sube tu Sábana SAP aquí mismo para habilitar la lectura automática de lotes.")
            archivo_bypass = st.file_uploader("📥 Cargar Sábana SAP (MB52)", type=["xlsx", "xls", "csv"], key="uploader_bypass_m19")
            if archivo_bypass:
                with st.spinner("Leyendo Sábana SAP..."):
                    try:
                        if archivo_bypass.name.endswith('.csv'): df_bypass = pd.read_csv(archivo_bypass, sep=';', encoding='latin1', on_bad_lines='skip')
                        else: df_bypass = pd.read_excel(archivo_bypass)
                        st.session_state['mem_sabana'] = df_bypass
                        st.success("✅ ¡Sábana conectada! El escáner de lotes ya está activo.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error leyendo el archivo: {e}")
        else:
            st.success("✅ Sábana SAP conectada en memoria. El escáner de lotes está 100% operativo.")
            if st.button("🗑️ Liberar Sábana de la Memoria", key="btn_clear_sabana"):
                del st.session_state['mem_sabana']
                st.rerun()

        if not df_traslados.empty: st.markdown(f"<div class='kpi-card kpi-azul'><div class='kpi-titulo'>🚚 Movimientos Históricos Registrados</div><p class='kpi-valor'>{len(df_traslados)}</p></div>", unsafe_allow_html=True)
        else: st.markdown(f"<div class='kpi-card kpi-azul'><div class='kpi-titulo'>🚚 Movimientos Históricos Registrados</div><p class='kpi-valor'>0</p></div>", unsafe_allow_html=True)
        
        st.markdown("---")
        st.markdown("### ➕ Inyector de Movimiento Interno")

        with st.expander("📍 1. DATOS DEL TRASLADO", expanded=True):
            col_espacio_t, col_limpiar_t = st.columns([3, 1])
            col_limpiar_t.button("🧹 VACIAR CASILLAS", on_click=limpiar_campos_traslados, key="btn_limpiar_traslados", use_container_width=True)

            t1, t2, t3 = st.columns(3)
            fk_t = st.session_state['form_key_m19_traslados']
            t_fecha = t2.date_input("🗓️ Fecha de Traslado", value=hoy_colombia, key=f"t_fecha_{fk_t}")
            semana_traslado = t_fecha.isocalendar()[1]
            t_semana = t1.text_input("📅 Semana del Año", value=str(semana_traslado), disabled=True, key=f"t_semana_{fk_t}")
            t_consecutivo = t3.text_input("🔢 Consecutivo", key=f"t_consecutivo_{fk_t}")

            st.markdown("<hr style='margin: 10px 0px; border: 1px solid #e2e8f0;'>", unsafe_allow_html=True)
            
            pistas_disponibles = ["PORI", "LUCI", "PLUC", "PDIV", "TEHO"]
            p1, p2 = st.columns(2)
            t_origen = p1.selectbox("🛫 Pista Origen", pistas_disponibles, index=0, key=f"t_origen_{fk_t}")
            t_destino = p2.selectbox("🛬 Pista Destino", pistas_disponibles, index=1 if len(pistas_disponibles) > 1 else 0, key=f"t_destino_{fk_t}")

            st.markdown("<hr style='margin: 10px 0px; border: 1px solid #e2e8f0;'>", unsafe_allow_html=True)

            lista_prods_limpia_t = set([p for p in lista_autorizada if len(p) > 3 and "🛑" not in p])
            lista_prods_ordenada_t = sorted(list(lista_prods_limpia_t))

            tr1, tr_mat, tr2, tr3 = st.columns([2, 1, 1, 1])
            t_producto = tr1.selectbox("🧪 Producto a Trasladar", lista_prods_ordenada_t, key=f"t_producto_{fk_t}")
            
            mat_item_tras = buscar_codigo_material(t_producto, mapeo_materiales)
            tr_mat.text_input("🔢 Cód. Material", value=mat_item_tras, disabled=True, key=f"t_mat_cod_{fk_t}_{t_producto}")

            t_cantidad = tr2.number_input("⚖️ Cantidad", min_value=0.0, step=1.0, key=f"t_cantidad_{fk_t}")
            t_unidad = tr3.selectbox("📦 Unidad", ["LITROS", "KILOS", "GALONES", "UNIDADES"], key=f"t_unidad_{fk_t}")

            tr4, tr5 = st.columns(2)
            opciones_obs = ["SIN NOVEDAD", "ANULACIÓN", "TRANSFORMACIÓN DE LOTE", "AJUSTE ENTRE LOTES", "OTRO"]
            t_observacion_sel = tr4.selectbox("📝 Observación", opciones_obs, key=f"t_obs_sel_{fk_t}")
            
            if t_observacion_sel == "OTRO": 
                t_observacion = tr4.text_input("📝 Especifique la observación:", key=f"t_obs_otro_{fk_t}")
            else: 
                t_observacion = t_observacion_sel

            lotes_disp = []
            df_sabana_memoria = st.session_state.get('mem_sabana')
            if not isinstance(df_sabana_memoria, pd.DataFrame):
                df_sabana_memoria = pd.DataFrame()
            
            prod_keywords = str(t_producto).strip().upper().split()
            prod_clave = prod_keywords[0] if prod_keywords else ""
            cod_clean = str(mat_item_tras).strip().lstrip('0')

            if not df_sabana_memoria.empty:
                col_lote_sap = next((c for c in df_sabana_memoria.columns if 'LOTE' in str(c).upper() and 'PROVEEDOR' not in str(c).upper()), None)
                col_mat_desc = next((c for c in df_sabana_memoria.columns if 'TEXTO' in str(c).upper() or 'DESC' in str(c).upper()), None)
                col_mat_cod = next((c for c in df_sabana_memoria.columns if 'MATERIAL' in str(c).upper() or 'ITEM' in str(c).upper() or 'CÓDIGO' in str(c).upper().replace('Ó','O') or 'COD' in str(c).upper()), None)
                col_alm_sap = next((c for c in df_sabana_memoria.columns if 'ALMACEN' in str(c).upper().replace('É','E') or 'PISTA' in str(c).upper()), None)
                col_sal_sap = next((c for c in df_sabana_memoria.columns if ('LIBRE' in str(c).upper() or 'SALDO' in str(c).upper()) and 'VALOR' not in str(c).upper()), None)

                if col_lote_sap and col_alm_sap:
                    pista_busqueda = str(t_origen).strip().upper()
                    pista_sap_alias = pista_busqueda
                    if pista_busqueda == "PLUC": pista_sap_alias = "LUCHA"
                    elif pista_busqueda == "PORI": pista_sap_alias = "ORIHUECA"
                    elif pista_busqueda == "PDIV": pista_sap_alias = "DIVAS"
                    elif pista_busqueda == "TEHO": pista_sap_alias = "TEHOBROMINA"

                    mask_pista = df_sabana_memoria[col_alm_sap].astype(str).str.upper().str.contains(pista_busqueda, na=False) | \
                                 df_sabana_memoria[col_alm_sap].astype(str).str.upper().str.contains(pista_sap_alias, na=False)
                    
                    mask_prod_cod = pd.Series(False, index=df_sabana_memoria.index)
                    if col_mat_cod and cod_clean and cod_clean != "S/N":
                        col_cod_limpia = df_sabana_memoria[col_mat_cod].astype(str).str.replace(".0", "", regex=False).str.strip().str.lstrip('0')
                        mask_prod_cod = col_cod_limpia.str.contains(cod_clean, regex=False, na=False)
                    
                    mask_prod_desc = pd.Series(False, index=df_sabana_memoria.index)
                    if col_mat_desc and prod_clave:
                        mask_prod_desc = df_sabana_memoria[col_mat_desc].astype(str).str.upper().str.contains(prod_clave, na=False)
                    
                    mask_prod = mask_prod_cod | mask_prod_desc
                    
                    df_filtro_sap = df_sabana_memoria[mask_pista & mask_prod]
                    
                    for _, row_s in df_filtro_sap.iterrows():
                        l_val = str(row_s[col_lote_sap]).strip()
                        s_val = row_s[col_sal_sap] if col_sal_sap else 0
                        
                        if l_val and l_val not in ["nan", "None", ""]:
                            try: s_val_str = f"{float(s_val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                            except: s_val_str = str(s_val)
                            lotes_disp.append(f"{l_val} (Saldo SAP: {s_val_str})")

            lotes_disp = list(dict.fromkeys(lotes_disp))

            t_lote_nuevo = ""
            with tr4:
                if t_observacion_sel == "TRANSFORMACIÓN DE LOTE":
                    t_lote_nuevo = st.text_input("🔄 NUEVO LOTE (Crear Manual):", help="Digite el número del lote resultante.", key=f"t_lote_nuevo_{fk_t}")
                elif t_observacion_sel == "AJUSTE ENTRE LOTES":
                    if lotes_disp:
                        lote_dest_sel = st.selectbox("🔄 LOTE DESTINO (SAP):", lotes_disp, key=f"t_lote_dest_sel_{fk_t}")
                        t_lote_nuevo = lote_dest_sel.split(" (Saldo")[0].strip()
                    else:
                        st.warning("⚠️ SAP no reporta otros lotes aquí.")
                        t_lote_nuevo = st.text_input("🔄 LOTE DESTINO (Manual):", key=f"t_lote_dest_man_{fk_t}")
            
            with tr5:
                opcion_metodo = st.selectbox("⚙️ Origen del Lote", ["📋 ELEGIR DE LA LISTA SAP", "✍️ ESCRIBIR MANUALMENTE"], key=f"metodo_lote_{fk_t}")
                
                if opcion_metodo == "📋 ELEGIR DE LA LISTA SAP":
                    if lotes_disp:
                        lote_seleccionado = st.selectbox("📦 Seleccione el Lote", lotes_disp, key=f"t_lote_sel_{fk_t}")
                        t_lote_origen = lote_seleccionado.split(" (Saldo")[0].strip()
                    else:
                        st.warning("⚠️ Sábana sin procesar o sin saldo.")
                        t_lote_origen = st.text_input("✍️ Digite Lote Manual (Forzado):", key=f"t_lote_man_force_{fk_t}")
                else:
                    t_lote_origen = st.text_input("✍️ Digite Lote Manual:", key=f"t_lote_man_{fk_t}")

            lote_final_print = t_lote_origen
            if t_observacion_sel in ["TRANSFORMACIÓN DE LOTE", "AJUSTE ENTRE LOTES"] and t_lote_nuevo.strip():
                lote_final_print = f"{t_lote_origen} ➔ {t_lote_nuevo.strip()}"

            st.markdown("<hr style='margin: 15px 0px; border: 1px solid #d4af37;'>", unsafe_allow_html=True)
            st.markdown("<p style='color: #0d1b2a; font-size: 14px; font-weight: 900; text-transform: uppercase;'>📋 Panel de Copiado Rápido (1-Clic para SAP)</p>", unsafe_allow_html=True)

            if t_observacion_sel in ["TRANSFORMACIÓN DE LOTE", "AJUSTE ENTRE LOTES"]:
                cpt_mat, cpt1, cpt2_o, cpt2_d, cpt3, cpt4 = st.columns(6)
                with cpt_mat: st.caption("🔢 MATERIAL"); st.code(mat_item_tras, language="text")
                with cpt1: st.caption("⚖️ CANTIDAD"); st.code(formatear_numero_sap(t_cantidad), language="text")
                with cpt2_o: st.caption("📦 L. ORIGEN"); st.code(t_lote_origen if t_lote_origen else "...", language="text")
                with cpt2_d: st.caption("📦 L. DESTINO"); st.code(t_lote_nuevo.strip() if t_lote_nuevo.strip() else "...", language="text")
                with cpt3: st.caption("🛫 ORIGEN"); st.code(t_origen, language="text")
                with cpt4: st.caption("🛬 DESTINO"); st.code(t_destino, language="text")
            else:
                cpt_mat, cpt1, cpt2, cpt3, cpt4 = st.columns(5)
                with cpt_mat: st.caption("🔢 MATERIAL"); st.code(mat_item_tras, language="text")
                with cpt1: st.caption("⚖️ CANTIDAD"); st.code(formatear_numero_sap(t_cantidad), language="text")
                with cpt2: st.caption("📦 LOTE"); st.code(lote_final_print if lote_final_print else "...", language="text")
                with cpt3: st.caption("🛫 ORIGEN"); st.code(t_origen, language="text")
                with cpt4: st.caption("🛬 DESTINO"); st.code(t_destino, language="text")

            st.markdown("<br>", unsafe_allow_html=True)
            btn_guardar_traslado = st.button("🚀 REGISTRAR TRASLADO EN LA BÓVEDA", type="primary", use_container_width=True)

            if btn_guardar_traslado:
                if not t_consecutivo.strip(): st.error("🚨 Debes ingresar un número de Consecutivo.")
                elif t_cantidad <= 0: st.error("🚨 La cantidad debe ser mayor a cero.")
                elif t_origen == t_destino and t_observacion_sel not in ["TRANSFORMACIÓN DE LOTE", "AJUSTE ENTRE LOTES", "OTRO"]: 
                    st.error("🚨 Para mover material dentro de la misma pista, la observación debe ser 'TRANSFORMACIÓN DE LOTE', 'AJUSTE ENTRE LOTES' u 'OTRO'.")
                elif not t_lote_origen or str(t_lote_origen).strip() == "": st.error("🚨 Debes especificar el Lote a trasladar.")
                elif t_observacion_sel in ["TRANSFORMACIÓN DE LOTE", "AJUSTE ENTRE LOTES"] and not t_lote_nuevo.strip(): st.error("🚨 Debes ingresar el LOTE DESTINO para el ajuste/transformación.")
                else:
                    try:
                        with st.spinner("Enviando traslado a la nube..."):
                            pista_combinada = f"{t_origen}-{t_destino}"
                            fecha_str = t_fecha.strftime("%d/%m/%Y")
                            cantidad_formateada = formatear_numero_sap(t_cantidad)

                            lote_tras_inject = f"'{lote_final_print}"
                            
                            nueva_fila_traslado = []
                            for h in encabezados_limpios_tras:
                                h_up = h.upper()
                                if "CONSECUTIVO" in h_up: nueva_fila_traslado.append(str(t_consecutivo).strip())
                                elif "FECHA" in h_up: nueva_fila_traslado.append(fecha_str)
                                elif "PROD" in h_up: nueva_fila_traslado.append(str(t_producto).upper().strip())
                                elif "CANT" in h_up: nueva_fila_traslado.append(cantidad_formateada)
                                elif "UNIDAD" in h_up: nueva_fila_traslado.append(str(t_unidad).upper())
                                elif "PISTA" in h_up: nueva_fila_traslado.append(pista_combinada)
                                elif "SEMANA" in h_up: nueva_fila_traslado.append(str(semana_traslado))
                                elif "OBSERVAC" in h_up: nueva_fila_traslado.append(str(t_observacion).strip())
                                elif "LOTE" in h_up: nueva_fila_traslado.append(lote_tras_inject)
                                else: nueva_fila_traslado.append("")
                            
                            gc_temp = inicializar_cliente_gspread()
                            sh_temp = gc_temp.open_by_url(URL_SHEET_TRASLADOS)
                            ws_write = sh_temp.worksheet(titulo_ws_traslados)

                            col_a = ws_write.col_values(1)
                            last_row = len(col_a)
                            while last_row > 0 and str(col_a[last_row-1]).strip() == "": last_row -= 1
                                
                            fila_destino = last_row + 1
                            rango_inyeccion = f"A{fila_destino}:{get_column_letter(len(nueva_fila_traslado))}{fila_destino}"
                            
                            try: ws_write.update(range_name=rango_inyeccion, values=[nueva_fila_traslado], value_input_option='USER_ENTERED')
                            except: ws_write.update(rango_inyeccion, [nueva_fila_traslado], value_input_option='USER_ENTERED')
                            
                        st.success(f"✅ ¡Operación registrada con éxito en la fila {fila_destino}!")
                        st.session_state['form_key_m19_traslados'] += 1
                        st.cache_data.clear(); st.rerun()
                    except Exception as e: st.error(f"Error al registrar operación: {e}")

        # --- VISOR HISTÓRICO DE TRASLADOS ---
        st.markdown("---")
        st.markdown("### 📋 Visor y Edición Histórica de Movimientos")
        
        if not df_traslados.empty:
            df_traslados_vista = df_traslados.copy()
            
            st.markdown("#### 🔍 Escáner de Filtrado")
            col_prod_t = next((c for c in df_traslados_vista.columns if "PRODUCTO" in c), None)
            
            if col_prod_t:
                productos_puros_t = set([str(x).strip().upper() for x in df_traslados_vista[col_prod_t].dropna() if str(x).strip() != ""])
                lista_productos_tabla_t = ["TODOS"] + sorted(list(productos_puros_t))
            else: lista_productos_tabla_t = ["TODOS"]
            
            f_col_t1, f_col_t2, f_col_t3 = st.columns([1.5, 2, 1])
            producto_filtro_t = f_col_t1.selectbox("🧪 Filtrar por Producto:", lista_productos_tabla_t, key="filtro_t_prod")
            st.markdown("<br>", unsafe_allow_html=True)
            
            if producto_filtro_t != "TODOS" and col_prod_t: df_traslados_vista = df_traslados_vista[df_traslados_vista[col_prod_t].str.upper() == producto_filtro_t]

            df_traslados_vista.insert(0, "🛡️ ACCIÓN", "✅ MANTENER")

            if "FECHA" in df_traslados_vista.columns:
                df_traslados_vista['FECHA_SORT'] = df_traslados_vista['FECHA'].apply(procesar_fecha_estricta)
                df_traslados_vista = df_traslados_vista.sort_values(by=['FECHA_SORT', 'FILA_EXCEL'], ascending=[False, False])
            else: df_traslados_vista = df_traslados_vista.sort_values(by=['FILA_EXCEL'], ascending=[False])

            st.caption("🔒 Haz doble clic en la columna '🛡️ ACCIÓN' para marcar y **ELIMINAR** un traslado de la Bóveda de Google Sheets.")

            columnas_vista_t = [c for c in df_traslados_vista.columns if c not in ['FILA_EXCEL', 'FECHA_SORT']]
            df_vista_t = df_traslados_vista[columnas_vista_t].copy()
            
            if "LOTE" in df_vista_t.columns: df_vista_t["LOTE"] = df_vista_t["LOTE"].astype(str).str.lstrip("'")
            
            cols_disabled_t = [c for c in df_vista_t.columns if c != "🛡️ ACCIÓN"]
            col_config_t = {"🛡️ ACCIÓN": st.column_config.SelectboxColumn("🛡️ ACCIÓN", help="Selecciona ELIMINAR para borrar esta fila.", options=["✅ MANTENER", "💥 ELIMINAR REGISTRO"], required=True)}
            
            for c in df_vista_t.columns:
                if c == "🛡️ ACCIÓN": continue
                c_up = c.upper()
                if "SEMANA" in c_up: col_config_t[c] = st.column_config.TextColumn("📅 SEM", width="small")
                elif "FECHA" in c_up: col_config_t[c] = st.column_config.TextColumn("🗓️ FECHA", width="medium")
                elif "PROD" in c_up: col_config_t[c] = st.column_config.TextColumn("🧪 PRODUCTO", width="large")
                elif "PISTA" in c_up: col_config_t[c] = st.column_config.TextColumn("📍 RUTA", width="medium")
                elif "CANT" in c_up: col_config_t[c] = st.column_config.TextColumn("⚖️ CANTIDAD", width="medium")
                elif "LOTE" in c_up: col_config_t[c] = st.column_config.TextColumn("📦 LOTE", width="medium")
                elif "OBSER" in c_up: col_config_t[c] = st.column_config.TextColumn("📝 OBS", width="medium")
                elif "CONSECUT" in c_up: col_config_t[c] = st.column_config.TextColumn("🔢 CONSECUTIVO", width="medium")

            df_editado_t = st.data_editor(df_vista_t, column_config=col_config_t, disabled=cols_disabled_t, hide_index=True, use_container_width=True, key="editor_traslados")

            st.markdown("<br>", unsafe_allow_html=True)
            
            # 💥 EJECUCIÓN BATCH PARA ELIMINACIÓN DE TRASLADOS
            if st.button("💾 EJECUTAR ELIMINACIÓN DE TRASLADOS EN DRIVE", type="primary", key="btn_del_traslados"):
                eliminaciones = [int(df_traslados_vista.iloc[i]['FILA_EXCEL']) for i in range(len(df_traslados_vista)) if "ELIMINAR REGISTRO" in str(df_editado_t.iloc[i]["🛡️ ACCIÓN"]).strip()]

                if eliminaciones:
                    with st.spinner("Eliminando registros de traslados en modo Batch (Cero bloqueos)..."):
                        try:
                            gc_temp = inicializar_cliente_gspread()
                            sh_temp = gc_temp.open_by_url(URL_SHEET_TRASLADOS)
                            ws_t = sh_temp.worksheet(titulo_ws_traslados)
                            
                            eliminaciones = sorted(list(set(eliminaciones)), reverse=True)
                            peticiones_borrado = []
                            for eli in eliminaciones:
                                peticiones_borrado.append({
                                    "deleteDimension": {
                                        "range": {
                                            "sheetId": ws_t.id,
                                            "dimension": "ROWS",
                                            "startIndex": eli - 1, 
                                            "endIndex": eli
                                        }
                                    }
                                })
                            sh_temp.batch_update({"requests": peticiones_borrado})
                            
                            st.success("✅ ¡Objetivo neutralizado! Los traslados han sido borrados del sistema mediante un solo barrido masivo.")
                            st.cache_data.clear(); st.rerun()
                        except Exception as e:
                            st.error(f"🚨 Error crítico al ejecutar borrado masivo: {e}")
        else: st.info("ℹ️ No marcaste ninguna fila con la acción de '💥 ELIMINAR REGISTRO'.")

    st.markdown("""<a href="#inicio-modulo-19" class="btn-ascensor" style="margin-top: 20px;">👆 VOLVER AL INICIO (ARRIBA) 👆</a>""", unsafe_allow_html=True)

if __name__ == "__main__":
    ejecutar()
