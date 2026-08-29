import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import gspread
import re
import math
import io
import openpyxl
from datetime import datetime, date
from oauth2client.service_account import ServiceAccountCredentials

# 🛰️ ENLACES NATIVOS
from modulos.utilidades import procesar_fecha_pesada
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =================================================================
# ⚙️ REGLAS DE NEGOCIO Y CONFIGURACIÓN ESTRATÉGICA (MODIFICABLES)
# =================================================================
TARIFAS_ST_BASE = {
    "TERCERO": 1583.0,
    "AFILIADO": 1510.0,
    "COOPERATIVA": 1510.0,
    "ORGANICO": 1337.0,
    "DEFAULT": 1337.0
}

# =================================================================
# 🔌 CONEXIÓN Y MOTORES DE FORMATO REGIONAL BLINDADOS
# =================================================================

def formato_latino(numero, decimales=0):
    if pd.isna(numero) or numero is None: return "0"
    try:
        num = float(numero)
        if num == 0: return "0"
        if decimales == 0: texto_us = f"{num:,.0f}"
        else: texto_us = f"{num:,.{decimales}f}"
        return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0"

def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except:
        return None

# 💥 LECTURA FINANCIERA EXACTA (Respeta los decimales reales)
def limpiar_tarifa_excel(val):
    if isinstance(val, (int, float)): return float(val)
    v = str(val).strip().replace("$", "").replace(" ", "").upper()
    if not v or v in ['-', 'NAN', 'NONE', '']: return 0.0
    
    s_clean = re.sub(r'[^\d\.,\-]', '', v)
    try:
        if '.' in s_clean and ',' in s_clean:
            if s_clean.rfind(',') > s_clean.rfind('.'): s_clean = s_clean.replace('.', '').replace(',', '.')
            else: s_clean = s_clean.replace(',', '')
        elif ',' in s_clean:
            if len(s_clean.split(',')[-1]) == 3: s_clean = s_clean.replace(',', '')
            else: s_clean = s_clean.replace(',', '.')
        elif '.' in s_clean:
            if s_clean.count('.') > 1: s_clean = s_clean.replace('.', '')
            elif len(s_clean.split('.')[-1]) == 3: s_clean = s_clean.replace('.', '')
        return float(s_clean) if s_clean else 0.0
    except:
        return 0.0

def normalizar_a_fecha_pura(val):
    try:
        res_nativo = procesar_fecha_pesada(val)
        if isinstance(res_nativo, (datetime, pd.Timestamp)): return res_nativo.date()
        if isinstance(res_nativo, date): return res_nativo
        return pd.to_datetime(str(res_nativo)).date()
    except: return None

# 💥 ESCUDO ANTI-ERRORES PARA FECHAS SUCIAS DE SAP
def parsear_fecha_sap(fecha_str):
    """Normaliza fechas de SAP y formatos comunes a date."""
    try:
        if fecha_str is None or pd.isna(fecha_str) or not str(fecha_str).strip():
            return None
    except (TypeError, ValueError):
        pass
    s = str(fecha_str).strip().replace("_", "-").replace("/", "-").replace(".", "-")
    try:
        return pd.to_datetime(s, dayfirst=True, errors="raise").date()
    except (ValueError, TypeError):
        return normalizar_a_fecha_pura(fecha_str)

@st.cache_data(show_spinner=False, ttl=600)
def cargar_bases_m18():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    try:
        boveda_act = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        
        # 1. CARGAR TABLA 1 (HISTÓRICO)
        datos_brutos = boveda_act.worksheet("TABLA 1").get_all_values()
        df_t1 = pd.DataFrame()
        if len(datos_brutos) > 5:
            idx_t1 = 4
            for i in range(min(8, len(datos_brutos))):
                if "FINCA" in [str(x).upper().strip() for x in datos_brutos[i]]:
                    idx_t1 = i
                    break
            
            encabezados = [str(c).upper().strip().replace("\n", "") for c in datos_brutos[idx_t1]]
            filas_limpias = [r + [""]*(len(encabezados) - len(r)) for r in datos_brutos[idx_t1+1:]]
            df_t1 = pd.DataFrame([r[:len(encabezados)] for r in filas_limpias], columns=encabezados)

        # 2. CARGAR TABLA 2 (PRODUCTORES)
        df_t2 = pd.DataFrame()
        try: 
            t2_raw = boveda_act.worksheet("TABLA 2").get_all_values()
            idx_t2 = next((i for i, r in enumerate(t2_raw) if "FINCA" in [str(x).upper().strip() for x in r]), 0)
            df_t2 = pd.DataFrame(t2_raw[idx_t2+1:], columns=[str(c).strip().upper() for c in t2_raw[idx_t2]])
        except: pass

        # 3. CARGAR CONFIGURACIÓN (TARIFAS ST)
        df_cfg = pd.DataFrame()
        try:
            cfg_raw = boveda_act.worksheet("Configuración").get_all_values()
            df_cfg = pd.DataFrame(cfg_raw[1:], columns=[str(c).upper().strip() for c in cfg_raw[0]])
        except: pass

        return df_t1, df_t2, df_cfg
    except Exception as e:
        raise Exception(f"Error de conexión con Google Drive: {e}")

# =================================================================
# 👑 RENDERIZADO VISUAL PRINCIPAL
# =================================================================

def ejecutar(*args, **kwargs):
    VERDE_INTENSO = '#143521'
    COLOR_NAVY = '#0d1b2a'
    COLOR_DORADO = '#d4af37'

    # 💥 BLOQUE CSS OPTIMIZADO
    css_maestro = f"""
    <style>
    .titulo-desglose {{ color: {COLOR_NAVY}; border-bottom: 3px solid {COLOR_DORADO}; padding-bottom: 5px; font-family: 'Arial Black'; margin-bottom: 15px;}}
    div[data-testid="stDataFrame"] {{ border: 3px solid {VERDE_INTENSO} !important; border-radius: 8px !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.1); overflow: hidden !important; }}
    .tarjeta-kpi {{ background: linear-gradient(135deg, {COLOR_NAVY} 0%, #1a365d 100%); border-left: 5px solid {COLOR_DORADO}; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.2); text-align: center; margin-bottom: 15px;}}
    .kpi-titulo {{ font-size: 13px; font-weight: bold; color: {COLOR_DORADO}; text-transform: uppercase; margin:0; letter-spacing: 1px; }}
    .kpi-valor {{ font-size: 26px; font-family: 'Arial Black'; margin: 5px 0 0 0; }}
    
    div[data-testid="stDateInput"] input,
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] {{
        background-color: #ffffff !important;
        border: 3px solid {VERDE_INTENSO} !important;
        border-radius: 6px !important;
    }}
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {{
        background-color: transparent !important;
        border: none !important;
    }}
    div[data-testid="stDateInput"] *, div[data-testid="stMultiSelect"] * {{
        color: #000000 !important;
        font-weight: bold !important;
    }}
    div[data-testid="stMainBlockContainer"] label p {{
        color: {COLOR_NAVY} !important;
        font-weight: 800 !important;
        text-transform: uppercase !important;
    }}
    </style>
    """
    
    st.markdown(css_maestro, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-desglose'>🔍 Módulo 18: Desglose y Auditoría de Facturación</h1>", unsafe_allow_html=True)
    st.write("Ingeniería inversa sobre la TABLA 1: Desglosa facturas ya ejecutadas para revelar el costo real de la mezcla química.")

    with st.spinner("Conectando con la Bóveda Maestra y calculando intervalos..."):
        df_t1, df_t2, df_cfg = cargar_bases_m18()

    if df_t1.empty:
        st.error("🚨 La Bóveda de Datos (TABLA 1) está vacía o inaccesible.")
        return

    # 1. IDENTIFICACIÓN DE COLUMNAS CLAVE
    col_finca = next((c for c in df_t1.columns if 'FINCA' in c or 'PROPIEDAD' in c), None)
    col_fecha = next((c for c in df_t1.columns if 'FECHA' in c or 'DATE' in c), None)
    col_ha = next((c for c in df_t1.columns if 'FUMIG' in c or 'NETA' in c or 'HECT' in c), None)
    col_coctel = next((c for c in df_t1.columns if 'COCTEL' in c or 'CÓCTEL' in c), None)
    col_precio_vuelo = next((c for c in df_t1.columns if 'COSTO' in c and 'AVI' in c and 'HA' in c.replace(" ", "")), None)
    col_costo_vuelo_finca = next((c for c in df_t1.columns if 'COSTO' in c and 'AVI' in c and 'FINCA' in c), None)
    col_valor_fact = next((c for c in df_t1.columns if 'VALOR' in c and 'FACTURAR' in c), None)

    if not all([col_finca, col_fecha, col_ha, col_precio_vuelo, col_costo_vuelo_finca, col_valor_fact]):
        st.error("🚨 Faltan columnas críticas en la TABLA 1. Verifique los nombres (FINCA, FECHA, ÁREA FUMIG, COSTO AVIÓN ($/ha), COSTO AVIÓN ($/finca), VALOR A FACTURAR).")
        return

    # 2. PREPROCESAMIENTO Y CÁLCULO DE DÍAS CICLO INTELIGENTE
    df_t1['FECHA_PURA'] = df_t1[col_fecha].apply(parsear_fecha_sap)
    df_t1 = df_t1.dropna(subset=['FECHA_PURA']).copy()
    df_t1['FINCA_CLEAN'] = df_t1[col_finca].astype(str).apply(lambda x: re.sub(r'[^A-Z0-9]', '', x.upper().strip()))
    
    df_fechas = df_t1[['FINCA_CLEAN', 'FECHA_PURA']].drop_duplicates().sort_values(by=['FINCA_CLEAN', 'FECHA_PURA'])
    df_fechas['FECHA_PREVIA'] = df_fechas.groupby('FINCA_CLEAN')['FECHA_PURA'].shift(1)
    df_fechas['DIAS_CICLO_REAL'] = (pd.to_datetime(df_fechas['FECHA_PURA']) - pd.to_datetime(df_fechas['FECHA_PREVIA'])).dt.days
    
    df_t1 = df_t1.merge(df_fechas, on=['FINCA_CLEAN', 'FECHA_PURA'], how='left')
    df_t1['DIAS_CICLO'] = df_t1['DIAS_CICLO_REAL'].fillna(14).astype(int)

    # 3. INTERFAZ DE FILTROS BLINDADA
    with st.container(border=True):
        st.markdown("#### 🎛️ Rango de Búsqueda y Selección")
        c1, c2 = st.columns(2)
        
        año_actual = datetime.now().year
        fecha_ini = c1.date_input("📅 Fecha Inicial", value=date(año_actual, 1, 1))
        fecha_fin = c2.date_input("📆 Fecha Final", value=date.today(), min_value=fecha_ini)

        fincas_disponibles = sorted([f for f in df_t1[col_finca].astype(str).str.upper().str.strip().unique().tolist() if f not in ['NAN', 'NONE', '']])
        fincas_sel = st.multiselect("📍 Seleccione las Fincas a desglosar (Deje vacío para analizarlas TODAS):", fincas_disponibles)

    if st.button("🔥 EJECUTAR DESGLOSE FINANCIERO", type="primary", use_container_width=True):
        
        mask_fechas = (df_t1['FECHA_PURA'] >= fecha_ini) & (df_t1['FECHA_PURA'] <= fecha_fin)
        df_operacion = df_t1[mask_fechas].copy()
        
        if fincas_sel:
            df_operacion = df_operacion[df_operacion[col_finca].astype(str).str.upper().str.strip().isin(fincas_sel)]

        if df_operacion.empty:
            st.warning("📭 No se encontraron facturas ejecutadas en ese rango de fechas.")
        else:
            with st.spinner("Desarmando la facturación pieza por pieza a ultra velocidad..."):
                resultados = []

                # 💥 ACELERADOR: Limpiar Tabla 2 UNA sola vez fuera del bucle
                if not df_t2.empty:
                    df_t2['FINCA_CLEAN_T2'] = df_t2.iloc[:, 0].astype(str).str.upper().apply(lambda x: re.sub(r'[^A-Z0-9]', '', x))

                # 💥 ACELERADOR DE BUCLE PARA MILES DE FILAS (to_dict)
                for row in df_operacion.to_dict('records'):
                    finca_raw = str(row[col_finca]).upper().strip()
                    finca_clean = row['FINCA_CLEAN']
                    
                    fecha_pura = row['FECHA_PURA']
                    fecha_str = fecha_pura.strftime("%d/%m/%Y") if pd.notna(fecha_pura) else "S/F"
                    
                    coctel = str(row.get(col_coctel, "S/N")).upper().strip() if col_coctel in row else "S/N"
                    ha_num = limpiar_tarifa_excel(row[col_ha])
                    dias_ciclo = row['DIAS_CICLO']

                    # 4. EXTRACCIÓN DE VALORES REALES
                    costo_vuelo_ha = limpiar_tarifa_excel(row[col_precio_vuelo])
                    costo_vuelo_finca = limpiar_tarifa_excel(row[col_costo_vuelo_finca])
                    costo_x_ha_facturado = limpiar_tarifa_excel(row[col_valor_fact])

                    if ha_num <= 0 or costo_x_ha_facturado <= 0: continue

                    # 5. CRUZAR TARIFA DE SERVICIO TÉCNICO (Cruce 100% perfecto)
                    tipo_productor = "TERCERO"
                    if not df_t2.empty:
                        match_t2 = df_t2[df_t2['FINCA_CLEAN_T2'] == finca_clean]
                        if not match_t2.empty and len(match_t2.columns) > 5:
                            tipo_productor = str(match_t2.iloc[0].iloc[5]).strip().upper()

                    if "COOP" in finca_raw or "EMPREBANCOOP" in finca_raw: 
                        tipo_productor = "COOPERATIVA"

                    st_base = 0.0
                    # Lectura desde configuración si existe
                    if not df_cfg.empty:
                        match_cfg = df_cfg[df_cfg.iloc[:, 0].astype(str).str.strip().str.upper() == tipo_productor]
                        if not match_cfg.empty and len(match_cfg.columns) > 4:
                            st_base = limpiar_tarifa_excel(match_cfg.iloc[0].iloc[4])

                    # 💥 FALLBACK DE REGLAS DE NEGOCIO (Evita el cero)
                    if st_base == 0:
                        st_base = TARIFAS_ST_BASE.get(tipo_productor, TARIFAS_ST_BASE["DEFAULT"])

                    # 6. INGENIERÍA INVERSA (El Desglose con escudo anti-negativos)
                    resultado_total = math.floor(costo_x_ha_facturado * ha_num)
                    
                    if costo_vuelo_finca == 0 and costo_vuelo_ha > 0:
                        costo_vuelo_finca = costo_vuelo_ha * ha_num

                    costo_st_total = math.floor(st_base * dias_ciclo * ha_num)
                    
                    # ESCUDO: Si hubo un error humano en SAP, no arroja mezcla negativa
                    costo_mezcla_total = max(0, resultado_total - costo_vuelo_finca - costo_st_total)

                    resultados.append({
                        "FINCA": finca_raw,
                        "FECHA": fecha_str,  
                        "HECTAREAS": ha_num,
                        "COCTEL": coctel,
                        "DIAS CICLO": dias_ciclo,
                        "PRECIO VUELO": costo_vuelo_ha,
                        "Costo ST ($)": costo_st_total,
                        "Costo Vuelo ($)": costo_vuelo_finca,
                        "Costo Mezcla ($)": costo_mezcla_total,
                        "Costo x Ha ($)": costo_x_ha_facturado,
                        "RESULTADO TOTAL ($)": resultado_total,
                        "_ORDEN_FECHA": fecha_pura 
                    })

                df_resultados = pd.DataFrame(resultados)

                if df_resultados.empty:
                    st.warning("⚠️ No se pudieron generar datos válidos. Verifique que las hectáreas y valores a facturar no sean cero en SAP.")
                    return

                # 💥 ORDENAMIENTO ALFABÉTICO Y CRONOLÓGICO OBLIGATORIO
                df_resultados = df_resultados.sort_values(by=["FINCA", "_ORDEN_FECHA"], ascending=[True, True]).drop(columns=["_ORDEN_FECHA"]).reset_index(drop=True)

                # 7. PRESENTACIÓN DE DATOS (KPIs 2x2 para UX Ejecutiva)
                t_st = df_resultados['Costo ST ($)'].sum()
                t_vu = df_resultados['Costo Vuelo ($)'].sum()
                t_mx = df_resultados['Costo Mezcla ($)'].sum()
                t_gr = df_resultados['RESULTADO TOTAL ($)'].sum()

                st.markdown("---")
                st.markdown("### 🎛️ Tablero de Resultados Financieros")
                
                # Fila 1: ST y Vuelo
                k1, k2 = st.columns(2)
                with k1: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>👨‍🔬 Total Serv. Tec</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_st, 0)}</p></div>", unsafe_allow_html=True)
                with k2: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>✈️ Total Vuelo</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_vu, 0)}</p></div>", unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)

                # Fila 2: Mezcla y Gran Total
                k3, k4 = st.columns(2)
                with k3: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>🧪 Total Mezcla Real</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_mx, 0)}</p></div>", unsafe_allow_html=True)
                with k4: st.markdown(f"<div class='tarjeta-kpi' style='border-left: 5px solid #00ff00;'><p class='kpi-titulo' style='color:#00ff00;'>🔥 FACTURADO (SAP)</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_gr, 0)}</p></div>", unsafe_allow_html=True)

                # Resumen Ejecutivo
                df_resumen_finca = df_resultados.groupby('FINCA', as_index=False)[
                    ['Costo ST ($)', 'Costo Vuelo ($)', 'Costo Mezcla ($)', 'RESULTADO TOTAL ($)']
                ].sum()

                tab1, tab2 = st.tabs(["📊 Detalles Económicos Fila x Fila", "📑 Resumen Ejecutivo por Finca"])
                
                # 💥 MOTOR VISUAL PREMIUM PARA RESULTADOS (Sombreados y Formatos)
                def estilo_detalles(row):
                    estilos = ['background-color: #ffffff; color: #0d1b2a; font-weight: 500;'] * len(row)
                    for i, col in enumerate(row.index):
                        if col == 'RESULTADO TOTAL ($)':
                            estilos[i] = 'background-color: #e6f4ea; color: #143521; font-weight: 900; border-left: 2px solid #28a745; text-align: right;'
                        elif 'Costo' in col or col in ['PRECIO VUELO', 'RECARGO ($/HA)']:
                            estilos[i] = 'background-color: #f8f9fa; color: #333333; font-weight: 600; text-align: right;'
                    return estilos

                def estilo_resumen(row):
                    estilos = ['background-color: #ffffff; color: #0d1b2a; font-weight: 500;'] * len(row)
                    for i, col in enumerate(row.index):
                        if col == 'RESULTADO TOTAL ($)':
                            estilos[i] = 'background-color: #fff9e6; color: #0d1b2a; font-weight: 900; border-left: 2px solid #d4af37; text-align: right;'
                        elif 'Costo' in col:
                            estilos[i] = 'background-color: #f8f9fa; color: #333333; font-weight: 600; text-align: right;'
                    return estilos

                formato_columnas = {col: lambda x: f"$ {x:,.0f}".replace(",", ".") for col in ["PRECIO VUELO", "Costo ST ($)", "Costo Vuelo ($)", "Costo Mezcla ($)", "Costo x Ha ($)", "RESULTADO TOTAL ($)"]}
                formato_columnas["HECTAREAS"] = lambda x: f"{x:.2f}"
                formato_columnas["DIAS CICLO"] = lambda x: f"{int(x)}"
                
                with tab1:
                    st.dataframe(
                        df_resultados.style.apply(estilo_detalles, axis=1).format(formato_columnas), 
                        use_container_width=True, 
                        hide_index=True
                    )

                with tab2:
                    if not df_resumen_finca.empty:
                        formato_dinero_res = {col: lambda x: f"$ {x:,.0f}".replace(",", ".") for col in ['Costo ST ($)', 'Costo Vuelo ($)', 'Costo Mezcla ($)', 'RESULTADO TOTAL ($)']}
                        st.dataframe(
                            df_resumen_finca.style.apply(estilo_resumen, axis=1).format(formato_dinero_res), 
                            use_container_width=True, 
                            hide_index=True
                        )
                    else: st.info("No hay datos para resumir.")

                # 8. EXPORTADOR EXCEL PROFESIONAL (VIP)
                st.markdown("<br>", unsafe_allow_html=True)
                buffer = io.BytesIO()
                
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_resultados.to_excel(writer, sheet_name='Detalle_Económico', index=False, startrow=3)
                    df_resumen_finca.to_excel(writer, sheet_name='Resumen_x_Finca', index=False, startrow=3)
                    
                    workbook = writer.book
                    
                    # Paleta de Estilos VIP
                    fill_titulo = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                    font_titulo = Font(color="FFFFFF", bold=True, size=14)
                    font_sub = Font(color="555555", italic=True, size=10)
                    
                    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
                    header_font = Font(color="D4AF37", bold=True)
                    borde = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                                   top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
                    
                    fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')

                    for sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        ws.sheet_view.showGridLines = False
                        max_r, max_c = ws.max_row, ws.max_column
                        
                        # Inyección de Títulos Combinados
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_c)
                        c_tit = ws.cell(row=1, column=1, value=f"REPORTE DE AUDITORÍA FINANCIERA — {sheet_name.upper().replace('_', ' ')}")
                        c_tit.fill = fill_titulo
                        c_tit.font = font_titulo
                        c_tit.alignment = Alignment(horizontal='center', vertical='center')

                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_c)
                        c_sub = ws.cell(row=2, column=1, value=f"Desglose de Costos e Ingeniería Inversa | Generado el: {fecha_actual}")
                        c_sub.font = font_sub
                        c_sub.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Auto-ajuste de columnas
                        column_headers = {}
                        for col_idx in range(1, max_c + 1):
                            letra_columna = openpyxl.utils.get_column_letter(col_idx)
                            header_val = ws.cell(row=4, column=col_idx).value
                            valores_columna = [ws.cell(row=r, column=col_idx).value for r in range(4, min(max_r, 50) + 1)]
                            ancho = max((len(str(v)) if v is not None else 0) for v in valores_columna) + 4
                            ws.column_dimensions[letra_columna].width = min(max(ancho, 14), 35)
                            column_headers[col_idx] = str(header_val).upper() if header_val else ""

                        for row in ws.iter_rows(min_row=4, max_row=max_r, min_col=1, max_col=max_c):
                            for cell in row:
                                cell.border = borde
                                if cell.row == 4:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                else:
                                    cell.alignment = Alignment(vertical='center')
                                    col_name = column_headers.get(cell.column, "")
                                    
                                    if isinstance(cell.value, (int, float)):
                                        if any(kw in col_name for kw in ["COSTO", "PRECIO", "RESULTADO", "TOTAL ($)", "RECARGO"]):
                                            cell.number_format = '"$"#,##0' 
                                        elif any(kw in col_name for kw in ["HECTAREAS", "VOLUMEN"]):
                                            cell.number_format = '#,##0.0'

                st.download_button(
                    label="💾 DESCARGAR AUDITORÍA GERENCIAL (EXCEL VIP)",
                    data=buffer.getvalue(),
                    file_name=f"Auditoria_Facturacion_{fecha_ini}_{fecha_fin}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.success("✅ Ingeniería Inversa completada y ordenada alfabéticamente. La estructura de costos ocultos ha sido revelada con exactitud matemática.")

if __name__ == "__main__":
    pass
