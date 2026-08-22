import streamlit as st
import pandas as pd
import gspread
from datetime import datetime, timedelta
import re
import io
import requests
import folium
import plotly.express as px
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# ⚙️ CONSTANTES CENTRALIZADAS (ÚNICA FUENTE DE VERDAD)
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

# =================================================================
# ⚡ MOTOR DE CONEXIÓN UNIFICADO (V42 VIP)
# =================================================================
@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

# =================================================================
# 🛡️ UTILIDADES DE PURIFICACIÓN Y KML
# =================================================================
def a_numero_limpio(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip().replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except: return 0.0

def procesar_fecha_pesada(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    s = str(val).strip()
    if s.replace('.', '', 1).isdigit(): 
        return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: return pd.to_datetime(s, errors='coerce')
    except: return pd.NaT

def limpiar_nombre(texto):
    txt = re.sub(r'[^\w]', '', str(texto).upper())
    txt = txt.replace("FINCA", "").replace("KML", "")
    return txt

def extraer_poligonos_kml(kml_bytes):
    try:
        texto = kml_bytes.decode("utf-8", errors="ignore")
        bloques = re.findall(r'<coordinates>(.*?)</coordinates>', texto, re.IGNORECASE | re.DOTALL)
        poligonos_finca = []
        for bloque in bloques:
            coordenadas_crudas = bloque.strip().split()
            puntos = []
            for coord in coordenadas_crudas:
                partes = coord.split(',')
                if len(partes) >= 2:
                    try:
                        lon = float(partes[0].strip())
                        lat = float(partes[1].strip())
                        puntos.append([lat, lon]) 
                    except: pass
            if len(puntos) >= 3: 
                poligonos_finca.append(puntos)
        return poligonos_finca
    except: return []

# =================================================================
# 🛰️ CONEXIÓN SATELITAL CLIMÁTICA AVANZADA
# =================================================================
@st.cache_data(show_spinner=False, ttl=3600)
def consultar_clima_avanzado(lat, lon):
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&past_days=90&forecast_days=7&daily=precipitation_sum&timezone=America/Bogota"
        headers = {"User-Agent": "AgroAereoTactico/1.0"}
        res = requests.get(url, headers=headers, timeout=15)
        
        if res.status_code == 200:
            data = res.json()
            if "daily" in data and "precipitation_sum" in data["daily"]:
                df_clima = pd.DataFrame({
                    'fecha': pd.to_datetime(data['daily']['time']),
                    'lluvia': [float(x) if x is not None else 0.0 for x in data['daily']['precipitation_sum']]
                })
                
                hoy_satelite = df_clima['fecha'].max() - pd.Timedelta(days=7)
                hace_30_dias = hoy_satelite - pd.Timedelta(days=30)
                
                lluvia_90d = float(df_clima[df_clima['fecha'] <= hoy_satelite]['lluvia'].sum())
                lluvia_30d = float(df_clima[(df_clima['fecha'] <= hoy_satelite) & (df_clima['fecha'] >= hace_30_dias)]['lluvia'].sum())
                lluvia_7d_futuro = float(df_clima[df_clima['fecha'] > hoy_satelite]['lluvia'].sum())
                
                df_clima['fecha_str'] = df_clima['fecha'].dt.strftime('%Y-%m-%d')
                clima_diario = dict(zip(df_clima['fecha_str'], df_clima['lluvia']))
                
                return lluvia_90d, lluvia_30d, lluvia_7d_futuro, clima_diario
    except Exception: pass
    return 0.0, 0.0, 0.0, {"error": True}

# =================================================================
# 💾 EXTRACCIÓN CACHEADA
# =================================================================
@st.cache_data(show_spinner=False, ttl=3600)
def cargar_historico_t1():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame()
    
    try:
        boveda = gc.open_by_url(URL_BOVEDA_MAESTRA)
        t1_vals = boveda.worksheet("TABLA 1").get_all_values()
        
        idx_t1 = 4
        for i in range(min(8, len(t1_vals))):
            fila_limpia = [str(x).upper().strip() for x in t1_vals[i]]
            if "Nº ORDEN" in fila_limpia or "FINCA" in fila_limpia:
                idx_t1 = i
                break
                
        df_t1 = pd.DataFrame(t1_vals[idx_t1+1:], columns=[str(c).upper().strip() for c in t1_vals[idx_t1]])
        
        col_fecha = next((c for c in df_t1.columns if 'FECHA' in c), 'FECHA')
        col_ha = next((c for c in df_t1.columns if 'NETA' in c or 'FUMIG' in c or 'HECT' in c), None)
        col_sector = next((c for c in df_t1.columns if 'SECTOR' in c), 'SECTOR')
        col_finca = next((c for c in df_t1.columns if 'FINCA' in c), 'FINCA')
        
        if col_fecha and col_ha:
            df_t1['FECHA_DT'] = df_t1[col_fecha].apply(procesar_fecha_pesada)
            df_t1 = df_t1.dropna(subset=['FECHA_DT'])
            df_t1['HA_CALCULO'] = df_t1[col_ha].apply(a_numero_limpio)
            df_t1['SECTOR_NOM'] = df_t1[col_sector].astype(str).str.upper().str.strip()
            df_t1['FINCA_NOM'] = df_t1[col_finca].astype(str).str.upper().str.strip()
            return df_t1
    except: pass
    return pd.DataFrame()

# =================================================================
# 📤 EXPORTADOR EXCEL VIP MULTI-HOJA
# =================================================================
def generar_excel_agronomico_vip(dict_dfs):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for sheet_name, df in dict_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            
            header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
            header_font = Font(color="D4AF37", bold=True)
            borde_fino = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                                top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
                                
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = borde_fino
                
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                    
                    if cell.row > 1:
                        cell.border = borde_fino
                        if isinstance(cell.value, (int, float)):
                            col_name = str(ws[col_letter + '1'].value).upper()
                            if "LLUVIA" in col_name or "DÍAS" in col_name or "RETORNO" in col_name or "PRONÓSTICO" in col_name or sheet_name == "Historial_Lluvias":
                                if col_name != "FECHA":
                                    cell.number_format = '#,##0.0'
                                
                ws.column_dimensions[col_letter].width = min(max_length + 4, 30)
                
    return buffer.getvalue()

# =================================================================
# 🚀 EJECUCIÓN PRINCIPAL
# =================================================================
def ejecutar(purificar_lote, extraer_numero):
    VERDE_INTENSO = '#143521'
    DORADO = '#d4af37'

    st.markdown(f"""
    <style>
    .titulo-agronomo {{ color: #0d1b2a; border-bottom: 3px solid #27AE60; padding-bottom: 5px; font-family: 'Arial Black'; }}
    [data-testid="column"] {{ display: flex !important; flex-direction: column !important; justify-content: flex-start !important; align-items: stretch !important; }}
    div[data-testid="stDataFrame"] {{ border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.08) !important; }}
    
    div[data-testid="stFileUploader"] > div {{ background-color: #ffffff !important; border: 2px solid {VERDE_INTENSO} !important; border-radius: 8px !important; box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important; }}
    div[data-testid="stFileUploader"] * {{ color: #000000 !important; font-weight: bold !important; }}
    div[data-testid="stFileUploader"] button, div[data-testid="stFileUploader"] button * {{ color: #ffffff !important; }}
    
    div[data-testid="stMainBlockContainer"] label p {{ color: #0d1b2a !important; font-weight: 800 !important; text-transform: uppercase !important; }}
    div[data-testid="stTabs"] button[role="tab"] {{ font-family: 'Arial Black', sans-serif; font-size: 14px; color: #0d1b2a; }}
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{ border-bottom-color: #27AE60; background-color: rgba(39, 174, 96, 0.1); }}
    
    [data-testid="stPlotlyChart"] {{ transition: transform 0.3s ease, box-shadow 0.3s ease !important; border-radius: 8px; }}
    [data-testid="stPlotlyChart"]:hover {{ transform: translateY(-4px) scale(1.015) !important; box-shadow: 0 12px 25px rgba(39, 174, 96, 0.25) !important; z-index: 10; }}
    </style>
    """, unsafe_allow_html=True)

    def tarjeta_kpi(titulo, valor, delta_texto="", color_delta="#28a745"):
        delta_html = f"<span style='font-size: 14px; color: {color_delta}; margin-left: 8px; vertical-align: middle; padding: 2px 6px; border-radius: 4px; background-color: rgba(255,255,255,0.1);'>{delta_texto}</span>" if delta_texto else ""
        return f"""
        <div style='background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); border-left: 5px solid #27AE60; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.15); margin-bottom: 20px; height: 100%; min-height: 85px; display: flex; flex-direction: column; justify-content: center;'>
            <p style='font-size: 11px; font-weight: bold; color: #27AE60; text-transform: uppercase; margin:0 0 5px 0; letter-spacing: 1px;'>{titulo}</p>
            <p style='font-size: 22px; font-family: "Arial Black", sans-serif; margin: 0; color: white; display: flex; align-items: center;'>{valor} {delta_html}</p>
        </div>
        """

    c_tit, c_sync = st.columns([3.5, 1.5])
    with c_tit:
        st.markdown("<h1 class='titulo-agronomo'>🗺️ Módulo 15: Centro de Mando Agronómico</h1>", unsafe_allow_html=True)
        st.write("Radar de ciclos biológicos, riesgo de lavado de producto y ventanas de vuelo satelitales.")
    with c_sync:
        st.write("")
        if st.button("🔄 Sincronizar Nube (Forzar Datos)", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()

    with st.container(border=True):
        st.markdown("### 📂 1. Inyección de Polígonos de Precisión")
        
        if "kml_reset_key" not in st.session_state:
            st.session_state.kml_reset_key = 0

        c_file, c_btn_trash = st.columns([3, 1])
        archivos_kml = c_file.file_uploader(
            "Arrastre aquí los archivos .kml de sus fincas", 
            type=['kml'], accept_multiple_files=True, 
            key=f"kml_uploader_{st.session_state.kml_reset_key}", label_visibility="collapsed"
        )

        with c_btn_trash:
            if st.button("🗑️ Vaciar Bandeja de KMLs", type="secondary", use_container_width=True):
                st.session_state.kml_reset_key += 1
                st.rerun()

        coor_estimadas = {
            "PALOMINO": [11.2442, -73.5623],
            "BURITACA": [11.2420, -73.7650],
            "GUACHACA": [11.2411, -73.8188],
            "CIENAGA": [11.0070, -74.2478],
            "RIO FRIO": [10.9000, -74.1667],
            "ORIHUECA": [10.7483, -74.1542],
            "CAÑO MOCHO": [10.7820, -74.1850],
            "LA CEIBA": [10.7350, -74.1620],
            "PALOMAR": [10.7210, -74.1150],
            "FLORIDA": [10.7650, -74.1320],
            "SEVILLA": [10.7667, -74.1500],
            "GUACAMAYAL": [10.7292, -74.1594],
            "TUCURINCA": [10.5842, -74.1489],
            "FUNDACION": [10.5208, -74.1833]
        }

        orden_logistico = [
            "PALOMINO", "BURITACA", "GUACHACA", "CIENAGA", "RIO FRIO", 
            "ORIHUECA", "CAÑO MOCHO", "LA CEIBA", "PALOMAR", "FLORIDA", 
            "SEVILLA", "GUACAMAYAL", "TUCURINCA", "FUNDACION"
        ]

        # 🎯 MEJORA TÁCTICA: MODO AUTOCALIBRADO CON ESCENARIOS (LIMPIEZA DE PANTALLA)
        escenario_sel = st.selectbox(
            "⚙️ Escenario de Calibración Satelital:",
            ["🤖 Autocalibrado Estándar (Recomendado por Defecto)", "🟢 Lectura Satelital Pura (Sin Ajuste - 1.0x)", "🔴 Escenario de Tormenta Convectiva Extrema"],
            index=0,
            help="Seleccione el escenario sin necesidad de ajustar deslizadores numéricos."
        )

        # Asignación de factores según el escenario
        if escenario_sel == "🟢 Lectura Satelital Pura (Sin Ajuste - 1.0x)":
            factor_norte, factor_sur = 1.0, 1.0
        elif escenario_sel == "🔴 Escenario de Tormenta Convectiva Extrema":
            factor_norte, factor_sur = 10.0, 6.0
        else: # Autocalibrado por defecto
            factor_norte, factor_sur = 6.0, 3.5

        # Expander opcional por si el analista experto desea un ajuste fino
        with st.expander("🛠️ Ajuste Fino Manual de Multiplicadores (Avanzado)"):
            c_cal1, c_cal2 = st.columns(2)
            factor_norte = c_cal1.slider("🌊 Multiplicador Zona Norte (Caribe)", min_value=1.0, max_value=20.0, value=factor_norte, step=0.5)
            factor_sur = c_cal2.slider("🍌 Multiplicador Zona Sur (Bananera/Fundación)", min_value=1.0, max_value=20.0, value=factor_sur, step=0.5)

        st.markdown("<br>", unsafe_allow_html=True)

        if st.button("🚀 BARRIDO SATELITAL Y EJECUCIÓN GENERAL", type="primary", use_container_width=True):
            with st.spinner("Decodificando satélites y construyendo base pluviométrica..."):
                
                df_t1 = cargar_historico_t1()
                if df_t1.empty:
                    st.error("🚨 No se pudo conectar a la base de datos operativa.")
                    st.stop()

                dict_poligonos_kml = {}
                if archivos_kml:
                    for f_kml in archivos_kml:
                        nombre_finca_kml = f_kml.name.upper().replace(".KML", "").strip()
                        poligonos = extraer_poligonos_kml(f_kml.read())
                        if poligonos:
                            dict_poligonos_kml[nombre_finca_kml] = poligonos

                fincas_unicas = df_t1['FINCA_NOM'].unique()
                analisis_fincas = []
                
                sectores_unicos = df_t1['SECTOR_NOM'].unique()
                cache_clima = {}
                api_fallo = False
                zona_norte_keywords = ["PALOMINO", "GUACHACA", "BURITACA", "DON DIEGO", "CIENAGA"]
                
                for sec, gps in coor_estimadas.items():
                    l_90, l_30, l_7f, dict_diario = consultar_clima_avanzado(gps[0], gps[1])
                    if "error" in dict_diario:
                        api_fallo = True
                        
                    multiplicador = factor_norte if any(k in sec for k in zona_norte_keywords) else factor_sur
                    
                    l_90 = max(0.0, l_90) * multiplicador
                    l_30 = max(0.0, l_30) * multiplicador
                    l_7f = max(0.0, l_7f) * multiplicador
                    dict_diario_ajustado = {k: max(0.0, v) * multiplicador for k, v in dict_diario.items() if k != "error"}
                    
                    cache_clima[sec] = (l_90, l_30, l_7f, dict_diario_ajustado)
                
                if api_fallo:
                    st.error("📡 SATÉLITE DESCONECTADO: Open-Meteo no devolvió la pluviometría. Los valores mostrarán 0 mm.")
                
                historico_clima = []
                for sec, datos in cache_clima.items():
                    dict_diario = datos[3] 
                    for d, val in dict_diario.items():
                        historico_clima.append({"FECHA": d, "SECTOR": sec, "LLUVIA (mm)": val})
                
                df_clima_raw = pd.DataFrame(historico_clima)
                df_clima_pivot = pd.DataFrame()
                cols_presentes_ordenadas = []
                
                if not df_clima_raw.empty:
                    df_clima_pivot = df_clima_raw.pivot_table(index="FECHA", columns="SECTOR", values="LLUVIA (mm)", aggfunc='mean').reset_index()
                    df_clima_pivot = df_clima_pivot.sort_values("FECHA", ascending=False)
                    df_clima_pivot['FECHA'] = pd.to_datetime(df_clima_pivot['FECHA']).dt.strftime('%d/%m/%Y')
                    
                    cols_presentes_ordenadas = [sec for sec in orden_logistico if sec in df_clima_pivot.columns]
                    cols_extra = [sec for sec in df_clima_pivot.columns if sec not in orden_logistico and sec != "FECHA"]
                    
                    df_clima_pivot = df_clima_pivot[['FECHA'] + cols_presentes_ordenadas + cols_extra]
                
                for finca in fincas_unicas:
                    if not finca or finca in ["NAN", "NONE", ""]: continue
                    
                    df_finca = df_t1[df_t1['FINCA_NOM'] == finca].sort_values(by='FECHA_DT')
                    fechas_vuelos = df_finca['FECHA_DT'].unique()
                    
                    sectores_frecuentes = df_finca['SECTOR_NOM'].value_counts()
                    sector_asociado = sectores_frecuentes.index[0] if not sectores_frecuentes.empty else "DESCONOCIDO"
                    
                    ultimo_vuelo = None
                    dias_ciclo = 30
                    if len(fechas_vuelos) >= 1:
                        ultimo_vuelo = pd.to_datetime(fechas_vuelos[-1])
                        if len(fechas_vuelos) >= 2:
                            vuelo_anterior = pd.to_datetime(fechas_vuelos[-2])
                            dias_ciclo = (ultimo_vuelo - vuelo_anterior).days
                    
                    if dias_ciclo <= 12:
                        estado = "🚨 CRÍTICO"
                        color_hex = "#cc0000"
                    elif dias_ciclo <= 20:
                        estado = "🟠 MODERADO"
                        color_hex = "#ff9900"
                    else:
                        estado = "🟢 CONTROLADO"
                        color_hex = "#27AE60"

                    lluvia_90d, lluvia_30d, lluvia_7d_futuro, clima_diario = 0.0, 0.0, 0.0, {}
                    if sector_asociado in cache_clima:
                        lluvia_90d, lluvia_30d, lluvia_7d_futuro, clima_diario = cache_clima[sector_asociado]
                    
                    alerta_epidemia = "Baja / Normal"
                    if lluvia_30d > 100.0: alerta_epidemia = "⚡ ALTA (Peligro Inminente)"

                    lluvia_aplicacion = 0.0
                    riesgo_lavado = "🟢 SEGURO"
                    if ultimo_vuelo and not api_fallo:
                        fecha_vuelo_str = ultimo_vuelo.strftime('%Y-%m-%d')
                        lluvia_aplicacion = clima_diario.get(fecha_vuelo_str, 0.0)
                        if lluvia_aplicacion >= 30.0:
                            riesgo_lavado = "🔴 ALTO (Lavado Probable)"
                        elif lluvia_aplicacion >= 10.0:
                            riesgo_lavado = "🟡 MODERADO"

                    if lluvia_7d_futuro >= 80.0:
                        ventana_vuelo = "🔴 CERRADA (No programar)"
                    elif lluvia_7d_futuro >= 30.0:
                        ventana_vuelo = "🟡 PRECAUCIÓN"
                    else:
                        ventana_vuelo = "🟢 ÓPTIMA"
                        
                    gps = coor_estimadas.get(sector_asociado, [10.7483, -74.1542])

                    analisis_fincas.append({
                        "FINCA": finca,
                        "SECTOR": sector_asociado,
                        "ÚLTIMA APLICACIÓN": ultimo_vuelo.strftime('%d/%m/%Y') if ultimo_vuelo else "Sin Registro",
                        "ÚLTIMO RETORNO (Días)": float(dias_ciclo),
                        "ESTADO CICLO": estado,
                        "LLUVIA DÍA APLIC. (mm)": float(lluvia_aplicacion),
                        "RIESGO DE LAVADO": riesgo_lavado,
                        "PRONÓSTICO 7D (mm)": float(lluvia_7d_futuro),
                        "VENTANA DE VUELO": ventana_vuelo,
                        "LLUVIA 90D (mm)": float(lluvia_90d),
                        "LLUVIA 30D (mm)": float(lluvia_30d),
                        "PRESIÓN HONGO": alerta_epidemia,
                        "COOR": gps,
                        "COLOR": color_hex
                    })

                st.markdown("---")
                df_maestro = pd.DataFrame(analisis_fincas)
                
                if not df_maestro.empty:
                    fincas_lavado = len(df_maestro[df_maestro['RIESGO DE LAVADO'].str.contains("ALTO")])
                    fincas_ventana_cerrada = len(df_maestro[df_maestro['VENTANA DE VUELO'].str.contains("CERRADA")])

                    k1, k2, k3 = st.columns(3)
                    with k1: st.markdown(tarjeta_kpi("Fincas Mapeadas", f"{len(df_maestro)}", "En radar", "#27AE60"), unsafe_allow_html=True)
                    with k2: st.markdown(tarjeta_kpi("🚨 Riesgo Lavado Químico", f"{fincas_lavado} Fincas", "Acortar ciclo", "#ff4b4b" if fincas_lavado > 0 else "#28a745"), unsafe_allow_html=True)
                    with k3: st.markdown(tarjeta_kpi("⛈️ Ventana Vuelo Cerrada", f"{fincas_ventana_cerrada} Zonas", "No volar (7 Días)", "#ff4b4b" if fincas_ventana_cerrada > 0 else "#28a745"), unsafe_allow_html=True)

                mapa_magdalena = folium.Map(
                    location=[10.7483, -74.1542], 
                    zoom_start=10, 
                    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                    attr='Esri World Imagery'
                )
                st.markdown("### 🛰️ Mapa de Calor y Decisiones (En Vivo)")
                
                sectores_dibujados = []
                kmls_usados = set() 

                for f_info in analisis_fincas:
                    finca_nom = f_info["FINCA"]
                    sector_nom = f_info["SECTOR"]
                    color_nodo = f_info["COLOR"]
                    
                    popup_text = f"""
                    <b>Finca:</b> {finca_nom} ({sector_nom})<br>
                    <b>Retorno Actual:</b> {f_info["ÚLTIMO RETORNO (Días)"]:.0f} Días<br>
                    <b>Riesgo Lavado:</b> {f_info["RIESGO DE LAVADO"]}<br>
                    <b>Ventana 7D:</b> {f_info["VENTANA DE VUELO"]}<br>
                    <b>Lluvia Mensual:</b> {f_info["LLUVIA 30D (mm)"]:.1f} mm
                    """
                    
                    f_norm = limpiar_nombre(finca_nom)
                    kml_clave = None
                    
                    for k in dict_poligonos_kml.keys():
                        k_norm = limpiar_nombre(k)
                        if (k_norm in f_norm or f_norm in k_norm) and len(k_norm) > 3:
                            kml_clave = k
                            kmls_usados.add(k) 
                            break
                    
                    if kml_clave:
                        lats_finca = []
                        lons_finca = []
                        for poligono in dict_poligonos_kml[kml_clave]:
                            folium.Polygon(
                                locations=poligono, color=color_nodo, weight=2, fill=True,
                                fill_color=color_nodo, fill_opacity=0.6,
                                tooltip=f"{finca_nom} | Ventana: {f_info['VENTANA DE VUELO']}",
                                popup=folium.Popup(popup_text, max_width=300)
                            ).add_to(mapa_magdalena)
                            lats_finca.extend([p[0] for p in poligono])
                            lons_finca.extend([p[1] for p in poligono])
                            
                        if lats_finca and lons_finca:
                            centro_lat = (min(lats_finca) + max(lats_finca)) / 2
                            centro_lon = (min(lons_finca) + max(lons_finca)) / 2
                            html_label = f"""<div style="font-size: 11px; font-weight: 900; color: #FFFFFF; text-shadow: 2px 2px 3px #000, -2px -2px 3px #000, 2px -2px 3px #000, -2px 2px 3px #000, 0px 0px 5px #000; white-space: nowrap; text-align: center; transform: translate(-50%, -50%);">{finca_nom}</div>"""
                            folium.Marker(location=[centro_lat, centro_lon], icon=folium.DivIcon(html=html_label)).add_to(mapa_magdalena)
                    else:
                        if sector_nom not in sectores_dibujados:
                            folium.CircleMarker(
                                location=f_info["COOR"], radius=15, color=color_nodo, fill=True,
                                fill_color=color_nodo, fill_opacity=0.8,
                                tooltip=f"Sector: {sector_nom}", popup=folium.Popup(f"Sector: {sector_nom} (Suba KML para detalle)", max_width=300)
                            ).add_to(mapa_magdalena)
                            html_label = f"""<div style="font-size: 12px; font-weight: 900; color: #FFFFFF; text-shadow: 2px 2px 3px #000, -2px -2px 3px #000, 2px -2px 3px #000, -2px 2px 3px #000;">{sector_nom}</div>"""
                            folium.Marker(location=[f_info["COOR"][0] + 0.01, f_info["COOR"][1]], icon=folium.DivIcon(html=html_label)).add_to(mapa_magdalena)
                            sectores_dibujados.append(sector_nom)

                for kml_clave, poligonos in dict_poligonos_kml.items():
                    if kml_clave not in kmls_usados:
                        lats_finca, lons_finca, color_gris = [], [], "#A0A0A0" 
                        for poligono in poligonos:
                            folium.Polygon(locations=poligono, color=color_gris, weight=2, fill=True, fill_color=color_gris, fill_opacity=0.4, tooltip=f"{kml_clave} | Inactiva").add_to(mapa_magdalena)
                            lats_finca.extend([p[0] for p in poligono])
                            lons_finca.extend([p[1] for p in poligono])
                        if lats_finca and lons_finca:
                            centro_lat, centro_lon = sum(lats_finca)/len(lats_finca), sum(lons_finca)/len(lons_finca)
                            html_label = f"""<div style="font-size: 10px; font-weight: 700; color: #CCCCCC; text-shadow: 1px 1px 2px #000; text-align: center; transform: translate(-50%, -50%);">{kml_clave} (Inactiva)</div>"""
                            folium.Marker(location=[centro_lat, centro_lon], icon=folium.DivIcon(html=html_label)).add_to(mapa_magdalena)

                st.components.v1.html(mapa_magdalena._repr_html_(), height=650)

                # ==================================================
                # 📋 PANELES SEGMENTADOS (TABS)
                # ==================================================
                st.markdown("<br>### 📋 Segmentación de Inteligencia Operativa", unsafe_allow_html=True)
                
                tab_maestra, tab_general, tab_lavado, tab_ventana, tab_clima = st.tabs([
                    "👑 Vista Maestra",
                    "📋 Estado del Ciclo", 
                    "⛈️ Riesgo Lavado", 
                    "🔭 Ventanas Vuelo", 
                    "🌧️ Historial Lluvias"
                ])
                
                def pintar_estado_ciclo(row):
                    if "CRÍTICO" in row['ESTADO CICLO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    if "MODERADO" in row['ESTADO CICLO']: return ['background-color: #fff3cd; color: #ff9900; font-weight:bold;'] * len(row)
                    return ['color: #155724;'] * len(row)
                    
                def pintar_estado_lavado(row):
                    if "ALTO" in row['RIESGO DE LAVADO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    if "MODERADO" in row['RIESGO DE LAVADO']: return ['background-color: #fff3cd; color: #ff9900; font-weight:bold;'] * len(row)
                    return ['color: #155724;'] * len(row)

                def pintar_estado_ventana(row):
                    if "CERRADA" in row['VENTANA DE VUELO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    if "PRECAUCIÓN" in row['VENTANA DE VUELO']: return ['background-color: #fff3cd; color: #ff9900; font-weight:bold;'] * len(row)
                    return ['color: #155724;'] * len(row)

                def pintar_maestra(row):
                    if "CRÍTICO" in row['ESTADO CICLO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    if "ALTO" in row['RIESGO DE LAVADO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    if "CERRADA" in row['VENTANA DE VUELO']: return ['background-color: #ffe6e6; color: #cc0000; font-weight:bold;'] * len(row)
                    return [''] * len(row)

                with tab_maestra:
                    st.markdown("#### 👑 Matriz Satelital Completa")
                    st.caption("Visión global de todos los indicadores agrometeorológicos. Idéntica a la exportación en Excel.")
                    df_vista_maestra = df_maestro.drop(columns=['COOR', 'COLOR']).copy()
                    df_vista_maestra = df_vista_maestra.sort_values(by=['ESTADO CICLO', 'FINCA'], ascending=[True, True])
                    
                    st.dataframe(
                        df_vista_maestra.style.apply(pintar_maestra, axis=1),
                        use_container_width=True, hide_index=True,
                        column_config={
                            "FINCA": st.column_config.TextColumn("📍 FINCA", width="medium"),
                            "SECTOR": st.column_config.TextColumn("🗺️ SECTOR", width="small"),
                            "ÚLTIMA APLICACIÓN": st.column_config.TextColumn("📅 ÚLTIMA APLICACIÓN", width="medium"),
                            "ÚLTIMO RETORNO (Días)": st.column_config.NumberColumn("⏱️ RETORNO", format="%.0f Días", width="small"),
                            "ESTADO CICLO": st.column_config.TextColumn("🚨 ESTADO CICLO", width="small"),
                            "LLUVIA DÍA APLIC. (mm)": st.column_config.NumberColumn("⛈️ LLU. APLIC.", format="%.1f mm", width="small"),
                            "RIESGO DE LAVADO": st.column_config.TextColumn("⚠️ LAVADO", width="small"),
                            "PRONÓSTICO 7D (mm)": st.column_config.NumberColumn("🔭 PRONÓSTICO", format="%.1f mm", width="small"),
                            "VENTANA DE VUELO": st.column_config.TextColumn("✈️ VENTANA", width="medium"),
                            "LLUVIA 90D (mm)": st.column_config.NumberColumn("🌧️ 90D", format="%.1f mm", width="small"),
                            "LLUVIA 30D (mm)": st.column_config.NumberColumn("🌧️ 30D", format="%.1f mm", width="small"),
                            "PRESIÓN HONGO": st.column_config.TextColumn("🍄 PRESIÓN HONGO", width="small")
                        }
                    )

                with tab_general:
                    df_general = df_maestro[['FINCA', 'SECTOR', 'ÚLTIMO RETORNO (Días)', 'ESTADO CICLO', 'LLUVIA 30D (mm)', 'PRESIÓN HONGO']].copy()
                    df_general = df_general.sort_values(by=['ESTADO CICLO'], ascending=True)
                    st.dataframe(
                        df_general.style.apply(pintar_estado_ciclo, axis=1), 
                        use_container_width=True, hide_index=True,
                        column_config={
                            "FINCA": st.column_config.TextColumn("📍 FINCA", width="medium"),
                            "SECTOR": st.column_config.TextColumn("🗺️ SECTOR", width="small"),
                            "ÚLTIMO RETORNO (Días)": st.column_config.NumberColumn("⏱️ RETORNO (DÍAS)", format="%.0f Días", width="small"),
                            "ESTADO CICLO": st.column_config.TextColumn("🚨 ESTADO", width="small"),
                            "LLUVIA 30D (mm)": st.column_config.NumberColumn("🌧️ LLUVIA 30D", format="%.1f mm", width="small"),
                            "PRESIÓN HONGO": st.column_config.TextColumn("🍄 RIESGO EPIDEMIOLÓGICO", width="medium")
                        }
                    )

                with tab_lavado:
                    df_lavado = df_maestro[['FINCA', 'SECTOR', 'ÚLTIMA APLICACIÓN', 'LLUVIA DÍA APLIC. (mm)', 'RIESGO DE LAVADO']].copy()
                    df_lavado = df_lavado.sort_values(by=['RIESGO DE LAVADO', 'LLUVIA DÍA APLIC. (mm)'], ascending=[True, False])
                    st.dataframe(
                        df_lavado.style.apply(pintar_estado_lavado, axis=1), 
                        use_container_width=True, hide_index=True,
                        column_config={
                            "FINCA": st.column_config.TextColumn("📍 FINCA", width="medium"),
                            "SECTOR": st.column_config.TextColumn("🗺️ SECTOR", width="small"),
                            "ÚLTIMA APLICACIÓN": st.column_config.TextColumn("📅 FECHA VUELO", width="small"),
                            "LLUVIA DÍA APLIC. (mm)": st.column_config.NumberColumn("⛈️ LLUVIA EXACTA", format="%.1f mm", width="small"),
                            "RIESGO DE LAVADO": st.column_config.TextColumn("⚠️ RIESGO DE LAVADO", width="medium")
                        }
                    )

                with tab_ventana:
                    df_ventana = df_maestro[['FINCA', 'SECTOR', 'PRONÓSTICO 7D (mm)', 'VENTANA DE VUELO']].copy()
                    df_ventana = df_ventana.sort_values(by=['VENTANA DE VUELO', 'PRONÓSTICO 7D (mm)'], ascending=[True, False])
                    st.dataframe(
                        df_ventana.style.apply(pintar_estado_ventana, axis=1), 
                        use_container_width=True, hide_index=True,
                        column_config={
                            "FINCA": st.column_config.TextColumn("📍 FINCA", width="medium"),
                            "SECTOR": st.column_config.TextColumn("🗺️ SECTOR", width="small"),
                            "PRONÓSTICO 7D (mm)": st.column_config.NumberColumn("🔭 LLUVIA ESPERADA (7D)", format="%.1f mm", width="small"),
                            "VENTANA DE VUELO": st.column_config.TextColumn("✈️ DECISIÓN LOGÍSTICA", width="medium")
                        }
                    )

                with tab_clima:
                    st.markdown("#### 🌧️ Registro Diario de Lluvias por Sector (Paneles Independientes de Norte a Sur)")
                    
                    if not df_clima_raw.empty:
                        df_chart = df_clima_raw.sort_values("FECHA")
                        
                        fig_lluvia = px.area(
                            df_chart, 
                            x="FECHA", 
                            y="LLUVIA (mm)", 
                            facet_col="SECTOR", 
                            facet_col_wrap=3, 
                            color="SECTOR",
                            title="<b>Comportamiento de Lluvia por Sector (Paneles Independientes)</b>",
                            category_orders={"SECTOR": cols_presentes_ordenadas + cols_extra}
                        )
                        
                        fig_lluvia.for_each_yaxis(lambda yaxis: yaxis.update(matches=None, showticklabels=True))
                        
                        fig_lluvia.update_layout(
                            plot_bgcolor='rgba(0,0,0,0)', 
                            paper_bgcolor='rgba(0,0,0,0)', 
                            height=800, 
                            showlegend=False,
                            margin=dict(t=50, b=20, l=10, r=10)
                        )
                        
                        fig_lluvia.for_each_annotation(lambda a: a.update(text=f"<b>{a.text.split('=')[-1]}</b>"))
                        
                        st.plotly_chart(fig_lluvia, use_container_width=True)
                        
                        cols_cfg_clima = {"FECHA": st.column_config.TextColumn("📅 FECHA")}
                        for col in df_clima_pivot.columns:
                            if col != "FECHA":
                                cols_cfg_clima[col] = st.column_config.NumberColumn(f"🗺️ {col}", format="%.1f mm")
                                
                        st.dataframe(df_clima_pivot, use_container_width=True, hide_index=True, column_config=cols_cfg_clima)
                    else:
                        st.warning("⚠️ No se encontraron datos de lluvia para los sectores operativos.")

                # ==================================================
                # 📥 EXPORTACIÓN MULTI-HOJA VIP
                # ==================================================
                st.markdown("<br>", unsafe_allow_html=True)
                dict_exportacion = {
                    "Matriz_Maestra": df_vista_maestra,
                    "Estado_Ciclos": df_general,
                    "Riesgo_Lavado": df_lavado,
                    "Ventanas_Vuelo_7D": df_ventana,
                    "Historial_Lluvias": df_clima_pivot if not df_clima_pivot.empty else pd.DataFrame()
                }
                
                excel_export = generar_excel_agronomico_vip(dict_exportacion)
                st.download_button(
                    label="💾 DESCARGAR INTELIGENCIA AGRONÓMICA COMPLETA (EXCEL VIP - 5 HOJAS)", 
                    data=excel_export, 
                    file_name=f"Inteligencia_Agronomica_Satelital.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    use_container_width=True
                )

if __name__ == "__main__":
    pass
