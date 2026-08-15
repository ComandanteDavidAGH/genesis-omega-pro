import streamlit as st
import pandas as pd
import gspread
import io
import re
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from oauth2client.service_account import ServiceAccountCredentials

# =================================================================
# ⚡ MOTORES DE CONEXIÓN Y ESTILIZADO DE REPORTES (ALTA VELOCIDAD)
# =================================================================

@st.cache_resource(show_spinner=False)
def inicializar_cliente_gspread():
    """ Centraliza la autenticación con Google Cloud una sola vez en RAM """
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except:
        return None

def estilar_reporte_excel(ws):
    """ Aplica la línea estética corporativa directamente al archivo binario """
    ws.sheet_view.showGridLines = True
    borde_fino = Border(
        left=Side(style='thin', color='D1D1D1'), right=Side(style='thin', color='D1D1D1'),
        top=Side(style='thin', color='D1D1D1'), bottom=Side(style='thin', color='D1D1D1')
    )
    fill_cabecera = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    fuente_cabecera = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    fuente_datos = Font(name="Arial", size=10)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = borde_fino
            if cell.row == 1:
                cell.fill = fill_cabecera
                cell.font = fuente_cabecera
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.font = fuente_datos
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) else '#,##0'
                    
    # Auto-ajuste milimétrico de columnas
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

# --- 🧪 TRADUCTOR SEGURO DE NÚMEROS ---
def a_numero_limpio(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip().replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except: return 0.0

# --- 🖨️ MOTOR EXTRACTOR ADAPTABLE (HISTÓRICO O SEMANAL) ---
def generar_reporte_filtrado(filtrar_semana=False, pestaña_nombre="TABLA 1"):
    url_maestra = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"
    try:
        gc = inicializar_cliente_gspread()
        if gc is None:
            return pd.DataFrame()
            
        sh = gc.open_by_url(url_maestra)
        worksheet = sh.worksheet(pestaña_nombre)
        
        data = worksheet.get_all_values()
        if not data or len(data) < 6: return pd.DataFrame()
        
        # Lectura alineada a la Fila 5 de su matriz
        encabezados = [str(c).upper().strip() for c in data[4]]
        filas_datos = data[5:]
        
        df = pd.DataFrame(filas_datos, columns=encabezados)
        
        # 🎯 FILTRO FRANCOTIRADOR: Columnas autorizadas para la empresa (Protección de costos y fórmulas)
        columnas_validas = [col for col in df.columns if any(c in col for c in ['ORDEN', 'BLOQUE', 'FINCA', 'SECTOR', 'BRUTA', 'FUMIG', 'COCTEL', 'FECHA', 'SEM', 'PILOTO', 'MODELO', 'PISTA'])]
        df_filtrado = df[columnas_validas].copy()
        df_filtrado.columns = [c.replace('\n', ' ').strip() for c in df_filtrado.columns]
        
        if filtrar_semana:
            df_filtrado['FECHA_DT'] = pd.to_datetime(df_filtrado['FECHA'], dayfirst=True, errors='coerce')
            df_filtrado = df_filtrado.dropna(subset=['FECHA_DT'])
            
            # Hora Colombia UTC-5
            hora_col = datetime.utcnow() + timedelta(hours=-5)
            fecha_limite = hora_col - timedelta(days=7)
            df_filtrado = df_filtrado[df_filtrado['FECHA_DT'] >= fecha_limite].copy()
            
            if df_filtrado.empty: return pd.DataFrame()
            
            df_filtrado['FECHA'] = df_filtrado['FECHA_DT'].dt.strftime('%d/%m/%Y')
            df_filtrado = df_filtrado.drop(columns=['FECHA_DT'], errors='ignore')
            
        return df_filtrado
    except Exception as e:
        st.error(f"🚨 Error en el procesamiento de datos: {str(e)}")
        return pd.DataFrame()

# --- 📡 INTERFAZ LINEAL CORPORATIVA ---
def ejecutar(*args, **kwargs):
    hora_colombia = datetime.utcnow() + timedelta(hours=-5)
    
    st.markdown("""
    <style>
    .titulo-principal-gov { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; text-align: center; }
    .sub-gov { text-align: center; font-style: italic; color: #64748b; margin-top: -10px; margin-bottom: 25px; }
    .card-descarga { background-color: #f8f9fa; border: 1px solid #dee2e6; padding: 15px; border-radius: 8px; border-top: 4px solid #0d1b2a; }
    div[data-testid="stTabs"] button[role="tab"] { font-family: 'Arial Black', sans-serif; font-size: 14px; color: #0d1b2a; }
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { border-bottom-color: #d4af37; background-color: rgba(212, 175, 55, 0.1); }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 class='titulo-principal-gov'>📜 Módulo 11: Manual de Gobierno Técnico</h1>", unsafe_allow_html=True)
    st.markdown("<p class='sub-gov'>Bóveda de Criterios Científicos, Extracción Segura y Código Fuente</p>", unsafe_allow_html=True)
    
    # Inicialización segura
    if "buffer_historico" not in st.session_state: st.session_state.buffer_historico = None
    if "rows_historico" not in st.session_state: st.session_state.rows_historico = 0
    if "buffer_semanal" not in st.session_state: st.session_state.buffer_semanal = None
    if "rows_semanal" not in st.session_state: st.session_state.rows_semanal = 0
    
    st.markdown("---")
    st.markdown("### 📤 Extractor de Datos Seguro para la Empresa")
    st.write("Archivos **100% limpios de costos financieros y fórmulas**, ideales para reportes a la gerencia externa.")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("<div class='card-descarga'>", unsafe_allow_html=True)
        st.markdown("#### 📂 Operación Inicial (Histórico)")
        if st.button("🚀 COMPILAR HISTÓRICO COMPLETO", key="btn_historico", use_container_width=True):
            with st.spinner("Descargando matriz completa y purificando columnas..."):
                df_hist = generar_reporte_filtrado(filtrar_semana=False)
                if not df_hist.empty:
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df_hist.to_excel(writer, index=False, sheet_name='Histórico Operaciones')
                        estilar_reporte_excel(writer.sheets['Histórico Operaciones'])
                    st.session_state.buffer_historico = buffer.getvalue()
                    st.session_state.rows_historico = len(df_hist)
                else:
                    st.session_state.buffer_historico = None
                    st.warning("⚠️ No se encontraron datos en la TABLA 1.")
        
        if st.session_state.buffer_historico is not None:
            st.success(f"✅ Compilados {st.session_state.rows_historico} registros históricos.")
            st.download_button("📥 DESCARGAR EXCEL MAESTRO PLANO", data=st.session_state.buffer_historico, file_name=f"Reporte_Historico_Operaciones_{hora_colombia.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
                    
    with col2:
        st.markdown("<div class='card-descarga'>", unsafe_allow_html=True)
        st.markdown("#### 📅 Operación Rutinaria (Últimos 7 Días)")
        if st.button("⚡ COMPILAR INFORMACIÓN SEMANAL", key="btn_semanal", type="primary", use_container_width=True):
            with st.spinner("Aislando misiones de los últimos 7 días..."):
                df_sem = generar_reporte_filtrado(filtrar_semana=True)
                if not df_sem.empty:
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        df_sem.to_excel(writer, index=False, sheet_name='Reporte Semanal')
                        estilar_reporte_excel(writer.sheets['Reporte Semanal'])
                    st.session_state.buffer_semanal = buffer.getvalue()
                    st.session_state.rows_semanal = len(df_sem)
                else:
                    st.session_state.buffer_semanal = None
                    st.warning("⚠️ No se detectaron misiones en los últimos 7 días.")
        
        if st.session_state.buffer_semanal is not None:
            st.success(f"✅ Purgadas {st.session_state.rows_semanal} misiones de esta semana.")
            st.download_button("📥 DESCARGAR EXCEL SEMANAL", data=st.session_state.buffer_semanal, file_name=f"Reporte_Semanal_Operaciones_{hora_colombia.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

    # =================================================================
    # 📚 LA BIBLIOTECA DE ALEJANDRÍA: ADN DEL SISTEMA
    # =================================================================
    st.markdown("<br><hr>", unsafe_allow_html=True)
    st.markdown("### 📚 Biblioteca de Alejandría: ADN del Sistema")
    st.caption("Documentación técnica y mapa de arquitectura para restauraciones de emergencia.")

    t_m3, t_m8, t_m19, t_panico = st.tabs(["⚙️ Mód. 3: Facturación", "📈 Mód. 8 y 10: BI Costos", "📦 Mód. 19: Logística", "💾 EL SANTO GRIAL"])

    with t_m3:
        st.markdown("#### 🧠 Lógica de Validación y Facturación")
        st.info("Este módulo es el corazón financiero. Cruza el inventario de SAP con las recetas agronómicas para determinar el costo exacto por hectárea.")
        st.markdown("""
        **1. El Reloj Satelital (Fallo de Zona Horaria)**
        * **Problema:** Los servidores de Streamlit operan en formato Universal (UTC). A las 7 PM de Colombia, el servidor cree que ya es el día siguiente, fechando mal las OS.
        * **Solución (Implementada):** Todo el módulo llama a `obtener_hora_colombia()` que fuerza matemáticamente un `timedelta(hours=-5)`. NUNCA usar `datetime.now()` estándar.

        **2. La Máquina del Tiempo (Tarifas Dinámicas)**
        * **Mapeo:** Lee exclusivamente la pestaña `MATRIZ_TARIFAS` de Google Sheets.
        * **Mecanismo:** El usuario selecciona la **Fecha de Vuelo**. El sistema extrae el *Año* y busca esa columna exacta en la matriz de Google Sheets. 
        * **Protocolo de Emergencia:** Si la columna del año no existe o borran la hoja, el sistema entra en modo de supervivencia usando los precios fijos del 2026 pre-programados.

        **3. Emparejamiento de Cócteles IA**
        * Utiliza un sistema de puntos. Recibe las siglas de SAP y las busca en la `DD_Mesclas`. Si encuentra coincidencias exactas suma 100 puntos, si es aproximada 40 puntos. El cóctel ganador es el que más se asemeje a la receta.
        """)

    with t_m8:
        st.markdown("#### 📊 Inteligencia de Costos e Históricos")
        st.success("Módulo de macro-análisis diseñado para proyectar la inflación tarifaria y auditar el costo promedio por hectárea consolidado.")
        st.markdown("""
        **1. Escudo Anti-Duplicidad**
        * Al unir la "TABLA 1" (Datos Vivos) y el "Histórico", el sistema elimina operaciones repetidas utilizando esta llave primaria de 5 variables:
          `[FECHA_DT, FINCA_MAESTRA, OS_MAESTRA, AREA_NUM, COCTEL_CLEAN]`

        **2. Aislamiento Finca por Finca (Regla de Oro de Intervalos)**
        * Para evitar el "Efecto Bolsa de Fechas" (que arruina el promedio de días entre vuelos si se mezclan pistas), el sistema calcula el ciclo **aislando los vuelos de una misma finca**.
        * El umbral de ruptura para considerar que es un "ciclo nuevo" es una inactividad mayor a **5 días**.

        **3. Filtro Master Data (Curva de Inflación)**
        * La gráfica lee en tiempo real el Excel de tarifas. Las columnas opcionales como "DIF" o "INCRE %" son ignoradas automáticamente por el comando `str(c).isdigit()`, filtrando solo los años.
        """)

    with t_m19:
        st.markdown("#### 🚚 Centro Logístico Unificado")
        st.warning("Manejo de inventario crudo y traslados. Operación directa de escritura sobre bases de datos secundarias.")
        st.markdown("""
        **1. Inyección Megazord (Entradas Múltiples)**
        * Permite agregar dinámicamente infinitas filas a una factura de ingreso antes de golpear la API de Google, evitando el bloqueo temporal de Google Sheets (Error 429 - Too Many Requests).

        **2. Excepción Táctica de Traslados (Misma Pista)**
        * Por regla general, no se permite trasladar un producto de Pista A hacia Pista A (bloqueo de errores de usuario).
        * Sin embargo, el código contiene un bypass si el usuario selecciona en la observación: `TRANSFORMACIÓN DE LOTE` u `OTRO`. Esto permite operaciones de re-empacado logístico interno en una misma base operativa.
        """)

    with t_panico:
        st.markdown("#### 💾 EL BOTÓN DE PÁNICO")
        st.write("Si el sistema se corrompe gravemente en el futuro y no hay conexión a internet, este archivo contiene el manual arquitectónico completo para reconstruir la lógica de la empresa.")
        
        texto_grial = f"""=====================================================
INFORME DE ARQUITECTURA TÉCNICA - GÉNESIS OMEGA PRO
=====================================================
Compilado: {hora_colombia.strftime('%Y-%m-%d %H:%M:%S')} (Hora Colombia)
Autor: IA Arquitecto + Comandante Omega

1. HUSOS HORARIOS (EL FANTASMA DE LAS 7PM)
Los servidores corren en UTC. Toda extracción de fecha debe pasar por:
datetime.utcnow() + timedelta(hours=-5)
¡Jamás usar datetime.now() puro!

2. MAQUINA DEL TIEMPO TARIFARIA (MÓDULO 3)
Ruta en Sheets: Pestaña 'MATRIZ_TARIFAS'.
Estructura Obligatoria: 
Columna A: PISTA
Columna B: EQUIPO_O_TOPE
Columnas X: Años en formato numérico (Ej: 2026).
El sistema cruzará la Fecha de Vuelo con la columna del Año.

3. MOTOR IA DE CÓCTELES
Se basa en sistema de puntuación. 
Lee la pestaña 'DD_Mesclas'.
El Acondicionador se ajusta dinámicamente a 0.06 si en la receta existe ZN, BT, ZT o ZITRON. De lo contrario se usa 0.02.
IMBIOSIL aplica 1.5 de dosis si la mezcla arranca por "IN".

4. ESCUDO ANTI-DUPLICIDAD DE HECTÁREAS (MÓDULO 8 y 10)
Se eliminan duplicados comparando simultáneamente: 
FECHA, FINCA_MAESTRA, OS_MAESTRA, AREA_NUM, COCTEL_CLEAN.

5. BYPASS DE LOGÍSTICA (MÓDULO 19)
Para enviar un traslado hacia la misma pista, la observación debe ser 
obligatoriamente "TRANSFORMACIÓN DE LOTE" u "OTRO".

> Si algo falla, revise siempre que los nombres de las pestañas en el 
> Google Sheets maestro no tengan espacios al final (Ej: 'TABLA 2 ' dará error).
=====================================================
"""
        st.download_button(
            label="🚨 DESCARGAR EL SANTO GRIAL (TXT MANUAL COMPLETO)",
            data=texto_grial,
            file_name=f"Copia_Seguridad_Arquitectura_{hora_colombia.strftime('%Y%m%d')}.txt",
            mime="text/plain",
            type="primary",
            use_container_width=True
        )

if __name__ == "__main__":
    pass
