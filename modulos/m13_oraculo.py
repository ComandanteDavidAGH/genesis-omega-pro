import streamlit as st
import pandas as pd
import gspread
from datetime import datetime, timedelta
import re
import io
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# ⚙️ CONSTANTES CENTRALIZADAS (ÚNICA FUENTE DE VERDAD)
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

# =================================================================
# ⚡ MOTOR DE CONEXIÓN UNIFICADO (V41)
# =================================================================
@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

# =================================================================
# 🛡️ UTILIDADES DE PURIFICACIÓN Y FORMATO
# =================================================================
def a_numero_limpio(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip().replace(',', '.')
        v = re.sub(r'[^\d\.,\-]', '', v)
        if not v: return 0.0
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except: 
        return 0.0

def procesar_fecha_pesada(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    s = str(val).strip()
    if s.replace('.', '', 1).isdigit(): 
        return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try: 
            return pd.to_datetime(s, format=fmt)
        except: 
            pass
    try: 
        return pd.to_datetime(s, errors='coerce')
    except: 
        return pd.NaT

def fmt_latino(val, decimales=1):
    try: 
        return f"{float(val):,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except: 
        return str(val)

def obtener_dosis_fertilizante(df_mezclas, fert_name):
    try:
        for col_idx in range(len(df_mezclas.columns) - 1):
            mask = df_mezclas.iloc[:, col_idx].astype(str).str.strip().str.upper() == fert_name
            if mask.any():
                val = pd.to_numeric(df_mezclas[mask].iloc[0, col_idx+1], errors='coerce')
                if pd.notna(val) and val > 0: return float(val)
    except: 
        pass
    return None 

def calcular_intervalo_pista(df_pista):
    """ Calcula cuántos días tarda un ciclo de fumigación en una pista """
    if df_pista.empty: return 14.0
    fechas = sorted(df_pista['FECHA_DT'].dt.date.unique())
    if len(fechas) < 2: return 14.0
    ciclos = 1
    inicios_ciclo = [fechas[0]]
    for i in range(1, len(fechas)):
        if (fechas[i] - fechas[i-1]).days > 5:
            ciclos += 1
            inicios_ciclo.append(fechas[i])
    if ciclos > 1:
        avg_int = sum([(inicios_ciclo[j] - inicios_ciclo[j-1]).days for j in range(1, ciclos)]) / (ciclos - 1)
        return float(avg_int)
    return 14.0

# =================================================================
# 🧠 CEREBRO QUÍMICO CALIBRADO
# =================================================================
def extraer_receta_completa(coctel_sel, df_mezclas, dict_fertilizantes_dinamico):
    coctel_u = str(coctel_sel).upper().strip().replace("+", " ").replace("-", " ")
    partes = coctel_u.split()
    base_coctel = partes[0] if len(partes) > 0 else ""
    aditivos = partes[1:] if len(partes) > 1 else []
    
    dict_prods = {}
    
    if not df_mezclas.empty:
        col_0_limpia = df_mezclas.iloc[:, 0].astype(str).str.upper().str.strip()
        rb = df_mezclas[col_0_limpia == base_coctel]
        for _, r in rb.iterrows():
            p = str(r.iloc[1]).strip().upper()
            d = a_numero_limpio(r.iloc[2])
            if d > 0 and p not in ['NAN', 'NONE', '']: dict_prods[p] = d

    for aditivo in aditivos:
        if aditivo in dict_fertilizantes_dinamico:
            nombre_fert = dict_fertilizantes_dinamico[aditivo]
            dosis_fert = obtener_dosis_fertilizante(df_mezclas, nombre_fert)
            
            if dosis_fert is not None:
                dict_prods[nombre_fert] = dict_prods.get(nombre_fert, 0.0) + dosis_fert
            elif aditivo == "NM": dict_prods["NATURAMIN WSP"] = 0.2
            elif aditivo == "ZN": dict_prods["ZINTRAC X LITRO SV"] = 0.5
            elif aditivo == "BT": dict_prods["BANATREL SC"] = 0.5
    
    if "SV" in coctel_u or "ACONDICIONADOR" in coctel_u:
        dict_prods["ACONDICIONADOR SV"] = 0.06 if any(x in coctel_u for x in ["ZN", "BT", "ZT", "ZITRON"]) else 0.02
        dict_prods["ADHERENTE SV"] = 0.13
        
    if base_coctel.startswith("IN") or "IMBIOSIL" in base_coctel: 
        dict_prods["IMBIOSIL O"] = 1.5

    return dict_prods

# =================================================================
# 💾 MOTOR DE EXTRACCIÓN CACHEADA
# =================================================================
@st.cache_data(show_spinner=False, ttl=600)
def extraer_datos_boveda_oraculo():
    gc = obtener_cliente_gspread_unificado()
    df_t1, df_mezclas = pd.DataFrame(), pd.DataFrame()
    if not gc: return df_t1, df_mezclas
    
    try:
        boveda = gc.open_by_url(URL_BOVEDA_MAESTRA)
        try:
            t1 = boveda.worksheet("TABLA 1").get_all_values()
            idx_t1 = 4
            for i in range(min(8, len(t1))):
                fila_limpia = [str(x).upper().strip() for x in t1[i]]
                if "Nº ORDEN" in fila_limpia or "FINCA" in fila_limpia or "VALOR A FACTURAR" in "".join(fila_limpia):
                    idx_t1 = i
                    break
            df_t1 = pd.DataFrame(t1[idx_t1+1:], columns=[str(c).upper().strip() for c in t1[idx_t1]]) if len(t1) > idx_t1 else pd.DataFrame()
        except Exception: pass
        
        try:
            data_mez = boveda.worksheet("DD_Mesclas").get_all_values()
            df_mezclas = pd.DataFrame(data_mez[1:], columns=[str(c).upper().strip() for c in data_mez[0]]) if len(data_mez) > 1 else pd.DataFrame()
        except Exception: pass
    except Exception: pass
    
    return df_t1, df_mezclas

# =================================================================
# 📤 EXPORTADOR EXCEL VIP
# =================================================================
def generar_excel_vip(df, sheet_name="Oraculo_Proyeccion"):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        
        header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        header_font = Font(color="D4AF37", bold=True)
        borde_fino = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
                            
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde_fino
            
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
                
                if cell.row > 1:
                    cell.border = borde_fino
                    if isinstance(cell.value, (int, float)):
                        col_name = str(ws[col_letter + '1'].value).upper()
                        if "SALDO" in col_name or "PROYECCIÓN" in col_name or "AUTONOMÍA" in col_name or "CICLO" in col_name:
                            cell.number_format = '#,##0.00'
                            
            ws.column_dimensions[col_letter].width = min(max_length + 4, 35)
            
    return buffer.getvalue()

# =================================================================
# 🚀 EJECUCIÓN PRINCIPAL
# =================================================================
def ejecutar(purificar_lote, extraer_numero):
    VERDE_INTENSO = '#143521'
    
    st.markdown(f"""
    <style>
    .titulo-oraculo {{ color: #0d1b2a; border-bottom: 3px solid #27AE60; padding-bottom: 5px; font-family: 'Arial Black'; }}
    
    [data-testid="column"] {{
        display: flex !important;
        flex-direction: column !important;
        justify-content: flex-start !important;
        align-items: stretch !important;
    }}

    div[data-testid="stDataFrame"] {{ border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.08) !important; }}
    
    div[data-testid="stSelectbox"] > div,
    div[data-testid="stSelectbox"] div[data-baseweb="select"],
    div[data-testid="stFileUploader"] > div,
    div[data-testid="stTextInput"] > div {{
        background-color: #ffffff !important;
        border: 2px solid {VERDE_INTENSO} !important;
        border-radius: 8px !important;
        box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important;
    }}
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stTextInput"] div[data-baseweb="input"] {{
        background-color: transparent !important;
        border: none !important;
    }}
    div[data-testid="stSelectbox"] *, div[data-testid="stFileUploader"] *, div[data-testid="stTextInput"] input {{
        color: #000000 !important;
        font-weight: bold !important;
    }}
    
    div[data-testid="stFileUploader"] button, div[data-testid="stFileUploader"] button * {{
        color: #ffffff !important;
    }}
    
    div[data-testid="stMainBlockContainer"] label p {{
        color: #0d1b2a !important;
        font-weight: 800 !important;
        text-transform: uppercase !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    c_tit, c_sync = st.columns([3.5, 1.5])
    with c_tit:
        st.markdown("<h1 class='titulo-oraculo'>🔮 El Oráculo: Predicción Cíclica de Rupturas</h1>", unsafe_allow_html=True)
        st.write("Análisis estacional del comportamiento epidemiológico cruzado con inventarios de SAP.")
    with c_sync:
        st.write("")
        if st.button("🔄 Sincronizar Nube (Forzar Datos)", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()

    # 💾 MEMORIA RAM DE SESIÓN PARA LA HOMOLOGACIÓN
    if 'traductor_sap' not in st.session_state:
        st.session_state['traductor_sap'] = {
            "PLUC": "FUMIGARAY",
            "PORI": "AEROPENOR",
            "LUCI": "GENESYS",
            "TEHO": "AVIL",
            "PDIV": "ASA"
        }

    # =================================================================
    # 🎯 AISLAMIENTO TÁCTICO: MODO CONFIGURACIÓN SAP
    # =================================================================
    modo_config = st.toggle("⚙️ MODO CONFIGURACIÓN: Homologación Avanzada de Almacenes SAP", value=False)
    
    if modo_config:
        with st.container(border=True):
            st.markdown("### ⚙️ Diccionario de Traducción (SAP ↔ Operación)")
            st.info("💡 **Visión Aislada:** SAP utiliza códigos técnicos que difieren de los nombres operativos. Empareja los códigos aquí; el sistema guardará tu configuración en la memoria RAM.")
            
            c_h1, c_h2, c_h3, c_h4, c_h5 = st.columns(5)
            st.session_state['traductor_sap']["PLUC"] = c_h1.text_input("PLUC (Dron)", value=st.session_state['traductor_sap']["PLUC"]).upper()
            st.session_state['traductor_sap']["PORI"] = c_h2.text_input("PORI (Dron)", value=st.session_state['traductor_sap']["PORI"]).upper()
            st.session_state['traductor_sap']["LUCI"] = c_h3.text_input("LUCI (Dron)", value=st.session_state['traductor_sap']["LUCI"]).upper()
            st.session_state['traductor_sap']["TEHO"] = c_h4.text_input("TEHO (Dron)", value=st.session_state['traductor_sap']["TEHO"]).upper()
            st.session_state['traductor_sap']["PDIV"] = c_h5.text_input("PDIV (Dron)", value=st.session_state['traductor_sap']["PDIV"]).upper()
            
            st.success("✅ Las homologaciones están sincronizadas. **Apaga este interruptor para volver al Radar del Oráculo.**")
        st.stop() # 🛑 PARADA EN SECO: Oculta el resto de la interfaz

    traductor_pistas = st.session_state['traductor_sap']

    # =================================================================
    # 🚀 INTERFAZ PRINCIPAL DEL ORÁCULO
    # =================================================================
    with st.container(border=True):
        st.markdown("### 📥 1. Radar de Existencias Actuales (SAP)")
        archivo_sap = st.file_uploader("Cargue la Sábana SAP actualizada (.xlsx o .csv)", type=['xlsx', 'csv'], key="sap_oraculo")
        
        st.markdown("### 📅 2. Parámetros de Predicción")
        col_mes, col_pista, col_profundidad, col_modo = st.columns([1, 1, 1.2, 1.2])
        
        meses_dict = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
        mes_actual = datetime.now().month
        mes_proyeccion = col_mes.selectbox("Mes a Proyectar:", list(meses_dict.keys()), index=mes_actual-1, format_func=lambda x: meses_dict[x])
        
        lista_pistas = ["TODAS", "PLUC", "PORI", "PDIV", "TEHO", "LUCI"]
        pista_objetivo = col_pista.selectbox("📍 Base Operativa (SAP):", lista_pistas)

        opciones_profundidad = ["Último Año (Tendencia Reciente)", "Últimos 2 Años", "Últimos 3 Años", "Histórico Completo"]
        profundidad_sel = col_profundidad.selectbox("🔍 Profundidad Histórica:", opciones_profundidad)
        
        modo_analisis = col_modo.selectbox("⚙️ Tipo de Análisis:", ["Estándar (Autonomía Días)", "Táctico (Autonomía Ciclos)"])

        st.markdown("<br>", unsafe_allow_html=True)

        if not archivo_sap:
            st.info("💡 Despliegue el archivo SAP para que el sistema evalúe el blindaje de las pistas.")
            return

        if st.button("🚀 EJECUTAR PREDICCIÓN DE RUPTURAS", type="primary", use_container_width=True):
            with st.spinner(f"Sincronizando lenguajes y analizando comportamiento agronómico..."):
                try:
                    # --- LECTURA DE SAP ---
                    if archivo_sap.name.lower().endswith('.xlsx') or archivo_sap.name.lower().endswith('.xls'):
                        df_sap = pd.read_excel(archivo_sap)
                    else:
                        try: df_sap = pd.read_csv(archivo_sap, sep=None, engine='python', encoding='utf-8')
                        except:
                            archivo_sap.seek(0)
                            df_sap = pd.read_csv(archivo_sap, sep=None, engine='python', encoding='latin1')

                    def purificar_columna(col_name):
                        return str(col_name).upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').strip()
                    
                    cols_limpias = [purificar_columna(c) for c in df_sap.columns]
                    
                    idx_cod = next((i for i, c in enumerate(cols_limpias) if 'MATERIAL' in c or 'COD' in c or 'ITEM' in c), None)
                    idx_prod = next((i for i, c in enumerate(cols_limpias) if ('TEXTO' in c or 'DESC' in c or 'PRODUCTO' in c or 'DENOMINACION' in c) and i != idx_cod), None)
                    idx_pista = next((i for i, c in enumerate(cols_limpias) if 'ALMACEN' in c or 'PISTA' in c or 'LGORT' in c), None)
                    idx_saldo = next((i for i, c in enumerate(cols_limpias) if 'LIBRE' in c or 'SALDO' in c or 'UTILIZACION' in c or 'LABST' in c), None)

                    if idx_prod is None or idx_pista is None or idx_saldo is None:
                        st.error(f"❌ Error de Radar: No se pudieron mapear las columnas críticas en SAP.")
                        return
                    
                    c_prod = df_sap.columns[idx_prod]
                    c_pista = df_sap.columns[idx_pista]
                    c_saldo = df_sap.columns[idx_saldo]
                    c_cod = df_sap.columns[idx_cod] if idx_cod is not None else None

                    if c_cod is not None and c_prod is not None:
                        df_sap['PRODUCTO_RADAR'] = df_sap[c_cod].astype(str).str.split('.').str[0].str.strip() + " | " + df_sap[c_prod].astype(str).str.upper().str.strip()
                    else:
                        df_sap['PRODUCTO_RADAR'] = df_sap[c_prod].astype(str).str.upper().str.strip()

                    df_sap['SALDO_FISICO'] = df_sap[c_saldo].apply(a_numero_limpio)
                    df_sap['PISTA_SAP'] = df_sap[c_pista].astype(str).str.upper().str.strip()
                    
                    df_sap_agrupado = df_sap.groupby(['PISTA_SAP', 'PRODUCTO_RADAR'])['SALDO_FISICO'].sum().reset_index()
                    df_sap_agrupado = df_sap_agrupado[df_sap_agrupado['SALDO_FISICO'] > 0]

                    if pista_objetivo != "TODAS":
                        df_sap_agrupado = df_sap_agrupado[df_sap_agrupado['PISTA_SAP'].str.contains(pista_objetivo, na=False)]

                    # --- LECTURA DE BÓVEDA ---
                    df_t1, df_mezclas = extraer_datos_boveda_oraculo()

                    if df_t1.empty or df_mezclas.empty:
                        st.error("🚨 ENLACE SATELITAL ROTO: No se pudo leer TABLA 1 o DD_Mesclas desde Google Drive.")
                        return

                    dict_fert = {}
                    if len(df_mezclas.columns) > 13:
                        for _, row in df_mezclas.iterrows():
                            f_n = str(row.iloc[12]).strip().upper() 
                            f_s = str(row.iloc[13]).strip().upper() 
                            if f_s and f_n not in ["", "NAN", "NONE", "FERTILIZANTES", "SIGLAS"]:
                                dict_fert[f_s] = f_n

                    col_fecha = next((c for c in df_t1.columns if 'FECHA' in c), 'FECHA')
                    col_ha = next((c for c in df_t1.columns if 'NETA' in c or 'FUMIG' in c or 'HECT' in c), None)
                    col_coctel = next((c for c in df_t1.columns if 'COCTEL' in c or 'CÓCTEL' in c or 'MEZCLA' in c), None)
                    col_pista_t1 = next((c for c in df_t1.columns if 'PISTA' in c or 'BASE' in c), None)

                    df_t1['FECHA_DT'] = df_t1[col_fecha].apply(procesar_fecha_pesada)
                    df_t1 = df_t1.dropna(subset=['FECHA_DT'])
                    df_t1['MES'] = df_t1['FECHA_DT'].dt.month
                    df_t1['AÑO'] = df_t1['FECHA_DT'].dt.year
                    df_t1['HA_CALCULO'] = df_t1[col_ha].apply(a_numero_limpio)
                    df_t1['PISTA_OPERATIVA'] = df_t1[col_pista_t1].astype(str).str.upper().str.strip()

                    año_actual_operacion = datetime.now().year
                    if profundidad_sel == "Último Año (Tendencia Reciente)":
                        df_t1 = df_t1[df_t1['AÑO'] >= (año_actual_operacion - 1)]
                    elif profundidad_sel == "Últimos 2 Años":
                        df_t1 = df_t1[df_t1['AÑO'] >= (año_actual_operacion - 2)]
                    elif profundidad_sel == "Últimos 3 Años":
                        df_t1 = df_t1[df_t1['AÑO'] >= (año_actual_operacion - 3)]
                    
                    max_date = df_t1['FECHA_DT'].max()
                    fecha_limite = max_date - timedelta(days=90)
                    df_reciente = df_t1[df_t1['FECHA_DT'] >= fecha_limite]
                    
                    ha_mensual_actual_pista = {}
                    if not df_reciente.empty:
                        ha_mensual_actual_pista = (df_reciente.groupby('PISTA_OPERATIVA')['HA_CALCULO'].sum() / 3.0).to_dict()

                    df_hist_mes = df_t1[df_t1['MES'] == mes_proyeccion].copy()
                    consumo_esperado_pista = {} 
                    ha_total_detectada = 0.0

                    intervalos_pista = {}
                    for pista_u in df_hist_mes['PISTA_OPERATIVA'].unique():
                        intervalos_pista[pista_u] = calcular_intervalo_pista(df_hist_mes[df_hist_mes['PISTA_OPERATIVA'] == pista_u])

                    if not df_hist_mes.empty:
                        ha_total_detectada = df_hist_mes['HA_CALCULO'].sum()
                        ha_hist_total_pista = df_hist_mes.groupby('PISTA_OPERATIVA')['HA_CALCULO'].sum().to_dict()

                        volumen_hist_total = {}
                        for _, row_c in df_hist_mes.iterrows():
                            pista_op = row_c['PISTA_OPERATIVA']
                            coctel_completo = str(row_c[col_coctel]).upper().strip()
                            ha_aplicadas = row_c['HA_CALCULO']
                            
                            if pista_op not in volumen_hist_total:
                                volumen_hist_total[pista_op] = {}

                            receta_dict = extraer_receta_completa(coctel_completo, df_mezclas, dict_fert)
                            for prod_quimico, dosis in receta_dict.items():
                                volumen_hist_total[pista_op][prod_quimico] = volumen_hist_total[pista_op].get(prod_quimico, 0) + (dosis * ha_aplicadas)

                        for pista_op, prods in volumen_hist_total.items():
                            ha_historicas_mes = ha_hist_total_pista.get(pista_op, 0)
                            ha_actuales_mes = ha_mensual_actual_pista.get(pista_op, ha_historicas_mes / df_hist_mes['AÑO'].nunique() if df_hist_mes['AÑO'].nunique() > 0 else 1)
                            
                            if pista_op not in consumo_esperado_pista:
                                consumo_esperado_pista[pista_op] = {}
                                
                            for prod, vol_hist in prods.items():
                                if ha_historicas_mes > 0:
                                    dosis_promedio_blended = vol_hist / ha_historicas_mes
                                    consumo_esperado_pista[pista_op][prod] = dosis_promedio_blended * ha_actuales_mes
                                else:
                                    consumo_esperado_pista[pista_op][prod] = 0.0

                    resultados = []
                    
                    for _, row_s in df_sap_agrupado.iterrows():
                        pista_sap = row_s['PISTA_SAP']
                        producto_sap_completo = str(row_s['PRODUCTO_RADAR']).upper().strip()
                        saldo = row_s['SALDO_FISICO']

                        consumo_mes_proyectado = 0.0
                        pista_t1_esperada = traductor_pistas.get(pista_sap, pista_sap)
                        pista_clave = next((k for k in consumo_esperado_pista.keys() if pista_t1_esperada in k or k in pista_t1_esperada), None)
                        
                        intervalo_real_dias = intervalos_pista.get(pista_clave, 14.0) if pista_clave else 14.0
                        
                        if pista_clave:
                            for p_receta, vol_mes in consumo_esperado_pista[pista_clave].items():
                                p_receta_clean = p_receta.replace(" ", "")
                                prod_sap_clean = producto_sap_completo.replace(" ", "")
                                if p_receta_clean in prod_sap_clean or prod_sap_clean in p_receta_clean:
                                    consumo_mes_proyectado += vol_mes

                        if modo_analisis == "Estándar (Autonomía Días)":
                            consumo_diario = consumo_mes_proyectado / 30 if consumo_mes_proyectado > 0 else 0
                            if consumo_diario > 0:
                                dias_autonomia = saldo / consumo_diario
                                if dias_autonomia <= 7: estado = "🚨 CRÍTICO (< 7 Días)"
                                elif dias_autonomia <= 21: estado = "⚠️ ALERTA (8-21 Días)"
                                else: estado = "✅ ÓPTIMO (> 21 Días)"
                            else:
                                dias_autonomia = 9999
                                estado = "✅ ÓPTIMO (Sin Consumo Histórico)"

                            resultados.append({
                                "📍 PISTA": pista_sap,
                                "🧪 CÓDIGO | PRODUCTO": producto_sap_completo,
                                "📦 SALDO (SAP)": saldo,
                                "📈 PROYECCIÓN MES (L/Kg)": round(consumo_mes_proyectado, 1),
                                "⏳ AUTONOMÍA (DÍAS)": round(dias_autonomia, 0),
                                "ESTADO": estado
                            })
                            
                        else: # Táctico (Ciclos)
                            ciclos_al_mes = 30.0 / intervalo_real_dias if intervalo_real_dias > 0 else 2.0
                            consumo_por_ciclo = consumo_mes_proyectado / ciclos_al_mes if ciclos_al_mes > 0 else 0
                            
                            if consumo_por_ciclo > 0:
                                ciclos_autonomia = saldo / consumo_por_ciclo
                                if ciclos_autonomia < 1.0: estado = "🚨 CRÍTICO (< 1 Ciclo)"
                                elif ciclos_autonomia <= 2.0: estado = "⚠️ ALERTA (1 a 2 Ciclos)"
                                else: estado = "✅ ÓPTIMO (> 2 Ciclos)"
                            else:
                                ciclos_autonomia = 99.9
                                estado = "✅ ÓPTIMO (Sin Consumo Histórico)"
                                
                            resultados.append({
                                "📍 PISTA": pista_sap,
                                "🧪 CÓDIGO | PRODUCTO": producto_sap_completo,
                                "📦 SALDO (SAP)": saldo,
                                "⏱️ DÍAS POR CICLO": round(intervalo_real_dias, 1),
                                "🔄 CONSUMO POR CICLO": round(consumo_por_ciclo, 1),
                                "🔋 CICLOS RESTANTES": round(ciclos_autonomia, 2),
                                "ESTADO": estado
                            })

                    df_oraculo = pd.DataFrame(resultados)

                    st.markdown("---")
                    if ha_total_detectada > 0:
                        st.success(f"✅ Motor Híbrido Activado: El sistema analizó el patrón histórico de {meses_dict[mes_proyeccion]} y lo ajustó al crecimiento en hectáreas.")
                    else:
                        st.warning(f"⚠️ El radar no encontró hectáreas operadas en el mes de {meses_dict[mes_proyeccion]} dentro de la base histórica seleccionada.")

                    st.markdown(f"### 🎯 Tablero de Mando: Proyección {modo_analisis.split(' ')[0]} para {meses_dict[mes_proyeccion]}")
                    
                    if df_oraculo.empty:
                        st.info("No se hallaron productos en SAP para la pista seleccionada.")
                    else:
                        def get_sort_weight(estado_str):
                            if "CRÍTICO" in estado_str: return 1
                            if "ALERTA" in estado_str: return 2
                            return 3

                        df_oraculo['SORT_WEIGHT'] = df_oraculo['ESTADO'].apply(get_sort_weight)
                        df_oraculo['SOLO_NOMBRE'] = df_oraculo['🧪 CÓDIGO | PRODUCTO'].apply(lambda x: x.split('|')[1].strip() if '|' in x else x)
                        df_oraculo = df_oraculo.sort_values(by=["📍 PISTA", "SORT_WEIGHT", "SOLO_NOMBRE"], ascending=[True, True, True])
                        df_oraculo = df_oraculo.drop(columns=['SORT_WEIGHT', 'SOLO_NOMBRE'])
                        
                        criticos = len(df_oraculo[df_oraculo['ESTADO'].str.contains("CRÍTICO")])
                        alertas = len(df_oraculo[df_oraculo['ESTADO'].str.contains("ALERTA")])
                        optimos = len(df_oraculo) - (criticos + alertas)
                        
                        c_k1, c_k2, c_k3 = st.columns(3)
                        c_k1.markdown(f"""
                        <div style="background-color: #ffe6e6; border-left: 5px solid #cc0000; padding: 10px; border-radius: 5px; height: 100%;">
                            <span style="color: #cc0000; font-weight: bold;">🚨 RUPTURA INMINENTE</span><br/>
                            <span style="font-size: 18px; color: #0d1b2a; font-weight: bold;">{criticos} Insumos</span>
                        </div>
                        """, unsafe_allow_html=True)
                        c_k2.markdown(f"""
                        <div style="background-color: #fff3cd; border-left: 5px solid #ffc107; padding: 10px; border-radius: 5px; height: 100%;">
                            <span style="color: #856404; font-weight: bold;">⚠️ ALERTA LOGÍSTICA</span><br/>
                            <span style="font-size: 18px; color: #0d1b2a; font-weight: bold;">{alertas} Insumos</span>
                        </div>
                        """, unsafe_allow_html=True)
                        c_k3.markdown(f"""
                        <div style="background-color: #d4edda; border-left: 5px solid #28a745; padding: 10px; border-radius: 5px; height: 100%;">
                            <span style="color: #155724; font-weight: bold;">✅ INVENTARIO SANO</span><br/>
                            <span style="font-size: 18px; color: #0d1b2a; font-weight: bold;">{optimos} Insumos</span>
                        </div>
                        """, unsafe_allow_html=True)
                        st.markdown("<br/>", unsafe_allow_html=True)
                        
                        # 💥 MOTOR VISUAL PREMIUM (Alineación y Sombreado Dinámico)
                        def pintar_oraculo(row):
                            estilos = [''] * len(row)
                            estado = str(row['ESTADO']).upper()
                            
                            # Fondo por nivel de alerta
                            if "CRÍTICO" in estado: 
                                base_style = 'background-color: #ffe6e6; color: #cc0000; font-weight: 900;'
                            elif "ALERTA" in estado: 
                                base_style = 'background-color: #fff3cd; color: #856404; font-weight: 900;'
                            else: 
                                base_style = 'background-color: #ffffff; color: #155724; font-weight: 600;'
                                
                            for i, col in enumerate(row.index):
                                cell_style = base_style
                                # Alineamos a la derecha todo lo que sean números
                                if col not in ['📍 PISTA', '🧪 CÓDIGO | PRODUCTO', 'ESTADO']:
                                    cell_style += ' text-align: right;'
                                else:
                                    cell_style += ' text-align: left;'
                                estilos[i] = cell_style
                                
                            return estilos

                        df_vista = df_oraculo.copy()
                        
                        # Formato estricto para mantener la tabla viva y ordenable
                        formato_oraculo = {}

                        if modo_analisis == "Estándar (Autonomía Días)":
                            formato_oraculo["📦 SALDO (SAP)"] = lambda x: f"{x:,.1f}".replace(',', '.') if pd.notna(x) else ""
                            formato_oraculo["📈 PROYECCIÓN MES (L/Kg)"] = lambda x: f"{x:,.1f}".replace(',', '.') if pd.notna(x) else ""
                            formato_oraculo["⏳ AUTONOMÍA (DÍAS)"] = lambda x: "∞" if x >= 9999 else f"{int(x)}"
                            
                            config_columnas = {
                                "📍 PISTA": st.column_config.TextColumn("📍 PISTA", width="small"),
                                "🧪 CÓDIGO | PRODUCTO": st.column_config.TextColumn("🧪 PRODUCTO", width="large"),
                                "📦 SALDO (SAP)": st.column_config.TextColumn("📦 SALDO (SAP)", width="medium"),
                                "📈 PROYECCIÓN MES (L/Kg)": st.column_config.TextColumn("📈 PROYECCIÓN MES", width="medium"),
                                "⏳ AUTONOMÍA (DÍAS)": st.column_config.TextColumn("⏳ AUTONOMÍA (DÍAS)", width="medium"),
                                "ESTADO": st.column_config.TextColumn("🛡️ ESTADO LOGÍSTICO", width="medium")
                            }
                        else:
                            formato_oraculo["📦 SALDO (SAP)"] = lambda x: f"{x:,.1f}".replace(',', '.') if pd.notna(x) else ""
                            formato_oraculo["⏱️ DÍAS POR CICLO"] = lambda x: f"{x:.1f}" if pd.notna(x) else ""
                            formato_oraculo["🔄 CONSUMO POR CICLO"] = lambda x: f"{x:,.1f}".replace(',', '.') if pd.notna(x) else ""
                            formato_oraculo["🔋 CICLOS RESTANTES"] = lambda x: "∞" if x >= 99.9 else f"{x:.2f}"
                            
                            config_columnas = {
                                "📍 PISTA": st.column_config.TextColumn("📍 PISTA", width="small"),
                                "🧪 CÓDIGO | PRODUCTO": st.column_config.TextColumn("🧪 PRODUCTO", width="large"),
                                "📦 SALDO (SAP)": st.column_config.TextColumn("📦 SALDO (SAP)", width="medium"),
                                "⏱️ DÍAS POR CICLO": st.column_config.TextColumn("⏱️ DÍAS POR CICLO", width="small"),
                                "🔄 CONSUMO POR CICLO": st.column_config.TextColumn("🔄 CONSUMO/CICLO", width="medium"),
                                "🔋 CICLOS RESTANTES": st.column_config.TextColumn("🔋 CICLOS RESTANTES", width="medium"),
                                "ESTADO": st.column_config.TextColumn("🛡️ ESTADO TÁCTICO", width="medium")
                            }

                        st.dataframe(
                            df_vista.style.apply(pintar_oraculo, axis=1).format(formato_oraculo), 
                            use_container_width=True, 
                            hide_index=True,
                            column_config=config_columnas
                        )
                            df_vista.style.apply(pintar_oraculo, axis=1), 
                            use_container_width=True, 
                            hide_index=True,
                            column_config=config_columnas
                        )

                        nombre_archivo = "Proyeccion_Lineal" if modo_analisis == "Estándar (Autonomía Días)" else "Proyeccion_Tactica_Ciclos"
                        excel_export = generar_excel_vip(df_vista, nombre_archivo)
                        
                        st.download_button(
                            label=f"💾 DESCARGAR INFORME {modo_analisis.split(' ')[0].upper()} (EXCEL VIP)", 
                            data=excel_export, 
                            file_name=f"{nombre_archivo}_{meses_dict[mes_proyeccion]}.xlsx", 
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                            use_container_width=True
                        )
                        
                except Exception as e:
                    st.error(f"🚨 Falla en los cálculos predictivos o estructura de datos: {e}")

if __name__ == "__main__":
    pass
