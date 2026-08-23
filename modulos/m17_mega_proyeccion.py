import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import gspread
import re
import math
import io
import openpyxl
from datetime import datetime, date
from oauth2client.service_account import ServiceAccountCredentials

# 🛰️ ENLACES NATIVOS
from modulos.utilidades import procesar_fecha_pesada
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =================================================================
# ⚙️ REGLAS DE NEGOCIO Y CONFIGURACIÓN ESTRATÉGICA
# =================================================================
TARIFA_VUELO_DEFAULT = 45000.0

REGLAS_FINANCIERAS = {
    "TERCERO":     {"M_MEZCLA": 1.451, "ST_BASE": 1583.0, "M_VUELO": 1.451},
    "AFILIADO":    {"M_MEZCLA": 1.164, "ST_BASE": 1510.0, "M_VUELO": 1.164},
    "COOPERATIVA": {"M_MEZCLA": 1.112, "ST_BASE": 1510.0, "M_VUELO": 1.164},
    "ORGANICO":    {"M_MEZCLA": 1.011, "ST_BASE": 1337.0, "M_VUELO": 1.011},
    "DEFAULT":     {"M_MEZCLA": 1.112, "ST_BASE": 1337.0, "M_VUELO": 1.112}
}

FALLBACK_FERTILIZANTES = {
    "ZN": "ZINTRAC X LITRO SV",
    "BT": "BANATREL SC",
    "NM": "NATURAMIN WSP",
    "QM": "QUELAMIX",
    "ZT": "ZITRON"
}

DOSIS_ACONDICIONADOR_ALTA = 0.06
DOSIS_ACONDICIONADOR_BAJA = 0.02
DOSIS_IMBIOSIL_ALTA = 1.5
DOSIS_IMBIOSIL_BAJA = 1.0
DOSIS_SPRAYFIX_ORG = 0.2

# =================================================================
# 🔌 CONEXIÓN Y MOTORES DE FORMATO REGIONAL BLINDADOS
# =================================================================

def formato_latino(numero, decimales=0):
    if pd.isna(numero) or numero is None: return "0"
    try:
        num = float(numero)
        if num == 0: return "0"
        if decimales == 0: texto_us = f"{num:,.0f}"
        else: texto_us = f"{num:,.{decimales}f}"
        return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0"

def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except:
        return None

def limpiar_tarifa_excel(val):
    if isinstance(val, (int, float)): return float(val)
    v = str(val).strip().replace("$", "").replace(" ", "").upper()
    if not v or v in ['-', 'NAN', 'NONE', '']: return 0.0
    
    s_clean = re.sub(r'[^\d\.,\-]', '', v)
    try:
        if '.' in s_clean and ',' in s_clean:
            if s_clean.rfind(',') > s_clean.rfind('.'): s_clean = s_clean.replace('.', '').replace(',', '.')
            else: s_clean = s_clean.replace(',', '')
        elif ',' in s_clean:
            if len(s_clean.split(',')[-1]) == 3: s_clean = s_clean.replace(',', '')
            else: s_clean = s_clean.replace(',', '.')
        elif '.' in s_clean:
            if s_clean.count('.') > 1: s_clean = s_clean.replace('.', '')
            elif len(s_clean.split('.')[-1]) == 3: s_clean = s_clean.replace('.', '')
        return float(s_clean) if s_clean else 0.0
    except:
        return 0.0

def normalizar_a_fecha_pura(val):
    try:
        res_nativo = procesar_fecha_pesada(val)
        if isinstance(res_nativo, (datetime, pd.Timestamp)): return res_nativo.date()
        if isinstance(res_nativo, date): return res_nativo
        return pd.to_datetime(str(res_nativo)).date()
    except: return None

def parsear_fecha_sap(fecha_str):
    try:
        if fecha_str is None or pd.isna(fecha_str) or not str(fecha_str).strip(): return None
    except: pass
    s = str(fecha_str).strip().replace("_", "-").replace("/", "-").replace(".", "-")
    try: return pd.to_datetime(s, dayfirst=True, errors="raise").date()
    except: return normalizar_a_fecha_pura(fecha_str)

@st.cache_data(show_spinner=False, ttl=60)
def cargar_bases_m17(url_boveda, url_precios, _supabase_client=None):
    gc = obtener_cliente_gspread_unificado()
    if not gc: return None, None, None, None, None, pd.DataFrame()
    
    df_mezclas, df_conf, df_dicc, df_t2, df_precios, df_t1 = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    try:
        boveda_recetas = gc.open_by_url(url_boveda)
        sh_precios = gc.open_by_url(url_precios)
        
        try: df_mezclas = pd.DataFrame(boveda_recetas.worksheet("DD_Mesclas").get_all_values()[1:], columns=boveda_recetas.worksheet("DD_Mesclas").get_all_values()[0])
        except: pass
        if not df_mezclas.empty: df_mezclas['COCTEL_CLEAN'] = df_mezclas.iloc[:, 0].astype(str).str.upper().str.replace(" ", "")

        try: df_conf = pd.DataFrame(boveda_recetas.worksheet("Configuración").get_all_values()[1:], columns=boveda_recetas.worksheet("Configuración").get_all_values()[0])
        except: pass
        
        if _supabase_client:
            try:
                res = _supabase_client.table("DICCIONARIO_SIGLAS").select("*").execute()
                if res.data:
                    df_dicc = pd.DataFrame(res.data)
                    df_dicc.columns = [str(c).upper().strip() for c in df_dicc.columns]
            except: pass

        if df_dicc.empty:
            try: 
                dicc_raw = boveda_recetas.worksheet("DICCIONARIO_SIGLAS").get_all_values()
                if dicc_raw: df_dicc = pd.DataFrame(dicc_raw[1:], columns=[str(c).upper().strip() for c in dicc_raw[0]])
            except: pass
        
        try: 
            t2_raw = boveda_recetas.worksheet("TABLA 2").get_all_values()
            idx_t2 = next((i for i, r in enumerate(t2_raw) if "FINCA" in [str(x).upper().strip() for x in r]), 0)
            df_t2 = pd.DataFrame(t2_raw[idx_t2+1:], columns=[str(c).strip() for c in t2_raw[idx_t2]])
        except: pass

        try:
            ws_datos = sh_precios.worksheet("DATOS") 
            datos_hoja = ws_datos.get_all_values()
            precios_consolidados = []
            if datos_hoja:
                idx_header, col_anio, col_prod = -1, -1, -1
                for i in range(min(10, len(datos_hoja))):
                    fila_upper = [str(x).upper().strip() for x in datos_hoja[i]]
                    if 'AÑO' in fila_upper and 'PRODUCTO' in fila_upper:
                        idx_header, col_anio, col_prod = i, fila_upper.index('AÑO'), fila_upper.index('PRODUCTO'); break
                
                if idx_header != -1:
                    for row in datos_hoja[idx_header+1:]:
                        if len(row) > max(col_anio, col_prod):
                            anio_str, str_prod = str(row[col_anio]).strip().upper(), str(row[col_prod]).strip().upper()
                            if anio_str and str_prod:
                                vals = []
                                for v in row[max(col_anio, col_prod) + 1:]:
                                    val_c = re.sub(r'[^\d\.,\-]', '', str(v).strip())
                                    if val_c:
                                        if '.' in val_c and ',' in val_c: val_c = val_c.replace('.', '').replace(',', '.') if val_c.rfind(',') > val_c.rfind('.') else val_c.replace(',', '')
                                        elif ',' in val_c: val_c = val_c.replace(',', '.')
                                        try: vals.append(float(val_c))
                                        except: pass
                                if vals: precios_consolidados.append({'AÑO': anio_str, 'PRODUCTO': str_prod, 'PRODUCTO_CLEAN': str_prod.replace(" ", ""), 'PRECIO_PROM': sum(vals)/len(vals)})
            df_precios = pd.DataFrame(precios_consolidados)
        except: pass

        try:
            t1_raw = boveda_recetas.worksheet("TABLA 1").get_all_values()
            if t1_raw:
                idx_t1 = next((i for i, r in enumerate(t1_raw) if "FINCA" in [str(x).upper().strip() for x in r]), 4)
                encabezados = [str(c).upper().strip() for c in t1_raw[idx_t1]]
                df_t1 = pd.DataFrame(t1_raw[idx_t1+1:], columns=encabezados)
                
                col_finca = next((c for c in encabezados if "FINCA" in c or "PROPIEDAD" in c), None)
                if not col_finca and len(encabezados) > 2: col_finca = encabezados[2]
                col_fecha = next((c for c in encabezados if "FECHA" in c or "DATE" in c), None)
                col_costo_ha = next((c for c in encabezados if "COSTO" in c and "AVI" in c and "$/HA" in c.replace(" ", "")), None)
                if not col_costo_ha: col_costo_ha = next((c for c in encabezados if "COSTO" in c and "$/HA" in c.replace(" ", "")), None)
                col_recargo = next((c for c in encabezados if "DOMINIC" in c or "RECARGO" in c), None)
                
                if col_finca and col_costo_ha:
                    def limp_num_col(val):
                        v = str(val).strip()
                        if not v or v == '-': return 0.0
                        v = re.sub(r'[^\d\.,\-]', '', v)
                        if not v: return 0.0
                        try:
                            if v.count('.') == 1 and v.count(',') == 0:
                                if len(v.split('.')[1]) == 3: v = v.replace('.', '')
                            if '.' in v and ',' in v:
                                if v.rfind(',') > v.rfind('.'): v = v.replace('.', '').replace(',', '.')
                                else: v = v.replace(',', '')
                            elif ',' in v: v = v.replace(',', '.')
                            f_val = float(v)
                            if f_val < 1000 and '.' in str(val) and len(str(val).split('.')[-1]) == 3: f_val = f_val * 1000
                            return f_val
                        except: return 0.0
                    
                    df_t1['F_CLEAN'] = df_t1[col_finca].astype(str).apply(lambda x: re.sub(r'[^A-Z0-9]', '', x.upper().strip()))
                    df_t1['VAL_COSTO_HA'] = df_t1[col_costo_ha].apply(limp_num_col)
                    df_t1['VAL_RECARGO_HA'] = df_t1[col_recargo].apply(limp_num_col) if col_recargo else 0.0
                    if col_fecha: df_t1['FECHA_CLEAN'] = df_t1[col_fecha].astype(str).str.strip()
        except: pass
                            
    except Exception as e: 
        raise Exception(f"Error de conexión: {e}")

    return df_mezclas, df_conf, df_dicc, df_t2, df_precios, df_t1

# =================================================================
# 🧠 MOTORES DE LÓGICA 100% ORIGINAL
# =================================================================

def limpiar_numero(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip().replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1: v = v.rsplit('.', 1)[0].replace('.', '') + '.' + v.rsplit('.', 1)[1]
        return float(v) if v else 0.0
    except: return 0.0

def calcular_historicos_finca_rapido(finca_n_clean, dict_t1, fecha_inicio, fecha_fin):
    df_finca = dict_t1.get(finca_n_clean, pd.DataFrame())
    if df_finca.empty: return 45000.0, 0.0 
        
    df_evaluar = df_finca
    
    # 💥 FILTRO DE FECHAS ULTRA-RÁPIDO (Se procesó antes del bucle)
    if 'FECHA_PURA' in df_finca.columns:
        mask_fechas = (df_finca['FECHA_PURA'] >= fecha_inicio) & (df_finca['FECHA_PURA'] <= fecha_fin)
        df_finca_fechas = df_finca[mask_fechas]
        if not df_finca_fechas.empty and not df_finca_fechas[df_finca_fechas['VAL_COSTO_HA'] > 1000].empty:
            df_evaluar = df_finca_fechas
            
    prom_vuelo, prom_recargo = 45000.0, 0.0
    df_valid_costos = df_evaluar[df_evaluar['VAL_COSTO_HA'] > 1000]
    if not df_valid_costos.empty:
        prom_vuelo = float(df_valid_costos['VAL_COSTO_HA'].mean())
        if pd.isna(prom_vuelo): prom_vuelo = 45000.0

    if 'VAL_RECARGO_HA' in df_evaluar.columns:
        df_recargos_validos = df_evaluar[df_evaluar['VAL_RECARGO_HA'] > 100]
        if not df_recargos_validos.empty:
            prom_recargo = float(df_recargos_validos['VAL_RECARGO_HA'].mean())
            if pd.isna(prom_recargo): prom_recargo = 0.0
            
    return prom_vuelo, prom_recargo

def extraer_receta_mega_rapida(coctel_sel, finca_sel_clean, dict_mezclas, dict_dicc, dict_t2, fallback_mezclas_dosis):
    coctel_u = str(coctel_sel).upper().strip().replace("+", " ").replace("-", " ")
    partes = coctel_u.split()
    base_coctel = partes[0] if partes else ""
    aditivos = partes[1:] if len(partes) > 1 else []
    
    dict_prods = {}
    es_organico = False
    
    match_f = dict_t2.get(finca_sel_clean, pd.DataFrame())
    if not match_f.empty and "ORGANIC" in str(match_f.iloc[0, 5]).upper(): es_organico = True

    base_buscar = f"{base_coctel}O" if es_organico and not base_coctel.endswith('O') else base_coctel

    rb = dict_mezclas.get(base_buscar, pd.DataFrame())
    if rb.empty and es_organico: rb = dict_mezclas.get(base_coctel, pd.DataFrame())
        
    if not rb.empty:
        for _, r in rb.iterrows():
            p, d = str(r.iloc[1]).strip().upper(), limpiar_numero(r.iloc[2])
            if d > 0 and p not in ['NAN', 'NONE', '']: dict_prods[p] = d

    for ad in aditivos:
        m_s = dict_dicc.get(ad, pd.DataFrame())
        if not m_s.empty:
            p_ad, d_ad = str(m_s.iloc[0]['PRODUCTO']).strip().upper(), limpiar_numero(m_s.iloc[0]['DOSIS'])
            if d_ad > 0 and p_ad not in ['NAN', 'NONE', '']: dict_prods[p_ad] = dict_prods.get(p_ad, 0.0) + d_ad

    for ad in aditivos:
        if ad in FALLBACK_FERTILIZANTES:
            p_fall = FALLBACK_FERTILIZANTES[ad]
            if not any(p_fall in k for k in dict_prods.keys()):
                d_fall = fallback_mezclas_dosis.get(p_fall, 0.5) 
                dict_prods[p_fall] = dict_prods.get(p_fall, 0.0) + d_fall

    for p in list(dict_prods.keys()):
        if "ACONDICIONADOR" in p: dict_prods[p] = DOSIS_ACONDICIONADOR_ALTA if any(x in coctel_u for x in ["ZN", "BT", "ZT", "ZITRON"]) else DOSIS_ACONDICIONADOR_BAJA
        elif "IMBIOSIL" in p.replace(" ", ""): dict_prods[p] = DOSIS_IMBIOSIL_ALTA if base_coctel.startswith("IN") or "IMBIOSIL" in base_coctel else DOSIS_IMBIOSIL_BAJA
        if es_organico and "ADHERENTE" in p: del dict_prods[p]
    if es_organico and not any("SPRAYFIX" in k for k in dict_prods.keys()): dict_prods["SPRAYFIX"] = DOSIS_SPRAYFIX_ORG
    
    return dict_prods

# 💥 GENERADOR DE EXCEL CON CACHÉ (Evita que la pantalla se congele al usar filtros)
@st.cache_data(show_spinner=False)
def generar_excel_gerencial(df_filtro, df_resumen_finca, df_insumos_raw):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_filtro.to_excel(writer, sheet_name='Detalle_Económico', index=False)
        df_resumen_finca.to_excel(writer, sheet_name='Resumen_x_Finca', index=False)
        if not df_insumos_raw.empty:
            df_insumos_raw[["🧪 PRODUCTO", "VOLUMEN ESTIMADO"]].to_excel(writer, sheet_name='Consumo_Insumos', index=False)
        
        workbook = writer.book
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.sheet_view.showGridLines = False
            max_r, max_c = ws.max_row, ws.max_column
            column_headers = {}
            
            for col_idx in range(1, max_c + 1):
                letra_columna = openpyxl.utils.get_column_letter(col_idx)
                header_val = ws.cell(row=1, column=col_idx).value
                valores_columna = [ws.cell(row=r, column=col_idx).value for r in range(1, min(max_r, 50) + 1)]
                ancho = max((len(str(v)) if v is not None else 0) for v in valores_columna) + 2
                ws.column_dimensions[letra_columna].width = min(max(ancho, 14), 35)
                column_headers[col_idx] = str(header_val).upper() if header_val else ""

            for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
                for cell in row:
                    cell.border = borde
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center')
                        col_name = column_headers.get(cell.column, "")
                        if isinstance(cell.value, (int, float)):
                            if "COSTO" in col_name or "PRECIO" in col_name or "RESULTADO" in col_name or "TOTAL" in col_name or "RECARGO" in col_name:
                                cell.number_format = '"$" #,##0' 
                            elif "HECTAREAS" in col_name or "VOLUMEN" in col_name:
                                cell.number_format = '#,##0.0'
    return buffer.getvalue()

# =================================================================
# 👑 RENDERIZADO VISUAL PRINCIPAL
# =================================================================

def ejecutar(supabase_client=None):
    VERDE_INTENSO = '#143521'
    COLOR_NAVY = '#0d1b2a'
    COLOR_DORADO = '#d4af37'

    css_maestro = f"""
    <style>
    .titulo-mega {{ color: {COLOR_NAVY}; border-bottom: 3px solid {COLOR_DORADO}; padding-bottom: 5px; font-family: 'Arial Black'; margin-bottom: 15px;}}
    div[data-testid="stDataEditor"], div[data-testid="stDataFrame"] {{ border: 2px solid {COLOR_NAVY} !important; border-radius: 8px !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.1); overflow: hidden !important; }}
    .tarjeta-kpi {{ background: linear-gradient(135deg, {COLOR_NAVY} 0%, #1a365d 100%); border-left: 5px solid {COLOR_DORADO}; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.2); text-align: center; margin-bottom: 15px;}}
    .kpi-titulo {{ font-size: 11px; font-weight: bold; color: {COLOR_DORADO}; text-transform: uppercase; margin:0; letter-spacing: 1px; }}
    .kpi-valor {{ font-size: 21px; font-family: 'Arial Black'; margin: 5px 0 0 0; }}
    div[data-testid="stTextInput"] > div, div[data-testid="stNumberInput"] > div, div[data-testid="stDateInput"] > div, div[data-testid="stMultiSelect"] div[data-baseweb="select"] {{ background-color: #ffffff !important; border: 2px solid {COLOR_NAVY} !important; border-radius: 6px !important; }}
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {{ background-color: transparent !important; border: none !important; }}
    div[data-testid="stTextInput"] input, div[data-testid="stNumberInput"] input, div[data-testid="stDateInput"] input, div[data-testid="stMultiSelect"] * {{ color: {COLOR_NAVY} !important; font-weight: bold !important; }}
    div[data-testid="stMainBlockContainer"] label p {{ color: {COLOR_NAVY} !important; font-weight: 800 !important; text-transform: uppercase !important; }}
    </style>
    """
    st.markdown(css_maestro, unsafe_allow_html=True)
    st.markdown("<h1 class='titulo-mega'>🚀 Módulo 17: Mega-Proyección Operativa</h1>", unsafe_allow_html=True)

    db_cargada = st.session_state.get('m17_db_cargada', False)
    if db_cargada and 'm17_t1' not in st.session_state:
        st.session_state['m17_db_cargada'] = False
        db_cargada = False

    with st.expander("🔌 CONEXIÓN A LAS MAESTRAS DE GOOGLE DRIVE Y BASE DE DATOS", expanded=not db_cargada):
        if db_cargada: st.success("✅ Bases de Datos conectadas y en Memoria RAM.")
        else: st.info("💡 Pega los enlaces de tus archivos de Google Sheets.")
            
        url_1 = st.text_input("🔗 Link Bóveda (Recetas, Fincas, Tabla 1):", value=st.session_state.get('m17_url1', ''))
        url_2 = st.text_input("🔗 Link Comparativo de Precios:", value=st.session_state.get('m17_url2', ''))

        if st.button("🔄 Conectar y Descargar", type="primary"):
            if url_1 and url_2:
                with st.spinner("Descargando información (Modo Original Protegido)..."):
                    try:
                        mez, conf, dicc, t2, prec, t1 = cargar_bases_m17(url_1, url_2, supabase_client)
                        st.session_state.update({'m17_mez':mez, 'm17_conf':conf, 'm17_dicc':dicc, 'm17_t2':t2, 'm17_prec':prec, 'm17_t1':t1, 'm17_url1':url_1, 'm17_url2':url_2, 'm17_db_cargada':True})
                        st.success("¡Extracción perfecta!")
                        st.rerun()
                    except Exception as e: st.error(str(e))
            else: st.warning("⚠️ Debes pegar ambos enlaces.")

    if not db_cargada: st.stop() 

    df_mezclas = st.session_state.get('m17_mez', pd.DataFrame())
    df_conf = st.session_state.get('m17_conf', pd.DataFrame())
    df_dicc = st.session_state.get('m17_dicc', pd.DataFrame())
    df_t2 = st.session_state.get('m17_t2', pd.DataFrame())
    df_precios = st.session_state.get('m17_prec', pd.DataFrame())
    df_t1 = st.session_state.get('m17_t1', pd.DataFrame())
    
    if 'm17_df_entrada_grid' not in st.session_state or 'DOMINICAL' not in st.session_state.m17_df_entrada_grid.columns:
        st.session_state.m17_df_entrada_grid = pd.DataFrame([{"FINCA": "", "HECTAREAS": "", "COCTEL": "", "FERTILIZANTE": "", "DIAS CICLO": "", "PRECIO VUELO": "", "DOMINICAL": False} for _ in range(1000)])

    st.markdown("### 📥 1. Pista de Aterrizaje Segura")
    df_edited = st.data_editor(st.session_state.m17_df_entrada_grid, key="m17_tabla_maestra_grid", use_container_width=True, hide_index=True,
        column_config={
            "FINCA": st.column_config.TextColumn("Finca"), "HECTAREAS": st.column_config.TextColumn("Hectáreas"), 
            "COCTEL": st.column_config.TextColumn("Cóctel"), "FERTILIZANTE": st.column_config.TextColumn("Fertilizante"),
            "DIAS CICLO": st.column_config.TextColumn("Días Ciclo"), "PRECIO VUELO": st.column_config.TextColumn("Precio Vuelo Manual (Opcional)"),
            "DOMINICAL": st.column_config.CheckboxColumn("¿Dom/Fest?", default=False),
        })

    st.markdown("---")
    st.markdown("### ⚙️ 2. Parámetros de Riesgo y Base Histórica")
    
    c_f1, c_f2, c_r1, c_r2 = st.columns(4)
    fecha_base_inicio = c_f1.date_input("📅 Rango Histórico (Desde)", value=date(date.today().year, 1, 1))
    fecha_base_fin = c_f2.date_input("📅 Rango Histórico (Hasta)", value=date(date.today().year, 12, 31), min_value=fecha_base_inicio)
    inflacion_proyectada = c_r1.number_input("📈 Inflación a Proyectar (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0)
    colchon_dias = c_r2.number_input("🛡️ Colchón de Días Ciclo", min_value=0, max_value=30, value=0, step=1)

    factor_inflacion = 1 + (inflacion_proyectada / 100)

    if st.button("🔥 EJECUTAR MEGA-PROYECCIÓN", type="primary", use_container_width=True):
        df_valid = df_edited.dropna(subset=['FINCA']).copy()
        df_valid = df_valid[df_valid['FINCA'].astype(str).str.strip() != ""]
        
        if df_valid.empty: st.error("⚠️ La tabla está vacía.")
        else:
            with st.spinner("⚡ Compilando Motores de Memoria (Hyper-Speed)..."):
                
                # 💥 ACELERADOR 1: PARSEAR FECHAS DE SAP UNA SOLA VEZ
                if not df_t1.empty and 'FECHA_CLEAN' in df_t1.columns:
                    if 'FECHA_PURA' not in df_t1.columns:
                        df_t1['FECHA_PURA'] = df_t1['FECHA_CLEAN'].apply(parsear_fecha_sap)

                # 💥 ACELERADOR 2: DICCIONARIOS EN RAM
                dict_t1 = dict(tuple(df_t1.groupby('F_CLEAN'))) if not df_t1.empty and 'F_CLEAN' in df_t1.columns else {}
                
                dict_t2 = {}
                col_prod_idx = 5
                if not df_t2.empty:
                    for i, c_name in enumerate(df_t2.columns):
                        c_clean = str(c_name).upper().replace('\n', ' ').strip()
                        if 'TIPO' in c_clean and 'PROD' in c_clean:
                            col_prod_idx = i
                            break 
                    df_t2_clean = df_t2.assign(F_CLEAN=df_t2.iloc[:, 0].astype(str).str.upper().apply(lambda x: re.sub(r"[^A-Z0-9]", "", x)))
                    dict_t2 = dict(tuple(df_t2_clean.groupby('F_CLEAN')))
                
                dict_conf = dict(tuple(df_conf.assign(TIPO=df_conf.iloc[:, 0].astype(str).str.strip().str.upper()).groupby('TIPO'))) if not df_conf.empty else {}
                dict_mezclas = dict(tuple(df_mezclas.assign(KEY=df_mezclas.iloc[:, 0].astype(str).str.upper().str.strip()).groupby('KEY'))) if not df_mezclas.empty else {}
                dict_dicc = dict(tuple(df_dicc.assign(KEY=df_dicc['SIGLA'].astype(str).str.upper().str.strip()).groupby('KEY'))) if not df_dicc.empty and 'SIGLA' in df_dicc.columns else {}
                dict_precios = dict(tuple(df_precios.groupby('PRODUCTO_CLEAN'))) if not df_precios.empty and 'PRODUCTO_CLEAN' in df_precios.columns else {}

                c_p_i, c_c_i = 8, 9
                dict_precios_conf = {}
                if not df_conf.empty:
                    for i in range(min(5, len(df_conf))):
                        r_c = [str(x).upper() for x in df_conf.iloc[i]]
                        if 'PRODUCTO' in r_c and 'COSTO' in r_c:
                            c_p_i, c_c_i = r_c.index('PRODUCTO'), r_c.index('COSTO'); break
                    for _, r in df_conf.iterrows():
                        prod = str(r.iloc[c_p_i]).strip().upper()
                        if prod:
                            costo = limpiar_numero(r.iloc[c_c_i])
                            dict_precios_conf[prod] = costo
                            dict_precios_conf[prod.replace(" ", "")] = costo

                fallback_mezclas_dosis = {}
                if not df_mezclas.empty:
                    for col_idx in range(len(df_mezclas.columns) - 1):
                        for _, r in df_mezclas.iterrows():
                            p_name = str(r.iloc[col_idx]).strip().upper()
                            if p_name and p_name not in fallback_mezclas_dosis:
                                val = limpiar_numero(r.iloc[col_idx+1])
                                if val > 0: fallback_mezclas_dosis[p_name] = val

            with st.spinner("Calculando proyección (Velocidad Nativa)..."):
                resultados = []
                log_volumetrico = {}

                for row in df_valid.to_dict('records'):
                    finca_n = str(row['FINCA']).strip().upper()
                    finca_n_clean = re.sub(r'[^A-Z0-9]', '', finca_n)
                    
                    ha_num = limpiar_numero(row['HECTAREAS'])
                    coctel_n = str(row['COCTEL']).strip().upper() if pd.notna(row['COCTEL']) else ""
                    fert_n = str(row.get('FERTILIZANTE', '')).strip().upper() if pd.notna(row.get('FERTILIZANTE')) and str(row.get('FERTILIZANTE')).strip().upper() != "NONE" else ""
                    coctel_combinado = f"{coctel_n} {fert_n}".strip()

                    dias_c = max(0, int(round(limpiar_numero(row["DIAS CICLO"])))) + int(colchon_dias)
                    precio_vuelo_manual = limpiar_numero(row["PRECIO VUELO"])
                    valor_dominical = row.get("DOMINICAL", False)
                    aplica_dominical = (valor_dominical is True or str(valor_dominical).strip().upper() in {"TRUE", "VERDADERO", "SI", "SÍ", "1"})

                    match_f = dict_t2.get(finca_n_clean, pd.DataFrame())
                    if ha_num == 0 and not match_f.empty:
                        ha_num = limpiar_numero(match_f.iloc[0].iloc[2])

                    if ha_num <= 0: continue

                    precio_vuelo_historico, recargo_historico = calcular_historicos_finca_rapido(finca_n_clean, dict_t1, fecha_base_inicio, fecha_base_fin)

                    precio_vuelo_final = precio_vuelo_historico if precio_vuelo_manual == 0 else precio_vuelo_manual
                    precio_vuelo_final = precio_vuelo_final * factor_inflacion
                    recargo_final_ha = (recargo_historico * factor_inflacion) if aplica_dominical else 0.0

                    tipo_prod = "TERCERO"
                    if not match_f.empty: 
                        tipo_prod = str(match_f.iloc[0].iloc[col_prod_idx]).strip().upper() if len(match_f.columns) > col_prod_idx else "TERCERO"
                    if "COOP" in finca_n or "EMPREBANCOOP" in finca_n: tipo_prod = "COOPERATIVA"

                    mult_m, st_base, mult_v = 1.112, 1337.0, 1.112
                    match_cfg = dict_conf.get(tipo_prod, pd.DataFrame())
                    if not match_cfg.empty:
                        mult_m = limpiar_numero(match_cfg.iloc[0].iloc[3])
                        st_base = limpiar_numero(match_cfg.iloc[0].iloc[4])
                        mult_v = limpiar_numero(match_cfg.iloc[0].iloc[6])
                    
                    if mult_m == 0 or st_base == 0:
                        if tipo_prod == "TERCERO": mult_m, st_base, mult_v = 1.451, 1583.0, 1.451
                        elif tipo_prod == "AFILIADO": mult_m, st_base, mult_v = 1.164, 1510.0, 1.164
                        elif tipo_prod == "COOPERATIVA": mult_m, st_base, mult_v = 1.112, 1510.0, 1.164
                        elif tipo_prod == "ORGANICO": mult_m, st_base, mult_v = 1.011, 1337.0, 1.011
                        else: mult_m, st_base, mult_v = 1.112, 1337.0, 1.112 

                    st_base = st_base * factor_inflacion
                    costo_mezcla_fila = 0.0

                    dict_receta = extraer_receta_mega_rapida(coctel_combinado, finca_n_clean, dict_mezclas, dict_dicc, dict_t2, fallback_mezclas_dosis)
                    
                    for p, d in dict_receta.items():
                        log_volumetrico[finca_n] = log_volumetrico.get(finca_n, {})
                        log_volumetrico[finca_n][p] = log_volumetrico[finca_n].get(p, 0.0) + (d * ha_num)

                        precio_unitario = 0.0
                        p_clean_spaces = p.replace(" ", "")
                        
                        if p in dict_precios_conf: precio_unitario = dict_precios_conf[p]
                        elif p_clean_spaces in dict_precios_conf: precio_unitario = dict_precios_conf[p_clean_spaces]
                        else:
                            if "NEMATI" in p:
                                for k_prod, v_cost in dict_precios_conf.items():
                                    if "NEMATI" in k_prod:
                                        precio_unitario = v_cost; break
                        
                        if precio_unitario == 0.0:
                            match_p = dict_precios.get(p_clean_spaces, pd.DataFrame())
                            if not match_p.empty: precio_unitario = match_p['PRECIO_PROM'].mean()

                        precio_unitario = precio_unitario * factor_inflacion
                        costo_mezcla_fila += (d * ha_num * precio_unitario * mult_m)

                    costo_st_fila = dias_c * st_base * ha_num
                    costo_vuelo_fila = precio_vuelo_final * ha_num 
                    costo_recargo_fila = recargo_final_ha * ha_num 

                    costo_mezcla_fila = 0.0 if pd.isna(costo_mezcla_fila) else float(costo_mezcla_fila)
                    costo_st_fila = 0.0 if pd.isna(costo_st_fila) else float(costo_st_fila)
                    costo_vuelo_fila = 0.0 if pd.isna(costo_vuelo_fila) else float(costo_vuelo_fila)
                    costo_recargo_fila = 0.0 if pd.isna(costo_recargo_fila) else float(costo_recargo_fila)

                    gran_total = math.floor(costo_mezcla_fila + costo_st_fila + costo_vuelo_fila + costo_recargo_fila + 0.5)
                    costo_ha = math.floor((gran_total / ha_num) + 0.5) if ha_num > 0 else 0

                    resultados.append({
                        "FINCA": finca_n, "HECTAREAS": ha_num, "COCTEL": coctel_combinado, "DIAS CICLO": dias_c, 
                        "PRECIO VUELO": precio_vuelo_final, "RECARGO ($/HA)": recargo_final_ha,
                        "Costo ST ($)": math.floor(costo_st_fila), "Costo Vuelo ($)": math.floor(costo_vuelo_fila), 
                        "Costo Recargo ($)": math.floor(costo_recargo_fila), "Costo Mezcla ($)": math.floor(costo_mezcla_fila),
                        "Costo x Ha ($)": costo_ha, "RESULTADO TOTAL ($)": gran_total
                    })

                df_resultados_final = pd.DataFrame(resultados)
                if not df_resultados_final.empty: df_resultados_final = df_resultados_final.sort_values(by="FINCA", ascending=True).reset_index(drop=True)

                st.session_state.m17_resultados = df_resultados_final
                st.session_state.m17_volumetria = log_volumetrico
                st.success("✅ Proyección completada (Ultra-Speed).")

    if 'm17_resultados' in st.session_state and not st.session_state.m17_resultados.empty:
        st.markdown("---")
        df_res = st.session_state.m17_resultados
        vol_dict = st.session_state.m17_volumetria
        
        fincas_procesadas = sorted(df_res['FINCA'].unique().tolist())
        
        st.markdown("### 🎛️ 3. Tablero de Mando y Filtros")
        fincas_filtro = st.multiselect("📍 Filtrar análisis por Finca(s) [Dejar vacío para ver TOTAL GENERAL]:", fincas_procesadas)
        
        if fincas_filtro:
            df_filtro = df_res[df_res['FINCA'].isin(fincas_filtro)]
            vol_dict_filtro = {k: v for k, v in vol_dict.items() if k in fincas_filtro}
        else:
            df_filtro = df_res
            vol_dict_filtro = vol_dict

        cons_vol_agrupado = {}
        for f, prods in vol_dict_filtro.items():
            for p, vol in prods.items():
                cons_vol_agrupado[p] = cons_vol_agrupado.get(p, 0.0) + vol
        
        t_st = df_filtro['Costo ST ($)'].sum()
        t_vu = df_filtro['Costo Vuelo ($)'].sum()
        t_re = df_filtro['Costo Recargo ($)'].sum() 
        t_mx = df_filtro['Costo Mezcla ($)'].sum()
        t_gr = df_filtro['RESULTADO TOTAL ($)'].sum()
        
        c1, c2, c3 = st.columns(3)
        with c1: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>👨‍🔬 Total Serv. Tec</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_st, 0)}</p></div>", unsafe_allow_html=True)
        with c2: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>✈️ Total Vuelo</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_vu, 0)}</p></div>", unsafe_allow_html=True)
        with c3: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>⚠️ Total Recargos</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_re, 0)}</p></div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        c4, c5 = st.columns(2)
        with c4: st.markdown(f"<div class='tarjeta-kpi'><p class='kpi-titulo'>🧪 Total Mezcla</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_mx, 0)}</p></div>", unsafe_allow_html=True)
        with c5: st.markdown(f"<div class='tarjeta-kpi' style='border-left: 5px solid #00ff00;'><p class='kpi-titulo' style='color:#00ff00;'>🔥 GRAN TOTAL</p><p class='kpi-valor'>$&nbsp;{formato_latino(t_gr, 0)}</p></div>", unsafe_allow_html=True)

        df_resumen_finca = df_filtro.groupby('FINCA', as_index=False)[['Costo ST ($)', 'Costo Vuelo ($)', 'Costo Recargo ($)', 'Costo Mezcla ($)', 'RESULTADO TOTAL ($)']].sum()

        tab1, tab2, tab3 = st.tabs(["📊 Detalles Económicos Fila x Fila", "📑 Resumen Ejecutivo por Finca", "📦 Auditoría Volumétrica de Insumos"])
        
        with tab1:
            df_view = df_filtro.copy()
            for col in ["PRECIO VUELO", "RECARGO ($/HA)", "Costo ST ($)", "Costo Vuelo ($)", "Costo Recargo ($)", "Costo Mezcla ($)", "Costo x Ha ($)", "RESULTADO TOTAL ($)"]:
                df_view[col] = df_view[col].apply(lambda x: f"$ {formato_latino(x, 0)}")
            st.dataframe(df_view, use_container_width=True, hide_index=True)

        with tab2:
            if not df_resumen_finca.empty:
                df_resumen_view = df_resumen_finca.copy()
                for col in ['Costo ST ($)', 'Costo Vuelo ($)', 'Costo Recargo ($)', 'Costo Mezcla ($)', 'RESULTADO TOTAL ($)']:
                    df_resumen_view[col] = df_resumen_view[col].apply(lambda x: f"$ {formato_latino(x, 0)}")
                st.dataframe(df_resumen_view, use_container_width=True, hide_index=True)
            else: st.info("No hay datos para resumir.")

        with tab3:
            df_insumos_raw = pd.DataFrame()
            if cons_vol_agrupado:
                df_insumos_raw = pd.DataFrame(list(cons_vol_agrupado.items()), columns=["🧪 PRODUCTO", "VOLUMEN ESTIMADO"]).sort_values("VOLUMEN ESTIMADO", ascending=False)
                df_insumos_vista = df_insumos_raw.copy()
                df_insumos_vista["📦 VOLUMEN ESTIMADO (L/Kg)"] = df_insumos_vista["VOLUMEN ESTIMADO"].apply(lambda x: formato_latino(x, 1))
                
                c_tbl, c_grf = st.columns([1, 1.2])
                with c_tbl: st.dataframe(df_insumos_vista[["🧪 PRODUCTO", "📦 VOLUMEN ESTIMADO (L/Kg)"]], use_container_width=True, hide_index=True)
                with c_grf:
                    df_grafica = df_insumos_raw.head(15).copy()
                    fig = px.bar(df_grafica, y="🧪 PRODUCTO", x="VOLUMEN ESTIMADO", text="VOLUMEN ESTIMADO", orientation='h', color="VOLUMEN ESTIMADO", color_continuous_scale="GnBu", title=f"Top 15 Insumos Proyectados")
                    fig.update_traces(texttemplate='%{text:,.1f}', textposition='outside', textfont_size=12)
                    fig.update_layout(yaxis={'categoryorder':'total ascending'}, plot_bgcolor='rgba(0,0,0,0)', margin=dict(r=100))
                    st.plotly_chart(fig, use_container_width=True)
            else: st.info("No hay datos de insumos químicos para las fincas seleccionadas.")

        st.markdown("<br>", unsafe_allow_html=True)
        
        # 💥 GENERAR EXCEL DESDE CACHÉ EN MILISEGUNDOS
        excel_data = generar_excel_gerencial(df_filtro, df_resumen_finca, df_insumos_raw if cons_vol_agrupado else pd.DataFrame())
        
        st.download_button(label="💾 DESCARGAR REPORTE GERENCIAL (EXCEL)", data=excel_data, file_name=f"MegaProyeccion_Operativa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

if __name__ == "__main__":
    pass
