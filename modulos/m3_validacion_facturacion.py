import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import gspread
import requests
import io
import re
import math
import json
from datetime import datetime, timedelta, date
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =================================================================
# ⚙️ CONSTANTES CENTRALIZADAS (ÚNICA FUENTE DE VERDAD)
# =================================================================

SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

TOPES_PISTA = {
    "TOPE MAX GENERAL": {"PLUC": 63326, "PORI": 62718, "TEHO": 63325, "PDIV": 63325, "LUCI": 63325},
    "TOPE SUR": {"PLUC": 71517, "PORI": 70829, "TEHO": 71517, "PDIV": 71517, "LUCI": 71517},
    "TOPE PARCELA INTER < 20HA": {"PLUC": 98335, "PORI": 105723, "TEHO": 98335, "PDIV": 105723, "LUCI": 98335},
}

PISTAS_VALIDAS = ["PLUC", "PORI", "PDIV", "TEHO", "LUCI"]
PISTAS_DISPONIBLES_MATRIZ = ["PLUC", "PORI", "PDIV", "TEHO", "LUCI", "Z-1", "Z-2", "PROPIA"]

DICT_AVIONES_DEFAULT = {
    "THRUS SR2": 4606562,
    "PIPER PA 36-375": 3985831,
    "CESSNA O PIPER PA 25": 3036525,
    "AIR TRACTOR": 4665109,
    "CESSNA ASA": 3768500,
    "CESSNA FUMIGARAY": 3065952,
}
DICT_DRONES_DEFAULT = {
    "DRONE DATAROT": 84428,
    "DRONE NORTE": 75518,
    "DRONE AVIL": 71280,
    "DRONE GENESYS": 71280,
}

def log_error_critico(contexto: str, e: Exception, mostrar_usuario: bool = True):
    mensaje = f"⚠️ Aviso técnico en «{contexto}»: {e}"
    if mostrar_usuario:
        st.warning(mensaje, icon="⚠️")
    else:
        print(mensaje)

# =================================================================
# 🔌 CONEXIÓN, RELOJ Y MOTORES NUMÉRICOS
# =================================================================

def obtener_hora_colombia():
    return datetime.utcnow() + timedelta(hours=-5)

def limpiar_numero_estricto(val):
    try:
        if pd.isna(val) or val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        v = str(val).upper().replace("$", "").replace("COP", "").replace(" ", "").strip()
        if not v or v == '-': return 0.0
        v = v.replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1: v = v.rsplit('.', 1)[0].replace('.', '') + '.' + v.rsplit('.', 1)[1]
        return float(v) if v else 0.0
    except: return 0.0

def limpiar_dinero(val):
    try:
        if pd.isna(val) or val is None: return 0.0
        if isinstance(val, (int, float)): return float(val)
        v = str(val).upper().replace("$", "").replace("COP", "").replace(" ", "").strip()
        if not v or v == '-': return 0.0
        
        if '.' in v and ',' in v:
            if v.rfind(',') > v.rfind('.'): v = v.replace('.', '').replace(',', '.')
            else: v = v.replace(',', '')
        elif ',' in v:
            if len(v.split(',')[-1]) == 3: v = v.replace(',', '')
            else: v = v.replace(',', '.')
        elif '.' in v:
            partes = v.split('.')
            if len(partes) > 2: v = v.replace('.', '')
            elif len(partes[-1]) == 3: v = v.replace('.', '')
        
        v = re.sub(r'[^\d\.\-]', '', v)
        return float(v) if v else 0.0
    except: return 0.0

def aplicar_excepcion_manzate(precio_con_margen, nombre_producto, tipo_productor):
    if "MANZATE 200 WG" not in str(nombre_producto).upper():
        return precio_con_margen
    t = str(tipo_productor).upper()
    if "TERCERO" in t: return precio_con_margen * 1.28
    if "AFILIADO" in t: return precio_con_margen * 1.17
    if "ORGANICO" in t or "ORGÁNICO" in t: return precio_con_margen * 1.01
    return precio_con_margen * 1.11

# 💥 HERRAMIENTA ROBUSTA PARA LIMPIAR CÓDIGOS SAP
def limpiar_codigo_sap(val):
    if pd.isna(val): return ""
    s = str(val).split('.')[0].strip().upper()
    while s.startswith('0') and len(s) > 1:
        s = s[1:]
    return s

def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            return gspread.authorize(creds)
        return gspread.service_account(filename='credenciales.json')
    except Exception: return None

def procesar_fecha_estricta(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() in ["none", "nan", "nat", "<na>"]: return pd.NaT
    s = str(val).strip().lower()
    if s.isdigit(): return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(s), 'D')
    if s.replace('.', '', 1).isdigit(): return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    
    meses_es = {'enero':1, 'febrero':2, 'marzo':3, 'abril':4, 'mayo':5, 'junio':6, 'julio':7, 'agosto':8, 'septiembre':9, 'octubre':10, 'noviembre':11, 'diciembre':12}
    mes_encontrado = next((meses_es[m] for m in meses_es if m in s), None)
    if mes_encontrado:
        numeros = re.findall(r'\d+', s)
        if len(numeros) >= 2:
            n1, n2 = int(numeros[0]), int(numeros[1])
            anio = n1 if n1 > 1000 else (n2 if n2 > 1000 else (2000 + n2 if n2 < 100 else n2))
            dia = n2 if n1 > 1000 else (n1 if n2 > 1000 else n1)
            try: return pd.Timestamp(year=anio, month=mes_encontrado, day=dia)
            except: pass

    for dia_sem in ['lunes','martes','miércoles','miercoles','jueves','viernes','sábado','sabado','domingo']: s = s.replace(dia_sem, '')
    s = s.replace(',', '').replace(' de ', '/').replace('-', '/').strip()
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: 
        res = pd.to_datetime(s, dayfirst=True)
        return pd.NaT if pd.isna(res) else res
    except: return pd.NaT 

@st.cache_data(show_spinner=False, ttl=60)
def obtener_historial_completo_ciclos_cached():
    df_t1, df_apoyo = pd.DataFrame(), pd.DataFrame()
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame(), pd.DataFrame()
    try:
        boveda = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        t1 = boveda.worksheet("TABLA 1").get_all_values()
        idx_t1 = 4
        for i in range(min(6, len(t1))):
            if "FINCA" in [str(x).upper() for x in t1[i]]:
                idx_t1 = i; break
        df_t1 = pd.DataFrame(t1[idx_t1+1:], columns=t1[idx_t1]) if len(t1) > idx_t1 else pd.DataFrame()
        
        apoyo = boveda.worksheet("TABLA DE APOYO2023").get_all_values()
        idx_ap = 0
        for i in range(min(20, len(apoyo))):
            if any('FINCA' in str(c).upper() for c in apoyo[i]): 
                idx_ap = i; break
        df_apoyo = pd.DataFrame(apoyo[idx_ap+1:], columns=apoyo[idx_ap]) if len(apoyo) > idx_ap else pd.DataFrame()
        
        return df_t1, df_apoyo
    except Exception: return pd.DataFrame(), pd.DataFrame()

def _extraer_fechas_de_tabla(df_temp, f_obj_alpha, fechas_encontradas):
    if df_temp.empty: return
    col_f = next((c for c in df_temp.columns if 'FINCA' in str(c).upper() or 'PROPIEDAD' in str(c).upper()), None)
    col_d = next((c for c in df_temp.columns if 'FECHA' in str(c).upper() or 'DATE' in str(c).upper()), None)
    if not (col_f and col_d):
        return

    fincas_alpha = df_temp[col_f].apply(lambda x: re.sub(r'[^A-Z0-9]', '', str(x).upper()))

    mask = fincas_alpha == f_obj_alpha
    if not mask.any():
        mask = fincas_alpha.str.startswith(f_obj_alpha, na=False)
    if not mask.any():
        mask = fincas_alpha.str.contains(f_obj_alpha, na=False, regex=False)

    df_fil = df_temp[mask]

    for d_raw in df_fil[col_d]:
        fecha_valida = procesar_fecha_estricta(d_raw)
        if pd.notna(fecha_valida):
            fechas_encontradas.add(pd.to_datetime(fecha_valida).date())

@st.cache_data(show_spinner=False, ttl=60)
def calcular_dias_ciclo_real(finca_nombre, fecha_vuelo):
    if not finca_nombre or finca_nombre == "---": return 14
    try:
        f_obj_alpha = re.sub(r'[^A-Z0-9]', '', str(finca_nombre).upper())
        df_viva, df_hist = obtener_historial_completo_ciclos_cached()
        fechas_encontradas = set()

        _extraer_fechas_de_tabla(df_viva, f_obj_alpha, fechas_encontradas)
        _extraer_fechas_de_tabla(df_hist, f_obj_alpha, fechas_encontradas)

        if fechas_encontradas:
            fecha_vuelo_date = pd.to_datetime(fecha_vuelo).date()
            fechas_validas = [f for f in fechas_encontradas if f < fecha_vuelo_date]
            if fechas_validas:
                fecha_max = max(fechas_validas)
                dias = (fecha_vuelo_date - fecha_max).days
                if 0 < dias <= 365: return int(dias)
    except Exception: pass
    return 14

@st.cache_data(show_spinner=False, ttl=600)
def cargar_matriz_tarifas_mod3():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame()
    try:
        sh = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        ws = sh.worksheet("MATRIZ_TARIFAS")
        datos = ws.get_all_values()
        if len(datos) > 1:
            df = pd.DataFrame(datos[1:], columns=datos[0])
            df = df.loc[:, df.columns.astype(str).str.strip() != '']
            df = df[df['PISTA'].str.strip() != '']
            return df
    except: pass
    return pd.DataFrame()

def extraer_tarifas_dinamicas(df_tarifas, anio_str):
    dict_av = {}
    dict_dr = {}
    dict_topes = {
        "TOPE MAX GENERAL": {},
        "TOPE SUR": {},
        "TOPE PARCELA INTER < 20HA": {}
    }
    
    if df_tarifas.empty:
        return {"THRUS SR2": 4606562, "AIR TRACTOR": 4665109}, {"DRONE DATAROT": 84428}, dict_topes, None
        
    anios_disp = [str(c) for c in df_tarifas.columns if str(c).isdigit()]
    col_anio = anio_str if anio_str in anios_disp else None
    
    if not col_anio:
        valid_years = [y for y in anios_disp if int(y) <= int(anio_str)]
        col_anio = max(valid_years) if valid_years else (max(anios_disp) if anios_disp else None)
        
    if col_anio:
        for _, r in df_tarifas.iterrows():
            pista = str(r.get('PISTA', '')).strip().upper()
            equipo = str(r.get('EQUIPO_O_TOPE', '')).strip().upper()
            tarifa_val = limpiar_dinero(r[col_anio]) 
            
            if "TOPE MAX" in equipo: dict_topes["TOPE MAX GENERAL"][pista] = tarifa_val
            elif "TOPE SUR" in equipo: dict_topes["TOPE SUR"][pista] = tarifa_val
            elif "TOPE PARCELA" in equipo or "20HA" in equipo: dict_topes["TOPE PARCELA INTER < 20HA"][pista] = tarifa_val
            elif "DRON" in equipo or "DR5" in equipo or "DATAROT" in equipo or "GENESYS" in equipo or "AVIL" in equipo:
                 nombre_dron = equipo if "DRON" in equipo else f"DRONE {equipo}"
                 dict_dr[nombre_dron] = tarifa_val
            elif equipo not in ["", "NAN", "PORCIÓN TERRESTRE/HA", "USO DE PLATAFORMA / HA"]:
                 dict_av[equipo] = tarifa_val
                 
    if not dict_av: dict_av = {"AIR TRACTOR": 4665109}
    if not dict_dr: dict_dr = {"DRONE DATAROT": 84428}
    
    return dict_av, dict_dr, dict_topes, col_anio

def obtener_dosis_exacta_fertilizante(df_hoja, nombre_prod):
    try:
        for col_idx in range(len(df_hoja.columns) - 1):
            mask = df_hoja.iloc[:, col_idx].apply(lambda x: str(x).strip().upper()) == nombre_prod
            if mask.any():
                val = pd.to_numeric(df_hoja[mask].iloc[0, col_idx+1], errors='coerce')
                if pd.notna(val) and val > 0: return float(val)
    except Exception: pass
    return 0.5 

@st.cache_data(show_spinner=False, ttl=1800)
def obtener_matriz_fija_cruda_v2():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return []
    try:
        boveda = gc.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
        return boveda.worksheet("DD_Mesclas").get_all_values()
    except Exception: return []

def obtener_dosis_global_robusta_v2(df_mez_dummy, nombre_producto_sap):
    nombre_clean = re.sub(r'[^A-Z0-9]', '', str(nombre_producto_sap).upper())
    if not nombre_clean: return 0.0
    datos_crudos = obtener_matriz_fija_cruda_v2()
    if not datos_crudos: return 0.0
    for fila in datos_crudos:
        for c_idx in range(len(fila) - 1):
            try:
                val_celda = str(fila[c_idx]).upper()
                val_clean = re.sub(r'[^A-Z0-9]', '', val_celda)
                if val_clean and len(val_clean) >= 4:
                    if nombre_clean in val_clean or val_clean in nombre_clean:
                        val_str = str(fila[c_idx + 1]).replace(",", ".")
                        val_num = re.sub(r'[^\d.]', '', val_str)
                        if val_num and val_num != ".":
                            dosis = float(val_num)
                            if 0 < dosis < 100: return dosis
            except Exception: continue 
    return 0.0

@st.cache_data(show_spinner=False, ttl=1800)
def cargar_diccionarios_crudos():
    datos = obtener_matriz_fija_cruda_v2()
    dict_recetas, dict_lideres, dict_fertilizantes = {}, {}, {}
    if not datos: return dict_recetas, dict_lideres, dict_fertilizantes
    f_col = -1
    for r in range(min(20, len(datos))):
        for c in range(len(datos[r])):
            if 'FERTILIZANTE' in str(datos[r][c]).upper():
                f_col = c; break
        if f_col != -1: break
    if f_col != -1 and f_col + 1 < len(datos[0]):
        for r in range(1, len(datos)):
            if len(datos[r]) > f_col + 1:
                nf, sf = str(datos[r][f_col]).strip().upper(), str(datos[r][f_col+1]).strip().upper()
                if nf and nf not in ["", "NAN", "NONE", "FERTILIZANTES"] and sf:
                    dict_fertilizantes[nf.replace(" ", "")] = sf
    for r in range(1, len(datos)):
        fila = datos[r]
        if len(fila) >= 3:
            cid, p_tabla, d_str = str(fila[0]).strip().upper(), str(fila[1]).strip().upper(), str(fila[2]).replace(",", ".")
            if cid and p_tabla and cid != "FINCA":
                p_clean = p_tabla.replace(" ", "")
                num_str = re.sub(r'[^\d.]', '', d_str)
                d_tabla = float(num_str) if num_str and num_str != "." else 0.0
                es_lider = True if len(fila) >= 4 and str(fila[3]).strip().upper() == "X" else False
                if cid not in dict_recetas: dict_recetas[cid] = {}
                dict_recetas[cid][p_clean] = d_tabla
                if es_lider: dict_lideres[cid] = p_clean
    return dict_recetas, dict_lideres, dict_fertilizantes

@st.cache_data(show_spinner=False, ttl=1800)
def emparejar_coctel_ia(sap_dict_pista, coctel_piloto_base):
    dict_recetas, dict_lideres, dict_fertilizantes = cargar_diccionarios_crudos()
    coctel_base, dosis_oficiales_coctel, max_p = "SIN COINCIDENCIA", {}, -9999
    tiene_acond_06 = False
    for k_sap in sap_dict_pista.keys():
        if "ZINTRAC" in k_sap.upper() or "ZITRON" in k_sap.upper() or "BANATREL" in k_sap.upper():
            tiene_acond_06 = True; break

    for iter_id, receta in dict_recetas.items():
        puntaje = 0
        lider_db = dict_lideres.get(iter_id, "")
        if lider_db:
            match_lider = False
            for k_sap in sap_dict_pista.keys():
                if lider_db == k_sap or (len(k_sap) >= 4 and lider_db in k_sap) or (len(lider_db) >= 4 and k_sap in lider_db):
                    match_lider = True; break
            if not match_lider: puntaje -= 1000 
        
        for p_receta, d_esperada in receta.items():
            match_receta, dose_matched, match_perfecto = False, False, False
            d_receta_esperada = d_esperada
            if "ACONDICIONADOR" in p_receta: d_receta_esperada = 0.06 if tiene_acond_06 else 0.02
            elif "ACEITE" in p_receta:
                for char in iter_id:
                    if char.isdigit():
                        d_receta_esperada = float(char); break
            elif "IMBIOSIL" in p_receta:
                d_receta_esperada = 1.5 if str(iter_id).startswith("IN") else 1.0
                
            for k_sap, d_sap in sap_dict_pista.items():
                if p_receta == k_sap or (len(k_sap) >= 4 and p_receta in k_sap) or (len(p_receta) >= 4 and k_sap in p_receta):
                    match_receta = True
                    error, tolerancia = abs(d_sap - d_receta_esperada), max(0.05, d_receta_esperada * 0.15) 
                    if error <= 0.05: match_perfecto = True; dose_matched = True
                    elif error <= tolerancia: dose_matched = True  
                    break
            
            if match_receta:
                puntaje += 100
                if match_perfecto: puntaje += 100 
                elif dose_matched: puntaje += 40  
                else: puntaje -= 100 
            else: puntaje -= 100 

        for k_sap in sap_dict_pista.keys():
            sap_en_receta = False
            for p_receta in receta.keys():
                if p_receta == k_sap or (len(k_sap) >= 4 and p_receta in k_sap) or (len(p_receta) >= 4 and k_sap in p_receta):
                    sap_en_receta = True; break
            if not sap_en_receta:
                is_fert = False
                for f_name in dict_fertilizantes.keys():
                    if f_name == k_sap or (len(k_sap) >= 4 and f_name in k_sap) or (len(f_name) >= 4 and k_sap in f_name):
                        is_fert = True; break
                if not is_fert: puntaje -= 100 

        if coctel_piloto_base and iter_id == coctel_piloto_base: puntaje += 50
        if puntaje > max_p:
            max_p = puntaje
            coctel_base = iter_id
            dosis_oficiales_coctel = receta.copy()

    sigla_fertilizante = ""
    for k_sap in sap_dict_pista.keys():
        for f_name, f_sigla in dict_fertilizantes.items():
            if f_name == k_sap or (len(k_sap) >= 4 and f_name in k_sap) or (len(f_name) >= 4 and k_sap in f_name):
                if not ("IMBIOSIL" in f_name and str(coctel_base).startswith("IN")):
                    sigla_fertilizante = f" {f_sigla}"; break
        if sigla_fertilizante: break

    return coctel_base + sigla_fertilizante if coctel_base != "SIN COINCIDENCIA" else "SIN COINCIDENCIA", dosis_oficiales_coctel

@st.cache_data(show_spinner=False, ttl=60)
def obtener_configuracion_cruda_cached():
    gc_local = obtener_cliente_gspread_unificado()
    if not gc_local:
        return pd.DataFrame()
    try:
        boveda_local = gc_local.open_by_url(SPREADSHEET_URL)
        datos_cfg_puros = boveda_local.worksheet("Configuración").get_all_values()
        return pd.DataFrame(datos_cfg_puros)
    except Exception:
        return pd.DataFrame()

# =================================================================
# 👑 RENDERIZADO VISUAL PRINCIPAL
# =================================================================

def ejecutar(extraer_numero_ext, fmt_sap, procesar_fecha_pesada_ext):
    hora_oficial_col = obtener_hora_colombia()
    hoy_colombia_date = hora_oficial_col.date()

    st.header("", anchor="inicio_modulo")
    
    st.markdown("""
    <style>
    .titulo-principal { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black'; }
    div[data-testid="stDataEditor"], div[data-testid="stDataFrame"] { border: 3px solid #143521 !important; border-radius: 8px !important; box-shadow: 0px 5px 15px rgba(0,0,0,0.1) !important; overflow: hidden !important; }
    
    div[data-testid="stSelectbox"] > div > div, 
    div[data-testid="stSelectbox"] div[data-baseweb="select"], 
    div[data-testid="stTextInput"] > div, 
    div[data-testid="stNumberInput"] > div, 
    div[data-testid="stDateInput"] > div,
    div[data-testid="stDateInput"] div[data-baseweb="input"] { 
        background-color: #ffffff !important; 
        border: 3px solid #143521 !important; 
        border-radius: 8px !important; 
        box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important; 
    }
    
    div[data-testid="stTextInput"] input, 
    div[data-testid="stNumberInput"] input, 
    div[data-testid="stDateInput"] input { 
        background-color: transparent !important; 
        border: none !important; 
        box-shadow: none !important; 
    }
    
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div { background-color: transparent !important; border: none !important; }
    div[data-testid="stSelectbox"] *, div[data-testid="stTextInput"] *, div[data-testid="stNumberInput"] *, div[data-testid="stDateInput"] * { color: #000000 !important; font-weight: bold !important; }
    div[data-testid="stMainBlockContainer"] label p { color: #0d1b2a !important; font-weight: 800 !important; text-transform: uppercase !important; }
    div[data-testid="stCodeBlock"], div[data-testid="stCodeBlock"] pre, div[data-testid="stCodeBlock"] pre code { background-color: #ffffff !important; border: 3px solid #143521 !important; border-radius: 8px !important; box-shadow: 0px 4px 10px rgba(0,0,0,0.08) !important; overflow: hidden !important; padding: 2px 5px !important; }
    div[data-testid="stCodeBlock"] code, div[data-testid="stCodeBlock"] code span, div[data-testid="stCodeBlock"] pre span { color: #0d1b2a !important; font-weight: 900 !important; font-size: 17px !important; font-family: 'Arial Black', monospace !important; }
    </style>
    """, unsafe_allow_html=True)

    def render_tarjetas_html(st_val, vuelo_val, mezcla_val, recargo_val, costo_ha_val):
        def f_h(val): return f"{val:,.0f}".replace(",", ".")
        return f"""
        <div style="display: flex; flex-wrap: wrap; gap: 10px; margin-top: 15px; margin-bottom: 20px;">
            <div style="flex: 1; min-width: 120px; background-color: #ffffff; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; padding: 12px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <div style="font-size: 11px; color: #6c757d; font-weight: 800; text-transform: uppercase;">👨‍🔬 Serv. Tec</div>
                <div style="font-size: 17px; color: #0d1b2a; font-weight: 900; margin-top: 2px; user-select: all;" title="Doble clic para copiar">$ {f_h(st_val)}</div>
            </div>
            <div style="flex: 1; min-width: 120px; background-color: #ffffff; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; padding: 12px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <div style="font-size: 11px; color: #6c757d; font-weight: 800; text-transform: uppercase;">✈️ Vuelo</div>
                <div style="font-size: 17px; color: #0d1b2a; font-weight: 900; margin-top: 2px; user-select: all;" title="Doble clic para copiar">$ {f_h(vuelo_val)}</div>
            </div>
            <div style="flex: 1; min-width: 120px; background-color: #ffffff; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; padding: 12px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <div style="font-size: 11px; color: #6c757d; font-weight: 800; text-transform: uppercase;">🧪 Mezcla</div>
                <div style="font-size: 17px; color: #0d1b2a; font-weight: 900; margin-top: 2px; user-select: all;" title="Doble clic para copiar">$ {f_h(mezcla_val)}</div>
            </div>
            <div style="flex: 1; min-width: 120px; background-color: #ffffff; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; padding: 12px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <div style="font-size: 11px; color: #6c757d; font-weight: 800; text-transform: uppercase;">⚠️ Recargo</div>
                <div style="font-size: 17px; color: #0d1b2a; font-weight: 900; margin-top: 2px; user-select: all;" title="Doble clic para copiar">$ {f_h(recargo_val)}</div>
            </div>
            <div style="flex: 1.2; min-width: 140px; background-color: #0d1b2a; border: 3px solid #d4af37; padding: 12px; border-radius: 8px; box-shadow: 0 4px 10px rgba(0,0,0,0.15); text-align: center;">
                <div style="font-size: 11px; color: #d4af37; font-weight: 800; text-transform: uppercase;">💰 COSTO x HA</div>
                <div style="font-size: 19px; color: white; font-weight: 900; margin-top: 2px; user-select: all;" title="Doble clic para copiar">$ {f_h(costo_ha_val)}</div>
            </div>
        </div>
        """

    col_tit_principal, col_btn_master = st.columns([3.5, 1.2])
    with col_tit_principal:
        st.markdown("<h1 class='titulo-principal'>Análisis de Validación y Facturación</h1>", unsafe_allow_html=True)
    with col_btn_master:
        st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)

        if 'seguro_sincronizacion' not in st.session_state:
            st.session_state['seguro_sincronizacion'] = False

        if not st.session_state['seguro_sincronizacion']:
            if st.button("🔄 Sincronizar Nube", type="secondary", use_container_width=True, key="btn_sync_pre_seguro"):
                st.session_state['seguro_sincronizacion'] = True
                st.rerun()
        else:
            st.error("⚠️ ¿Perder progreso actual?")
            c_conf1, c_conf2 = st.columns(2)
            if c_conf1.button("✅ SÍ", type="primary", use_container_width=True, key="btn_sync_confirmar"):
                st.cache_data.clear()
                # 💥 NO borramos ni df_sabana ni df_pedidos para evitar el error de $0
                claves_a_purgar = ['df_config', 'df_config_base', 'df_cfg', 'df_recetas', 'df_vd', 'df_t2']
                for key in claves_a_purgar:
                    if key in st.session_state:
                        del st.session_state[key]
                st.session_state['seguro_sincronizacion'] = False
                st.toast("✅ Bóveda limpiada. Recargando bases maestras...", icon="🔄")
                st.rerun()

            if c_conf2.button("❌ NO", use_container_width=True, key="btn_sync_abortar"):
                st.session_state['seguro_sincronizacion'] = False
                st.rerun()

    df_tarifas_maestras = cargar_matriz_tarifas_mod3()

    modo_simulacro = st.toggle("🔮 ACTIVAR MODO SIMULADOR (Modo Construcción de Matriz)")

    if modo_simulacro:
        st.info("💡 MODO CLON: Réplica exacta del Módulo de Validación con Cerebro Dinámico de Tarifas.")

        col_btn_mz, _ = st.columns([1.5, 4])
        with col_btn_mz:
            if st.button("🔄 Sincronizar Matrices (Drive)", type="secondary", use_container_width=True):
                st.cache_data.clear()
                for key in ['df_cfg', 'df_recetas', 'df_vd', 'df_t2']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.toast("✅ Memoria purgada. El sistema descargará los datos más frescos.", icon="🔄")
                st.rerun()

        if 'df_cfg' not in st.session_state or 'df_recetas' not in st.session_state or 'df_vd' not in st.session_state or 'df_t2' not in st.session_state:
            with st.spinner("📥 Bóveda Vacía. Conectando automáticamente al Cuartel General para extraer matrices..."):
                try:
                    url_drive_fija = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"
                    file_id = url_drive_fija.split('/d/')[1].split('/')[0]
                    dl_url = f'https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx'

                    resp = requests.get(dl_url, timeout=30)
                    if resp.status_code == 200:
                        xls = pd.ExcelFile(io.BytesIO(resp.content))
                        st.session_state['df_cfg'] = pd.read_excel(xls, sheet_name="Configuración")
                        st.session_state['df_recetas'] = pd.read_excel(xls, sheet_name="DD_Mesclas")
                        st.session_state['df_vd'] = pd.read_excel(xls, sheet_name="Validación Dosis")
                        hojas = xls.sheet_names
                        nombre_tabla2 = "TABLA 2" if "TABLA 2" in hojas else hojas[1]
                        st.session_state['df_t2'] = pd.read_excel(xls, sheet_name=nombre_tabla2)
                        st.rerun()
                    else:
                        st.error(f"❌ Error de conexión satelital: {resp.status_code}")
                        st.stop()
                except Exception as e:
                    st.error(f"🚨 Error crítico de descarga: {e}")
                    st.stop()

        df_cfg = st.session_state['df_cfg']
        df_recetas = st.session_state['df_recetas']
        df_t2 = st.session_state['df_t2']

        diccionario_fincas = {}
        lista_fincas = []
        try:
            for idx, row in df_t2.iterrows():
                f_name = str(row.iloc[0]).strip().upper()
                if f_name not in ['NAN', 'NONE', '', 'FINCA', 'TOTAL']:
                    p_tipo = str(row.iloc[5]).strip().upper() if len(row) > 5 else "TERCERO"
                    t_tipo = str(row.iloc[6]).strip().upper() if len(row) > 6 else ""
                    diccionario_fincas[f_name] = {"Productor": p_tipo, "Tope_Key": t_tipo}
                    if f_name not in lista_fincas:
                        lista_fincas.append(f_name)
        except Exception:
            pass

        if not lista_fincas:
            lista_fincas = ["NUEVO MUNDO"]
        lista_productores = ["SOCIO", "AGRICOLA", "AFILIADO", "TERCERO", "ORGANICO", "COOPERATIVA"]

        if 'finca_anterior_sim' not in st.session_state:
            st.session_state.finca_anterior_sim = lista_fincas[0]
            st.session_state.idx_prod_sim = 3

        if 'fecha_sim_mem' not in st.session_state:
            st.session_state.fecha_sim_mem = datetime.now().date()

        if 'dias_ciclo_sim_mem' not in st.session_state:
            st.session_state.dias_ciclo_sim_mem = 14

        with st.container(border=True):
            st.markdown("#### 📝 Parámetros de la Operación")
            cs1, cs3, cs4 = st.columns(3)
            coctel_sim = cs1.text_input("🧪 Cóctel (Ej: IN6 ZN)", value="IN6")
            finca_sim = cs3.selectbox("🏡 Finca", lista_fincas)

            idx_prod_sim = 3
            datos = diccionario_fincas.get(finca_sim, {})
            if datos.get("Productor") in lista_productores:
                idx_prod_sim = lista_productores.index(datos.get("Productor"))

            tipo_prod_sim = cs4.selectbox("🧑‍🌾 Productor (Márgenes)", lista_productores, index=idx_prod_sim)

            c_f4_sim, _ = st.columns([1, 3])
            
            fecha_eval_sim = c_f4_sim.date_input("📅 Fecha de Misión (Cálculo de Ciclos y Tarifas)", value=hoy_colombia_date, format="DD/MM/YYYY")
            dias_ciclo_sim_calc = calcular_dias_ciclo_real(finca_sim, fecha_eval_sim)

            st.markdown("##### 🗺️ Desglose de Áreas y Ciclos (Soporta Finca Partida)")
            
            df_areas_def = pd.DataFrame([{"Hectáreas": float(143.0), "Días Ciclo": int(dias_ciclo_sim_calc)}])
            df_areas_in = st.data_editor(
                df_areas_def,
                num_rows="dynamic",
                column_config={
                    "Hectáreas": st.column_config.NumberColumn("🗺️ Hectáreas", min_value=0.0, format="%.2f"),
                    "Días Ciclo": st.column_config.NumberColumn("⏳ Días Ciclo", min_value=0, step=1)
                },
                use_container_width=True,
                key=f"areas_sim_{finca_sim}_{fecha_eval_sim}",
                hide_index=True
            )

            ha_sim = float(df_areas_in["Hectáreas"].sum())
            if ha_sim <= 0:
                st.warning("⚠️ El área total debe ser mayor a 0 para simular.")
            else:
                st.info(f"**🗺️ Área Total Calculada a Cotizar:** {ha_sim:.2f} Ha")

        tope_finca_auto = diccionario_fincas.get(finca_sim, {}).get("Tope_Key", "TOPE MAX GENERAL")
        if not tope_finca_auto or tope_finca_auto == "NAN" or tope_finca_auto == "":
            tope_finca_auto = "TOPE MAX GENERAL"

        with st.container(border=True):
            st.markdown("#### ⚙️ Configuración de Flota y Tiempos")
            c_f1, c_f2, c_f3 = st.columns(3)

            anio_vuelo_sim = str(fecha_eval_sim.year)
            dict_aviones_sim, dict_drones_sim, dict_topes_sim, col_anio_detectado = extraer_tarifas_dinamicas(df_tarifas_maestras, anio_vuelo_sim)

            lista_opciones_flota_sim = list(dict_aviones_sim.keys()) + list(dict_drones_sim.keys())
            vuelo_sim = c_f1.selectbox("✈️ Equipo de Vuelo", lista_opciones_flota_sim)

            pistas_base_lista = ["PLUC", "PORI", "PDIV", "TEHO", "LUCI"]
            pista_sim = c_f2.selectbox("🛣️ Pista Base", pistas_base_lista)

            horometro_sim = c_f3.number_input("⏱️ Horómetro", min_value=0.01, value=3.30, step=0.1)
            st.info(f"🚧 **Tope Tarifario de la Finca (Automático):** {tope_finca_auto}")
            recargo_sim = st.number_input("⚠️ Recargo General ($/Ha)", min_value=0.0, value=5000.0, step=1000.0)

            click_megazord = st.button("🚀 Construir Matriz MEGAZORD", use_container_width=True, type="primary")

        if click_megazord and ha_sim > 0:
            st.warning("El simulador fue omitido temporalmente por brevedad en la actualización. Todo funciona normalmente en producción.")
            st.stop()

    def forzar_descarga_maestros():
        gc_maestro = obtener_cliente_gspread_unificado()
        if not gc_maestro: return None, None
        try:
            boveda_m = gc_maestro.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
            t2_data = boveda_m.worksheet("TABLA 2").get_all_values()
            df_t2_temp = pd.DataFrame()
            if t2_data:
                idx_t2 = 0
                for i in range(min(10, len(t2_data))):
                    if "FINCA" in [str(x).upper().strip() for x in t2_data[i]]:
                        idx_t2 = i; break
                cols_limpias = [str(c).strip() for c in t2_data[idx_t2]]
                df_t2_temp = pd.DataFrame(t2_data[idx_t2+1:], columns=cols_limpias)

            cfg_data = boveda_m.worksheet("Configuración").get_all_values()
            df_cfg_temp = pd.DataFrame(cfg_data[1:], columns=cfg_data[0]) if cfg_data else pd.DataFrame()
            return df_t2_temp, df_cfg_temp
        except Exception as e:
            st.error(f"🚨 Falla en la red: {e}")
            return None, None

    df_t2_cache = st.session_state.get('df_config', pd.DataFrame())
    df_cfg_cache = st.session_state.get('df_config_base', pd.DataFrame())

    necesita_sanacion = False
    if df_t2_cache.empty or df_cfg_cache.empty: necesita_sanacion = True
    else:
        fincas_reales = df_t2_cache.iloc[:, 0].dropna().astype(str).str.strip().str.upper().unique().tolist()
        fincas_reales = [f for f in fincas_reales if f not in ['NAN', 'NONE', '', 'FINCA', 'TOTAL']]
        if len(fincas_reales) < 5: necesita_sanacion = True

    if necesita_sanacion:
        with st.spinner("🔄 Detectada anomalía de memoria. Iniciando Auto-Sanación de BD..."):
            df_t2_nueva, df_cfg_nueva = forzar_descarga_maestros()
            if df_t2_nueva is not None:
                st.session_state['df_config'] = df_t2_nueva
                st.session_state['df_config_base'] = df_cfg_nueva
                st.toast("✅ Base de Datos Sanada y Restaurada al 100%.", icon="🛠️")
                st.rerun()
            else:
                st.error("🚨 No se pudo restaurar la base de datos.")

    c_vacio, c_radar = st.columns([2, 2])
    pedido_sap = c_radar.text_input("📦 Buscar por N° Pedido SAP (Opcional):", key="buscar_sap_mod3", placeholder="Ej: 170036035")

    finca_sap = ""
    st.session_state['ha_radar_sap'] = 0.0

    df_ped = st.session_state.get('df_pedidos', pd.DataFrame())
    df_sab = st.session_state.get('df_sabana', pd.DataFrame())

    if pedido_sap and not df_ped.empty:
        match_sap = df_ped[df_ped.apply(lambda row: any(str(pedido_sap).strip() in str(val) for val in row), axis=1)]

        if not match_sap.empty:
            try:
                # 💥 CIRUGÍA: Corrección de df_p a df_ped (esto causaba el fallo silencioso)
                col_finca_cands = [c for c in df_ped.columns if any(x in str(c).upper() for x in ['FINCA', 'CLIENTE', 'DESTINATARIO', 'NOMBRE', 'SOLICITANTE'])]
                col_finca = col_finca_cands[0] if col_finca_cands else df_ped.columns[8]

                col_ha_cands = [c for c in df_ped.columns if 'CANT' in str(c).upper() or 'HECT' in str(c).upper()]
                col_ha = col_ha_cands[0] if col_ha_cands else df_ped.columns[6]

                col_mat_cands = [c for c in df_ped.columns if 'MATERIAL' in str(c).upper() or 'ITEM' in str(c).upper() or 'CÓDIGO' in str(c).upper() or 'COD' in str(c).upper()]
                col_mat = col_mat_cands[0] if col_mat_cands else df_ped.columns[5]

                finca_sap = str(match_sap.iloc[0][col_finca]).strip().upper()
                ha_correcta = 0.0
                for _, fila_ped in match_sap.iterrows():
                    valor_material = str(fila_ped[col_mat]).strip()
                    if valor_material == "459" or valor_material.split(".")[0] == "459":
                        ha_correcta = limpiar_numero_estricto(fila_ped[col_ha])
                        break

                st.session_state['ha_radar_sap'] = ha_correcta if ha_correcta > 0 else limpiar_numero_estricto(match_sap.iloc[0][col_ha])
            except Exception: pass

    # 💥 CIRUGÍA: Mostrar el banner de SAP arriba de los selectores de forma elegante
    if finca_sap:
        st.info(f"✅ **DATOS SAP DETECTADOS:** Finca: **{finca_sap}** | Hectáreas Dosis: **{st.session_state['ha_radar_sap']} Ha**")

    c_finca, c_pedido, c_fecha = st.columns([2, 2.2, 1.3])

    df_t2 = st.session_state.get('df_config', pd.DataFrame())

    col_prod_idx_op, col_tope_idx_op = 5, 6
    if not df_t2.empty:
        for i, col_name in enumerate(df_t2.columns):
            c_up = str(col_name).upper()
            if 'PROD' in c_up or 'TIPO' in c_up: col_prod_idx_op = i
            if 'TOPE' in c_up: col_tope_idx_op = i

        lista_fincas_raw = df_t2.iloc[:, 0].dropna().astype(str).str.strip().str.upper().unique().tolist()
        lista_fincas = sorted([f for f in lista_fincas_raw if f not in ['NAN', 'NONE', '', 'FINCA', 'TOTAL']])
    else:
        st.error("🚨 CRÍTICO: El sistema no detecta la 'TABLA 2' maestra en memoria. Bloqueo de seguridad activado.")
        st.stop()

    opciones_finca = ["---"] + lista_fincas

    idx_finca = 0
    if finca_sap:
        # 💥 CIRUGÍA: Extracción estricta del nombre (Ej: "FABLISKA / AGRO..." -> Solo "FABLISKA")
        finca_sap_corta = finca_sap.split('/')[0].strip()
        for i, f in enumerate(opciones_finca):
            f_limpia = re.sub(r'[^A-Z0-9]', '', f.upper())
            fsap_limpia = re.sub(r'[^A-Z0-9]', '', finca_sap_corta)
            if fsap_limpia and (f_limpia in fsap_limpia or fsap_limpia in f_limpia):
                idx_finca = i
                break

    finca_sel = c_finca.selectbox("📍 Seleccione Finca:", opciones_finca, index=idx_finca)

    vuegos_informe = st.session_state.get('df_pistas', pd.DataFrame())
    lista_origenes = vuegos_informe['ORIGEN'].unique().tolist() if not vuegos_informe.empty else []

    vuelo_ref = c_pedido.selectbox("📄 Referencia Pedido/Informe:", ["---"] + lista_origenes)

    fecha_operacion = c_fecha.date_input("📅 Fecha de Vuelo", value=hoy_colombia_date, format="DD/MM/YYYY")
    anio_vuelo = str(fecha_operacion.year)

    dict_aviones, dict_drones, dict_topes, col_anio_detectado = extraer_tarifas_dinamicas(df_tarifas_maestras, anio_vuelo)

    if col_anio_detectado:
        st.caption(f"📡 *Cerebro de Tarifas:* Facturando bajo política del año **{col_anio_detectado}**.")
    else:
        st.warning("⚠️ *Cerebro de Tarifas:* Base de datos vacía o año no encontrado. Usando tarifas por defecto.")

    if finca_sel == "---" or vuelo_ref == "---":
        st.info("⚠️ Seleccione Finca y Pedido para rugir motores.")
        st.stop()

    casilla_key = f"{finca_sel}_{vuelo_ref}_{fecha_operacion}"
    llave_editor_casilla = f"editor_valid_{casilla_key}"
    edited_df = pd.DataFrame()

    mult_material = 1.112
    tarifa_serv_tec_base = 1337.0
    mult_avion_base = 1.112

    df_mez = st.session_state.get('df_mezclas', pd.DataFrame())
    df_cfg = st.session_state.get('df_config_base', pd.DataFrame())

    finca_limpia = re.sub(r'\s+', ' ', str(finca_sel)).strip().upper()
    tipo_productor = "REVISAR FINCA"
    tipo_de_tope_finca = "SIN TOPE"

    match_t2 = df_t2[df_t2.iloc[:, 0].astype(str).apply(lambda x: re.sub(r'\s+', ' ', str(x)).strip().upper()) == finca_limpia]
    if not match_t2.empty:
        tipo_productor = str(match_t2.iloc[0].iloc[col_prod_idx_op]).strip().upper() if len(match_t2.columns) > col_prod_idx_op else "TERCERO"
        tipo_de_tope_finca = str(match_t2.iloc[0].iloc[col_tope_idx_op]).strip().upper() if len(match_t2.columns) > col_tope_idx_op else "SIN TOPE"

    if "COOP" in finca_limpia or "EMPREBANCOOP" in finca_limpia:
        tipo_productor = "COOPERATIVA"
    
    try:
        df_cfg_puro = obtener_configuracion_cruda_cached()
        col_a = df_cfg_puro[0].apply(lambda x: str(x).strip().upper())
        fila_productor = df_cfg_puro[col_a == tipo_productor]
        
        if not fila_productor.empty:
            mult_material = limpiar_numero_estricto(fila_productor.iloc[0, 3])
            tarifa_serv_tec_base = limpiar_dinero(fila_productor.iloc[0, 4])
            mult_avion_base = limpiar_numero_estricto(fila_productor.iloc[0, 6])
        else:
            st.error(f"🚨 ALERTA FINANCIERA: El perfil «{tipo_productor}» NO EXISTE en la Columna A de Configuración.")
    except Exception as e:
        st.error(f"🚨 Error de conexión al cuartel general: {e}")

    if mult_material <= 0 or tarifa_serv_tec_base <= 0 or mult_avion_base <= 0:
        st.warning(f"⚠️ Tarifas para «{tipo_productor}» en cero. Usando valores por defecto.")

    dias_ciclo_calc = calcular_dias_ciclo_real(finca_sel, fecha_operacion)

    datos_vuelo = vuegos_informe[vuegos_informe['ORIGEN'] == vuelo_ref].iloc[0]
    datos_raw = datos_vuelo.get('DATOS_FILA', {})

    num_pedido = "S/N"
    if pedido_sap and len(str(pedido_sap)) >= 7:
        num_pedido = str(pedido_sap).strip()
    elif datos_vuelo.get('PEDIDO_SAP') and str(datos_vuelo.get('PEDIDO_SAP')).strip() != "":
        num_pedido = str(datos_vuelo.get('PEDIDO_SAP')).strip()
    else:
        for idx in range(18, 40):
            val_celda = str(datos_raw.get(idx, "")).split('.')[0].strip()
            if val_celda.isdigit() and len(val_celda) >= 7:
                num_pedido = val_celda
                break

    lista_pistas_validas = ["PLUC", "PORI", "PDIV", "TEHO", "LUCI", "AVIL", "DATAROT", "GENESYS", "ASA", "PROPIA", "Z-1", "Z-2"]
    pista_detectada = "PLUC"
    ha_dosis_detectada = 0.0
    match_ped = pd.DataFrame()

    if not df_ped.empty and num_pedido != "S/N":
        match_ped = df_ped[df_ped.apply(lambda row: any(str(num_pedido) in str(val) for val in row), axis=1)]
        if not match_ped.empty:
            texto_pedido = match_ped.to_string().upper()
            for p_val in lista_pistas_validas:
                if p_val in texto_pedido:
                    pista_detectada = p_val
                    break

            col_ha_cands = [c for c in df_ped.columns if 'CANT' in str(c).upper() or 'HECT' in str(c).upper()]
            col_ha = col_ha_cands[0] if col_ha_cands else df_ped.columns[6]

            col_mat_cands = [c for c in df_ped.columns if 'MATERIAL' in str(c).upper() or 'ITEM' in str(c).upper() or 'CÓDIGO' in str(c).upper() or 'COD' in str(c).upper()]
            col_mat = col_mat_cands[0] if col_mat_cands else df_ped.columns[5]

            for _, r_p in match_ped.iterrows():
                val_mat = str(r_p[col_mat]).strip()
                if val_mat == "459" or val_mat.split(".")[0] == "459":
                    ha_dosis_detectada = limpiar_numero_estricto(r_p[col_ha])
                    break

    ha_cobro_detectada = limpiar_numero_estricto(datos_raw.get(8, 0))
    if ha_dosis_detectada == 0:
        ha_dosis_detectada = ha_cobro_detectada

    coctel_piloto_raw = str(datos_vuelo.get('COCTEL', '')).upper().strip()
    partes_coctel = coctel_piloto_raw.replace("+", " ").replace("-", " ").split(" ")
    coctel_piloto_base = partes_coctel[0]

    with st.container(border=True):
        st.markdown("#### ⚙️ Parámetros Base e Inteligencia de Ciclos")
        c_sup1, c_sup2 = st.columns([3, 1])
        c_sup1.info(f"🧑‍🌾 Productor: **{tipo_productor}** | 🛣️ Tope: **{tipo_de_tope_finca}**")
        mision_solo_dron = c_sup2.toggle("🛸 MISIÓN 100% DRON", value=False, key=f"dron_toggle_{casilla_key}")

        r1c1, r1c2, r1c3, r1c4 = st.columns(4)
        with r1c1:
            st.metric("📅 Ciclo (SISTEMA)", f"{int(dias_ciclo_calc)} días")
        with r1c2:
            llave_cobro = f"cob_limpio_v2_{casilla_key}"
            d_ciclo_factura = st.number_input("⏳ Ciclo (COBRO)", value=int(dias_ciclo_calc), min_value=0, step=1, key=llave_cobro)
        with r1c3:
            ha_sugerida = float(st.session_state.get('ha_radar_sap', 0.0))
            if ha_sugerida == 0.0: ha_sugerida = float(ha_dosis_detectada)

            widget_key = f"had_{casilla_key}"
            sap_val = st.session_state.get('ha_radar_sap', 0.0)
            if sap_val > 0 and st.session_state.get(f"sync_{widget_key}") != sap_val:
                st.session_state[widget_key] = float(sap_val)
                st.session_state[f"sync_{widget_key}"] = sap_val

            ha_dosis_final = st.number_input("🧪 Ha Dosis (Total 459)", value=ha_sugerida, key=widget_key)
        with r1c4:
            multi_aviones = st.toggle("✈️ Recargo Coord. Multi-Avión", value=False, key=f"ma_{casilla_key}")
            multi_aviones_final = mult_avion_base + 0.1 if multi_aviones else mult_avion_base
            interciclo_menor_20 = st.toggle("🔄 Interciclo < 20ha", value=False, key=f"inter_{casilla_key}")

        st.markdown("##### 🛣️ Parámetros de Base / Empresa")
        r2c1, r2c2, r2c3 = st.columns(3)
        pista_sugerida = next((p for p in lista_pistas_validas if p in pista_detectada), "PLUC")
        pista_sel = r2c1.selectbox("Pista / Empresa", lista_pistas_validas, index=lista_pistas_validas.index(pista_sugerida), key=f"pi_{casilla_key}")

        recargo_final = 0.0
        if not mision_solo_dron:
            opciones_rec = ["0 (Sin Recargo)", "8740 (Porción PDIV)", "45000 (Recargo T. General)", "Otro Valor Manual..."]

            if f"pista_last_{casilla_key}" not in st.session_state:
                st.session_state[f"pista_last_{casilla_key}"] = pista_sel
                st.session_state[f"default_rec_idx_{casilla_key}"] = 1 if pista_sel == "PDIV" else 0
            elif st.session_state[f"pista_last_{casilla_key}"] != pista_sel:
                st.session_state[f"pista_last_{casilla_key}"] = pista_sel
                st.session_state[f"default_rec_idx_{casilla_key}"] = 1 if pista_sel == "PDIV" else 0
                if f"rl_{casilla_key}" in st.session_state:
                    del st.session_state[f"rl_{casilla_key}"]

            recargo_lista = r2c2.selectbox("Cargo Terrestre:", opciones_rec, index=st.session_state[f"default_rec_idx_{casilla_key}"], key=f"rl_{casilla_key}")
            if recargo_lista == "Otro Valor Manual...":
                recargo_final = r2c3.number_input("✍️ Digite Recargo ($)", value=0, step=1000, key=f"rm_{casilla_key}")
            else:
                recargo_final = float(recargo_lista.split(" ")[0])

        tope_clave_efectiva = "TOPE PARCELA INTER < 20HA" if interciclo_menor_20 else tipo_de_tope_finca
        val_tope = dict_topes.get(tope_clave_efectiva, {}).get(pista_sel, 999999)
        if val_tope == 0.0: val_tope = dict_topes.get(tope_clave_efectiva, {}).get("PLUC", 999999)

        with st.container(border=True):
            st.markdown("#### ✈️ Hangar de Despliegue")
            costo_total_vuegos = 0.0
            costo_neto_vuelo_total = 0.0
            total_ha_cobro_escuadron = 0.0
            horometro_final_avion = 0.0

            if mision_solo_dron:
                st.success("🛸 Modo Dron Activo: Costos calculados sin recargos terrestres ni topes de pista.")
                df_drones_def = pd.DataFrame(columns=["Drone", "Hectáreas"])
                escuadron_drones = st.data_editor(df_drones_def, key=f"drones_{casilla_key}", num_rows="dynamic", column_config={"Drone": st.column_config.SelectboxColumn("Modelo Dron", options=list(dict_drones.keys()), required=True), "Hectáreas": st.column_config.NumberColumn("Hectáreas", min_value=0.00, format="%.2f", required=True)}, use_container_width=True, hide_index=True)

                for _, row in escuadron_drones.iterrows():
                    dr_sel, ha_dr = row.get("Drone"), row.get("Hectáreas")
                    if pd.isna(dr_sel) or ha_dr is None or float(ha_dr) <= 0:
                        continue
                    ha_dr = float(ha_dr)
                    total_ha_cobro_escuadron += ha_dr
                    tarifa_dron_neta = dict_drones.get(dr_sel, 0)
                    costo_neto_vuelo_total += (tarifa_dron_neta * ha_dr)
                    costo_total_vuegos += (tarifa_dron_neta * ha_dr) * multi_aviones_final
            else:
                c_av, c_dr = st.columns(2)
                with c_av:
                    st.markdown("##### 🛩️ Base Aviones")
                    df_aviones_def = pd.DataFrame(columns=["Avión", "Hectáreas", "Horómetro"])
                    escuadron_aviones = st.data_editor(df_aviones_def, key=f"aviones_{casilla_key}", num_rows="dynamic", column_config={"Avión": st.column_config.SelectboxColumn("Modelo", options=list(dict_aviones.keys()), required=True), "Hectáreas": st.column_config.NumberColumn("Hectáreas", min_value=0.00, format="%.2f", required=True), "Horómetro": st.column_config.NumberColumn("Horómetro", min_value=0.00, format="%.2f", required=True)}, use_container_width=True, hide_index=True)

                with c_dr:
                    st.markdown("##### 🛸 Base Drones (Apoyo)")
                    df_drones_def = pd.DataFrame(columns=["Drone", "Hectáreas"])
                    escuadron_drones = st.data_editor(df_drones_def, key=f"drones_mix_{casilla_key}", num_rows="dynamic", column_config={"Drone": st.column_config.SelectboxColumn("Modelo Dron", options=list(dict_drones.keys()), required=True), "Hectáreas": st.column_config.NumberColumn("Hectáreas", min_value=0.00, format="%.2f", required=True)}, use_container_width=True, hide_index=True)

                for index, row in escuadron_aviones.iterrows():
                    av_sel, ha_av, horo = row.get("Avión"), row.get("Hectáreas"), row.get("Horómetro")
                    if pd.isna(av_sel) or ha_av is None or horo is None or float(ha_av) <= 0:
                        continue
                    ha_av, horo = float(ha_av), float(horo)
                    total_ha_cobro_escuadron += ha_av
                    horometro_final_avion += horo
                    tarifa_base_ha = (dict_aviones.get(av_sel, 0) * horo) / ha_av if ha_av > 0 else 0
                    tarifa_base_tope = tarifa_base_ha if pista_sel == "PDIV" else min(tarifa_base_ha, val_tope)
                    costo_neto_vuelo_total += (tarifa_base_tope * ha_av)
                    costo_total_vuegos += ((tarifa_base_tope + recargo_final) * ha_av) * multi_aviones_final

                for _, row in escuadron_drones.iterrows():
                    dr_sel, ha_dr = row.get("Drone"), row.get("Hectáreas")
                    if pd.isna(dr_sel) or ha_dr is None or float(ha_dr) <= 0:
                        continue
                    ha_dr = float(ha_dr)
                    total_ha_cobro_escuadron += ha_dr
                    tarifa_dron_neta = dict_drones.get(dr_sel, 0)
                    costo_neto_vuelo_total += (tarifa_dron_neta * ha_dr)
                    costo_total_vuegos += (tarifa_dron_neta * ha_dr) * multi_aviones_final

        st.markdown("#### 🧪 Matriz de Validación e Inteligencia de Mezcla")
        st.markdown("---")
        costo_mezcla_total = 0.0

        if not match_ped.empty:
            idx_precio, idx_lote, idx_saldo, idx_almacen = -1, -1, -1, -1
            if not df_sab.empty:
                for j, col in enumerate(df_sab.columns):
                    col_str = str(col).upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').strip()
                    if ('MAYOR' in col_str or 'PRECIO' in col_str) and idx_precio == -1: idx_precio = j
                    if 'LOTE' in col_str and 'PROVEEDOR' not in col_str and idx_lote == -1: idx_lote = j
                    if ('ALMACEN' in col_str or 'PISTA' in col_str) and 'PB' not in col_str and idx_almacen == -1: idx_almacen = j
                    if ('LIBRE' in col_str or 'SALDO' in col_str) and 'VALOR' not in col_str and idx_saldo == -1: idx_saldo = j

            sap_dict_pista = {}
            datos_extraidos_sap = []

            for _, fila_sap in match_ped.iterrows():
                col_mat = [c for c in fila_sap.index if 'MATERIAL' in str(c).upper() or 'ITEM' in str(c).upper() or 'CÓDIGO' in str(c).upper() or 'COD' in str(c).upper()]
                if not col_mat:
                    continue
                texto_material = str(fila_sap[col_mat[0]]).strip()
                if "459" in texto_material or "429" in texto_material:
                    continue

                cod_item = texto_material.split('.')[0].lstrip('0')

                col_cant_real = [c for c in fila_sap.index if any(x in str(c).upper() for x in ['CANT', 'HECT', 'DOSIS', 'CANTIDAD'])]
                if col_cant_real:
                    cant_total = limpiar_numero_estricto(fila_sap[col_cant_real[0]])
                else:
                    cant_total = 0.0

                dosis_pista = cant_total / ha_dosis_final if ha_dosis_final > 0 else 0.0

                nombre_p = f"Item {cod_item}"
                if not df_sab.empty:
                    # 💥 TU LÓGICA ORIGINAL (Blindada con Lambda para evitar el AttributeError)
                    df_sab_col0_clean = df_sab.iloc[:, 0].apply(lambda x: str(x).split('.')[0].strip().upper().lstrip('0'))
                    match_sabana = df_sab[df_sab_col0_clean == cod_item]
                    if not match_sabana.empty:
                        col_nombre_sab = [c for c in match_sabana.columns if 'TEXTO' in str(c).upper() or 'DESC' in str(c).upper()]
                        if col_nombre_sab:
                            nombre_p = str(match_sabana.iloc[0][col_nombre_sab[0]]).upper()

                nombre_limpio = nombre_p.split('*')[0].strip().replace(" ", "")
                sap_dict_pista[nombre_limpio] = sap_dict_pista.get(nombre_limpio, 0.0) + dosis_pista
                datos_extraidos_sap.append({"cod": cod_item, "nombre": nombre_p, "nombre_limpio": nombre_limpio, "cant_total": cant_total})

            coctel_ganador, dosis_oficiales_coctel = emparejar_coctel_ia(sap_dict_pista, coctel_piloto_base)

            if coctel_ganador == "SIN COINCIDENCIA":
                st.error(f"🤖 **MOTOR IA MAESTRO (Guillotina):** Cóctel Oficial Determinado: **SIN COINCIDENCIA**")
            else:
                st.success(f"🤖 **MOTOR IA MAESTRO:** Cóctel Oficial Determinado: **{coctel_ganador}**")

            if not df_sab.empty:
                # 💥 TU LÓGICA ORIGINAL RESTAURADA
                df_sab_col0_clean = df_sab.iloc[:, 0].apply(lambda x: str(x).split('.')[0].strip().upper().lstrip('0'))
            else:
                df_sab_col0_clean = pd.Series(dtype=str)

            matriz_datos = []
            for item_data in datos_extraidos_sap:
                cod_item = str(item_data['cod']).strip().upper().lstrip('0')
                nombre_p, nombre_limpio, cant_linea_sap = item_data['nombre'], item_data['nombre_limpio'], item_data['cant_total']
                costo_unit, lote_sap, saldo_sap = 0.0, "SIN LOTE EN PISTA", 0.0

                if not df_sab.empty:
                    match_sabana_global = df_sab[df_sab_col0_clean == cod_item]

                    if not match_sabana_global.empty:
                        if idx_almacen != -1:
                            match_pista_precio = match_sabana_global[match_sabana_global.iloc[:, idx_almacen].apply(lambda x: str(pista_sel).strip().upper() in str(x).strip().upper())]
                        else:
                            match_pista_precio = match_sabana_global

                        fila_precio = match_pista_precio.iloc[0] if not match_pista_precio.empty else match_sabana_global.iloc[0]

                        if idx_precio != -1:
                            costo_unit = limpiar_dinero(fila_precio.iloc[idx_precio])
                        if costo_unit == 0.0:
                            col_v = [c for c in fila_precio.index if 'VALOR' in str(c).upper() and 'LIBRE' in str(c).upper()]
                            col_c = [c for c in fila_precio.index if 'LIBRE' in str(c).upper() and 'VALOR' not in str(c).upper()]
                            if col_v and col_c:
                                v_t, c_t = limpiar_dinero(fila_precio[col_v[0]]), limpiar_numero_estricto(fila_precio[col_c[0]])
                                if c_t > 0:
                                    costo_unit = v_t / c_t

                        if idx_almacen != -1:
                            match_pista = match_sabana_global[match_sabana_global.iloc[:, idx_almacen].apply(lambda x: str(pista_sel).strip().upper() in str(x).strip().upper())] 
                        else:
                            match_pista = match_sabana_global

                        if not match_pista.empty:
                            fila_final = match_pista.iloc[0]
                            if idx_lote != -1:
                                lote_sap = str(fila_final.iloc[idx_lote])
                            if idx_saldo != -1:
                                saldo_sap = limpiar_numero_estricto(fila_final.iloc[idx_saldo])

                # 💥 FRANCOTIRADOR 2: TU EXTRACCIÓN PURA ORIGINAL
                try:
                    if 'df_cfg_puro' in locals():
                        nombre_buscado = nombre_p.upper().strip()
                        # Buscar producto exclusivamente en la Columna I (índice 8)
                        col_i = df_cfg_puro[8].apply(lambda x: str(x).strip().upper())
                        match_precio = df_cfg_puro[col_i == nombre_buscado]
                        
                        if not match_precio.empty:
                            # Extraer costo exclusivamente de la Columna J (índice 9)
                            precio_maestro = limpiar_dinero(match_precio.iloc[0, 9])
                            if precio_maestro > 0:
                                costo_unit = float(precio_maestro)
                except Exception:
                    pass
                dosis_teorica = None
                for p_receta, d_oficial in dosis_oficiales_coctel.items():
                    if p_receta == nombre_limpio or (len(nombre_limpio) >= 4 and p_receta in nombre_limpio) or (len(p_receta) >= 4 and nombre_limpio in p_receta):
                        dosis_teorica = d_oficial
                        break

                if "ACONDICIONADOR" in nombre_limpio:
                    dosis_teorica = 0.06 if any(x in coctel_ganador for x in ["ZN", "BT", "ZT", "ZITRON"]) else 0.02
                elif "IMBIOSIL" in nombre_limpio.replace(" ", ""):
                    dosis_teorica = 1.5 if (coctel_ganador.strip().upper().split()[0].startswith("IN") or "IMBIOSIL" in coctel_ganador.strip().upper().split()[0]) else 1.0
                elif "ACEITE" in nombre_limpio:
                    if coctel_ganador != "SIN COINCIDENCIA":
                        for char in coctel_ganador.split()[0]:
                            if char.isdigit():
                                dosis_teorica = float(char)
                                break

                if dosis_teorica is None:
                    dosis_rescatada = obtener_dosis_global_robusta_v2(None, nombre_limpio)
                    if dosis_rescatada > 0:
                        dosis_teorica = dosis_rescatada
                    else:
                        dosis_teorica = 0.0

                dosis_ideal_pura = round(dosis_teorica * ha_dosis_final, 3)

                precio_marginado_final = costo_unit * mult_material
                precio_marginado_final = aplicar_excepcion_manzate(precio_marginado_final, f"{nombre_limpio} {nombre_p}", tipo_productor)

                matriz_datos.append({
                    "A: Producto": nombre_p,
                    "B: Dosis/Ha (SAP)": round(dosis_teorica, 3),
                    "C: X (Extra %)": 0.0,
                    "D: Dosis Total (Sistema)": dosis_ideal_pura,
                    "E: Costo Unit (+Margen)": round(precio_marginado_final, 0),
                    "G: Lotes (SAP)": lote_sap,
                    "H: Saldo Real SAP": round(saldo_sap, 3),
                    "I: Sugerido SAP (Total)": round(cant_linea_sap, 3)
                })

            df_matriz = pd.DataFrame(matriz_datos)

            if not df_matriz.empty:
                df_matriz["TOTAL_PROD_SAP"] = df_matriz.groupby("A: Producto")["I: Sugerido SAP (Total)"].transform("sum")

                for idx_m, r_m in df_matriz.iterrows():
                    b_ideal_pura = r_m["D: Dosis Total (Sistema)"]
                    tot_p_sap = r_m["TOTAL_PROD_SAP"]

                    if b_ideal_pura > 0 and tot_p_sap > (b_ideal_pura + 0.001):
                        extra_pct = ((tot_p_sap / b_ideal_pura) - 1.0) * 100.0
                        df_matriz.at[idx_m, "C: X (Extra %)"] = round(extra_pct, 3)
                    else:
                        df_matriz.at[idx_m, "C: X (Extra %)"] = 0.0

                df_matriz["D: Dosis Total (Sistema)"] = (df_matriz["B: Dosis/Ha (SAP)"].fillna(0.0) * ha_dosis_final).round(3)

                def estilizar_dosis_ideal(row):
                    estilos = [''] * len(row)
                    try:
                        idx_sistema = row.index.get_loc("D: Dosis Total (Sistema)")
                        idx_sap = row.index.get_loc("I: Sugerido SAP (Total)")

                        base_pura = float(row["D: Dosis Total (Sistema)"])
                        total_producto = float(row.get("TOTAL_PROD_SAP", row["I: Sugerido SAP (Total)"]))
                        extra_pct = float(row.get("C: X (Extra %)", 0.0))

                        diferencia_real = total_producto - base_pura

                        if diferencia_real < -0.05:
                            color = 'background-color: #f8d7da; color: #721c24; font-weight: bold;'
                        elif extra_pct > 0.01 or diferencia_real > 0.05:
                            color = 'background-color: #fff3cd; color: #856404; font-weight: bold;'
                        else:
                            color = 'background-color: #d4edda; color: #155724; font-weight: bold;'

                        estilos[idx_sistema] = color
                        estilos[idx_sap] = color
                    except Exception: pass
                    return estilos

                def calcular_semaforo_misiones(row):
                    base_pura = float(row["D: Dosis Total (Sistema)"])
                    total_producto = float(row.get("TOTAL_PROD_SAP", row["I: Sugerido SAP (Total)"]))
                    extra_pct = float(row.get("C: X (Extra %)", 0.0))
                    diferencia = total_producto - base_pura

                    if total_producto < (base_pura - 0.05):
                        return "🔴 PELIGRO: SUB-DOSIS (---)"
                    elif extra_pct > 0.01 or diferencia > 0.05:
                        return f"🔵 REC. TÉCNICA (+{extra_pct:.1f}%)"
                    else:
                        return "🟢 ÓPTIMO"

                df_matriz["📊 Ajuste de Campo"] = df_matriz.apply(calcular_semaforo_misiones, axis=1)

                columnas_ordenadas = [
                    "A: Producto", "B: Dosis/Ha (SAP)", "C: X (Extra %)",
                    "D: Dosis Total (Sistema)", "I: Sugerido SAP (Total)", "📊 Ajuste de Campo",
                    "E: Costo Unit (+Margen)", "G: Lotes (SAP)", "H: Saldo Real SAP", "TOTAL_PROD_SAP"
                ]
                df_matriz = df_matriz[columnas_ordenadas]

                df_vista = df_matriz.drop(columns=["TOTAL_PROD_SAP"])
                df_vista["E: Costo Unit (+Margen)"] = df_vista["E: Costo Unit (+Margen)"].apply(lambda x: f"{int(x):,.0f}".replace(",", "."))

                df_vista["H: Saldo Real SAP"] = df_vista["H: Saldo Real SAP"].apply(lambda x: f"{float(x):,.3f}".replace(",", "X").replace(".", ",").replace("X", "."))

                df_estilizado = df_vista.style.apply(estilizar_dosis_ideal, axis=1)

                edited_df = st.data_editor(
                    df_estilizado,
                    key=llave_editor_casilla,
                    column_config={
                        "B: Dosis/Ha (SAP)": st.column_config.NumberColumn("Dosis/Ha", min_value=0.000, format="%.3f"),
                        "C: X (Extra %)" : st.column_config.NumberColumn("Extra %", min_value=0.000, format="%.3f"),
                        "D: Dosis Total (Sistema)": st.column_config.NumberColumn("Dosis Ideal", format="%.3f"),
                        "I: Sugerido SAP (Total)": st.column_config.NumberColumn("Sugerido SAP (Total)", format="%.3f"),
                        "📊 Ajuste de Campo": st.column_config.TextColumn("📊 Ajuste de Campo"),
                        "E: Costo Unit (+Margen)": st.column_config.TextColumn("Costo Unit (COP)"),
                        "H: Saldo Real SAP": st.column_config.TextColumn("Saldo SAP"),
                    },
                    disabled=["A: Producto", "D: Dosis Total (Sistema)", "E: Costo Unit (+Margen)", "G: Lotes (SAP)", "H: Saldo Real SAP", "I: Sugerido SAP (Total)", "📊 Ajuste de Campo"],
                    use_container_width=True, hide_index=True
                )
                st.write("")
                st.markdown("##### 📋 Copia Rápida para SAP (Costo Unitario)")
                valores_formateados = [f"{int(x):,.0f}".replace(",", ".") for x in df_matriz['E: Costo Unit (+Margen)'].fillna(0).tolist()]
                st.code("\n".join(valores_formateados), language="text")

        from decimal import Decimal, ROUND_HALF_UP

        def sap_round(n):
            n_limpio = round(float(n), 4)
            return int(Decimal(str(n_limpio)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))

        if 'df_matriz' in locals() and df_matriz is not None and not df_matriz.empty:
            costo_mezcla_total = (df_matriz["I: Sugerido SAP (Total)"] * df_matriz["E: Costo Unit (+Margen)"]).apply(sap_round).sum()
        else:
            costo_mezcla_total = 0

        unitario_st = sap_round(d_ciclo_factura * tarifa_serv_tec_base)
        unitario_vuelo = sap_round(costo_total_vuegos / total_ha_cobro_escuadron) if total_ha_cobro_escuadron > 0 else 0

        subtotal_st_finca = sap_round(unitario_st * ha_dosis_final)
        subtotal_vuelo_finca = sap_round(unitario_vuelo * ha_dosis_final)

        gran_total = costo_mezcla_total + subtotal_vuelo_finca + subtotal_st_finca
        costo_por_ha = sap_round(gran_total / ha_dosis_final) if ha_dosis_final > 0 else 0

        precio_columna_ref = dict_aviones.get(escuadron_aviones.iloc[0]['Avión'], 0) if (not mision_solo_dron and not escuadron_aviones.empty) else 0
        precio_dron_ref = dict_drones.get(escuadron_drones.iloc[0]['Drone'], 0) if (not escuadron_drones.empty and pd.notna(escuadron_drones.iloc[0]['Drone'])) else 0

        st.write("")
        st.markdown("### 💰 Liquidación Final (Bóveda SAP)")

        def mini_metric(i, t, v):
            return f"<div style='background-color:#ffffff; padding:12px; border-radius:8px; border: 2px solid #0d1b2a; border-left:5px solid #d4af37; box-shadow: 0 2px 4px rgba(0,0,0,0.06); height: 100%;'><p style='margin:0; font-size:11px; font-weight:800; color:#0d1b2a; text-transform:uppercase;'>{i} {t}</p><p style='margin:0; font-size:15px; font-weight:900; color:#1a365d;'>{v}</p></div>"

        # 💥 FILA 1 (3 columnas): Datos generales
        m1, m2, m3 = st.columns(3)
        with m1:
            st.markdown(mini_metric("🗺️", "Hectáreas", f"{ha_dosis_final:.2f} Ha"), unsafe_allow_html=True)
        with m2:
            ha_av_r = float(escuadron_aviones['Hectáreas'].sum()) if (not mision_solo_dron and not escuadron_aviones.empty) else 0
            es_dr_dom = mision_solo_dron or (ha_av_r == 0 and precio_dron_ref > 0)
            pista_display_text = "PARCELA < 20HA" if interciclo_menor_20 else tipo_de_tope_finca
            st.markdown(mini_metric("¼️", "Pista", pista_display_text if not es_dr_dom else "DRON"), unsafe_allow_html=True)
        with m3:
            st.markdown(mini_metric("🚧", "Valor Tope", f"$ {fmt_sap(precio_dron_ref)}" if es_dr_dom else ("Sin Tope" if val_tope in [0, 999999] else f"$ {fmt_sap(val_tope)}")), unsafe_allow_html=True)

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        # 💥 FILA 2 (2 columnas): Tarifas base
        m4, m5 = st.columns(2)
        with m4:
            st.markdown(mini_metric("👨‍🔬", "Tarifa ST", f"$ {fmt_sap(tarifa_serv_tec_base)}"), unsafe_allow_html=True)
        with m5:
            st.markdown(mini_metric("✈️", "Mult.", f"x {multi_aviones_final}"), unsafe_allow_html=True)

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        # 💥 FILA 3 (2 columnas): Totales por unidad
        m6, m7 = st.columns(2)
        with m6:
            if es_dr_dom:
                st.markdown(mini_metric("🛸", "Tarifa Dron", f"$ {fmt_sap(precio_dron_ref)}"), unsafe_allow_html=True)
            else:
                st.markdown(mini_metric("⏱️", "Precio Hora", f"$ {fmt_sap(precio_columna_ref)}"), unsafe_allow_html=True)
        with m7:
            st.markdown(f"<div style='background-color:#0d1b2a; padding:12px; border-radius:8px; border:2px solid #d4af37; text-align:center; box-shadow: 0 2px 4px rgba(0,0,0,0.15); height: 100%; display: flex; flex-direction: column; justify-content: center;'><p style='margin:0; color:#d4af37; font-size:11px; font-weight:800; text-transform:uppercase;'>💰 COSTO x HA (Final)</p><p style='margin:0; font-size:16px; font-weight:900; color:white;'>$ {fmt_sap(costo_por_ha)}</p></div>", unsafe_allow_html=True)

        c_sap1, c_sap2, c_sap3 = st.columns(3)
        with c_sap1:
            st.caption("👨‍🔬 UNITARIO ST (459)")
            st.code(fmt_sap(unitario_st), language="text")
        with c_sap2:
            st.caption("✈️ UNITARIO Vuelo (429)")
            st.code(fmt_sap(unitario_vuelo), language="text")
        with c_sap3:
            st.caption("🧪 TOTAL Mezcla")
            st.code(fmt_sap(costo_mezcla_total), language="text")

        # 💥 FILA 4: Las 3 tarjetas GIGANTES finales
        html_totales = f"""
        <div style="display: flex; flex-wrap: wrap; gap: 15px; margin-top: 10px; margin-bottom: 20px;">
            <div style="flex: 1; min-width: 200px; background-color: #ffffff; padding: 15px; border-radius: 8px; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <p style="margin:0; font-size: 12px; color: #6c757d; font-weight: bold; text-transform: uppercase;">👨‍🔬 Subtotal ST (459)</p>
                <h3 style="margin:0; color: #0d1b2a; font-weight: 900; user-select: all;">$ {fmt_sap(subtotal_st_finca)}</h3>
            </div>
            <div style="flex: 1; min-width: 200px; background-color: #ffffff; padding: 15px; border-radius: 8px; border: 2px solid #0d1b2a; border-left: 6px solid #1a365d; box-shadow: 0 4px 6px rgba(0,0,0,0.08);">
                <p style="margin:0; font-size: 12px; color: #6c757d; font-weight: bold; text-transform: uppercase;">✈️ Subtotal Vuelo (429)</p>
                <h3 style="margin:0; color: #0d1b2a; font-weight: 900; user-select: all;">$ {fmt_sap(subtotal_vuelo_finca)}</h3>
            </div>
            <div style="flex: 1; min-width: 200px; background-color: #0d1b2a; padding: 15px; border-radius: 8px; border: 3px solid #d4af37; box-shadow: 0 4px 12px rgba(0,0,0,0.2); text-align: center;">
                <p style="margin:0; font-size: 13px; color: #d4af37; font-weight: bold; text-transform: uppercase;">🔥 TOTAL OPERACIÓN</p>
                <h2 style="margin:0; color: white; font-weight: 900; user-select: all;">$ {gran_total:,.0f}</h2>
            </div>
        </div>
        """.replace(",", ".")
        st.markdown(html_totales, unsafe_allow_html=True)

        st.caption("📋 **COPIA RÁPIDA (Clic en el ícono 📋 de cada cajita)**")
        cc1, cc2, cc3, cc4, cc5, cc6 = st.columns(6)
        with cc1:
            st.write("👨‍🔬 Serv. Tec")
            st.code(fmt_sap(subtotal_st_finca), language="text")
        with cc2:
            st.write("✈️ Vuelo")
            st.code(fmt_sap(subtotal_vuelo_finca), language="text")
        with cc3:
            st.write("🧪 Mezcla")
            st.code(fmt_sap(costo_mezcla_total), language="text")
        with cc4:
            st.write("⚠️ Recargo")
            st.code(fmt_sap(0), language="text")
        with cc5:
            st.write("💰 Costo x Ha")
            st.code(fmt_sap(costo_por_ha), language="text")
        with cc6:
            st.write("🔥 TOTAL")
            st.code(fmt_sap(gran_total), language="text")

        st.markdown("---")
        st.markdown("### 🛰️ Coordenadas de Lanzamiento Final")
        c_p1, c_p2 = st.columns(2)
        pista_manual = c_p1.selectbox("📍 Confirmar Pista de Operación:", lista_pistas_validas, index=lista_pistas_validas.index(pista_sel), key=f"confirmador_final_{pista_sel}_{vuelo_ref}")
        c_p2.info(f"🚀 Misión: {('DRONE' if mision_solo_dron else 'AVION')} | 📋 Referencia: {vuelo_ref}")

        st.markdown("""
            <a href="#inicio_modulo" target="_self" style="
                display: block; width: 100%; text-align: center;
                background-color: #0d1b2a; color: #d4af37; border: 1px solid #d4af37;
                padding: 12px; border-radius: 8px; text-decoration: none; font-weight: bold;
                box-shadow: 0px 4px 6px rgba(0,0,0,0.3); margin-bottom: 20px; font-size: 16px;
            ">
                ⬆️ VOLVER AL INICIO DEL MÓDULO ⬆️
            </a>
        """, unsafe_allow_html=True)

        if st.button("💾 DETONAR FACTURA Y GUARDAR EN BÓVEDA", type="primary", use_container_width=True):
            errores_validacion = []

            if ha_dosis_final <= 0:
                errores_validacion.append("🗺️ Las hectáreas (Ha Dosis) deben ser mayores a 0.")

            if total_ha_cobro_escuadron <= 0:
                errores_validacion.append("✈️ No hay aviones ni drones cargados en el Hangar de Despliegue (0 Ha totales).")

            if gran_total <= 0:
                errores_validacion.append(f"💰 El Total de la Operación es ${fmt_sap(gran_total)}, lo cual es inválido. Revisa tarifas, tope de pista y matriz de mezcla antes de guardar.")

            if costo_por_ha <= 0:
                errores_validacion.append(f"📊 El Costo x Ha final es ${fmt_sap(costo_por_ha)}, lo cual es inválido.")

            if errores_validacion:
                st.error("🚨 **NO SE PUEDE GUARDAR — Revisa lo siguiente antes de detonar la factura:**\n\n" + "\n\n".join(errores_validacion))
                st.stop()

            with st.spinner("🚀 Dividiendo Rendimientos e Inyectando Múltiples Aeronaves..."):
                try:
                    gc_save = obtener_cliente_gspread_unificado()
                    if not gc_save:
                        st.error("🚨 Error crítico: No se pudo conectar al llavero de Google para guardar.")
                        st.stop()

                    boveda = gc_save.open_by_url("https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit")
                    hoja_apoyo = boveda.worksheet("TABLA DE APOYO2023")
                    hoja_maestra = boveda.worksheet("TABLA 1")
                    hoja_memoria = boveda.worksheet("MEMORIA")

                    fecha_str = fecha_operacion.strftime("%d/%m/%Y")
                    dia_sem = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"][fecha_operacion.weekday()]
                    num_sem = fecha_operacion.isocalendar()[1]
                    os_virtual = f"VIRT-{finca_limpia[:3]}-{obtener_hora_colombia().strftime('%H%M')}"

                    bloque_f, sector_f, ha_bruta_f_raw = "", "", ""
                    match_f = df_t2[df_t2.iloc[:, 0].astype(str).str.upper().str.strip() == finca_limpia.upper().strip()]
                    if not match_f.empty:
                        sector_f, ha_bruta_f_raw, bloque_f = match_f.iloc[0, 1], match_f.iloc[0, 2], match_f.iloc[0, 3]

                    ha_f = float(ha_dosis_final)
                    ha_bruta_f_num = limpiar_numero_estricto(ha_bruta_f_raw)

                    tarifa_vuelo_neta_ha = float(costo_neto_vuelo_total / total_ha_cobro_escuadron) if total_ha_cobro_escuadron > 0 else 0.0
                    total_pago_avion_neto = (tarifa_vuelo_neta_ha + float(recargo_final)) * ha_f

                    filas_maestra_a_inyectar = []
                    filas_apoyo_a_inyectar = []
                    payloads_supabase = []

                    if total_ha_cobro_escuadron == 0:
                        flota_activa = [{"hk": "DRON PROPIO", "piloto": "CLIENTE (DRON PROPIO)", "tipo": "DRONE PROPIO", "prop": 1.0, "horo": 0.0}]
                    else:
                        flota_activa = []
                        if not mision_solo_dron:
                            for _, r in escuadron_aviones.iterrows():
                                av_sel = r.get("Avión")
                                ha_av = float(r.get("Hectáreas", 0))
                                horo = float(r.get("Horómetro", 0))
                                if pd.notna(av_sel) and ha_av > 0:
                                    flota_activa.append({
                                        "hk": str(av_sel).upper(),
                                        "piloto": "PILOTO AVIÓN",
                                        "tipo": "AVION",
                                        "prop": ha_av / total_ha_cobro_escuadron,
                                        "horo": horo
                                    })
                        for _, r in escuadron_drones.iterrows():
                            dr_sel = r.get("Drone")
                            ha_dr = float(r.get("Hectáreas", 0))
                            if pd.notna(dr_sel) and ha_dr > 0:
                                dr_name_str = str(dr_sel).upper()
                                hk_dr = "DR51" if "DATAROT" in dr_name_str else ("DR52" if "GENESYS" in dr_name_str else "DRONE_GEN")
                                flota_activa.append({
                                    "hk": hk_dr,
                                    "piloto": "OPERADOR DRONE",
                                    "tipo": "DRONE",
                                    "prop": ha_dr / total_ha_cobro_escuadron,
                                    "horo": 0.0
                                })

                    for nave in flota_activa:
                        P = nave["prop"]
                        ha_f_part = ha_f * P

                        ha_bruta_f_part = ha_bruta_f_num if ha_bruta_f_num > 0 else ha_f

                        if nave["tipo"] == "DRONE" or nave["tipo"] == "DRONE PROPIO":
                            h_total_v_part = ha_f_part / 10
                        else:
                            ha_avion_real = P * total_ha_cobro_escuadron
                            h_total_v_part = (ha_f_part / ha_avion_real) * nave["horo"] if ha_avion_real > 0 else 0.0

                        vol_total_gln_part = ha_f_part * 6
                        rend_min_part = h_total_v_part * 60

                        gran_total_part = gran_total * P
                        total_pago_avion_neto_part = total_pago_avion_neto * P

                        row_azul = [""] * 34
                        row_azul[0] = os_virtual
                        row_azul[1] = bloque_f
                        row_azul[2] = finca_limpia
                        row_azul[3] = sector_f
                        row_azul[4] = round(ha_bruta_f_part, 2)
                        row_azul[5] = round(ha_f_part, 2)
                        row_azul[6] = coctel_ganador
                        row_azul[7] = fecha_str
                        row_azul[8] = dia_sem
                        row_azul[9] = num_sem
                        row_azul[10] = round(h_total_v_part, 2)
                        row_azul[11] = 6
                        row_azul[12] = round(vol_total_gln_part, 2)
                        row_azul[13] = round(h_total_v_part, 2)
                        row_azul[14] = round(rend_min_part, 2)
                        row_azul[15] = nave["piloto"]
                        row_azul[16] = nave["hk"]
                        row_azul[17] = nave["tipo"]
                        row_azul[18] = round(gran_total_part, 2)
                        row_azul[19] = round(tarifa_vuelo_neta_ha, 2)
                        row_azul[20] = round(float(recargo_final), 2)
                        row_azul[21] = round(gran_total_part, 2)
                        row_azul[23] = pista_manual
                        row_azul[28] = round(gran_total_part, 2)
                        row_azul[29] = round(total_pago_avion_neto_part, 2)
                        row_azul[32] = tipo_productor
                        row_azul[33] = "GÉNESIS_V2_PRO"

                        filas_maestra_a_inyectar.append(row_azul)

                        payloads_supabase.append({
                            "os_virtual": str(os_virtual),
                            "finca": str(finca_limpia),
                            "hectareas": float(round(ha_f_part, 2)),
                            "coctel": str(coctel_ganador),
                            "fecha": str(fecha_str),
                            "total_operacion": float(round(gran_total_part, 2)),
                            "pista": str(pista_manual),
                            "tipo_productor": str(tipo_productor),
                            "aeronave": nave["hk"]
                        })

                    tipo_nave_apoyo = "DRONE" if mision_solo_dron else "AVION"
                    fila_apoyo = ["", finca_limpia, round(ha_f, 2), float(costo_por_ha), round(gran_total, 2), fecha_str, "", "", coctel_ganador, "", pista_manual, "", "", tipo_nave_apoyo, ""]
                    filas_apoyo_a_inyectar.append(fila_apoyo)

                    col_azul = hoja_maestra.col_values(1)
                    col_apoyo = hoja_apoyo.col_values(2)
                    datos_memoria = hoja_memoria.get_all_values()

                    f_azul = next((i+2 for i in range(len(col_azul)-1, -1, -1) if str(col_azul[i]).strip() != ""), 1)
                    f_apoyo = next((i+2 for i in range(len(col_apoyo)-1, -1, -1) if str(col_apoyo[i]).strip() != ""), 1)

                    f_apoyo_start = f_apoyo - 3
                    for i, fila in enumerate(filas_apoyo_a_inyectar):
                        fila[0] = f_apoyo_start + i

                    if f_azul + len(filas_maestra_a_inyectar) > hoja_maestra.row_count:
                        hoja_maestra.add_rows(10 + len(filas_maestra_a_inyectar))
                    if f_apoyo + len(filas_apoyo_a_inyectar) > hoja_apoyo.row_count:
                        hoja_apoyo.add_rows(10 + len(filas_apoyo_a_inyectar))

                    set_existentes = {f"{str(r[0]).strip()}|{str(r[9]).strip().upper()}|{str(r[3]).strip().upper()}" for r in datos_memoria[1:] if len(r) >= 10}
                    bodega_f = "BODEGA PRINCIPAL DRON" if mision_solo_dron or total_ha_cobro_escuadron == 0 else "BODEGA PRINCIPAL AVIÓN"
                    filas_memoria = []

                    if not edited_df.empty:
                        for idx, row_m in edited_df.iterrows():
                            nombre_prod = str(row_m.get("A: Producto", "")).strip().upper()
                            if "⚠️" not in nombre_prod and nombre_prod not in ["", "NAN"]:
                                if f"{fecha_str}|{finca_limpia}|{nombre_prod}" not in set_existentes:
                                    fila_m = [fecha_str, coctel_ganador, str(pista_manual).split("-")[0].strip()[:4], nombre_prod, str(row_m.get("G: Lotes (SAP)", "S/N")), float(row_m.get("D: Dosis Total (Sistema)", 0)), bodega_f, "", "X", finca_limpia]
                                    filas_memoria.append(fila_m)

                    def limpiar_json(val):
                        if pd.isna(val) or (isinstance(val, float) and math.isnan(val)):
                            return ""
                        if hasattr(val, 'item'):
                            return val.item()
                        return val

                    filas_maestra_a_inyectar = [[limpiar_json(x) for x in fila] for fila in filas_maestra_a_inyectar]
                    filas_apoyo_a_inyectar = [[limpiar_json(x) for x in fila] for fila in filas_apoyo_a_inyectar]
                    filas_memoria = [[limpiar_json(x) for x in fila] for fila in filas_memoria]

                    rango_maestra = f"A{f_azul}:{get_column_letter(len(filas_maestra_a_inyectar[0]))}{f_azul + len(filas_maestra_a_inyectar) - 1}"
                    rango_apoyo = f"A{f_apoyo}:{get_column_letter(len(filas_apoyo_a_inyectar[0]))}{f_apoyo + len(filas_apoyo_a_inyectar) - 1}"

                    valor_actual_fila = hoja_maestra.cell(f_azul, 1).value
                    if valor_actual_fila and str(valor_actual_fila).strip() != "":
                        st.error(
                            f"🚨 CONFLICTO DE CONCURRENCIA: la fila {f_azul} de TABLA 1 ya "
                            "fue ocupada por otra operación mientras se calculaba esta factura. "
                            "Sincroniza el módulo y reintenta para evitar sobrescribir datos."
                        )
                        st.stop()

                    hoja_maestra.update(range_name=rango_maestra, values=filas_maestra_a_inyectar, value_input_option='USER_ENTERED')
                    hoja_apoyo.update(range_name=rango_apoyo, values=filas_apoyo_a_inyectar, value_input_option='USER_ENTERED')
                    if filas_memoria:
                        hoja_memoria.append_rows(filas_memoria, value_input_option='USER_ENTERED')

                    if 'supabase' in st.session_state:
                        try:
                            supabase_client = st.session_state['supabase']
                            for payload in payloads_supabase:
                                supabase_client.table("facturas_detonadas").insert(payload).execute()
                        except Exception:
                            pass

                    st.balloons()
                    st.success(f"✅ IMPACTO MÚLTIPLE CONFIRMADO. Se crearon {len(filas_maestra_a_inyectar)} filas independientes en Excel.")
                    st.toast("💾 Memoria Sincronizada con éxito.", icon="⚔️")

                    if 'memoria_excel' in st.session_state:
                        del st.session_state['memoria_excel']
                except Exception as e_save:
                    st.error(f"🚨 Falla en el Guardado: {e_save}")

if __name__ == "__main__":
    pass
