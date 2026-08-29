import streamlit as st
import logging
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import gspread
import re
import math
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
logging.basicConfig(level=logging.WARNING, filename="app_errors.log",
                     format="%(asctime)s %(levelname)s %(message)s")
# =================================================================
# ⚙️ CONSTANTES CENTRALIZADAS (ÚNICA FUENTE DE VERDAD)
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"
URL_BOVEDA_HISTORICA = "https://docs.google.com/spreadsheets/d/16OZdiWwW7nLHyZBEnhiKlDTDttR7Tjhn37O9zm6wJOk/edit"

# =================================================================
# 🛡️ BLOQUE 1: UTILIDADES Y FORMATO
# =================================================================

def formato_latino(numero, decimales=0):
    if pd.isna(numero) or numero is None: return "0"
    try:
        num = float(numero)
        if num == 0: return "0"
        if decimales == 0: texto_us = f"{num:,.0f}"
        else: texto_us = f"{num:,.{decimales}f}"
        return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0"

def formato_gerencial_latino(numero):
    if pd.isna(numero) or numero == 0: return "$ 0"
    if numero >= 1_000_000: return f"$ {numero / 1_000_000:,.1f} M".replace(".", "X").replace(",", ".").replace("X", ",")
    elif numero >= 1_000: return f"$ {numero / 1_000:,.0f} K".replace(",", ".")
    else: return f"$ {formato_latino(numero, 0)}"

def limpiar_encabezados(df):
    df.columns = [str(col).upper().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').strip() for col in df.columns]
    df = df.loc[:, ~df.columns.duplicated(keep='first')]
    if "" in df.columns: df = df.drop(columns=[""])
    return df

def estandarizar_base(df):
    renombres = {}
    for col in df.columns:
        col_u = str(col).upper().replace('\n', ' ').strip()
        if 'FINCA' in col_u and 'COSTO' in col_u: continue
        if 'FACTURAR' in col_u and 'PRODUCTOR' in col_u: renombres[col] = 'COSTO_MAESTRO'
        elif 'FUMIG' in col_u and 'AREA' in col_u: renombres[col] = 'AREA_MAESTRA'
        elif 'AVION' in col_u and '/HA' in col_u: renombres[col] = 'AVION_MAESTRO'
        elif 'DOMINIC' in col_u and '/HA' in col_u: renombres[col] = 'DOMINIC_MAESTRO'
        elif not ('FINCA_MAESTRA' in renombres.values()) and (col_u == 'FINCA' or col_u == 'PROPIEDAD'): renombres[col] = 'FINCA_MAESTRA'
        elif not ('FECHA_MAESTRA' in renombres.values()) and col_u == 'FECHA': renombres[col] = 'FECHA_MAESTRA'
        elif not ('OS_MAESTRA' in renombres.values()) and ("Nº ORDEN" in col_u or "ORDEN DE" in col_u or "OS" == col_u): renombres[col] = 'OS_MAESTRA'
        elif not ('COCTEL_MAESTRO' in renombres.values()) and col_u in ['COCTEL', 'CÓCTEL']: renombres[col] = 'COCTEL_MAESTRO'
    df.rename(columns=renombres, inplace=True)
    return df

def limpiar_area(val):
    try:
        if isinstance(val, (int, float)): return float(val)
        v = str(val).strip()
        if not v: return 0.0
        v = v.replace(',', '.')
        v = re.sub(r'[^\d\.\-]', '', v)
        if v.count('.') > 1:
            partes = v.rsplit('.', 1)
            v = partes[0].replace('.', '') + '.' + partes[1]
        return float(v) if v else 0.0
    except Exception: return 0.0

def limpiar_dinero(val):
    if pd.isna(val) or val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    v = str(val).upper().replace("$", "").replace("COP", "").replace(" ", "").strip()
    if not v or v == '-': return 0.0
    try:
        if '.' in v and ',' in v:
            if v.rfind(',') > v.rfind('.'):
                v = v.replace('.', '').replace(',', '.')
            else:
                v = v.replace(',', '')
        elif ',' in v:
            if len(v.split(',')[-1]) == 3:
                v = v.replace(',', '')
            else:
                v = v.replace(',', '.')
        elif '.' in v:
            if v.count('.') > 1 or len(v.split('.')[-1]) == 3:
                v = v.replace('.', '')
        return float(v)
    except Exception:
        return 0.0

def fecha_fallback(val):
    return pd.to_datetime(val, errors='coerce', dayfirst=True)

# =================================================================
# ⚙️ BLOQUE 2: MOTORES DE CONEXIÓN UNIFICADOS (Bypass Nube)
# =================================================================
@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    # En Módulo 8 (Solo lectura) apagamos la dependencia de credenciales
    return None  

# =================================================================
# 📦 BLOQUE 3: EXTRACCIÓN DE DATOS Y MODELO (Extracción Directa)
# =================================================================
@st.cache_data(show_spinner=False, ttl=600)
def cargar_fuentes_maestras_bi(_descargar_matriz_rapida=None):
    df_vivos = pd.DataFrame()
    df_historico = pd.DataFrame()
    
    # 💥 1. DESCARGA BÓVEDA ACTUAL (Bypass sin credenciales)
    try:
        file_id_act = URL_BOVEDA_MAESTRA.split('/d/')[1].split('/')[0]
        dl_url_act = f'https://docs.google.com/spreadsheets/d/{file_id_act}/export?format=xlsx'
        resp_act = requests.get(dl_url_act, timeout=30)
        
        if resp_act.status_code == 200:
            xls_act = pd.ExcelFile(io.BytesIO(resp_act.content))
            df_raw_act = pd.read_excel(xls_act, sheet_name="TABLA 1", header=None)
            datos_brutos_act = df_raw_act.fillna("").values.tolist()
            
            if len(datos_brutos_act) > 5:
                columnas_t1 = ["OS", "BLOQUE", "FINCA", "SECTOR", "AREA_BRUTA", "AREA_FUMIG", "COCTEL", "FECHA", "DIA", "SEMANA", "H_TOTAL", "GLN_HA", "VOL_TOTAL", "REND_HR", "REND_MIN", "PILOTO", "HK", "MODELO", "COSTO_AVION", "COSTO_HA", "DOMINICAL_HA", "COSTO_FINCA", "VALOR_FACTURAR", "PISTA", "INC_2026", "LIMITE", "ALERTA", "VAR_PCT", "COSTO_TOTAL", "PAGO_AVION"]
                filas_limpias = [list(r) + [""]*(len(columnas_t1) - len(r)) for r in datos_brutos_act[5:]]
                df_vivos = pd.DataFrame([r[:len(columnas_t1)] for r in filas_limpias], columns=columnas_t1)
                df_vivos.rename(columns={'AREA_FUMIG': 'AREA_MAESTRA', 'COSTO_HA': 'AVION_MAESTRO', 'DOMINICAL_HA': 'DOMINIC_MAESTRO', 'FINCA': 'FINCA_MAESTRA', 'FECHA': 'FECHA_MAESTRA', 'OS': 'OS_MAESTRA', 'COCTEL': 'COCTEL_MAESTRO'}, inplace=True)
                df_vivos['ORIGEN_BI'] = 'ACTUAL'
    except Exception as e:
        pass

    # 💥 2. DESCARGA BÓVEDA HISTÓRICA (Bypass sin credenciales)
    try:
        file_id_hist = URL_BOVEDA_HISTORICA.split('/d/')[1].split('/')[0]
        dl_url_hist = f'https://docs.google.com/spreadsheets/d/{file_id_hist}/export?format=xlsx'
        resp_hist = requests.get(dl_url_hist, timeout=30)
        
        if resp_hist.status_code == 200:
            xls_hist = pd.ExcelFile(io.BytesIO(resp_hist.content))
            df_raw_hist = pd.read_excel(xls_hist, sheet_name="Datos", header=None)
            datos_brutos_hist = df_raw_hist.fillna("").values.tolist()
            
            if len(datos_brutos_hist) > 0:
                df_historico = pd.DataFrame(datos_brutos_hist[1:], columns=datos_brutos_hist[0])
                df_historico = estandarizar_base(limpiar_encabezados(df_historico))
                df_historico['ORIGEN_BI'] = 'HISTORICO'
            
            ws_hist_name = next((s for s in xls_hist.sheet_names if "HISTORICO" in s.upper() and "PISTA" in s.upper()), None)
            if ws_hist_name:
                df_pistas_antiguas = pd.read_excel(xls_hist, sheet_name=ws_hist_name, header=None)
                datos_pistas_antiguas = df_pistas_antiguas.fillna("").values.tolist()
                if len(datos_pistas_antiguas) > 1:
                    df_hp = pd.DataFrame(datos_pistas_antiguas[1:], columns=datos_pistas_antiguas[0])
                    df_hp = limpiar_encabezados(df_hp)
                    
                    col_anio = next((c for c in df_hp.columns if 'AÑO' in c or 'ANO' in c or 'YEAR' in c), None)
                    col_mes = next((c for c in df_hp.columns if 'MES' in c or 'FECHA' in c), None)
                    col_pista = next((c for c in df_hp.columns if 'PISTA' in c or 'ALM' in c or 'BASE' in c), None)
                    col_ha = next((c for c in df_hp.columns if 'HECTA' in c or 'AREA' in c or 'SUMA' in c or 'CANT' in c), None)
                    
                    if col_anio and col_mes and col_pista and col_ha:
                        df_hp_clean = pd.DataFrame()
                        def parse_mes(m):
                            try:
                                m_str = str(m).upper().strip()[:3]
                                meses = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}
                                if m_str in meses: return meses[m_str]
                                return int(float(m))
                            except: return 12
                        df_hp_clean['AÑO'] = pd.to_numeric(df_hp[col_anio], errors='coerce').fillna(2017).astype(int)
                        df_hp_clean['MES_NUM'] = df_hp[col_mes].apply(parse_mes).astype(int)
                        df_hp_clean['AREA_MAESTRA'] = df_hp[col_ha].apply(limpiar_area)
                        df_hp_clean['PISTA'] = df_hp[col_pista].astype(str).str.upper().str.strip()
                        df_hp_clean['FINCA_MAESTRA'] = 'HISTORICO_SAP'
                        df_hp_clean['OS_MAESTRA'] = 'HIST-' + df_hp_clean.index.astype(str)
                        df_hp_clean['COCTEL_MAESTRO'] = 'COCTEL_HISTORICO'
                        df_hp_clean['HK'] = 'HK-HIST'
                        df_hp_clean['MODELO'] = 'AVION_HISTORICO'
                        df_hp_clean['ORIGEN_BI'] = 'HISTORICO_ANTIGUO'
                        df_hp_clean['FECHA_MAESTRA'] = df_hp_clean.apply(lambda r: f"28/{int(r['MES_NUM']):02d}/{int(r['AÑO'])}", axis=1)
                        df_historico = pd.concat([df_historico, df_hp_clean], ignore_index=True)
    except Exception as e:
        pass

    return df_vivos, df_historico

@st.cache_data(show_spinner=False, ttl=600)
def cargar_matriz_tarifas():
    try:
        file_id = URL_BOVEDA_MAESTRA.split('/d/')[1].split('/')[0]
        dl_url = f'https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx'
        resp = requests.get(dl_url, timeout=30)
        if resp.status_code == 200:
            xls = pd.ExcelFile(io.BytesIO(resp.content))
            df_raw = pd.read_excel(xls, sheet_name="MATRIZ_TARIFAS", header=None)
            datos = df_raw.fillna("").values.tolist()
            if len(datos) > 1:
                df = pd.DataFrame(datos[1:], columns=datos[0])
                df = df.loc[:, df.columns.astype(str).str.strip() != '']
                df = df[df['PISTA'].str.strip() != '']
                return df
    except Exception as e:
        pass
    return pd.DataFrame()

# =================================================================
# 🚀 BLOQUE 4: NÚCLEO OPERATIVO Y VISUAL DEL DASHBOARD
# =================================================================
def ejecutar(supabase_client=None, descargar_matriz_rapida=None, extraer_numero_app=None, procesar_fecha_pesada_app=None, **kwargs):
    
    def fmt_latino(val, decimales=2):
        try: return f"{float(val):,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception: return str(val) if val is not None else ""

    def fmt_dinero(val):
        try: return f"$ {float(val):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception: return f"$ {val}"

    st.header("", anchor="inicio_modulo")

    st.markdown("""
    <style>
    .titulo-principal { color: #0d1b2a; border-bottom: 3px solid #d4af37; padding-bottom: 5px; font-family: 'Arial Black', sans-serif; }
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] { border: 3px solid #0d1b2a !important; border-radius: 8px !important; overflow: hidden !important; }
    .hud-bi { background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); border-left: 5px solid #d4af37; padding: 15px; border-radius: 8px; color: white; box-shadow: 0px 4px 10px rgba(0,0,0,0.15); margin-bottom: 25px; }
    .hud-bi-title { font-size: 11px; font-weight: bold; color: #d4af37; text-transform: uppercase; margin:0; letter-spacing: 1px; }
    .hud-bi-value { font-size: 22px; font-family: 'Arial Black', sans-serif; margin: 5px 0 0 0; }
    
    [data-testid="column"] {
        display: flex !important;
        flex-direction: column !important;
        justify-content: flex-start !important;
        align-items: stretch !important;
    }
    
    div[data-testid="stSelectbox"] > div,
    div[data-testid="stSelectbox"] div[data-baseweb="select"],
    div[data-testid="stMultiSelect"] > div,
    div[data-testid="stMultiSelect"] div[data-baseweb="select"],
    div[data-testid="stDateInput"] > div {
        border: 2px solid #0d1b2a !important;
        border-radius: 8px !important;
        background-color: #ffffff !important;
        box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important;
    }
    
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stMultiSelect"] div[data-baseweb="select"] > div,
    div[data-testid="stDateInput"] div[data-baseweb="baseInput"] {
        background-color: transparent !important;
        border: none !important;
    }
    
    div[data-testid="stSelectbox"] *,
    div[data-testid="stMultiSelect"] *,
    div[data-testid="stDateInput"] input,
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input {
        color: #0d1b2a !important;
        font-weight: bold !important;
    }
    
    span[data-baseweb="tag"] {
        background-color: #d4af37 !important;
        color: #000000 !important;
    }
    span[data-baseweb="tag"] span {
        color: #000000 !important;
    }

    section[data-testid="stMain"] div[role="radiogroup"] {
        border: 2px solid #0d1b2a !important;
        border-radius: 8px !important;
        padding: 10px !important;
        background-color: #ffffff !important;
        box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important;
    }
    section[data-testid="stMain"] div[role="radiogroup"] label {
        color: #0d1b2a !important;
        font-weight: 900 !important;
    }

    div[data-testid="stMainBlockContainer"] label p {
        color: #0d1b2a !important;
        font-weight: 800 !important;
    }

    div[data-testid="stPlotlyChart"] {
        transition: transform 0.3s cubic-bezier(0.25, 0.8, 0.25, 1), box-shadow 0.3s ease !important;
        border-radius: 12px !important;
        background-color: #ffffff !important;
        padding: 8px !important;
        border: 1px solid #e2e8f0 !important;
    }
    div[data-testid="stPlotlyChart"]:hover {
        transform: scale(1.02) !important;
        z-index: 9999 !important;
        position: relative !important;
        box-shadow: 0px 16px 32px rgba(0, 0, 0, 0.2) !important;
        border: 2px solid #d4af37 !important;
    }

    .battle-panel {
        background-color: #ffffff; border-radius: 12px; padding: 20px;
        box-shadow: 0 8px 15px rgba(0,0,0,0.1); border-top: 6px solid;
        margin-bottom: 20px; transition: transform 0.3s ease;
    }
    .battle-panel:hover { transform: translateY(-5px); }
    .battle-title { font-size: 18px; font-weight: 900; text-transform: uppercase; margin-bottom: 15px; text-align: center; }
    .battle-metric-container { display: flex; justify-content: space-between; border-bottom: 1px solid #e2e8f0; padding: 10px 0; }
    .battle-metric-label { font-size: 13px; color: #4a5568; font-weight: bold; } 
    .battle-metric-value { font-size: 18px; font-weight: 900; }
    
    div[data-testid="stTabs"] button[role="tab"] { font-family: 'Arial Black', sans-serif; font-size: 14px; color: #0d1b2a; }
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] { border-bottom-color: #d4af37; background-color: rgba(212, 175, 55, 0.1); }
    </style>
    """, unsafe_allow_html=True)

    c_tit, c_sync = st.columns([3.5, 1.5])
    with c_tit:
        st.markdown("<h1 class='titulo-principal'>📊 Radar Operativo y Financiero <span style='font-size:14px; color:#d4af37;'>(v41.0)</span></h1>", unsafe_allow_html=True)
    with c_sync:
        st.write("")
        if st.button("🔄 Sincronizar Nube (Forzar Datos)", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()
    st.caption(f"🕒 Última carga: {datetime.now().strftime('%H:%M:%S')} (caché de 10 min)")

    try:
        if procesar_fecha_pesada_app is None: procesar_fecha_pesada_app = fecha_fallback
            
        df_vivos, df_historico = cargar_fuentes_maestras_bi(descargar_matriz_rapida)
        if df_vivos.empty and df_historico.empty:
            st.warning("⚠️ Los sistemas de almacenamiento están vacíos.")
            return

        super_base_bi = pd.concat([df_historico, df_vivos], ignore_index=True)
        if 'FINCA_MAESTRA' not in super_base_bi.columns or 'FECHA_MAESTRA' not in super_base_bi.columns:
            st.error("🚨 Columnas críticas estructurales ausentes en la Bóveda.")
            return

        for col_req in ['COSTO_MAESTRO', 'AVION_MAESTRO', 'DOMINIC_MAESTRO', 'AREA_MAESTRA', 'OS_MAESTRA', 'COCTEL_MAESTRO', 'HK', 'MODELO', 'PISTA', 'REND_HR', 'H_PROPORCIONAL', 'SEMANA']:
            if col_req not in super_base_bi.columns: super_base_bi[col_req] = 0.0 if col_req not in ['OS_MAESTRA', 'COCTEL_MAESTRO', 'HK', 'MODELO', 'PISTA'] else ""

        super_base_bi['FINCA_MAESTRA'] = super_base_bi['FINCA_MAESTRA'].astype(str).str.strip().str.upper()
        super_base_bi['OS_MAESTRA'] = super_base_bi['OS_MAESTRA'].astype(str).str.strip().str.upper()
        super_base_bi['COCTEL_MAESTRO'] = super_base_bi['COCTEL_MAESTRO'].astype(str).str.strip().str.upper()
        super_base_bi['COCTEL_CLEAN'] = super_base_bi['COCTEL_MAESTRO']
        
        if 'HK' in super_base_bi.columns: super_base_bi['HK'] = super_base_bi['HK'].astype(str).str.strip().str.upper()
        if 'MODELO' in super_base_bi.columns: super_base_bi['MODELO'] = super_base_bi['MODELO'].astype(str).str.strip().str.upper()

        col_pista = next((c for c in super_base_bi.columns if any(k in str(c).upper() for k in ["PISTA", "ALMACEN", "CENTRO"])), None)
        if col_pista: super_base_bi[col_pista] = super_base_bi[col_pista].astype(str).str.strip().str.upper()

        def aplicar_fecha_robusta(row):
            if row.get('ORIGEN_BI') == 'HISTORICO_ANTIGUO':
                return pd.to_datetime(row.get('FECHA_MAESTRA'), format='%d/%m/%Y', errors='coerce')
            else:
                try: return procesar_fecha_pesada_app(row.get('FECHA_MAESTRA'))
                except Exception: return pd.NaT

        super_base_bi['FECHA_DT'] = super_base_bi.apply(aplicar_fecha_robusta, axis=1)
        super_base_bi = super_base_bi.dropna(subset=['FECHA_DT'])
        
        super_base_bi['FECHA_DT'] = pd.to_datetime(super_base_bi['FECHA_DT'])
        super_base_bi['AÑO'] = super_base_bi['FECHA_DT'].dt.year.astype(int)
        super_base_bi['MES'] = super_base_bi['FECHA_DT'].dt.month.astype(int)
        super_base_bi['TRIMESTRE'] = super_base_bi['FECHA_DT'].dt.quarter.astype(int)
        
        super_base_bi['AREA_NUM'] = super_base_bi['AREA_MAESTRA'].apply(limpiar_area)
        if 'REND_HR' in super_base_bi.columns: super_base_bi['REND_HR'] = super_base_bi['REND_HR'].apply(limpiar_area)
        if 'H_PROPORCIONAL' in super_base_bi.columns: super_base_bi['H_PROPORCIONAL'] = super_base_bi['H_PROPORCIONAL'].apply(limpiar_area)
        if 'SEMANA' in super_base_bi.columns: super_base_bi['SEMANA'] = pd.to_numeric(super_base_bi['SEMANA'], errors='coerce').fillna(0).astype(int)

        super_base_bi = super_base_bi.drop_duplicates(
            subset=['FECHA_DT', 'FINCA_MAESTRA', 'OS_MAESTRA', 'AREA_NUM', 'COCTEL_CLEAN', 'HK'],
            keep='last'
        ).reset_index(drop=True)

        def calcular_costo_real(r):
            if r.get('ORIGEN_BI') == 'ACTUAL':
                tarifa = limpiar_dinero(r.get('AVION_MAESTRO', 0)) + limpiar_dinero(r.get('DOMINIC_MAESTRO', 0))
                area = float(r.get('AREA_NUM', 0))
                return tarifa * area
            else:
                return limpiar_dinero(r.get('COSTO_MAESTRO', 0))
                
        super_base_bi['COSTO_NUM'] = super_base_bi.apply(calcular_costo_real, axis=1)
        super_base_bi['AVION_NUM'] = super_base_bi['AVION_MAESTRO'].apply(limpiar_dinero) + super_base_bi['DOMINIC_MAESTRO'].apply(limpiar_dinero)
        
        super_base_bi['HA_CON_COSTO'] = np.where(super_base_bi['COSTO_NUM'] > 0, super_base_bi['AREA_NUM'], 0)

        total_ha_historicas = super_base_bi['AREA_NUM'].sum()
        ha_con_costo_global = super_base_bi['HA_CON_COSTO'].sum()
        costo_medio_historico = (super_base_bi['COSTO_NUM'].sum() / ha_con_costo_global) if ha_con_costo_global > 0 else 0
        total_ordenes_auditadas = super_base_bi['OS_MAESTRA'].nunique()

        hb1, hb2, hb3 = st.columns(3)
        with hb1: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>ÁREA HISTÓRICA CUBIERTA</p><p class='hud-bi-value'>🗺️ {total_ha_historicas:,.1f} ha</p></div>", unsafe_allow_html=True)
        with hb2: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>TARIFA MEDIA HISTÓRICA</p><p class='hud-bi-value'>💰 $ {formato_latino(costo_medio_historico, 0)}</p></div>", unsafe_allow_html=True)
        with hb3: st.markdown(f"<div class='hud-bi'><p class='hud-bi-title'>ÓRDENES DE SERVICIO AUDITADAS</p><p class='hud-bi-value'>🛰️ {total_ordenes_auditadas:,} OS</p></div>", unsafe_allow_html=True)

        meses_nom = {1:"01-Ene", 2:"02-Feb", 3:"03-Mar", 4:"04-Abr", 5:"05-May", 6:"06-Jun", 7:"07-Jul", 8:"08-Ago", 9:"09-Sep", 10:"10-Oct", 11:"11-Nov", 12:"12-Dic"}
        meses_nom_largo = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
        super_base_bi['MES_NMB'] = super_base_bi['MES'].apply(lambda x: meses_nom.get(x, "Desconocido"))

        st.markdown("### 🎛️ Centro de Comando y Filtros")

        c_tog1, c_tog2 = st.columns(2)
        ver_boveda_tarifas = c_tog1.toggle("🏦 ACTIVAR BÓVEDA DE TARIFAS MAESTRAS", value=False, key="toggle_boveda_tarifas_m8")
        modo_historico_global = c_tog2.toggle("🕰️ ACTIVAR VISOR MACRO-HISTÓRICO", value=False, key="toggle_macro_m8", disabled=ver_boveda_tarifas)

        # ---------------------------------------------------------
        # 🏦 MODO 1: BÓVEDA DE TARIFAS MAESTRAS
        # ---------------------------------------------------------
        if ver_boveda_tarifas:
            with st.container(border=True):
                df_tarifas = cargar_matriz_tarifas()
                if not df_tarifas.empty:
                    st.success("🏦 **BÓVEDA DE TARIFAS:** Estás viendo en exclusiva la matriz central de precios.")
                    
                    cols_base = [c for c in df_tarifas.columns if not str(c).isdigit()]
                    cols_anios = [c for c in df_tarifas.columns if str(c).isdigit()]
                    
                    anios_seleccionados = st.multiselect("📅 Selecciona los años a incluir en el reporte:", cols_anios, default=cols_anios)
                    
                    if not anios_seleccionados:
                        st.warning("⚠️ Selecciona al menos un año para visualizar la matriz.")
                    else:
                        df_tarifas_filtro = df_tarifas[cols_base + anios_seleccionados].copy()
                        df_mostrar = df_tarifas_filtro.copy()
                        
                        renombres = {
                            'PISTA': '📍 BASE OPERATIVA', 
                            'EQUIPO_O_TOPE': '🛩️ TIPO DE EQUIPO / CONCEPTO'
                        }
                        for anio in anios_seleccionados:
                            renombres[anio] = f"📅 {anio}"
                            
                        df_mostrar.rename(columns=renombres, inplace=True)
                        
                        def estilo_matriz(row):
                            estilos = []
                            for col in row.index:
                                if col in ['📍 BASE OPERATIVA', '🛩️ TIPO DE EQUIPO / CONCEPTO']:
                                    estilos.append('background-color: #F1F5F9; color: #0D1B2A; font-weight: 800; border-right: 2px solid #CBD5E1;')
                                else:
                                    estilos.append('background-color: #FFFFFF; color: #1A365D; font-weight: 600; text-align: right;')
                            return estilos

                        st.dataframe(
                            df_mostrar.style.apply(estilo_matriz, axis=1), 
                            use_container_width=True, 
                            hide_index=True
                        )
                        
                        buf_tarifas = io.BytesIO()
                        df_export_t = df_tarifas_filtro.copy()
                        
                        for col in anios_seleccionados:
                            df_export_t[col] = df_export_t[col].apply(limpiar_dinero)

                        with pd.ExcelWriter(buf_tarifas, engine='openpyxl') as writer:
                            df_export_t.to_excel(writer, sheet_name='Tarifario_Maestro', index=False, startrow=3)
                            ws_t = writer.sheets['Tarifario_Maestro']
                            
                            ws_t['A1'] = "REPORTE MÁSTER: MATRIZ DE TARIFAS AUTORIZADAS"
                            ws_t['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                            ws_t['A1'].fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                            ws_t['A1'].alignment = Alignment(horizontal='center', vertical='center')
                            ws_t.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df_export_t.columns))
                            
                            ws_t['A3'] = f"Actualizado a: {datetime.now().strftime('%d/%m/%Y')}"
                            ws_t['A3'].font = Font(italic=True, color="555555", bold=True)
                            ws_t.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(df_export_t.columns))

                            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                            header_font = Font(bold=True, color="000000")
                            for col_num in range(1, len(df_export_t.columns) + 1):
                                cell = ws_t.cell(row=4, column=col_num)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                            ws_t.column_dimensions['A'].width = 12
                            ws_t.column_dimensions['B'].width = 30
                            for col_num in range(3, len(df_export_t.columns) + 1):
                                ws_t.column_dimensions[get_column_letter(col_num)].width = 16
                                
                            for r_idx in range(5, len(df_export_t) + 5):
                                for c_idx in range(1, len(df_export_t.columns) + 1):
                                    cell = ws_t.cell(row=r_idx, column=c_idx)
                                    if c_idx > len(cols_base): 
                                        if cell.value != "" and cell.value is not None and cell.value != 0:
                                            cell.number_format = '$#,##0'
                                    cell.alignment = Alignment(horizontal='left') 

                        st.download_button(
                            label="💾 DESCARGAR TARIFARIO MÁSTER VIP (EXCEL)", 
                            data=buf_tarifas.getvalue(), 
                            file_name=f"Tarifario_Oficial_MasterData_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        try:
                            df_melt = df_export_t.melt(id_vars=['PISTA', 'EQUIPO_O_TOPE'], value_vars=anios_seleccionados, var_name='AÑO', value_name='TARIFA')
                            df_melt['TARIFA'] = pd.to_numeric(df_melt['TARIFA'], errors='coerce').fillna(0)
                            df_melt['LABEL'] = df_melt['PISTA'] + " - " + df_melt['EQUIPO_O_TOPE']
                            
                            fig_tarifas = px.line(df_melt, x='AÑO', y='TARIFA', color='LABEL', markers=True, title="<b>Curva de Inflación Tarifaria Autorizada</b>")
                            fig_tarifas.update_layout(xaxis_title="Año", yaxis_title="Tarifa Autorizada (COP $)", hovermode="x unified")
                            st.plotly_chart(fig_tarifas, use_container_width=True)
                        except Exception: pass
                else:
                    st.warning("⚠️ No se detectó la pestaña 'MATRIZ_TARIFAS' en el Drive o está vacía.")
            
            st.stop()

        # ---------------------------------------------------------
        # 🕰️ MODO 2: VISOR MACRO-HISTÓRICO
        # ---------------------------------------------------------
        elif modo_historico_global:
            st.success("🌐 **MODO MACRO ACTIVADO:** Tienes el control total de la historia operativa. Selecciona el rango de tiempo a analizar.")
            
            cm1, cm2 = st.columns(2)
            fecha_macro_ini = cm1.date_input("F. INICIAL (MACRO):", value=date(2017, 1, 1), min_value=date(2017, 1, 1), max_value=date(2030, 12, 31), key="m8_mac_ini")
            fecha_macro_fin = cm2.date_input("F. FINAL (MACRO):", value=date(2026, 12, 31), min_value=date(2017, 1, 1), max_value=date(2030, 12, 31), key="m8_mac_fin")
            
            df_macro = super_base_bi[(super_base_bi['AREA_NUM'] > 0) & 
                                     (super_base_bi['FECHA_DT'].dt.date >= fecha_macro_ini) & 
                                     (super_base_bi['FECHA_DT'].dt.date <= fecha_macro_fin)].copy()
                                     
            if not col_pista: col_pista = "PISTA"

            st.markdown("---")
            st.markdown(f"#### 📅 Evolución Anual de Hectáreas por Base")
            
            if df_macro.empty:
                st.warning("⚠️ No hay datos históricos registrados en este rango específico de fechas.")
            else:
                pivot_anual = pd.pivot_table(df_macro, values='AREA_NUM', index='AÑO', columns=col_pista, aggfunc='sum', fill_value=0)
                pivot_anual['TOTAL AÑO'] = pivot_anual.sum(axis=1)
                pivot_anual.loc['TOTAL HISTÓRICO'] = pivot_anual.sum(axis=0)
                st.dataframe(pivot_anual.style.format(fmt_latino).background_gradient(cmap="Blues", axis=None), use_container_width=True)

                df_graf = pivot_anual.drop('TOTAL HISTÓRICO', errors='ignore').drop(columns=['TOTAL AÑO'], errors='ignore')
                fig_macro = px.line(df_graf, markers=True, title="<b>Curva Histórica de Aplicación por Pista</b>")
                fig_macro.update_layout(xaxis_title="Año", yaxis_title="Hectáreas Netas")
                st.plotly_chart(fig_macro, use_container_width=True)

                st.markdown("#### 📆 Desglose Detallado: Año y Mes")
                df_macro = df_macro.sort_values(['AÑO', 'MES'])
                pivot_mes = pd.pivot_table(df_macro, values='AREA_NUM', index=['AÑO', 'MES_NMB'], columns=col_pista, aggfunc='sum', fill_value=0, sort=False)
                pivot_mes['TOTAL MES'] = pivot_mes.sum(axis=1)
                st.dataframe(pivot_mes.style.format(fmt_latino).background_gradient(cmap="YlGn", axis=None), use_container_width=True)

                st.markdown("---")
                buffer_macro = io.BytesIO()
                
                df_exp_anual = pivot_anual.reset_index()
                df_exp_mes = pivot_mes.reset_index()
                
                rango_txt_macro = f"{fecha_macro_ini.day}/{fecha_macro_ini.month}/{fecha_macro_ini.year} ⸺ {fecha_macro_fin.day}/{fecha_macro_fin.month}/{fecha_macro_fin.year}"

                with pd.ExcelWriter(buffer_macro, engine='openpyxl') as writer:
                    df_exp_anual.to_excel(writer, sheet_name='Resumen_Anual', index=False, startrow=3)
                    df_exp_mes.to_excel(writer, sheet_name='Desglose_Mensual', index=False, startrow=3)
                    
                    for s_name, df_sheet in [('Resumen_Anual', df_exp_anual), ('Desglose_Mensual', df_exp_mes)]:
                        ws = writer.sheets[s_name]
                        
                        ws['A1'] = f"REPORTE MACRO-HISTÓRICO - {s_name.replace('_', ' ').upper()}"
                        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                        ws['A1'].fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df_sheet.columns))
                        
                        ws['A3'] = f"Período Analizado: {rango_txt_macro}"
                        ws['A3'].font = Font(italic=True, color="555555", bold=True)
                        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(df_sheet.columns))
                        
                        header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                        header_font = Font(bold=True, color="000000")
                        for col_num in range(1, len(df_sheet.columns) + 1):
                            cell = ws.cell(row=4, column=col_num)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            ws.column_dimensions[get_column_letter(col_num)].width = 16
                        
                        for r_idx in range(5, len(df_sheet) + 5):
                            for c_idx in range(1, len(df_sheet.columns) + 1):
                                cell = ws.cell(row=r_idx, column=c_idx)
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='center')

                st.download_button(
                    label="💾 DESCARGAR HISTÓRICO MACRO EN EXCEL VIP",
                    data=buffer_macro.getvalue(),
                    file_name=f"Reporte_Macro_{fecha_macro_ini.strftime('%Y%m%d')}_{fecha_macro_fin.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        # ---------------------------------------------------------
        # 📊 MODO 3: DASHBOARD OPERATIVO Y FINANCIERO ESTÁNDAR
        # ---------------------------------------------------------
        else:
            c1, c2, c3, c4 = st.columns([1.5, 1.0, 1.0, 1.5])
            vista_seleccionada = c1.radio("VISTA OPERATIVA:", ["📊 Resumen Gerencial", "📅 Mapa Semanal", "📈 Dashboard Ejecutivo"], horizontal=True, key="m8_v_final_v40")
            
            fecha_sel_ini = c2.date_input("F. INICIAL:", value=date(2026, 1, 1), min_value=date(2017, 1, 1), max_value=date(2030, 12, 31), key="m8_dat_ini_v40")
            fecha_sel_fin = c3.date_input("F. FINAL:", value=date(2026, 12, 31), min_value=date(2017, 1, 1), max_value=date(2030, 12, 31), key="m8_dat_fin_v40")
            
            pistas_disp = sorted([str(x) for x in super_base_bi[col_pista].dropna().unique() if x != ""]) if col_pista else []
            pistas_sel = c4.multiselect("📍 BASES (PISTAS MÚLTIPLES):", pistas_disp, default=pistas_disp, key="m8_pista_v40")

            if vista_seleccionada != "📈 Dashboard Ejecutivo":
                cc1, cc2, cc3 = st.columns(3)
                mostrar_horas = cc1.checkbox("⏱️ MOSTRAR HORAS", value=True, key="m8_h_v40")
                calcular_rend_prom = cc2.checkbox("🚀 MOSTRAR RENDIMIENTO (ha/hr)", value=True, key="m8_r_v40")
                agrupar_avion = cc3.toggle("✈️ DESGLOSAR POR FLOTA", value=False, key="m8_f_v40")

            df_filt = super_base_bi[(super_base_bi['FECHA_DT'].dt.date >= fecha_sel_ini) & (super_base_bi['FECHA_DT'].dt.date <= fecha_sel_fin)].copy()
            
            if pistas_sel and col_pista:
                df_filt = df_filt[df_filt[col_pista].isin(pistas_sel)]
            else:
                st.warning("⚠️ Selecciona al menos una Base Operativa para generar el radar.")
                return
                
            if df_filt.empty:
                st.warning(f"⚠️ No hay registros de vuelo para las pistas seleccionadas en este rango de fechas.")
                return
                
            st.markdown("---")
            rango_txt = f"{fecha_sel_ini.day} de {meses_nom_largo.get(fecha_sel_ini.month, '')} {fecha_sel_ini.year} ⸺ {fecha_sel_fin.day} de {meses_nom_largo.get(fecha_sel_fin.month, '')} {fecha_sel_fin.year}"
            
            if vista_seleccionada == "📈 Dashboard Ejecutivo":
                st.markdown(f"#### 📈 Dashboard Ejecutivo y BI Financiero")
                st.caption(f"🗓️ *{rango_txt}*")
                
                df_drones = df_filt[df_filt['MODELO'].str.contains('DRON', na=False, case=False) | df_filt['HK'].str.startswith('DRON', na=False)]
                df_aviones = df_filt[~df_filt.index.isin(df_drones.index)]

                total_ha = df_filt['AREA_NUM'].sum()
                total_vuelos = len(df_filt)
                
                ha_drones = df_drones['AREA_NUM'].sum()
                vuelos_drones = len(df_drones)
                costo_tot_drones = df_drones['COSTO_NUM'].sum()
                ha_con_costo_dr = df_drones['HA_CON_COSTO'].sum()
                prom_costo_dr = costo_tot_drones / ha_con_costo_dr if ha_con_costo_dr > 0 else 0

                ha_aviones = df_aviones['AREA_NUM'].sum()
                vuelos_aviones = len(df_aviones)
                costo_tot_aviones = df_aviones['COSTO_NUM'].sum()
                ha_con_costo_av = df_aviones['HA_CON_COSTO'].sum()
                prom_costo_av = costo_tot_aviones / ha_con_costo_av if ha_con_costo_av > 0 else 0
                
                st.markdown("### ⚔️ Batalla de Escuadrones: ✈️ Aviones vs 🛸 Drones")
                col_av, col_dr = st.columns(2)
                
                with col_av:
                    st.markdown(f"""
                    <div class='battle-panel' style='border-top-color: #2F75B5;'>
                        <div class='battle-title' style='color: #2F75B5;'>✈️ Flota de Aviones</div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>TOTAL FACTURADO ($)</span>
                            <span class='battle-metric-value' style='color: #2F75B5;'>{fmt_dinero(costo_tot_aviones)}</span>
                        </div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>TARIFA PROMEDIO ($/ha)</span>
                            <span class='battle-metric-value' style='color: #2F75B5;'>{fmt_dinero(prom_costo_av)}</span>
                        </div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>HECTÁREAS APLICADAS</span>
                            <span class='battle-metric-value'>{fmt_latino(ha_aviones, 1)} ha</span>
                        </div>
                        <div class='battle-metric-container' style='border-bottom: none;'>
                            <span class='battle-metric-label'>MISIONES COMPLETADAS</span>
                            <span class='battle-metric-value'>{vuelos_aviones}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                with col_dr:
                    simular_dr = st.toggle("🔮 Activar Simulador de Proyección (Drones)", value=False)
                    if simular_dr:
                        pct_aumento = st.slider("📈 Aumento Proyectado de Hectáreas (%)", 0, 500, 50, 5)
                        ha_drones_final = ha_drones * (1 + (pct_aumento / 100.0))
                        costo_tot_drones_final = ha_drones_final * prom_costo_dr
                        
                        cm1, cm2 = st.columns(2)
                        cm1.metric("📍 Ha Actuales (Real)", f"{fmt_latino(ha_drones, 1)} ha")
                        cm2.metric(f"🚀 Ha Proyectadas (+{pct_aumento}%)", f"{fmt_latino(ha_drones_final, 1)} ha", delta=f"{fmt_latino(ha_drones_final - ha_drones, 1)} ha")
                        st.markdown("<br>", unsafe_allow_html=True)
                        
                        lbl_ha = f"HECTÁREAS (Proyectado +{pct_aumento}%)"
                        lbl_dinero = "FACTURACIÓN PROYECTADA ($)"
                        color_borde = "#e67e22" 
                        titulo_panel = "🛸 Drones (MODO SIMULACIÓN)"
                    else:
                        ha_drones_final = ha_drones
                        costo_tot_drones_final = costo_tot_drones
                        lbl_ha = "HECTÁREAS APLICADAS"
                        lbl_dinero = "TOTAL FACTURADO ($)"
                        color_borde = "#27AE60" 
                        titulo_panel = "🛸 Flota de Drones"

                    st.markdown(f"""
                    <div class='battle-panel' style='border-top-color: {color_borde};'>
                        <div class='battle-title' style='color: {color_borde};'>{titulo_panel}</div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>{lbl_dinero}</span>
                            <span class='battle-metric-value' style='color: {color_borde};'>{fmt_dinero(costo_tot_drones_final)}</span>
                        </div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>TARIFA PROMEDIO ($/ha)</span>
                            <span class='battle-metric-value' style='color: {color_borde};'>{fmt_dinero(prom_costo_dr)}</span>
                        </div>
                        <div class='battle-metric-container'>
                            <span class='battle-metric-label'>{lbl_ha}</span>
                            <span class='battle-metric-value'>{fmt_latino(ha_drones_final, 1)} ha</span>
                        </div>
                        <div class='battle-metric-container' style='border-bottom: none;'>
                            <span class='battle-metric-label'>MISIONES COMPLETADAS</span>
                            <span class='battle-metric-value'>{vuelos_drones}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                st.write("")
                
                df_dash = df_filt.groupby(col_pista).agg(
                    VUELOS=(col_pista, 'count'),
                    HECTAREAS=('AREA_NUM', 'sum'),
                    COSTO_TOTAL=('COSTO_NUM', 'sum'),
                    HA_CON_COSTO=('HA_CON_COSTO', 'sum')
                ).reset_index()
                
                df_dash['% VUELOS'] = (df_dash['VUELOS'] / total_vuelos) * 100 if total_vuelos > 0 else 0
                df_dash['% HECTAREAS'] = (df_dash['HECTAREAS'] / total_ha) * 100 if total_ha > 0 else 0
                df_dash = df_dash.sort_values(by='HECTAREAS', ascending=False)

                g1, g2 = st.columns(2)
                df_dash['TXT_PCT'] = df_dash['% HECTAREAS'].apply(lambda x: f"{x:.1f}%".replace(".", ","))
                
                fig_pie = px.pie(df_dash, values='VUELOS', names=col_pista, hole=0.5, 
                                 title="<b>Distribución de Vuelos por Pista</b>", color_discrete_sequence=px.colors.qualitative.Prism)
                fig_pie.update_traces(textposition='inside', textinfo='percent+label', texttemplate='<b>%{label}</b><br>%{percent}')
                fig_pie.update_layout(
                    separators=",.", 
                    showlegend=False, 
                    margin=dict(t=50, b=50, l=60, r=60), 
                    height=400,
                    uniformtext_minsize=11, 
                    uniformtext_mode='hide'
                )
                g1.plotly_chart(fig_pie, use_container_width=True)

                fig_bar = px.bar(df_dash.sort_values('HECTAREAS', ascending=True), 
                                 x='HECTAREAS', y=col_pista, orientation='h',
                                 title="<b>Volumen de Hectáreas por Base</b>",
                                 text='TXT_PCT',
                                 color='HECTAREAS', color_continuous_scale='Blues')
                fig_bar.update_layout(
                    separators=",.", 
                    xaxis_title="Hectáreas Netas", 
                    yaxis_title="", 
                    coloraxis_showscale=False, 
                    margin=dict(t=50, b=50, l=40, r=40), 
                    height=400
                )
                g2.plotly_chart(fig_bar, use_container_width=True)

                # =========================================================
                # 💎 REDISEÑO VIP DEL RANKING DE AERONAVES (TABLA EJECUTIVA)
                # =========================================================
                st.markdown("---")
                st.markdown("##### 🏆 Ranking: Impacto de Costos por Aeronave (Facturación a la Empresa)")
                
                df_hk = df_filt.groupby(['HK']).agg(
                    MISIONES=('HK', 'count'),
                    HECTAREAS=('AREA_NUM', 'sum'),
                    COSTO_TOTAL=('COSTO_NUM', 'sum'),
                    HA_CON_COSTO=('HA_CON_COSTO', 'sum')
                ).reset_index()
                
                df_hk['TARIFA_PROM'] = np.where(df_hk['HA_CON_COSTO'] > 0, df_hk['COSTO_TOTAL'] / df_hk['HA_CON_COSTO'], 0)
                df_hk = df_hk.sort_values('COSTO_TOTAL', ascending=False).reset_index(drop=True)
                
                # Renderizado con encabezados limpios y formatters nativos de Streamlit
                st.dataframe(
                    df_hk[['HK', 'MISIONES', 'HECTAREAS', 'COSTO_TOTAL', 'TARIFA_PROM']],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "HK": st.column_config.TextColumn(
                            "🛩️ MATRÍCULA / HK",
                            help="Matrícula o identificador único de la aeronave o dron"
                        ),
                        "MISIONES": st.column_config.NumberColumn(
                            "🛰️ MISIONES",
                            format="%d"
                        ),
                        "HECTAREAS": st.column_config.NumberColumn(
                            "🗺️ HECTÁREAS NETAS",
                            format="%.2f ha"
                        ),
                        "COSTO_TOTAL": st.column_config.ProgressColumn(
                            "💰 FACTURACIÓN TOTAL ($)",
                            format="$ %,.0f",
                            min_value=0,
                            max_value=float(df_hk['COSTO_TOTAL'].max()) if not df_hk.empty else 100
                        ),
                        "TARIFA_PROM": st.column_config.NumberColumn(
                            "📊 TARIFA PROM. ($/HA)",
                            format="$ %,.0f"
                        )
                    }
                )

            elif vista_seleccionada == "📊 Resumen Gerencial":
                st.markdown(f"#### 📑 Consolidado Gerencial")
                st.caption(f"🗓️ *{rango_txt}*")
                tabla_final = []
                total_hr_gral, total_ha_gral, total_costo_gral, total_ha_costo_gral = 0, 0, 0, 0
                
                col_rend = 'REND_HR' if 'REND_HR' in df_filt.columns else ('H_PROPORCIONAL' if 'H_PROPORCIONAL' in df_filt.columns else 'AREA_NUM')

                if agrupar_avion:
                    df_gerencia = df_filt.groupby([col_pista, 'HK', 'AÑO', 'MES_NMB']).agg(
                        REND_HR=(col_rend, 'sum'), 
                        AREA_FUMIG=('AREA_NUM', 'sum'),
                        COSTO_TOT=('COSTO_NUM', 'sum'),
                        HA_CON_COSTO=('HA_CON_COSTO', 'sum')
                    ).reset_index()
                    
                    for pista in sorted(df_gerencia[col_pista].unique()):
                        df_pista = df_gerencia[df_gerencia[col_pista] == pista]
                        sum_hr_pista = df_pista['REND_HR'].sum()
                        sum_ha_pista = df_pista['AREA_FUMIG'].sum()
                        sum_costo_pista = df_pista['COSTO_TOT'].sum()
                        sum_hcc_pista = df_pista['HA_CON_COSTO'].sum()
                        
                        fila_pista = {'NIVEL': f"📍 BASE: {pista}", 'AVIÓN (HK)': '', 'AÑO': 'TOTAL', 'MES': ''}
                        if mostrar_horas or calcular_rend_prom: fila_pista['REND (hr)'] = sum_hr_pista
                        fila_pista['ÁREA FUMIG (ha)'] = sum_ha_pista
                        if calcular_rend_prom: fila_pista['PROMEDIO (ha/hr)'] = sum_ha_pista / sum_hr_pista if sum_hr_pista > 0 else 0.0
                        
                        fila_pista['COSTO TOTAL ($)'] = sum_costo_pista
                        fila_pista['TARIFA PROM ($/ha)'] = sum_costo_pista / sum_hcc_pista if sum_hcc_pista > 0 else 0.0
                        tabla_final.append(fila_pista)
                        
                        for hk in sorted(df_pista['HK'].unique()):
                            datos_hk = df_pista[df_pista['HK'] == hk].sort_values(by=['AÑO'])
                            sum_hr_hk = datos_hk['REND_HR'].sum()
                            sum_ha_hk = datos_hk['AREA_FUMIG'].sum()
                            sum_costo_hk = datos_hk['COSTO_TOT'].sum()
                            sum_hcc_hk = datos_hk['HA_CON_COSTO'].sum()
                            
                            modelo = str(df_filt[df_filt['HK'] == hk]['MODELO'].iloc[0]).upper() if not df_filt[df_filt['HK'] == hk].empty else ""
                            es_dron = "DRON" in modelo or hk.startswith("DRON")
                            emoji = "🛸 DRON:" if es_dron else "✈️ AVION:"
                            
                            fila_hk = {'NIVEL': '', 'AVIÓN (HK)': f"{emoji} {hk}", 'AÑO': 'Total Flota', 'MES': ''}
                            if mostrar_horas or calcular_rend_prom: fila_hk['REND (hr)'] = sum_hr_hk
                            fila_hk['ÁREA FUMIG (ha)'] = sum_ha_hk
                            if calcular_rend_prom: fila_hk['PROMEDIO (ha/hr)'] = sum_ha_hk / sum_hr_hk if sum_hr_hk > 0 else 0.0
                            
                            fila_hk['COSTO TOTAL ($)'] = sum_costo_hk
                            fila_hk['TARIFA PROM ($/ha)'] = sum_costo_hk / sum_hcc_hk if sum_hcc_hk > 0 else 0.0
                            tabla_final.append(fila_hk)
                            
                            for _, row in datos_hk.iterrows():
                                fila_mes = {'NIVEL': '', 'AVIÓN (HK)': '', 'AÑO': f"  ↳ {row['AÑO']}", 'MES': f"{row['MES_NMB']}"}
                                if mostrar_horas or calcular_rend_prom: fila_mes['REND (hr)'] = row['REND_HR']
                                fila_mes['ÁREA FUMIG (ha)'] = row['AREA_FUMIG']
                                if calcular_rend_prom: fila_mes['PROMEDIO (ha/hr)'] = row['AREA_FUMIG'] / row['REND_HR'] if row['REND_HR'] > 0 else 0.0
                                
                                fila_mes['COSTO TOTAL ($)'] = row['COSTO_TOT']
                                fila_mes['TARIFA PROM ($/ha)'] = row['COSTO_TOT'] / row['HA_CON_COSTO'] if row['HA_CON_COSTO'] > 0 else 0.0
                                tabla_final.append(fila_mes)
                                
                        total_hr_gral += sum_hr_pista
                        total_ha_gral += sum_ha_pista
                        total_costo_gral += sum_costo_pista
                        total_ha_costo_gral += sum_hcc_pista
                        
                    fila_tot = {'NIVEL': '👑 TOTAL GENERAL', 'AVIÓN (HK)': '', 'AÑO': '', 'MES': ''}
                    if mostrar_horas or calcular_rend_prom: fila_tot['REND (hr)'] = total_hr_gral
                    fila_tot['ÁREA FUMIG (ha)'] = total_ha_gral
                    if calcular_rend_prom: fila_tot['PROMEDIO (ha/hr)'] = total_ha_gral / total_hr_gral if total_hr_gral > 0 else 0.0
                    
                    fila_tot['COSTO TOTAL ($)'] = total_costo_gral
                    fila_tot['TARIFA PROM ($/ha)'] = total_costo_gral / total_ha_costo_gral if total_ha_costo_gral > 0 else 0.0
                    tabla_final.append(fila_tot)
                    
                else:
                    df_gerencia = df_filt.groupby([col_pista, 'AÑO', 'MES_NMB']).agg(
                        REND_HR=(col_rend, 'sum'), 
                        AREA_FUMIG=('AREA_NUM', 'sum'),
                        COSTO_TOT=('COSTO_NUM', 'sum'),
                        HA_CON_COSTO=('HA_CON_COSTO', 'sum')
                    ).reset_index()
                    
                    for pista in sorted(df_gerencia[col_pista].unique()):
                        datos_pista = df_gerencia[df_gerencia[col_pista] == pista].sort_values(by=['AÑO', 'MES_NMB'])
                        sum_hr = datos_pista['REND_HR'].sum()
                        sum_ha = datos_pista['AREA_FUMIG'].sum()
                        sum_costo = datos_pista['COSTO_TOT'].sum()
                        sum_hcc = datos_pista['HA_CON_COSTO'].sum()
                        
                        fila_sub = {'NIVEL': f"📍 BASE: {pista}", 'AÑO': 'TOTAL', 'MES': ''}
                        if mostrar_horas or calcular_rend_prom: fila_sub['REND (hr)'] = sum_hr
                        fila_sub['ÁREA FUMIG (ha)'] = sum_ha
                        if calcular_rend_prom: fila_sub['PROMEDIO (ha/hr)'] = sum_ha / sum_hr if sum_hr > 0 else 0.0
                        
                        fila_sub['COSTO TOTAL ($)'] = sum_costo
                        fila_sub['TARIFA PROM ($/ha)'] = sum_costo / sum_hcc if sum_hcc > 0 else 0.0
                        tabla_final.append(fila_sub)
                        
                        for _, row in datos_pista.iterrows():
                            fila_mes = {'NIVEL': '', 'AÑO': f"  ↳ {row['AÑO']}", 'MES': f"{row['MES_NMB']}"}
                            if mostrar_horas or calcular_rend_prom: fila_mes['REND (hr)'] = row['REND_HR']
                            fila_mes['ÁREA FUMIG (ha)'] = row['AREA_FUMIG']
                            if calcular_rend_prom: fila_mes['PROMEDIO (ha/hr)'] = row['AREA_FUMIG'] / row['REND_HR'] if row['REND_HR'] > 0 else 0.0
                            
                            fila_mes['COSTO TOTAL ($)'] = row['COSTO_TOT']
                            fila_mes['TARIFA PROM ($/ha)'] = row['COSTO_TOT'] / row['HA_CON_COSTO'] if row['HA_CON_COSTO'] > 0 else 0.0
                            tabla_final.append(fila_mes)
                            
                        total_hr_gral += sum_hr
                        total_ha_gral += sum_ha
                        total_costo_gral += sum_costo
                        total_ha_costo_gral += sum_hcc
                        
                    fila_tot = {'NIVEL': '👑 TOTAL GENERAL', 'AÑO': '', 'MES': ''}
                    if mostrar_horas or calcular_rend_prom: fila_tot['REND (hr)'] = total_hr_gral
                    fila_tot['ÁREA FUMIG (ha)'] = total_ha_gral
                    if calcular_rend_prom: fila_tot['PROMEDIO (ha/hr)'] = total_ha_gral / total_hr_gral if total_hr_gral > 0 else 0.0
                    
                    fila_tot['COSTO TOTAL ($)'] = total_costo_gral
                    fila_tot['TARIFA PROM ($/ha)'] = total_costo_gral / total_ha_costo_gral if total_ha_costo_gral > 0 else 0.0
                    tabla_final.append(fila_tot)

                df_visual = pd.DataFrame(tabla_final)
                
                cols_base = ['NIVEL']
                if agrupar_avion: cols_base.append('AVIÓN (HK)')
                cols_base.extend(['AÑO', 'MES'])
                
                cols_op = cols_base.copy()
                if mostrar_horas or calcular_rend_prom: cols_op.append('REND (hr)')
                cols_op.append('ÁREA FUMIG (ha)')
                if calcular_rend_prom: cols_op.append('PROMEDIO (ha/hr)')
                
                cols_fin = cols_base.copy()
                cols_fin.extend(['COSTO TOTAL ($)', 'TARIFA PROM ($/ha)'])
                
                df_operativo = df_visual[cols_op].copy()
                df_financiero = df_visual[cols_fin].copy()
                
                fmt_op = {'ÁREA FUMIG (ha)': fmt_latino}
                if mostrar_horas or calcular_rend_prom: fmt_op['REND (hr)'] = fmt_latino
                if calcular_rend_prom: fmt_op['PROMEDIO (ha/hr)'] = fmt_latino
                
                fmt_fin = {'COSTO TOTAL ($)': fmt_dinero, 'TARIFA PROM ($/ha)': fmt_dinero}

                def aplicar_estilos_originales(row):
                    if "BASE:" in str(row['NIVEL']): return ['background-color: #d1ecf1; font-weight: bold; color: #0c5460;'] * len(row)
                    elif "TOTAL GENERAL" in str(row['NIVEL']): return ['background-color: #c3e6cb; font-weight: bold; color: #155724;'] * len(row)
                    elif 'AVIÓN (HK)' in row and ("✈️" in str(row.get('AVIÓN (HK)','')) or "🛸" in str(row.get('AVIÓN (HK)',''))):
                        return ['background-color: #f8f9fa; font-weight: bold; color: #212529;'] * len(row)
                    return [''] * len(row)

                tab_op, tab_fin = st.tabs(["🛩️ CONSOLIDADO OPERATIVO", "💰 CONSOLIDADO FINANCIERO"])
                
                with tab_op:
                    st.markdown("##### 🛩️ Desglose Operativo Puro")
                    st.caption("Perfecto para compartir: Muestra rendimiento físico de las aeronaves ocultando las tarifas monetarias.")
                    st.dataframe(df_operativo.style.apply(aplicar_estilos_originales, axis=1).format(fmt_op), use_container_width=True, hide_index=True)
                    
                with tab_fin:
                    st.markdown("##### 💰 Desglose Financiero Confidencial")
                    st.caption("Uso gerencial: Muestra el impacto en pesos facturados por cada base y matrícula.")
                    st.dataframe(df_financiero.style.apply(aplicar_estilos_originales, axis=1).format(fmt_fin), use_container_width=True, hide_index=True)

            else:
                matriz = pd.pivot_table(df_filt, values='AREA_NUM', index=['AÑO', 'MES_NMB'], columns='PISTA', aggfunc='sum', fill_value=0)
                matriz['TOTAL'] = matriz.sum(axis=1)
                
                st.markdown(f"#### 🛩️ Evolución de Hectáreas por Año y Mes")
                st.caption(f"🗓️ *{rango_txt}*")
                st.dataframe(matriz.style.format(fmt_latino).background_gradient(cmap="YlGn", axis=None), use_container_width=True)

            st.markdown("---")
            buffer_rep = io.BytesIO()
            rango_label = f"{fecha_sel_ini.strftime('%Y%m%d')}_{fecha_sel_fin.strftime('%Y%m%d')}"
            
            if vista_seleccionada == "📈 Dashboard Ejecutivo":
                wb = Workbook()
                ws = wb.active
                ws.title = "Dashboard Ejecutivo"
                
                fill_header = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                font_header = Font(color="FFFFFF", bold=True)
                fill_tot = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                font_tot = Font(color="000000", bold=True)
                borde = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"),
                               top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))
                align_center = Alignment(horizontal='center', vertical='center')
                
                ws['B2'] = "REPORTE GERENCIAL: RADAR DE HECTÁREAS, MISIONES Y COSTOS"
                ws['B2'].font = Font(size=14, bold=True, color="0D1B2A")
                ws['B3'] = f"Período Analizado: {rango_txt}"
                ws['B3'].font = Font(italic=True, color="555555")
                
                df_export = df_dash.copy()
                total_vuelos_exp = df_export['VUELOS'].sum()
                total_ha_exp = df_export['HECTAREAS'].sum()
                total_costo_exp = df_export['COSTO_TOTAL'].sum()
                total_hcc_exp = df_export['HA_CON_COSTO'].sum()
                
                headers = ['BASE OPERATIVA', 'TOTAL MISIONES', 'HECTÁREAS NETAS', 'COSTO TOTAL ($)', 'TARIFA PROM ($/ha)']
                start_row = 6
                for col_idx, header in enumerate(headers, start=2):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                    cell.border = borde
                    
                curr_row = start_row + 1
                for _, row in df_export.iterrows():
                    ws.cell(row=curr_row, column=2, value=row[col_pista]).border = borde
                    ws.cell(row=curr_row, column=3, value=row['VUELOS']).number_format = '#,##0'
                    ws.cell(row=curr_row, column=3).border = borde
                    ws.cell(row=curr_row, column=4, value=row['HECTAREAS']).number_format = '#,##0.00'
                    ws.cell(row=curr_row, column=4).border = borde
                    ws.cell(row=curr_row, column=5, value=row['COSTO_TOTAL']).number_format = '$#,##0'
                    ws.cell(row=curr_row, column=5).border = borde
                    tarifa_base = row['COSTO_TOTAL'] / row['HA_CON_COSTO'] if row['HA_CON_COSTO'] > 0 else 0
                    ws.cell(row=curr_row, column=6, value=tarifa_base).number_format = '$#,##0'
                    ws.cell(row=curr_row, column=6).border = borde
                    curr_row += 1
                    
                ws.cell(row=curr_row, column=2, value="TOTAL GENERAL").fill = fill_tot
                ws.cell(row=curr_row, column=2).font = font_tot
                ws.cell(row=curr_row, column=2).border = borde
                ws.cell(row=curr_row, column=3, value=total_vuelos_exp).fill = fill_tot
                ws.cell(row=curr_row, column=3).font = font_tot
                ws.cell(row=curr_row, column=3).border = borde
                ws.cell(row=curr_row, column=3).number_format = '#,##0'
                ws.cell(row=curr_row, column=4, value=total_ha_exp).fill = fill_tot
                ws.cell(row=curr_row, column=4).font = font_tot
                ws.cell(row=curr_row, column=4).border = borde
                ws.cell(row=curr_row, column=4).number_format = '#,##0.00'
                ws.cell(row=curr_row, column=5, value=total_costo_exp).fill = fill_tot
                ws.cell(row=curr_row, column=5).font = font_tot
                ws.cell(row=curr_row, column=5).border = borde
                ws.cell(row=curr_row, column=5).number_format = '$#,##0'
                tarifa_total = total_costo_exp / total_hcc_exp if total_hcc_exp > 0 else 0
                ws.cell(row=curr_row, column=6, value=tarifa_total).fill = fill_tot
                ws.cell(row=curr_row, column=6).font = font_tot
                ws.cell(row=curr_row, column=6).border = borde
                ws.cell(row=curr_row, column=6).number_format = '$#,##0'
                
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                
                data_len = len(df_export)
                if data_len > 0:
                    cats = Reference(ws, min_col=2, min_row=start_row+1, max_row=start_row+data_len)
                    
                    bar_chart = BarChart()
                    bar_chart.type = "bar"
                    bar_chart.style = 11
                    bar_chart.title = "Costo Total por Base Operativa ($)"
                    data_costo = Reference(ws, min_col=5, min_row=start_row, max_row=start_row+data_len)
                    bar_chart.add_data(data_costo, titles_from_data=True)
                    bar_chart.set_categories(cats)
                    bar_chart.legend = None
                    bar_chart.dataLabels = DataLabelList()
                    bar_chart.dataLabels.showVal = True
                    ws.add_chart(bar_chart, "H5")
                    
                    pie_chart = DoughnutChart()
                    pie_chart.title = "Distribución de Hectáreas"
                    pie_chart.style = 2
                    data_ha = Reference(ws, min_col=4, min_row=start_row, max_row=start_row+data_len)
                    pie_chart.add_data(data_ha, titles_from_data=True)
                    pie_chart.set_categories(cats)
                    pie_chart.dataLabels = DataLabelList()
                    pie_chart.dataLabels.showPercent = True 
                    pie_chart.dataLabels.showCatName = False
                    ws.add_chart(pie_chart, "H20")
                
                wb.save(buffer_rep)

            else:
                with pd.ExcelWriter(buffer_rep, engine='openpyxl') as writer:
                    if vista_seleccionada == "📊 Resumen Gerencial":
                        sheets_data = [
                            ('Resumen_Operativo', df_operativo, ['REND (hr)', 'ÁREA FUMIG (ha)', 'PROMEDIO (ha/hr)']),
                            ('Resumen_Financiero', df_financiero, ['COSTO TOTAL ($)', 'TARIFA PROM ($/ha)'])
                        ]
                        
                        for s_name, df_sheet, num_cols in sheets_data:
                            df_sheet.to_excel(writer, sheet_name=s_name, index=False, startrow=3)
                            ws = writer.sheets[s_name]
                            
                            ws['A1'] = f"REPORTE GERENCIAL - {s_name.replace('_', ' ').upper()}"
                            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                            ws['A1'].fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df_sheet.columns))
                            
                            ws['A3'] = f"Período Analizado: {rango_txt}"
                            ws['A3'].font = Font(italic=True, color="333333", bold=True)
                            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(df_sheet.columns))

                            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                            header_font = Font(bold=True, color="000000")
                            for col_num in range(1, len(df_sheet.columns) + 1):
                                cell = ws.cell(row=4, column=col_num)
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                ws.column_dimensions[get_column_letter(col_num)].width = 22
                            
                            for r_idx in range(5, len(df_sheet) + 5):
                                nivel_val = str(ws.cell(row=r_idx, column=1).value)
                                
                                bg_color = None
                                is_bold = False
                                
                                if "BASE:" in nivel_val:
                                    bg_color = "D1ECF1"
                                    is_bold = True
                                elif "TOTAL GENERAL" in nivel_val:
                                    bg_color = "C3E6CB"
                                    is_bold = True
                                    
                                for c_idx in range(1, len(df_sheet.columns) + 1):
                                    cell = ws.cell(row=r_idx, column=c_idx)
                                    
                                    if bg_color:
                                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                                    if is_bold:
                                        cell.font = Font(bold=True)
                                        
                                    col_name = df_sheet.columns[c_idx - 1]
                                    if col_name in ['REND (hr)', 'ÁREA FUMIG (ha)', 'PROMEDIO (ha/hr)']:
                                        try:
                                            if cell.value != "" and cell.value is not None:
                                                cell.value = float(cell.value)
                                                cell.number_format = '#,##0.00'
                                                cell.alignment = Alignment(horizontal='center')
                                        except Exception: pass
                                    elif col_name in ['COSTO TOTAL ($)', 'TARIFA PROM ($/ha)']:
                                        try:
                                            if cell.value != "" and cell.value is not None:
                                                cell.value = float(cell.value)
                                                cell.number_format = '$#,##0'
                                                cell.alignment = Alignment(horizontal='center')
                                        except Exception: pass
                            
                    elif vista_seleccionada == "📅 Mapa Semanal":
                        matriz.to_excel(writer, sheet_name='Mapa_Mensual', startrow=3)
                        ws = writer.sheets['Mapa_Mensual']
                        
                        ws['A1'] = "MAPA MENSUAL DE HECTÁREAS"
                        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                        ws['A1'].fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
                        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(matriz.columns)+1)
                        
                        ws['A3'] = f"Período Analizado: {rango_txt}"
                        ws['A3'].font = Font(italic=True, color="333333", bold=True)
                        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(matriz.columns)+1)
                        
                        header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
                        for col_num in range(1, len(matriz.columns) + 2):
                            cell = ws.cell(row=4, column=col_num)
                            cell.fill = header_fill
                            cell.font = Font(bold=True)
                            ws.column_dimensions[get_column_letter(col_num)].width = 15
                            
                        for row in ws.iter_rows(min_row=5):
                            for cell in row:
                                if cell.column > 1:
                                    try:
                                        if cell.value != "" and cell.value is not None:
                                            cell.value = float(cell.value)
                                            cell.number_format = '#,##0.00'
                                            cell.alignment = Alignment(horizontal='center')
                                    except Exception: pass
                
            st.download_button(
                label="💾 DESCARGAR REPORTE EJECUTIVO EN EXCEL",
                data=buffer_rep.getvalue(),
                file_name=f"Reporte_Ejecutivo_BI_{rango_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )                        

    except Exception as e:
        logging.error(f"Fallo procesando el reporte: {e}", exc_info=True)
        st.error("🚨 Ocurrió un error interno generando el reporte. El equipo técnico ya fue notificado.")

if __name__ == "__main__":
    pass
