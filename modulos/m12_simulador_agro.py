import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import gspread
import io
import re
from datetime import datetime, date
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# ⚙️ CONSTANTES CENTRALIZADAS
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

# =================================================================
# ⚡ MOTOR DE CONEXIÓN UNIFICADO
# =================================================================
@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

# =================================================================
# 🛡️ UTILIDADES DE PURIFICACIÓN Y FORMATO
# =================================================================
def limpiar_orden_extrema(val):
    if pd.isna(val) or str(val).strip() == "": return "SIN_ORDEN"
    v = str(val).upper().strip()
    v = re.sub(r'\s+', '', v) 
    if v.endswith('.0'): v = v[:-2] 
    return v

def limpiar_cantidad(val):
    if isinstance(val, pd.Series): val = val.iloc[0]
    if pd.isna(val) or str(val).strip() == "": return 0.0
    try:
        texto = str(val).replace(" ", "").strip()
        if "," in texto and "." in texto:
            if texto.rfind(".") > texto.rfind(","): texto = texto.replace(",", "")
            else: texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        return float(texto)
    except Exception:
        return 0.0

def limpiar_moneda(val):
    if isinstance(val, pd.Series): val = val.iloc[0]
    if pd.isna(val) or str(val).strip() == "": return 0.0
    try:
        texto = str(val).upper().replace("$", "").replace("COP", "").replace(" ", "").strip()
        if "." in texto and "," in texto:
            if texto.rfind(".") > texto.rfind(","): texto = texto.replace(",", "")
            else: texto = texto.replace(".", "").replace(",", ".")
        else:
            sep = "." if "." in texto else ("," if "," in texto else None)
            if sep:
                if texto.count(sep) > 1:
                    texto = texto.replace(sep, "")
                elif len(texto.split(sep)[-1]) == 3: 
                    texto = texto.replace(sep, "")
                else: 
                    texto = texto.replace(sep, ".")
        return float(texto) if texto else 0.0
    except Exception:
        return 0.0

def parsear_fecha_robusta(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    s = str(val).strip().lower()
    if s.isdigit(): return pd.to_datetime('1899-12-30') + pd.to_timedelta(int(s), 'D')
    meses = {'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12}
    match1 = re.search(r'(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})', s)
    if match1:
        dia_str, mes_str, anio_str = match1.groups()
        if mes_str in meses: return pd.to_datetime(f"{anio_str}-{meses[mes_str]:02d}-{int(dia_str):02d}")
    match2 = re.search(r'([a-z]+)\s+(\d{1,2}),\s+(\d{4})', s)
    if match2:
        mes_str, dia_str, anio_str = match2.groups()
        if mes_str in meses: return pd.to_datetime(f"{anio_str}-{meses[mes_str]:02d}-{int(dia_str):02d}")
    try: 
        return pd.to_datetime(s.split(" ")[0], dayfirst=True, errors='coerce')
    except Exception: 
        return pd.NaT

def purificar_datos_vuelo(eq_raw, pista_raw):
    eq = str(eq_raw).upper()
    p = str(pista_raw).upper()
    if "DRON" in eq or "DRONE" in eq:
        if "DATAROT" in eq or "PLUC" in p: return "DRONE DATAROT", "PLUC"
        if "NORTE" in eq or "PDIV" in p: return "DRONE NORTE", "PDIV"
        if "AVIL" in eq or "TEHO" in p: return "DRONE AVIL", "TEHO"
        if "GENESYS" in eq or "LUCI" in p: return "DRONE GENESYS", "LUCI"
        return "DRONE GENESYS", "LUCI" 
    if "TRUSH" in eq or "THRUS" in eq or "OMANDER" in eq: return "THRUS SR2", "AEROPENORT"
    if "PAWNEE" in eq or "BRAVO" in eq or "PIPER PA 36" in eq: return "PIPER PA 36-375", "AEROPENORT"
    if "AIR TRACTOR" in eq or "TRACTOR" in eq or "TOR" in eq: return "AIR TRACTOR", "FUMIGARAY"
    if "CESSNA" in eq or "PIPER PA 25" in eq:
        if "ASA" in p or "ASA" in eq: return "CESSNA ASA", "ASA"
        if "FUMIGARAY" in p or "FUMIGARAY" in eq: return "CESSNA FUMIGARAY", "FUMIGARAY"
        return "CESSNA O PIPER PA 25", "AEROPENORT"
    return "IGNORAR", "IGNORAR"

def formato_latino(numero, decimales=0):
    if pd.isna(numero) or numero is None: return "0"
    try:
        num = float(numero)
        if num == 0: return "0"
        if decimales == 0: texto_us = f"{num:,.0f}"
        else: texto_us = f"{num:,.{decimales}f}"
        return texto_us.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0"

# =================================================================
# 📦 EXTRACCIÓN DE DATOS BLINDADA E INTEGRACIÓN DE CONFIGURACIÓN
# =================================================================
@st.cache_data(show_spinner=False, ttl=600)
def extraer_datos_boveda():
    gc = obtener_cliente_gspread_unificado()
    df_t1, df_t2 = pd.DataFrame(), pd.DataFrame()
    dict_tarifas_conf = {}
    if not gc: return df_t1, df_t2, dict_tarifas_conf
    
    try:
        boveda = gc.open_by_url(URL_BOVEDA_MAESTRA)
        
        try:
            t1 = boveda.worksheet("TABLA 1").get_all_values()
            idx_t1 = 4
            for i in range(min(8, len(t1))):
                fila_limpia = [str(x).upper().strip() for x in t1[i]]
                if "Nº ORDEN" in fila_limpia or "FINCA" in fila_limpia or "VALOR A FACTURAR" in "".join(fila_limpia):
                    idx_t1 = i
                    break
            df_t1 = pd.DataFrame(t1[idx_t1+1:], columns=t1[idx_t1]) if len(t1) > idx_t1 else pd.DataFrame()
        except Exception: pass
        
        try:
            hojas = [ws.title for ws in boveda.worksheets()]
            nombre_t2 = "TABLA 2" if "TABLA 2" in hojas else hojas[1]
            t2 = boveda.worksheet(nombre_t2).get_all_values()
            df_t2 = pd.DataFrame(t2[1:], columns=t2[0]) if len(t2)>1 else pd.DataFrame()
        except Exception: pass

        try:
            if "Configuración" in [ws.title for ws in boveda.worksheets()]:
                conf_data = boveda.worksheet("Configuración").get_all_values()
                if len(conf_data) > 1:
                    df_conf = pd.DataFrame(conf_data[1:], columns=conf_data[0])
                    for _, row in df_conf.iterrows():
                        key_eq = str(row.iloc[0]).strip().upper()
                        val_m = limpiar_moneda(row.iloc[1]) if len(row) > 1 else 0.0
                        if key_eq and val_m > 0:
                            dict_tarifas_conf[key_eq] = val_m
        except Exception: pass
        
    except Exception: pass
    
    return df_t1, df_t2, dict_tarifas_conf

# =================================================================
# 💾 EXPORTADOR EXCEL MULTI-HOJA GERENCIAL
# =================================================================
def generar_excel_multi_hoja(df_filtrado_base, df_diario_agrupado, t_real, t_ideal, t_perdido, porcentaje_fuga):
    buffer = io.BytesIO()
    
    nombres_meses = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
    df_mes = df_filtrado_base.copy()
    df_mes["Mes_Num"] = df_mes["Fecha_DT"].dt.month.fillna(1).astype(int)
    
    df_mensual_base = df_mes.groupby("Mes_Num").agg({
        "Hectareas": "sum",
        "Total Real Facturado": "sum",
        "Total Simulado Ideal": "sum",
        "Lucro Cesante": "sum"
    }).reset_index()
    
    df_mensual_base["Mes de Operación"] = df_mensual_base["Mes_Num"].map(nombres_meses)
    df_mensual_base["Tarifa Real Prom/Ha"] = df_mensual_base["Total Real Facturado"] / df_mensual_base["Hectareas"]
    df_mensual_base["Tarifa Ideal Prom/Ha"] = df_mensual_base["Total Simulado Ideal"] / df_mensual_base["Hectareas"]
    df_mensual_base["Brecha Financiera/Ha"] = df_mensual_base["Tarifa Ideal Prom/Ha"] - df_mensual_base["Tarifa Real Prom/Ha"]
    
    df_mensual_final = df_mensual_base[["Mes de Operación", "Hectareas", "Tarifa Real Prom/Ha", "Tarifa Ideal Prom/Ha", "Brecha Financiera/Ha", "Total Real Facturado", "Total Simulado Ideal", "Lucro Cesante"]].copy()
    df_mensual_final = df_mensual_final.rename(columns={"Hectareas": "Total Hectáreas"})

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_mensual_final.to_excel(writer, sheet_name="Resumen_Ejecutivo_Mensual", index=False, startrow=5)
        ws1 = writer.sheets["Resumen_Ejecutivo_Mensual"]
        
        df_diario_renamed = df_diario_agrupado.copy().rename(columns={
            "Hectareas": "Total Ha",
            "Tarifa Real Prom/Ha": "Tarifa Real ($/Ha)",
            "Tarifa Ideal Prom/Ha": "Tarifa Ideal ($/Ha)",
            "Brecha por Ha": "Brecha ($/Ha)",
            "Total Real Facturado": "Cobro Real Total",
            "Total Simulado Ideal": "Total Costo OS Ideal",
            "Lucro Cesante": "Brecha Financiera Total"
        })
        df_diario_renamed.to_excel(writer, sheet_name="Detalle_Diario_Auditoria", index=False, startrow=5)
        ws2 = writer.sheets["Detalle_Diario_Auditoria"]

        fill_header = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        font_header = Font(color="D4AF37", bold=True)
        borde = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"),
                       top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))

        ws1.cell(row=1, column=1, value="📊 RESUMEN GENERAL DIRECTIVO: CONSOLIDADO MENSUAL").font = Font(size=14, bold=True, color="0D1B2A")
        ws1.cell(row=3, column=1, value=f"💰 Cobro Real Acumulado: $ {t_real:,.0f}").font = Font(bold=True)
        ws1.cell(row=3, column=4, value=f"📈 Costo Real OS Ideal: $ {t_ideal:,.0f}").font = Font(bold=True)
        ws1.cell(row=3, column=7, value=f"⚠️ Brecha Operativa: $ {t_perdido:,.0f} ({porcentaje_fuga:.1f}%)").font = Font(bold=True, color="C00000")

        for col_num in range(1, len(df_mensual_final.columns) + 1):
            cell = ws1.cell(row=6, column=col_num)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws1.column_dimensions[get_column_letter(col_num)].width = 22

        for r in range(7, len(df_mensual_final) + 7):
            ws1.cell(row=r, column=2).number_format = '#,##0.0' 
            for c in range(3, 9): 
                ws1.cell(row=r, column=c).number_format = '"$"#,##0'
            for c in range(1, 9): ws1.cell(row=r, column=c).border = borde

        ws2.cell(row=1, column=1, value="📋 INFORME ESPECÍFICO: AUDITORÍA CRONOLÓGICA DIARIA").font = Font(size=14, bold=True, color="0D1B2A")
        
        for col_num in range(1, len(df_diario_renamed.columns) + 1):
            cell = ws2.cell(row=6, column=col_num)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.column_dimensions[get_column_letter(col_num)].width = 18

        for r in range(7, len(df_diario_renamed) + 7):
            ws2.cell(row=r, column=6).number_format = '#,##0.0' 
            for c in range(7, 13): 
                ws2.cell(row=r, column=c).number_format = '"$"#,##0'
            for c in range(1, 13): ws2.cell(row=r, column=c).border = borde

    return buffer.getvalue()

# =================================================================
# 🛩️ MOTOR DEL SIMULADOR PRINCIPAL
# =================================================================
def ejecutar(procesar_fecha_pesada=None, extraer_numero=None):
    VERDE_INTENSO = '#143521'
    DORADO = '#d4af37'

    st.markdown(f"""
    <style>
    .titulo-simulador {{ color: #0d1b2a; border-bottom: 3px solid {DORADO}; padding-bottom: 5px; font-family: 'Arial Black'; }}
    
    [data-testid="column"] {{ display: flex !important; flex-direction: column !important; justify-content: flex-start !important; align-items: stretch !important; }}
    
    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {{
        border: 3px solid #0d1b2a !important; 
        border-radius: 8px !important; 
        overflow: hidden !important; 
        box-shadow: 0px 4px 10px rgba(0,0,0,0.08) !important;
    }}

    div[data-testid="stSelectbox"] > div,
    div[data-testid="stDateInput"] > div,
    div[data-testid="stNumberInput"] > div,
    div[data-testid="stTextInput"] > div {{
        border: 2px solid {VERDE_INTENSO} !important;
        border-radius: 8px !important;
        background-color: #ffffff !important;
        box-shadow: 0px 4px 8px rgba(0,0,0,0.06) !important;
        overflow: hidden !important;
    }}
    
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stDateInput"] div[data-baseweb="input"],
    div[data-testid="stNumberInput"] div[data-baseweb="input"],
    div[data-testid="stTextInput"] div[data-baseweb="input"] {{
        background-color: transparent !important;
        border: none !important;
    }}

    div[data-testid="stSelectbox"] *,
    div[data-testid="stDateInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextInput"] input {{
        color: #0d1b2a !important;
        font-weight: 900 !important;
    }}
    
    div[data-testid="stDateInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextInput"] input {{
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }}

    div[data-testid="stMainBlockContainer"] label p {{
        color: #0d1b2a !important;
        font-weight: 800 !important;
        text-transform: uppercase !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    c_t, c_btn = st.columns([3, 1])
    with c_t:
        st.markdown("<h1 class='titulo-simulador'>🛩️ Simulador Financiero Libre (OS Unificada)</h1>", unsafe_allow_html=True)
        st.caption("Consolidación Matemática y Proyección Gerencial VIP (v42.0)")
    with c_btn:
        st.write("")
        if st.button("🔄 FORZAR RECARGA RAM", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()

    with st.spinner("📥 Extrayendo y purificando matriz desde Google Sheets..."):
        df_base, df_t2_raw, dict_tarifas_conf = extraer_datos_boveda()

    if df_base.empty:
        st.error("🚨 Error de enlace: TABLA 1 no contiene registros o está desconectada.")
        return

    cols_limpias = []
    for c in df_base.columns:
        c_str = str(c).upper().replace('\n', ' ').replace('\r', '').strip()
        c_str = ' '.join(c_str.split())
        cols_limpias.append(c_str)
    df_base.columns = cols_limpias

    col_fecha = "FECHA" if "FECHA" in df_base.columns else df_base.columns[0]
    col_finca = "FINCA" if "FINCA" in df_base.columns else df_base.columns[1]
    col_pista = "PISTA" if "PISTA" in df_base.columns else df_base.columns[2]
    col_avion = "MODELO" if "MODELO" in df_base.columns else df_base.columns[3]
    
    col_orden_matches = [c for c in df_base.columns if "ORDEN" in c]
    col_orden = col_orden_matches[0] if col_orden_matches else df_base.columns[0]

    col_ha_matches = [c for c in df_base.columns if "ÁREA FUMIG" in c or "AREA FUMIG" in c]
    col_ha = col_ha_matches[0] if col_ha_matches else df_base.columns[4]

    col_rend_matches = [c for c in df_base.columns if "RENDIMIENTO" in c and "HORA" in c]
    col_rend_h = col_rend_matches[0] if col_rend_matches else df_base.columns[5]

    col_vuelo_matches = [c for c in df_base.columns if "COSTO AVI" in c and "$/HA" in c]
    col_vuelo = col_vuelo_matches[0] if col_vuelo_matches else df_base.columns[6]

    df_sim = df_base[[col_fecha, col_finca, col_pista, col_avion, col_ha, col_rend_h, col_vuelo, col_orden]].copy().reset_index(drop=True)
    renombres = {col_fecha: "Fecha", col_finca: "Finca", col_pista: "Pista_Raw", col_avion: "Equipo_Raw", col_ha: "Hectareas", col_rend_h: "RendimientoHoras", col_vuelo: "CobroReal", col_orden: "Nº ORDEN RAW"}
    df_sim = df_sim.rename(columns=renombres)

    df_sim["Nº ORDEN"] = df_sim["Nº ORDEN RAW"].apply(limpiar_orden_extrema)

    mask_valida = (df_sim["Finca"].astype(str).str.strip() != "") & (df_sim["Equipo_Raw"].astype(str).str.strip() != "")
    df_sim = df_sim[mask_valida].reset_index(drop=True)

    res_vuelo = [purificar_datos_vuelo(eq, p) for eq, p in zip(df_sim["Equipo_Raw"], df_sim["Pista_Raw"])]
    df_sim["Equipo"] = [r[0] for r in res_vuelo]
    df_sim["Pista"] = [r[1] for r in res_vuelo]
    
    df_sim["Hectareas"] = df_sim["Hectareas"].apply(limpiar_cantidad)
    df_sim["RendimientoHoras"] = df_sim["RendimientoHoras"].apply(limpiar_cantidad)
    df_sim["CobroReal"] = df_sim["CobroReal"].apply(limpiar_moneda)
    df_sim['Fecha_DT'] = df_sim["Fecha"].apply(parsear_fecha_robusta)
    
    mask_final = (df_sim["Hectareas"] > 0) & (df_sim["Equipo"] != "IGNORAR") & (df_sim['Fecha_DT'].notna())
    df_sim = df_sim[mask_final].reset_index(drop=True)

    if df_sim.empty:
        st.warning("⚠️ No hay registros matemáticamente válidos en la TABLA 1.")
        return

    min_date = df_sim['Fecha_DT'].min().date()
    max_date = df_sim['Fecha_DT'].max().date()
    
    opciones_finca = ["🌍 TODAS LAS FINCAS"] + sorted(df_sim["Finca"].dropna().unique().tolist())
    
    FLOTA_OFICIAL_POR_PISTA = {
        "AEROPENORT": ["THRUS SR2", "PIPER PA 36-375", "CESSNA O PIPER PA 25"],
        "FUMIGARAY": ["AIR TRACTOR", "CESSNA FUMIGARAY"],
        "ASA": ["CESSNA ASA"],
        "PLUC": ["DRONE DATAROT"],
        "PDIV": ["DRONE NORTE"],
        "TEHO": ["DRONE AVIL"],
        "LUCI": ["DRONE GENESYS"]
    }
    
    opciones_pista = ["🛣️ TODAS LAS PISTAS"] + list(FLOTA_OFICIAL_POR_PISTA.keys())
    lista_aviones_maestra = ["THRUS SR2", "PIPER PA 36-375", "CESSNA O PIPER PA 25", "AIR TRACTOR", "CESSNA ASA", "CESSNA FUMIGARAY", "DRONE DATAROT", "DRONE GENESYS", "DRONE NORTE", "DRONE AVIL"]

    if 'tarifas_simulador' not in st.session_state:
        st.session_state.tarifas_simulador = {}

    tarifas_base_oficiales = {
        "THRUS SR2": 4606562.0, "PIPER PA 36-375": 3985831.0, "CESSNA O PIPER PA 25": 3036525.0,
        "AIR TRACTOR": 4665109.0, "CESSNA ASA": 3666600.0, "CESSNA FUMIGARAY": 3065952.0,
        "DRONE DATAROT": 84427.0, "DRONE GENESYS": 71280.0, "DRONE NORTE": 75518.0, "DRONE AVIL": 71280.0
    }

    if dict_tarifas_conf:
        for k_conf, v_conf in dict_tarifas_conf.items():
            if k_conf in tarifas_base_oficiales:
                tarifas_base_oficiales[k_conf] = v_conf

    for k, v in tarifas_base_oficiales.items():
        if k not in st.session_state.tarifas_simulador:
            st.session_state.tarifas_simulador[k] = float(v)

    with st.container(border=True):
        st.markdown("#### 🎛️ Filtros de Escenario Gerencial")
        f1, f2, f3, f4 = st.columns([1, 1, 1.5, 1.5])
        
        fecha_ini = f1.date_input("📅 F. Inicial", value=min_date)
        fecha_fin = f2.date_input("📆 F. Final", value=max_date)
        finca_sel = f3.selectbox("📍 Finca Target", opciones_finca)
        pista_sel = f4.selectbox("🛣️ Pista", opciones_pista)
        
        if pista_sel != "🛣️ TODAS LAS PISTAS":
            pista_limpia = pista_sel.replace("🛣️ ", "").strip().upper()
            lista_aviones_dinamica = FLOTA_OFICIAL_POR_PISTA.get(pista_limpia, [])
        else:
            lista_aviones_dinamica = lista_aviones_maestra
            
        st.markdown("---")
        st.markdown("#### 🛩️ Gestor de Tarifas Base de Flota y Drones (Conexión Dinámica)")
        
        equipos_a_mostrar = [av for av in lista_aviones_dinamica if av != "✈️ TODOS LOS EQUIPOS"]
        if not equipos_a_mostrar:
            st.info("📭 Seleccione una pista para visualizar y calibrar el costo por hora.")
        else:
            for avion_editar in equipos_a_mostrar:
                c_nombre, c_precio = st.columns([1.5, 2])
                emoji_equipo = "🛸" if "DRONE" in avion_editar.upper() else "🛩️"
                c_nombre.markdown(f"<div style='margin-top: 5px; font-weight: bold; color: #1a365d; font-size: 15px;'>{emoji_equipo} {avion_editar}</div>", unsafe_allow_html=True)
                
                tarifa_actual_num = float(st.session_state.tarifas_simulador.get(avion_editar, 0.0))
                tarifa_inicial_formateada = f"$ {tarifa_actual_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                
                tarifa_usuario = c_precio.text_input("Tarifa", value=tarifa_inicial_formateada, key=f"in_bl_{avion_editar.replace(' ', '_')}", label_visibility="collapsed")
                
                if tarifa_usuario != tarifa_inicial_formateada:
                    try:
                        limpio = tarifa_usuario.replace("$", "").replace(" ", "").strip()
                        if "," in limpio and "." in limpio: limpio = limpio.replace(".", "").replace(",", ".")
                        elif "." in limpio and len(limpio.split(".")[-1]) == 3: limpio = limpio.replace(".", "")
                        elif "," in limpio: limpio = limpio.replace(",", ".")
                        st.session_state.tarifas_simulador[avion_editar] = float(limpio)
                        st.rerun()
                    except Exception: pass

        tarifas_aviones = st.session_state.tarifas_simulador

    def consolidar_os(df):
        records = []
        for orden, sub_df in df.groupby("Nº ORDEN"):
            ha_sum = float(sub_df["Hectareas"].sum())
            rend_list = [float(x) for x in sub_df["RendimientoHoras"] if pd.notna(x) and float(x) > 0]
            
            if not rend_list:
                h_tot = 0.0
            elif len(set([round(r, 4) for r in rend_list])) == 1:
                h_tot = rend_list[0]
            else:
                h_tot = sum(rend_list)
                
            records.append({
                "Nº ORDEN": orden,
                "Ha_OS_Total": ha_sum,
                "Horas_OS_Total": h_tot,
                "Fincas_En_La_OS": " | ".join(sub_df["Finca"].unique())
            })
        return pd.DataFrame(records)

    df_os_resumen = consolidar_os(df_sim)

    mask_filtro = (df_sim["Fecha_DT"].dt.date >= fecha_ini) & (df_sim["Fecha_DT"].dt.date <= fecha_fin)
    if finca_sel != "🌍 TODAS LAS FINCAS":
        mask_filtro = mask_filtro & (df_sim["Finca"] == finca_sel)
    if pista_sel != "🛣️ TODAS LAS PISTAS":
        mask_filtro = mask_filtro & (df_sim["Pista"] == pista_sel.replace("🛣️ ", ""))

    df_filtrado = df_sim[mask_filtro].copy().reset_index(drop=True)

    if df_filtrado.empty:
        st.warning("📭 No hay vuelos registrados con esos criterios de búsqueda.")
        return

    df_filtrado = df_filtrado.merge(df_os_resumen, on="Nº ORDEN", how="left")

    df_filtrado["Tarifa_Aplicada"] = df_filtrado["Equipo"].map(tarifas_aviones)
    df_filtrado["Fecha Operación"] = df_filtrado["Fecha_DT"].dt.strftime("%Y-%m-%d")
    df_filtrado["Semana"] = df_filtrado["Fecha_DT"].dt.isocalendar().week.apply(lambda x: f"Semana {x:02d}")
    df_filtrado["Total Real Facturado"] = df_filtrado["CobroReal"] * df_filtrado["Hectareas"]

    def calcular_tarifa_ideal_unificada(row):
        tarifa_hora = float(row["Tarifa_Aplicada"]) if pd.notna(row["Tarifa_Aplicada"]) else 0.0
        ha_totales_os = float(row["Ha_OS_Total"]) if (pd.notna(row["Ha_OS_Total"]) and row["Ha_OS_Total"] > 0) else float(row["Hectareas"])
        horas_totales_os = float(row["Horas_OS_Total"]) if (pd.notna(row["Horas_OS_Total"]) and row["Horas_OS_Total"] > 0) else float(row["RendimientoHoras"])

        if ha_totales_os > 0 and horas_totales_os > 0:
            return (horas_totales_os * tarifa_hora) / ha_totales_os
        return 0.0

    df_filtrado["Tarifa Ideal Prom/Ha"] = df_filtrado.apply(calcular_tarifa_ideal_unificada, axis=1)
    df_filtrado["Total Simulado Ideal"] = df_filtrado["Tarifa Ideal Prom/Ha"] * df_filtrado["Hectareas"]
    df_filtrado["Lucro Cesante"] = df_filtrado["Total Simulado Ideal"] - df_filtrado["Total Real Facturado"]

    st.markdown("---")
    ver_cebo = st.toggle("🩺 MODO AUDITOR TÉCNICO: Radiografía del Motor (Verificador de OS Unificadas)", value=False)
    
    if ver_cebo:
        with st.container(border=True):
            st.markdown("#### 🔍 Verificación Estricta de Órdenes de Servicio (OS)")
            st.caption("Esta tabla técnica detalla cómo el sistema sumó y vinculó las fincas que compartían número de Orden para el prorrateo exacto.")
            
            df_cebo = df_filtrado.groupby(["Nº ORDEN", "Fincas_En_La_OS", "Equipo", "Fecha Operación"]).agg(
                Horas_Calculadas_OS=("Horas_OS_Total", "max"),
                Suma_Hectareas_OS=("Ha_OS_Total", "max"),
                Tarifa_Ideal_Final_Ha=("Tarifa Ideal Prom/Ha", "mean")
            ).reset_index()
            
            col_cfg_cebo = {
                "Nº ORDEN": st.column_config.TextColumn("🛰️ Nº ORDEN"),
                "Fincas_En_La_OS": st.column_config.TextColumn("📍 FINCAS UNIFICADAS"),
                "Equipo": st.column_config.TextColumn("🛩️ EQUIPO"),
                "Fecha Operación": st.column_config.TextColumn("📅 FECHA"),
                "Horas_Calculadas_OS": st.column_config.NumberColumn("⏱️ HORAS TOTALES OS", format="%.3f hrs"),
                "Suma_Hectareas_OS": st.column_config.NumberColumn("🗺️ HECTÁREAS TOTALES OS", format="%.2f ha"),
                "Tarifa_Ideal_Final_Ha": st.column_config.NumberColumn("🎯 TARIFA IDEAL UNIFICADA", format="$ %,.0f")
            }

            st.dataframe(df_cebo, use_container_width=True, hide_index=True, column_config=col_cfg_cebo)
        st.stop() 

    df_agrupado = df_filtrado.groupby(["Fecha Operación", "Semana", "Pista", "Finca", "Equipo"]).agg({
        "Hectareas": "sum",
        "Total Real Facturado": "sum",
        "Total Simulado Ideal": "sum",
        "Lucro Cesante": "sum"
    }).reset_index()
    
    df_agrupado["Tarifa Real Prom/Ha"] = df_agrupado["Total Real Facturado"] / df_agrupado["Hectareas"]
    df_agrupado["Tarifa Ideal Prom/Ha"] = df_agrupado["Total Simulado Ideal"] / df_agrupado["Hectareas"]
    df_agrupado["Brecha por Ha"] = df_agrupado["Tarifa Ideal Prom/Ha"] - df_agrupado["Tarifa Real Prom/Ha"]

    df_agrupado = df_agrupado[["Fecha Operación", "Semana", "Pista", "Finca", "Equipo", "Hectareas", "Tarifa Real Prom/Ha", "Tarifa Ideal Prom/Ha", "Brecha por Ha", "Total Real Facturado", "Total Simulado Ideal", "Lucro Cesante"]]
    df_agrupado = df_agrupado.sort_values(by=["Finca", "Fecha Operación"]).reset_index(drop=True)

    st.markdown("### 💎 Impacto Financiero de la Operación")
    
    t_real = df_agrupado["Total Real Facturado"].sum()
    t_ideal = df_agrupado["Total Simulado Ideal"].sum()
    t_perdido = df_agrupado["Lucro Cesante"].sum()
    porcentaje_fuga = ((t_ideal / t_real) - 1) * 100 if t_real > 0 else 0

    def f_h(val): return f"{val:,.0f}".replace(",", ".")

    html_cards = f"""
    <div style="display: flex; flex-wrap: wrap; gap: 10px; margin-top: 15px; margin-bottom: 20px;">
        <div style="flex: 1; min-width: 180px; background-color: #f8f9fa; border-left: 4px solid #0D1B2A; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="font-size: 12px; color: #6c757d; font-weight: bold; text-transform: uppercase;">Cobro Real Facturado</div>
            <div style="font-size: 20px; color: #0D1B2A; font-weight: 900; margin-top: 4px;">$ {f_h(t_real)}</div>
        </div>
        <div style="flex: 1; min-width: 180px; background-color: #f8f9fa; border-left: 4px solid #D4AF37; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <div style="font-size: 12px; color: #6c757d; font-weight: bold; text-transform: uppercase;">Costo Base OS Ideal</div>
            <div style="font-size: 20px; color: #0D1B2A; font-weight: 900; margin-top: 4px;">$ {f_h(t_ideal)}</div>
        </div>
        <div style="flex: 1.2; min-width: 200px; background-color: #0D1B2A; border: 2px solid #ff4d4d; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.2); text-align: center;">
            <div style="font-size: 12px; color: #ff4d4d; font-weight: bold; text-transform: uppercase;">⚠️ Brecha Total (Lucro Cesante)</div>
            <div style="font-size: 22px; color: white; font-weight: 900; margin-top: 4px;">$ {f_h(t_perdido)} <span style="font-size: 13px; color: #ff4d4d;">({porcentaje_fuga:.1f}%)</span></div>
        </div>
    </div>
    """
    st.markdown(html_cards, unsafe_allow_html=True)

    st.markdown("### 📋 Resumen Detallado y Auditoría Financiera")
    
    df_visual = df_agrupado.copy()
    df_visual["Fecha Operación"] = pd.to_datetime(df_visual["Fecha Operación"], format='%Y-%m-%d', errors='coerce').dt.strftime('%d/%m/%Y')

    def color_fuga(val):
        if pd.isna(val): return ''
        if val > 0: return 'color: #D32F2F; font-weight: bold;'
        elif val < 0: return 'color: #198754; font-weight: bold;'
        return 'color: #424242;'

    col_cfg = {
        "Fecha Operación": st.column_config.TextColumn("📅 FECHA"),
        "Semana": st.column_config.TextColumn("📆 SEMANA"),
        "Pista": st.column_config.TextColumn("🛫 PISTA"),
        "Finca": st.column_config.TextColumn("📍 FINCA"),
        "Equipo": st.column_config.TextColumn("🛩️ EQUIPO"),
        "Hectareas": st.column_config.NumberColumn("🗺️ HECTÁREAS", format="%.2f"),
        "Tarifa Real Prom/Ha": st.column_config.NumberColumn("💰 TARIFA REAL", format="$ %,.0f"),
        "Tarifa Ideal Prom/Ha": st.column_config.NumberColumn("🎯 TARIFA IDEAL", format="$ %,.0f"),
        "Brecha por Ha": st.column_config.NumberColumn("⚖️ BRECHA/HA", format="$ %,.0f"),
        "Total Real Facturado": st.column_config.NumberColumn("💵 TOTAL REAL", format="$ %,.0f"),
        "Total Simulado Ideal": st.column_config.NumberColumn("🚀 TOTAL IDEAL", format="$ %,.0f"),
        "Lucro Cesante": st.column_config.NumberColumn("📉 LUCRO CESANTE", format="$ %,.0f")
    }

    st.dataframe(
        df_visual.style.map(color_fuga, subset=['Lucro Cesante', 'Brecha por Ha']), 
        use_container_width=True, 
        height=400, 
        hide_index=True,
        column_config=col_cfg
    )

    st.markdown("---")
    st.markdown("### 📈 Dashboard Analítico de Tendencias")

    # =================================================================
    # 📊 GRÁFICO 1: EVOLUCIÓN CRONOLÓGICA (AHORA AGRUPADO POR SEMANA)
    # =================================================================
    # Agrupamos por Semana en lugar de Día para limpiar el ruido visual
    df_tendencia = df_agrupado.groupby("Semana").agg({
        "Tarifa Real Prom/Ha": "mean",
        "Tarifa Ideal Prom/Ha": "mean"
    }).reset_index().sort_values(by="Semana")
    
    # 💥 MAGIA VISUAL: Transformar "Semana 05" a "S5"
    df_tendencia["Semana Corta"] = df_tendencia["Semana"].apply(lambda x: f"S{int(str(x).replace('Semana ', ''))}")

    fig_tarifas = go.Figure()

    fig_tarifas.add_trace(go.Scatter(
        x=df_tendencia["Semana Corta"],
        y=df_tendencia["Tarifa Ideal Prom/Ha"],
        mode='lines',
        name='Costo Base OS Ideal',
        line=dict(color='#d4af37', width=2, dash='dot'),
        fill='tozeroy', 
        fillcolor='rgba(212, 175, 55, 0.15)',
        hovertemplate='Ideal: $ %{y:,.0f}/ha<extra></extra>'
    ))

    fig_tarifas.add_trace(go.Scatter(
        x=df_tendencia["Semana Corta"],
        y=df_tendencia["Tarifa Real Prom/Ha"],
        mode='lines+markers',
        name='Cobro Real Facturado',
        line=dict(color='#0d1b2a', width=3),
        marker=dict(size=6, color='#0d1b2a', line=dict(color='white', width=1)),
        hovertemplate='Real: $ %{y:,.0f}/ha<extra></extra>'
    ))

    fig_tarifas.update_layout(
        title="<b>Evolución Promedio Semanal: Cobro Real vs Costo Ideal</b>",
        title_font=dict(color="#0d1b2a", size=16, family="Arial Black"),
        height=400,
        plot_bgcolor="#f8fafc", paper_bgcolor="#ffffff",
        # Al ser textos cortos (S1, S2), quitamos la inclinación (tickangle) para que se lean derechos
        xaxis=dict(showgrid=False, title="", tickfont=dict(size=11, color='#555555')), 
        yaxis=dict(showgrid=True, gridcolor="#e2e8f0", zeroline=False, title="Valor Promedio por Hectárea ($)", tickformat="$,.0f"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=50, r=20, t=70, b=40),
        hovermode="x unified"
    )
    st.plotly_chart(fig_tarifas, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # =================================================================
    # 📊 GRÁFICO 2: FUGA OPERATIVA (CASCADA FINANCIERA / WATERFALL)
    # =================================================================
    df_lucro_sem = df_agrupado.groupby("Semana")["Lucro Cesante"].sum().reset_index().sort_values(by="Semana")
    
    # 💥 Aplicamos la misma reducción (S1, S2) a la cascada
    df_lucro_sem["Semana Corta"] = df_lucro_sem["Semana"].apply(lambda x: f"S{int(str(x).replace('Semana ', ''))}")
    
    fig_lucro = go.Figure(go.Waterfall(
        name="Fuga Operativa",
        orientation="v",
        measure=["relative"] * len(df_lucro_sem),
        x=df_lucro_sem["Semana Corta"],
        y=df_lucro_sem["Lucro Cesante"],
        text=df_lucro_sem["Lucro Cesante"],
        texttemplate='<b>$%{text:,.0f}</b>',
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Impacto Semanal: $ %{y:,.0f}<extra></extra>",
        connector={"line": {"color": "#b3b3b3", "width": 1.5, "dash": "dot"}},
        increasing={"marker": {"color": "#dc3545"}}, 
        decreasing={"marker": {"color": "#28a745"}}, 
        totals={"marker": {"color": "#0d1b2a"}}
    ))

    fig_lucro.update_layout(
        title="<b>Acumulación de Fuga Operativa Semanal (Efecto Cascada)</b>",
        title_font=dict(color="#0d1b2a", size=16, family="Arial Black"),
        height=450,
        plot_bgcolor='#f8fafc', paper_bgcolor='#ffffff',
        # Eje X derecho y limpio
        xaxis=dict(showgrid=False, title="", tickfont=dict(size=11, color='#555555')),
        yaxis=dict(
            showgrid=True, gridcolor='#e2e8f0', 
            zeroline=True, zerolinecolor='#0d1b2a', zerolinewidth=2,
            title="Monto Acumulado de Fuga ($)", tickformat="$,.0f"
        ),
        margin=dict(l=50, r=20, t=70, b=40),
        showlegend=False
    )
    
    st.plotly_chart(fig_lucro, use_container_width=True)

    

    st.markdown("---")
    st.markdown("### 📤 Exportar Datos Consolidados Autorizados")

    buffer_excel = generar_excel_multi_hoja(df_filtrado, df_agrupado, t_real, t_ideal, t_perdido, porcentaje_fuga)

    st.download_button(
        label="💾 DESCARGAR REPORTE MULTI-HOJA COMPLETO (EXCEL GERENCIAL)",
        data=buffer_excel,
        file_name=f"Reporte_Simulador_Agro_OS_{fecha_ini}_{fecha_fin}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.success("🏁 Proceso completado. La interfaz opera con Formato Gerencial Dinámico y conexión activa a Configuración.")

if __name__ == "__main__":
    pass
