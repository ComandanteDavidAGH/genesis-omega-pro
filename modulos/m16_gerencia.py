import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import gspread
import re
import io
from datetime import datetime, date
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =================================================================
# ⚙️ CONSTANTES Y MOTOR DE CONEXIÓN UNIFICADO
# =================================================================
URL_BOVEDA_MAESTRA = "https://docs.google.com/spreadsheets/d/1gTu6mAec1qJrxAhw7F-Gl3fVcHaIOnmFUJQYFgqARP4/edit"

COLOR_NAVY = '#0d1b2a'
COLOR_DORADO = '#d4af37'
COLOR_VERDE = '#143521'

@st.cache_resource(show_spinner=False)
def obtener_cliente_gspread_unificado():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if "gcp_service_account" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_service_account"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    if "gcp_credentials" in st.secrets:
        try:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets["gcp_credentials"]), scope)
            return gspread.authorize(creds)
        except Exception: pass
    try:
        return gspread.service_account(filename='credenciales.json')
    except Exception:
        return None

# =================================================================
# 🛡️ UTILIDADES DE PURIFICACIÓN Y LIMPIEZA FINANCIERA
# =================================================================
def limpiar_tarifa_excel(val):
    if isinstance(val, (int, float)): return float(val)
    v = str(val).strip().replace("$", "").replace(" ", "").upper()
    if not v or v in ['-', 'NAN', 'NONE', '']: return 0.0
    
    s_clean = re.sub(r'[^\d\.,\-]', '', v)
    try:
        if '.' in s_clean and ',' in s_clean:
            if s_clean.rfind(',') > s_clean.rfind('.'): s_clean = s_clean.replace('.', '').replace(',', '.')
            else: s_clean = s_clean.replace(',', '')
        elif ',' in s_clean:
            if len(s_clean.split(',')[-1]) == 3: s_clean = s_clean.replace(',', '')
            else: s_clean = s_clean.replace(',', '.')
        elif '.' in s_clean:
            if s_clean.count('.') > 1: s_clean = s_clean.replace('.', '')
            elif len(s_clean.split('.')[-1]) == 3: s_clean = s_clean.replace('.', '')
        return float(s_clean) if s_clean else 0.0
    except:
        return 0.0

def procesar_fecha_pesada(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    s = str(val).strip()
    if s.replace('.', '', 1).isdigit(): 
        return pd.to_datetime('1899-12-30') + pd.to_timedelta(float(s), 'D')
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try: return pd.to_datetime(s, format=fmt)
        except: pass
    try: return pd.to_datetime(s, errors='coerce')
    except: return pd.NaT

def normalizar_a_fecha_pura(val):
    try:
        res_nativo = procesar_fecha_pesada(val)
        if isinstance(res_nativo, (datetime, pd.Timestamp)): return res_nativo.date()
        if isinstance(res_nativo, date): return res_nativo
        return pd.to_datetime(str(res_nativo)).date()
    except: return None

def es_cooperativa(finca_nombre):
    f_up = str(finca_nombre).upper().strip()
    # Filtro estricto: Solo entidades que sean verdaderamente Cooperativas
    patrones_coop = [
        'BANAFRUCOOP', 'COOMULBANANO', 'COOBAMAG', 'EMPREBANCOOP', 
        'COOBAFRIO', 'COOP'
    ]
    return any(p in f_up for p in patrones_coop)

# =================================================================
# 💾 EXTRACCIÓN CACHEADA DE DATOS OPERATIVOS (ESTRUCTURA ESTÁNDAR)
# =================================================================
@st.cache_data(show_spinner=False, ttl=120)
def cargar_datos_gerenciales():
    gc = obtener_cliente_gspread_unificado()
    if not gc: return pd.DataFrame()
    
    try:
        boveda_act = gc.open_by_url(URL_BOVEDA_MAESTRA)
        datos_brutos = boveda_act.worksheet("TABLA 1").get_all_values()
        
        if len(datos_brutos) > 5:
            columnas_t1 = [
                "OS", "BLOQUE", "FINCA", "SECTOR", "AREA_BRUTA", "AREA_FUMIG", "COCTEL", 
                "FECHA", "DIA", "SEMANA", "H_TOTAL", "GLN_HA", "VOL_TOTAL", "REND_HR", 
                "REND_MIN", "PILOTO", "HK", "MODELO", "COSTO_AVION", "COSTO_HA", 
                "DOMINICAL_HA", "COSTO_FINCA", "VALOR_FACTURAR", "PISTA"
            ]
            filas_limpias = [r + [""]*(len(columnas_t1) - len(r)) for r in datos_brutos[5:]]
            df = pd.DataFrame([r[:len(columnas_t1)] for r in filas_limpias], columns=columnas_t1)
            
            df['FINCA'] = df['FINCA'].astype(str).str.strip().str.upper()
            df['FECHA_FILTRABLE'] = df['FECHA'].apply(normalizar_a_fecha_pura)
            
            def clasificar_tec(row):
                texto = f"{str(row.get('PILOTO',''))} {str(row.get('HK',''))} {str(row.get('MODELO',''))}".upper()
                if 'DRON' in texto or 'DR5' in texto: return 'DRONE'
                return 'AVIÓN'
            
            df['TECNOLOGIA'] = df.apply(clasificar_tec, axis=1)
            df['TIPO_ENTIDAD'] = df['FINCA'].apply(lambda x: 'COOPERATIVAS' if es_cooperativa(x) else 'ESPECIALES / PILOTOS')
            
            df['COSTO_TOTAL_HA'] = df['VALOR_FACTURAR'].apply(limpiar_tarifa_excel)
            df['COSTO_VUELO_HA'] = df['COSTO_HA'].apply(limpiar_tarifa_excel)
            
            df['COSTO_VUELO_HA'] = df['COSTO_VUELO_HA'].apply(lambda x: x * 1000 if 0 < x < 2500 else x)
            df['COSTO_TOTAL_HA'] = df['COSTO_TOTAL_HA'].apply(lambda x: x * 1000 if 0 < x < 2500 else x)
            df['COSTO_VUELO_HA'] = df['COSTO_VUELO_HA'].apply(lambda x: 75000 if x > 150000 else x)
            
            df['OPERADOR_DRON'] = df['HK'].astype(str).str.strip() + " - " + df['PISTA'].astype(str).str.strip()
            
            return df.dropna(subset=['FECHA_FILTRABLE'])
        return pd.DataFrame()
    except Exception: return pd.DataFrame()

# =================================================================
# ⚙️ MOTOR EXCEL PROFESIONAL (DISEÑO VIP)
# =================================================================
def generar_excel_maestro(df_total, df_vuelo):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_total.to_excel(writer, index=False, sheet_name='Facturación Total', startrow=3)
        df_vuelo.to_excel(writer, index=False, sheet_name='Tarifa Vuelo Pura', startrow=3)
        
        fill_titulo = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        font_titulo = Font(color="FFFFFF", bold=True, size=14)
        font_sub = Font(color="555555", italic=True, size=10)
        
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_font = Font(color="D4AF37", bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        borde_fino = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
        
        font_rojo = Font(color="C00000", bold=True) 
        font_verde = Font(color="00B050", bold=True) 
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            
            ws.merge_cells("A1:G1")
            ws["A1"] = f"REPORTE GERENCIAL COMPARATIVO — {sheet_name.upper()}"
            ws["A1"].fill = fill_titulo
            ws["A1"].font = font_titulo
            ws["A1"].alignment = align_center

            ws.merge_cells("A2:G2")
            ws["A2"] = f"Análisis de eficiencia Dron vs Avión | Generado el: {fecha_actual}"
            ws["A2"].font = font_sub
            ws["A2"].alignment = align_left
            
            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 28
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 16
            
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = borde_fino
                
            for row in range(5, ws.max_row + 1):
                ws[f'D{row}'].number_format = '"$"#,##0'
                ws[f'E{row}'].number_format = '"$"#,##0'
                
                celda_dif = ws[f'F{row}']
                celda_efi = ws[f'G{row}']
                
                celda_dif.number_format = '"$"#,##0'
                celda_efi.number_format = '0.0%' 
                
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws[f'{col_letter}{row}'].border = borde_fino
                
                if isinstance(celda_dif.value, (int, float)):
                    if celda_dif.value > 0:
                        celda_dif.font = font_verde
                        celda_efi.font = font_verde
                    elif celda_dif.value < 0:
                        celda_dif.font = font_rojo
                        celda_efi.font = font_rojo
                
    return output.getvalue()

def construir_grafico_comparativo(df_datos, titulo_grafico):
    if df_datos.empty: return None
    df_plot = df_datos.copy().reset_index(drop=True)
    
    if 'Diferencia ($)' not in df_plot.columns:
        df_plot['Diferencia ($)'] = df_plot['AVIÓN'] - df_plot['DRONE']
        
    df_plot = df_plot.sort_values(by='Diferencia ($)', ascending=True)
    
    # 💥 AJUSTE 1: Aumentamos el límite de lectura de 25 a 38 caracteres
    df_plot['FINCA_CORTA'] = df_plot['FINCA'].astype(str).apply(lambda x: x[:38] + '...' if len(x) > 38 else x)
    
    colores = ['#28a745' if val > 0 else '#dc3545' for val in df_plot['Diferencia ($)']]
    
    fig = go.Figure()

    fig.add_trace(go.Bar(
        y=df_plot['FINCA_CORTA'],
        x=df_plot['Diferencia ($)'],
        orientation='h',
        marker=dict(color=colores, line=dict(color='#0d1b2a', width=0.5)),
        text=df_plot['Diferencia ($)'],
        texttemplate='<b>$ %{text:,.0f}</b>',
        textposition='auto',
        textfont=dict(size=10, color='white'),
        hovertext=df_plot['FINCA'],
        customdata=np.stack((df_plot['DRONE'], df_plot['AVIÓN']), axis=-1),
        hovertemplate=(
            "<div style='font-family: Arial;'><b>🏡 Finca: %{hovertext}</b><br><br>"
            "🛸 Costo Dron: $ %{customdata[0]:,.0f}<br>"
            "✈️ Costo Avión: $ %{customdata[1]:,.0f}<br>"
            "<b>⚖️ Impacto: $ %{x:,.0f} / ha</b></div><extra></extra>"
        )
    ))

    fig.update_layout(
        title=f"<b>{titulo_grafico}</b>",
        title_font=dict(color="#0d1b2a", size=14, family="Arial Black"),
        height=550, 
        plot_bgcolor='#f8fafc',
        paper_bgcolor='#ffffff',
        xaxis=dict(
            title="← Dron Costoso (Rojo)  |  Dron Rentable (Verde) →", 
            title_font=dict(size=11, color="#555555", family="Arial Black"),
            tickformat="$,.0f", 
            showgrid=True, gridcolor='#e2e8f0',
            zeroline=True, zerolinecolor='#0d1b2a', zerolinewidth=2
        ),
        yaxis=dict(
            title="", 
            showgrid=False,
            tickfont=dict(size=9, color='#0d1b2a', family='Arial'), 
            automargin=False 
        ),
        # 💥 AJUSTE 2: Ampliamos el margen izquierdo (l) de 180 a 260 píxeles
        margin=dict(l=260, r=30, t=40, b=40), 
        bargap=0.15, 
        showlegend=False
    )
    return fig

# =================================================================
# 👑 RENDERIZADO VISUAL
# =================================================================
def ejecutar(*args, **kwargs):

    st.markdown(f"""
    <style>
    /* TITULO ALINEADO */
    .titulo-contenedor {{ display: flex; align-items: flex-start; gap: 15px; border-bottom: 3px solid {COLOR_DORADO}; padding-bottom: 10px; margin-bottom: 15px; }}
    .titulo-icono {{ font-size: 34px; line-height: 1.2; }}
    .titulo-texto {{ display: flex; flex-direction: column; }}
    .titulo-gerencial-txt {{ color: {COLOR_NAVY}; margin: 0; font-weight: 900; letter-spacing: 0.5px; line-height: 1.2; font-size: 26px; font-family: 'Arial Black', sans-serif; }}
    .titulo-caption {{ color: #555555; font-size: 14px; margin: 4px 0 0 0; font-weight: 600; text-transform: uppercase; }}
    
    /* 💥 CIRUGÍA UNIVERSAL: Cajas de Fecha */
    
    /* 1. Atacamos la caja exterior universal de Streamlit */
    .stDateInput > div {{
        background-color: #e6f4ea !important; /* Fondo verde tenue */
        border: 2px solid {COLOR_VERDE} !important; /* Borde oscuro */
        border-radius: 8px !important;
    }}
    
    /* 2. Forzamos transparencia en la caja interior que bloqueaba el color */
    .stDateInput > div > div {{
        background-color: transparent !important;
        border: none !important;
    }}
    
    /* 3. NÚMEROS: Negros, grandes y en negrita extrema */
    .stDateInput input {{
        color: #000000 !important;
        font-weight: 900 !important;
        font-size: 16px !important;
        -webkit-text-fill-color: #000000 !important; /* Fuerza el color en todos los navegadores */
        background-color: transparent !important;
    }}
    
    /* 4. Etiquetas "Desde" y "Hasta" */
    .stDateInput label p {{
        color: {COLOR_VERDE} !important;
        font-weight: 800 !important;
        font-size: 13px !important;
    }}

    /* Tablas y Pestañas */
    div[data-testid="stDataFrame"] {{ border: 2px solid {COLOR_NAVY} !important; border-radius: 8px !important; overflow: hidden !important; }}
    div[data-testid="stTabs"] button[role="tab"] {{ font-weight: 800; font-size: 13px; color: {COLOR_NAVY}; }}
    div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{ border-bottom-color: {COLOR_DORADO}; background-color: rgba(212, 175, 55, 0.08); }}
    </style>
    """, unsafe_allow_html=True)

    def tarjeta_kpi(titulo, valor, delta_texto="", color_delta="#28a745"):
        delta_html = f"<span style='font-size: 13px; color: {color_delta}; margin-left: 6px; font-weight:bold;'>{delta_texto}</span>" if delta_texto else ""
        return f"""
        <div style='background: linear-gradient(135deg, #0d1b2a 0%, #1a365d 100%); 
             border-left: 4px solid {COLOR_DORADO}; padding: 14px 16px; 
             border-radius: 8px; color: white; box-shadow: 0px 3px 8px rgba(0,0,0,0.1); 
             margin-bottom: 15px; min-height: 92px; display: flex; flex-direction: column; justify-content: center;'>
             <p style='font-size: 11px; font-weight: 800; color: {COLOR_DORADO}; text-transform: uppercase; margin:0 0 4px 0;'>{titulo}</p>
             <p style='font-size: 20px; font-weight: 900; margin: 0; color: white; display: flex; align-items: center; flex-wrap: wrap;'>{valor} {delta_html}</p>
         </div>
         """

    c_tit, c_sync = st.columns([3.5, 1.5])
    with c_tit:
        st.markdown("""
        <div class='titulo-contenedor'>
            <div class='titulo-icono'>⚖️</div>
            <div class='titulo-texto'>
                <h2 class='titulo-gerencial-txt'>MÓDULO 16: COMPARATIVO GERENCIAL (DRON VS AVIÓN)</h2>
                <p class='titulo-caption'>Análisis de costos y brecha de eficiencia: Cooperativas vs. Fincas Especiales y Lotes Piloto.</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with c_sync:
        st.write("")
        st.write("")
        if st.button("🔄 Sincronizar Base Datos", use_container_width=True, type="primary"):
            st.cache_data.clear()
            st.rerun()

    with st.container(border=True):
        st.markdown("#### 📅 Parámetros de Análisis")
        c_f1, c_f2 = st.columns(2)
        fecha_inicio = c_f1.date_input("Desde:", value=date(2026, 1, 1))
        fecha_fin = c_f2.date_input("Hasta:", value=date(2026, 12, 31))

    df_raw = cargar_datos_gerenciales()
    if df_raw.empty:
        st.warning("⚠️ No se detectan registros en la base maestra.")
        return

    df_base = df_raw[(df_raw['FECHA_FILTRABLE'] >= fecha_inicio) & (df_raw['FECHA_FILTRABLE'] <= fecha_fin)].copy()
    if df_base.empty:
        st.error("❌ No se encontraron registros de vuelo para el rango seleccionado.")
        return

    df_aviones = df_base[df_base['TECNOLOGIA'] == 'AVIÓN'].copy()
    df_drones = df_base[df_base['TECNOLOGIA'] == 'DRONE'].copy()

    df_vuelo_avion = df_aviones.groupby(['FINCA', 'TIPO_ENTIDAD'])['COSTO_VUELO_HA'].max().reset_index().rename(columns={'COSTO_VUELO_HA': 'AVIÓN'})
    df_vuelo_dron = df_drones.groupby(['FINCA', 'TIPO_ENTIDAD', 'OPERADOR_DRON'])['COSTO_VUELO_HA'].max().reset_index().rename(columns={'COSTO_VUELO_HA': 'DRONE'})
    m_comp_v = pd.merge(df_vuelo_dron, df_vuelo_avion, on=['FINCA', 'TIPO_ENTIDAD'], how='inner')
    
    if not m_comp_v.empty:
        m_comp_v = m_comp_v[['FINCA', 'TIPO_ENTIDAD', 'OPERADOR_DRON', 'AVIÓN', 'DRONE']]
        m_comp_v.rename(columns={'OPERADOR_DRON': 'EQUIPO DRON'}, inplace=True)
        m_comp_v['Diferencia ($)'] = m_comp_v['AVIÓN'] - m_comp_v['DRONE']
        m_comp_v['Eficiencia (%)'] = m_comp_v['Diferencia ($)'] / m_comp_v['AVIÓN']

    df_total_avion = df_aviones.groupby(['FINCA', 'TIPO_ENTIDAD'])['COSTO_TOTAL_HA'].max().reset_index().rename(columns={'COSTO_TOTAL_HA': 'AVIÓN'})
    df_total_dron = df_drones.groupby(['FINCA', 'TIPO_ENTIDAD', 'OPERADOR_DRON'])['COSTO_TOTAL_HA'].max().reset_index().rename(columns={'COSTO_TOTAL_HA': 'DRONE'})
    m_comp_t = pd.merge(df_total_dron, df_total_avion, on=['FINCA', 'TIPO_ENTIDAD'], how='inner')

    if not m_comp_t.empty:
        m_comp_t = m_comp_t[['FINCA', 'TIPO_ENTIDAD', 'OPERADOR_DRON', 'AVIÓN', 'DRONE']]
        m_comp_t.rename(columns={'OPERADOR_DRON': 'EQUIPO DRON'}, inplace=True)
        m_comp_t['Diferencia ($)'] = m_comp_t['AVIÓN'] - m_comp_t['DRONE']
        m_comp_t['Eficiencia (%)'] = m_comp_t['Diferencia ($)'] / m_comp_t['AVIÓN'] 

    st.markdown("---")
    if not m_comp_v.empty:
        ahorro_prom_vuelo = m_comp_v['Diferencia ($)'].mean()
        eficiencia_prom_vuelo = m_comp_v['Eficiencia (%)'].mean() * 100
        fincas_coop = m_comp_v[m_comp_v['TIPO_ENTIDAD'] == 'COOPERATIVAS']['FINCA'].nunique()
        fincas_indep = m_comp_v[m_comp_v['TIPO_ENTIDAD'] == 'ESPECIALES / PILOTOS']['FINCA'].nunique()

        k1, k2, k3 = st.columns(3)
        with k1: st.markdown(tarjeta_kpi("Cobertura Mapeada", f"{fincas_coop} Coop / {fincas_indep} Especiales", "Fincas Cruzadas", "#d4af37"), unsafe_allow_html=True)
        with k2: st.markdown(tarjeta_kpi("Brecha Promedio Vuelo", f"$ {ahorro_prom_vuelo:,.0f} /ha".replace(",", "."), "Ahorro Dron vs Avión", "#28a745" if ahorro_prom_vuelo >= 0 else "#dc3545"), unsafe_allow_html=True)
        with k3: st.markdown(tarjeta_kpi("Eficiencia Financiera", f"{eficiencia_prom_vuelo:.1f}%", "vs Tarifa Avión", "#28a745" if eficiencia_prom_vuelo >= 0 else "#dc3545"), unsafe_allow_html=True)

def aplicar_estilo_premium(row):
        estilos = [''] * len(row)
        dif = row.get('Diferencia ($)', 0)
        
        base_style = 'background-color: #ffffff; color: #0d1b2a; font-weight: 600;'
        
        for i, col in enumerate(row.index):
            cell_style = base_style
            if col in ['AVIÓN', 'DRONE', 'Diferencia ($)', 'Eficiencia (%)']:
                cell_style += ' text-align: right;'
            
            if col == 'Diferencia ($)':
                if dif < 0: cell_style = 'background-color: #ffe5e5; color: #dc3545; font-weight: 900; text-align: right;'
                elif dif > 0: cell_style = 'background-color: #e6f4ea; color: #28a745; font-weight: 900; text-align: right;'
            elif col == 'Eficiencia (%)':
                if dif < 0: cell_style = 'color: #dc3545; font-weight: 900; text-align: right;'
                elif dif > 0: cell_style = 'color: #28a745; font-weight: 900; text-align: right;'
                
            estilos[i] = cell_style
        return estilos

    columnas_ui = {
        "FINCA": st.column_config.TextColumn("🏡 FINCA", width="medium"),
        "TIPO_ENTIDAD": st.column_config.TextColumn("🤝 PERFIL", width="small"),
        "EQUIPO DRON": st.column_config.TextColumn("🛸 OPERADOR DRON", width="medium"),
        "AVIÓN": st.column_config.NumberColumn("✈️ TARIFA AVIÓN", format="$ %d", width="small"),
        "DRONE": st.column_config.NumberColumn("🛸 TARIFA DRON", format="$ %d", width="small"),
        "Diferencia ($)": st.column_config.NumberColumn("⚖️ AHORRO ($)", format="$ %d", width="small"),
        "Eficiencia (%)": st.column_config.NumberColumn("📈 EFICIENCIA", format="%.1f %%", width="small")
    }
    
    # 💥 CREAMOS LA TERCERA PESTAÑA PARA GRÁFICOS
    tab_vuelo, tab_total, tab_graficos = st.tabs([
        "✈️ Matrices de Vuelo (Datos)", 
        "💰 Matrices de Facturación (Datos)", 
        "📊 Centro de Análisis Gráfico"
    ])

    # ==========================================
    # 1. PESTAÑA: SOLO DATOS DE VUELO
    # ==========================================
    with tab_vuelo:
        st.success("🔬 Matriz Detallada: Tarifas de Vuelo Pura (Cooperativas vs Especiales)")
        if not m_comp_v.empty:
            df_print_v = m_comp_v.copy()
            df_print_v['Eficiencia (%)'] = df_print_v['Eficiencia (%)'] * 100
            st.dataframe(
                df_print_v.style.apply(aplicar_estilo_premium, axis=1), 
                use_container_width=True, 
                hide_index=True,
                column_config=columnas_ui
            )
        else:
            st.warning("📌 No hay datos cruzados en el rango seleccionado.")

    # ==========================================
    # 2. PESTAÑA: SOLO DATOS DE FACTURACIÓN
    # ==========================================
    with tab_total:
        st.info("📊 Matriz Detallada: Impacto Macro en Facturación Total")
        if not m_comp_t.empty:
            df_print_t = m_comp_t.copy()
            df_print_t['Eficiencia (%)'] = df_print_t['Eficiencia (%)'] * 100
            st.dataframe(
                df_print_t.style.apply(aplicar_estilo_premium, axis=1), 
                use_container_width=True, 
                hide_index=True,
                column_config=columnas_ui
            )
        else:
            st.warning("📌 No hay datos cruzados en el rango seleccionado.")

    # ==========================================
    # 3. PESTAÑA: CENTRO EXCLUSIVO DE GRÁFICOS
    # ==========================================
    with tab_graficos:
        st.markdown("### 🎛️ Panel de Visualización Estratégica")
        
        # Selector para ver gráficos de Vuelo Puro o de Facturación Total
        tipo_grafico = st.radio("Seleccione la Métrica a Graficar:", ["✈️ Tarifa Vuelo Pura", "💰 Facturación Total"], horizontal=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        # 💥 SUB-PESTAÑAS INDEPENDIENTES PARA COOPERATIVAS Y PILOTOS
        sub_tab_coop, sub_tab_indep = st.tabs(["🌾 Ver Cooperativas y Asociativas", "🍌 Ver Fincas Especiales y Lotes Piloto"])
        
        # DATOS PARA COOPERATIVAS
        with sub_tab_coop:
            if tipo_grafico == "✈️ Tarifa Vuelo Pura" and not m_comp_v.empty:
                m_coop = m_comp_v[m_comp_v['TIPO_ENTIDAD'] == 'COOPERATIVAS'].copy()
                if not m_coop.empty:
                    fig_c = construir_grafico_comparativo(m_coop, "Cooperativas: Vuelo Puro (Avión vs Dron)")
                    st.plotly_chart(fig_c, use_container_width=True)
                else: st.info("No se registraron fincas cooperativas.")
            
            elif tipo_grafico == "💰 Facturación Total" and not m_comp_t.empty:
                m_coop_t = m_comp_t[m_comp_t['TIPO_ENTIDAD'] == 'COOPERATIVAS'].copy()
                if not m_coop_t.empty:
                    fig_c_t = construir_grafico_comparativo(m_coop_t, "Cooperativas: Facturación Total (Avión vs Dron)")
                    st.plotly_chart(fig_c_t, use_container_width=True)
                else: st.info("No se registraron fincas cooperativas.")

        # DATOS PARA FINCAS ESPECIALES
        with sub_tab_indep:
            if tipo_grafico == "✈️ Tarifa Vuelo Pura" and not m_comp_v.empty:
                m_indep = m_comp_v[m_comp_v['TIPO_ENTIDAD'] == 'ESPECIALES / PILOTOS'].copy()
                if not m_indep.empty:
                    fig_i = construir_grafico_comparativo(m_indep, "Especiales/Piloto: Vuelo Puro (Avión vs Dron)")
                    st.plotly_chart(fig_i, use_container_width=True)
                else: st.info("No se registraron fincas especiales.")
                
            elif tipo_grafico == "💰 Facturación Total" and not m_comp_t.empty:
                m_indep_t = m_comp_t[m_comp_t['TIPO_ENTIDAD'] == 'ESPECIALES / PILOTOS'].copy()
                if not m_indep_t.empty:
                    fig_i_t = construir_grafico_comparativo(m_indep_t, "Especiales/Piloto: Facturación Total (Avión vs Dron)")
                    st.plotly_chart(fig_i_t, use_container_width=True)
                else: st.info("No se registraron fincas especiales.")

    # ==========================================
    # BOTÓN DE EXCEL (FUERA DE LAS PESTAÑAS)
    # ==========================================
    if not m_comp_t.empty and not m_comp_v.empty:
        st.markdown("---")
        excel_data = generar_excel_maestro(m_comp_t, m_comp_v)
        st.download_button(
            label="💾 DESCARGAR REPORTE GERENCIAL EN EXCEL", 
            data=excel_data, 
            file_name=f"Reporte_Eficiencia_Avion_vs_Dron.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            type="primary",
            use_container_width=True
        )

if __name__ == "__main__":
    pass
